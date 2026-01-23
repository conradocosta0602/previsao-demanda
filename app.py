"""
Sistema de Previsão de Demanda - Aplicação Flask
Interface web para upload de dados, processamento e geração de relatórios
"""

import os
from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
from datetime import datetime

# Importar módulos do core
from core.data_adapter import DataAdapter
from core.stockout_handler import processar_stockouts_dataframe, calcular_metricas_stockout
from core.method_selector import MethodSelector
from core.forecasting_models import get_modelo
from core.aggregator import CDAggregator, gerar_previsoes_cd
from core.reporter import ExcelReporter, gerar_nome_arquivo, preparar_resultados
from core.replenishment_calculator import processar_reabastecimento
from core.smart_alerts import generate_alerts_for_forecast
from core.outlier_detector import auto_clean_outliers
from core.validation import validate_series
from core.seasonality_detector import detect_seasonality
from core.auto_logger import get_auto_logger
from core.ruptura_sanitizer import RupturaSanitizer

# Configuração do Flask
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máximo


def processar_previsao(arquivo_excel: str,
                      meses_previsao: int = 6,
                      granularidade: str = 'semanal',
                      filiais_filtro: list = None,
                      produtos_filtro: list = None) -> dict:
    """
    Pipeline completo de processamento de previsão de demanda
    Trabalha com dados diários do banco de dados

    Args:
        arquivo_excel: Caminho para arquivo Excel com dados diários
                      Colunas esperadas: data, cod_empresa, codigo, und_venda, qtd_venda, padrao_compra
        meses_previsao: Número de períodos para prever (semanas ou meses, conforme granularidade)
        granularidade: 'semanal' (padrão) ou 'mensal'
        filiais_filtro: Lista de códigos de filiais para filtrar (None = todas)
        produtos_filtro: Lista de códigos de produtos para filtrar (None = todos)

    Returns:
        Dicionário com todos os resultados e arquivo Excel gerado
    """
    alertas = []

    # 1. CARREGAR E VALIDAR DADOS DIÁRIOS
    print("1. Carregando dados diários do banco...")
    adapter = DataAdapter(arquivo_excel)

    valido, mensagens = adapter.carregar_e_validar()
    if not valido:
        raise ValueError(f"Dados inválidos: {'; '.join(mensagens)}")

    # Converter para formato processável (agregação semanal/mensal)
    print(f"   Agregando dados: {granularidade}")
    df = adapter.converter_para_formato_legado(
        granularidade=granularidade,
        filiais=filiais_filtro,
        produtos=produtos_filtro
    )

    # Adicionar metadados
    metadados = {
        'granularidade': granularidade,
        'filiais_disponiveis': adapter.get_lista_filiais(),
        'produtos_disponiveis': adapter.get_lista_produtos(),
        'estatisticas': adapter.estatisticas_dados()
    }

    # Adicionar mensagens de validação como alertas
    for msg in mensagens:
        alertas.append({'tipo': 'info', 'mensagem': msg})

    print(f"   [OK] {len(df)} registros ({granularidade}) carregados")

    # 2. PREPARAR COLUNAS NECESSÁRIAS
    if 'Mes_Projetado' not in df.columns:
        df['Mes_Projetado'] = False
        df['Taxa_Progresso'] = 1.0
        df['Vendas_Original'] = df['Vendas']

    # 3. TRATAR STOCKOUTS
    print("3. Tratando stockouts...")
    df = processar_stockouts_dataframe(df)
    df_rupturas = calcular_metricas_stockout(df)

    # 3.5 TREINAR SELETOR ML DE MÉTODOS
    print("3.5 Treinando seletor inteligente de métodos...")
    ml_selector = None
    try:
        from core.ml_selector import MLMethodSelector
        ml_selector = MLMethodSelector()

        # Treinar com histórico
        stats_treino = ml_selector.treinar_com_historico(df[['SKU', 'Loja', 'Mes', 'Vendas_Corrigidas']].rename(columns={'Vendas_Corrigidas': 'Vendas'}))

        if stats_treino['sucesso']:
            print(f"   [OK] Modelo ML treinado com {stats_treino['series_validas']} séries")
            print(f"   [OK] Acurácia: {stats_treino['acuracia']:.1%}")
        else:
            print(f"   [AVISO] ML não treinado - usando seletor baseado em regras")
            ml_selector = None
    except Exception as e:
        print(f"   [AVISO] Erro ao treinar ML: {e}")
        ml_selector = None

    # 4. GERAR PREVISÕES POR LOJA + SKU
    print("4. Gerando previsões por loja...")
    previsoes_lojas = []
    metodos_utilizados = []
    caracteristicas_series = []
    todos_alertas = []  # Lista de todos os alertas gerados

    combinacoes = df.groupby(['Loja', 'SKU']).size().reset_index()[['Loja', 'SKU']]

    for _, row in combinacoes.iterrows():
        loja = row['Loja']
        sku = row['SKU']

        # Filtrar dados
        df_filtro = df[(df['Loja'] == loja) & (df['SKU'] == sku)].copy()
        df_filtro = df_filtro.sort_values('Mes')

        if len(df_filtro) < 3:
            alertas.append({
                'tipo': 'warning',
                'mensagem': f'Dados insuficientes para {loja}/{sku}'
            })
            continue

        vendas = df_filtro['Vendas_Corrigidas'].tolist()
        datas = df_filtro['Mes'].tolist()
        serie = pd.Series(vendas)

        # TRATAMENTO ESPECIAL PARA SÉRIES CURTAS
        from core.short_series_handler import ShortSeriesHandler
        short_handler = ShortSeriesHandler()
        classificacao_serie = short_handler.classificar_serie(serie)

        # Se série é muito curta, usar tratamento especializado
        if len(serie) < 6:  # Menos de 6 meses
            sugestao_adaptativa = short_handler.sugerir_metodo_adaptativo(serie)

            recomendacao = {
                'metodo': sugestao_adaptativa['metodo_sugerido'],
                'confianca': sugestao_adaptativa['confianca_estimada'],
                'razao': sugestao_adaptativa['razao'] + f" ({len(serie)} meses de histórico)",
                'caracteristicas': {
                    'serie_curta': True,
                    'classificacao': classificacao_serie['categoria'],
                    'tendencia': sugestao_adaptativa['analise_tendencia']['tipo']
                },
                'alternativas': []
            }
        # Selecionar método (usar ML se disponível e série suficientemente longa)
        elif ml_selector and ml_selector.is_trained and classificacao_serie['permite_ml']:
            # Usar seletor ML
            metodo_ml, confianca_ml = ml_selector.selecionar_metodo(serie)

            # Criar recomendação compatível
            recomendacao = {
                'metodo': metodo_ml,
                'confianca': confianca_ml,
                'razao': f'Selecionado por ML (confiança: {confianca_ml:.1%})',
                'caracteristicas': ml_selector.extrair_caracteristicas(serie),
                'alternativas': []
            }
        else:
            # Usar seletor baseado em regras (fallback)
            selector = MethodSelector(vendas, datas)
            recomendacao = selector.recomendar_metodo()

        # Registrar método
        metodos_utilizados.append({
            'Local': loja,
            'SKU': sku,
            'Metodo': recomendacao['metodo'],
            'Confianca': recomendacao['confianca'],
            'Razao': recomendacao['razao'],
            'Alternativas': ', '.join(recomendacao.get('alternativas', []))
        })

        # Registrar características
        caract = recomendacao['caracteristicas']
        caracteristicas_series.append({
            'Local': loja,
            'SKU': sku,
            'Intermitente': 'Sim' if caract.get('intermitente') else 'Não',
            'Tendencia': caract.get('tipo_tendencia', 'none'),
            'Sazonalidade': 'Sim' if caract.get('tem_sazonalidade') else 'Não',
            'Volatilidade': caract.get('volatilidade', 'N/A'),
            'N_Periodos': caract.get('n_periodos', len(vendas))
        })

        # Gerar previsão
        try:
            modelo = get_modelo(recomendacao['metodo'])
            modelo.fit(vendas)
            previsoes = modelo.predict(meses_previsao)
        except Exception as e:
            # Fallback
            modelo = get_modelo('Média Móvel Simples')
            modelo.fit(vendas)
            previsoes = modelo.predict(meses_previsao)

        # Calcular métricas de acurácia (WMAPE + BIAS)
        wmape = None
        bias = None
        try:
            from core.accuracy_metrics import evaluate_model_accuracy
            if len(vendas) >= 9:  # Mínimo para validação confiável
                accuracy = evaluate_model_accuracy(
                    vendas,
                    recomendacao['metodo'],
                    horizon=1
                )
                wmape = accuracy['wmape']
                bias = accuracy['bias']
        except Exception:
            pass  # Se falhar, continua sem métricas

        # Gerar datas futuras
        ultima_data = datas[-1]
        for h in range(meses_previsao):
            data_previsao = ultima_data + pd.DateOffset(months=h + 1)
            previsoes_lojas.append({
                'Loja': loja,
                'SKU': sku,
                'Mes_Previsao': data_previsao,
                'Previsao': round(previsoes[h], 1),
                'Metodo': recomendacao['metodo'],
                'Confianca': recomendacao['confianca'],
                'WMAPE': round(wmape, 1) if wmape is not None else None,
                'BIAS': round(bias, 2) if bias is not None else None
            })

        # Gerar alertas inteligentes
        try:
            # Preparar modelo_info com métricas e parâmetros
            modelo_info = {
                **modelo.params,  # Parâmetros do modelo (outliers, sazonalidade, etc)
                'wmape': wmape,
                'bias': bias
            }

            # Gerar alertas (sem estoque_atual e custo_unitario por enquanto)
            alertas_resultado = generate_alerts_for_forecast(
                sku=sku,
                loja=loja,
                historico=vendas,
                previsao=previsoes,
                modelo_params=modelo_info,
                estoque_atual=None,  # TODO: Integrar com dados de estoque
                lead_time_dias=7,  # Padrão: 7 dias
                custo_unitario=None  # TODO: Integrar com dados de custo
            )

            # Adicionar alertas à lista global
            todos_alertas.extend(alertas_resultado['alertas'])

        except Exception as e:
            # Se falhar geração de alertas, não quebra o processo
            print(f"Aviso: Não foi possível gerar alertas para {loja}/{sku}: {e}")

    df_previsoes_lojas = pd.DataFrame(previsoes_lojas)
    df_metodos = pd.DataFrame(metodos_utilizados)
    df_caracteristicas = pd.DataFrame(caracteristicas_series)

    # 4.5 APLICAR EVENTOS SAZONAIS
    print("4.5 Aplicando eventos sazonais...")
    try:
        from core.event_manager import EventManager

        manager = EventManager()
        df_previsoes_lojas = manager.aplicar_eventos_na_previsao(df_previsoes_lojas)

        # Contar quantos eventos foram aplicados
        eventos_aplicados = df_previsoes_lojas[df_previsoes_lojas['Evento_Aplicado'].notna()]
        if len(eventos_aplicados) > 0:
            print(f"   OK {len(eventos_aplicados)} previsoes ajustadas por eventos sazonais")
            eventos_unicos = eventos_aplicados['Evento_Aplicado'].unique()
            for evento in eventos_unicos:
                count = len(eventos_aplicados[eventos_aplicados['Evento_Aplicado'] == evento])
                print(f"     - {evento}: {count} ajustes")
    except Exception as e:
        print(f"Aviso: Não foi possível aplicar eventos sazonais: {e}")

    # 5. GERAR PREVISÕES DO CD
    print("5. Gerando previsões do CD...")
    skus = df['SKU'].unique().tolist()
    df_previsoes_cd = gerar_previsoes_cd(df, skus, meses_previsao)

    # Adicionar métodos do CD
    for sku in skus:
        aggregator = CDAggregator(df)
        previsao_cd = aggregator.gerar_previsao_cd(sku, meses_previsao)

        if 'erro' not in previsao_cd and previsao_cd['metodo']:
            metodos_utilizados.append({
                'Local': 'CD',
                'SKU': sku,
                'Metodo': previsao_cd['metodo'],
                'Confianca': previsao_cd['confianca'],
                'Razao': previsao_cd.get('razao', ''),
                'Alternativas': ''
            })

            caract = previsao_cd.get('caracteristicas', {})
            caracteristicas_series.append({
                'Local': 'CD',
                'SKU': sku,
                'Intermitente': 'Sim' if caract.get('intermitente') else 'Não',
                'Tendencia': caract.get('tipo_tendencia', 'none'),
                'Sazonalidade': 'Sim' if caract.get('tem_sazonalidade') else 'Não',
                'Volatilidade': caract.get('volatilidade', 'N/A'),
                'N_Periodos': caract.get('n_periodos', 0)
            })

    df_metodos = pd.DataFrame(metodos_utilizados)
    df_caracteristicas = pd.DataFrame(caracteristicas_series)

    # 6. COMPILAR RESUMO
    print("6. Compilando resumo...")

    # Contagem de métodos
    contagem_metodos = df_metodos['Metodo'].value_counts().to_dict()

    # Período histórico
    data_min = df['Mes'].min().strftime('%Y-%m')
    data_max = df['Mes'].max().strftime('%Y-%m')

    resumo = {
        'total_skus': len(skus),
        'total_lojas': len(df['Loja'].unique()),
        'total_combinacoes': len(combinacoes),
        'periodo_historico': f"{data_min} a {data_max}",
        'meses_previsao': meses_previsao,
        'meses_com_ruptura': int(df_rupturas['meses_com_ruptura'].sum()) if 'meses_com_ruptura' in df_rupturas.columns else 0,
        'taxa_ruptura_media': float(df_rupturas['taxa_ruptura'].mean()) if 'taxa_ruptura' in df_rupturas.columns else 0,
        'vendas_perdidas': float(df_rupturas['vendas_perdidas'].sum()) if 'vendas_perdidas' in df_rupturas.columns else 0,
        'contagem_metodos': contagem_metodos
    }

    # 7. GERAR EXCEL
    print("7. Gerando relatório Excel...")

    nome_arquivo = gerar_nome_arquivo()
    caminho_saida = os.path.join(app.config['OUTPUT_FOLDER'], nome_arquivo)

    resultados = preparar_resultados(
        df_previsoes_lojas,
        df_previsoes_cd,
        df_metodos,
        df_rupturas,
        df_caracteristicas,
        resumo,
        alertas
    )

    reporter = ExcelReporter()
    reporter.gerar_relatorio(resultados, caminho_saida)

    # 8. PREPARAR DADOS PARA GRÁFICO
    print("8. Preparando dados para gráfico...")

    # Dados históricos por loja
    dados_historicos_lojas = df.groupby(['Mes', 'Loja', 'SKU'])['Vendas_Corrigidas'].sum().reset_index()
    dados_historicos_lojas['Mes'] = dados_historicos_lojas['Mes'].dt.strftime('%Y-%m')

    # Dados históricos agregados (total)
    dados_historicos_total = df.groupby('Mes')['Vendas_Corrigidas'].sum().reset_index()
    dados_historicos_total['Mes'] = dados_historicos_total['Mes'].dt.strftime('%Y-%m')

    # Dados históricos por SKU (agregado todas lojas)
    dados_historicos_sku = df.groupby(['Mes', 'SKU'])['Vendas_Corrigidas'].sum().reset_index()
    dados_historicos_sku['Mes'] = dados_historicos_sku['Mes'].dt.strftime('%Y-%m')

    # Converter para formato JSON
    grafico_data = {
        'historico_lojas': dados_historicos_lojas.to_dict('records'),
        'historico_total': dados_historicos_total.to_dict('records'),
        'historico_sku': dados_historicos_sku.to_dict('records'),
        'previsoes_lojas': df_previsoes_lojas.to_dict('records'),
        'lojas': sorted(df['Loja'].unique().tolist()),
        'skus': sorted(skus)
    }

    # Converter datas de previsão para string
    for item in grafico_data['previsoes_lojas']:
        if 'Mes_Previsao' in item and hasattr(item['Mes_Previsao'], 'strftime'):
            item['Mes_Previsao'] = item['Mes_Previsao'].strftime('%Y-%m')

    # 9. PREPARAR DADOS PARA COMPARAÇÃO YoY (Ano contra Ano)
    print("9. Preparando dados para comparação YoY...")

    # Agrupar por mês (MM) ignorando ano para comparação
    df_copy = df.copy()
    df_copy['Mes_Numero'] = df_copy['Mes'].dt.month
    df_copy['Ano'] = df_copy['Mes'].dt.year

    # Dados históricos agregados por mês do ano
    historico_por_mes = df_copy.groupby(['Ano', 'Mes_Numero'])['Vendas_Corrigidas'].sum().reset_index()

    # Preparar dados de comparação YoY
    comparacao_yoy = []

    # Obter meses únicos de previsão
    meses_previsao_unicos = sorted(list(set([p['Mes_Previsao'] for p in grafico_data['previsoes_lojas']])))

    # Para cada mês de previsão, buscar o mesmo mês do ano anterior
    for mes_prev in meses_previsao_unicos:
        try:
            mes_prev_date = pd.to_datetime(mes_prev)
            mes_num = mes_prev_date.month
            ano_anterior = mes_prev_date.year - 1

            # Buscar demanda do mesmo mês no ano anterior
            demanda_ano_anterior = historico_por_mes[
                (historico_por_mes['Ano'] == ano_anterior) &
                (historico_por_mes['Mes_Numero'] == mes_num)
            ]

            if not demanda_ano_anterior.empty:
                valor_ano_anterior = float(demanda_ano_anterior['Vendas_Corrigidas'].iloc[0])
            else:
                valor_ano_anterior = None

            # Buscar previsão agregada para este mês
            previsao_agregada = sum([
                p['Previsao'] for p in grafico_data['previsoes_lojas']
                if p['Mes_Previsao'] == mes_prev
            ])

            # Calcular variação percentual
            if valor_ano_anterior and valor_ano_anterior > 0:
                variacao_percentual = ((previsao_agregada - valor_ano_anterior) / valor_ano_anterior) * 100
            else:
                variacao_percentual = None

            comparacao_yoy.append({
                'mes': mes_prev,
                'mes_nome': mes_prev_date.strftime('%b/%y'),
                'demanda_ano_anterior': valor_ano_anterior,
                'previsao_atual': previsao_agregada,
                'variacao_percentual': round(variacao_percentual, 1) if variacao_percentual is not None else None
            })
        except:
            pass

    grafico_data['comparacao_yoy'] = comparacao_yoy

    # 9.1. PREPARAR DADOS POR FORNECEDOR/ITEM
    print("9.1. Preparando dados por fornecedor/item...")

    # Função para padronizar nomenclatura dos métodos conforme documentação
    def padronizar_metodo(metodo):
        """
        Converte nomenclatura interna para nomenclatura da documentação.
        Agora simplificado pois method_selector já retorna os nomes corretos.
        """
        if not metodo:
            return 'N/A'

        # Mapeamento apenas para casos legados ou compatibilidade retroativa
        # Os 7 métodos oficiais: SMA, WMA, EMA, Regressão com Tendência, Decomposição Sazonal, TSB, AUTO
        mapeamento = {
            # Casos legados (caso algum código antigo ainda use)
            'média móvel simples': 'SMA',
            'média móvel ponderada': 'WMA',
            'suavização exponencial simples': 'EMA',
            'regressão linear': 'Regressão com Tendência',
            'média móvel sazonal': 'Decomposição Sazonal',
            'holt-winters': 'Decomposição Sazonal',
            'holt': 'Regressão com Tendência',
            'croston': 'TSB',  # Croston foi SUBSTITUÍDO por TSB
            'sba': 'TSB',
        }

        # Tentar match case-insensitive
        metodo_lower = metodo.lower()

        # Se já está no formato correto, retornar direto
        if metodo in ['SMA', 'WMA', 'EMA', 'Regressão com Tendência', 'Decomposição Sazonal', 'TSB', 'AUTO']:
            return metodo

        # Caso contrário, tentar conversão
        return mapeamento.get(metodo_lower, metodo)

    # Mapeamento SKU → Fornecedor (baseado na documentação INTEGRACAO_ARQUIVOS.md)
    def get_fornecedor_by_sku(sku):
        try:
            # Extrair número do SKU (ex: PROD_001 → 1)
            sku_num = int(sku.split('_')[1])
            if 1 <= sku_num <= 34:
                return 'FORNECEDOR_A'
            elif 35 <= sku_num <= 67:
                return 'FORNECEDOR_B'
            elif 68 <= sku_num <= 100:
                return 'FORNECEDOR_C'
            else:
                return 'OUTROS'
        except:
            return 'OUTROS'

    # Agregar previsões por SKU (soma de todas as lojas)
    previsoes_por_sku = {}
    for prev in grafico_data['previsoes_lojas']:
        sku = prev['SKU']
        if sku not in previsoes_por_sku:
            previsoes_por_sku[sku] = {
                'previsao_total': 0,
                'metodo': prev['Metodo']
            }
        previsoes_por_sku[sku]['previsao_total'] += prev['Previsao']

    # Calcular demanda do MESMO PERÍODO DO ANO ANTERIOR por SKU
    # Igual à lógica da comparação YoY mensal (linhas 398-434)
    print(f"9.1. Calculando demanda do mesmo período do ano anterior por SKU...")

    # Preparar histórico agregado por SKU, Ano e Mês
    df_copy['Ano'] = df_copy['Mes'].dt.year
    df_copy['Mes_Numero'] = df_copy['Mes'].dt.month

    # Para cada mês de previsão, buscar o mesmo mês do ano anterior
    meses_previsao_unicos = sorted(list(set([p['Mes_Previsao'] for p in grafico_data['previsoes_lojas']])))

    # Calcular demanda ano anterior por SKU (somando os mesmos meses do ano anterior)
    demanda_ano_anterior_sku = {}

    for sku in skus:
        total_ano_anterior = 0
        for mes_prev in meses_previsao_unicos:
            try:
                mes_prev_date = pd.to_datetime(mes_prev)
                mes_num = mes_prev_date.month
                ano_anterior = mes_prev_date.year - 1

                # Buscar demanda do mesmo mês no ano anterior para este SKU
                demanda_mes = df_copy[
                    (df_copy['SKU'] == sku) &
                    (df_copy['Ano'] == ano_anterior) &
                    (df_copy['Mes_Numero'] == mes_num)
                ]['Vendas_Corrigidas'].sum()

                total_ano_anterior += demanda_mes
            except:
                pass

        demanda_ano_anterior_sku[sku] = total_ano_anterior

    print(f"   Comparando previsões com mesmos meses do ano anterior")
    print(f"   Exemplo: Se prevendo Jul-Dez/2024, compara com Jul-Dez/2023")

    # Montar dados por fornecedor/item
    dados_fornecedor_item = []
    for sku in sorted(skus):
        fornecedor = get_fornecedor_by_sku(sku)
        previsao_dados = previsoes_por_sku.get(sku, {'previsao_total': 0, 'metodo': 'N/A'})
        previsao_total = previsao_dados['previsao_total']
        metodo = previsao_dados['metodo']

        # Demanda do mesmo período do ano anterior
        demanda_anterior = demanda_ano_anterior_sku.get(sku, 0)

        # Calcular variação percentual YoY (igual à comparação mensal)
        if demanda_anterior > 0:
            variacao_pct = ((previsao_total - demanda_anterior) / demanda_anterior) * 100
        else:
            variacao_pct = None

        dados_fornecedor_item.append({
            'Fornecedor': fornecedor,
            'SKU': sku,
            'Demanda_Prevista': round(previsao_total, 1),
            'Demanda_Ano_Anterior': round(demanda_anterior, 1),
            'Variacao_YoY_Percentual': round(variacao_pct, 1) if variacao_pct is not None else None,
            'Metodo_Estatistico': padronizar_metodo(metodo)
        })

    grafico_data['fornecedor_item'] = dados_fornecedor_item

    # 10. CALCULAR MÉTRICAS ADICIONAIS PARA O RESUMO
    print("10. Calculando métricas adicionais...")

    # Calcular taxa YoY usando SOMA DAS PREVISÕES INDIVIDUAIS
    # Isso é o mesmo cálculo que aparece na tabela YoY da interface
    print("10.1. Calculando YoY das previsões individuais...")

    # Somar todas as previsões e todas as demandas do ano anterior da tabela YoY
    total_previsao_agregada = sum([item['previsao_atual'] for item in comparacao_yoy if item['previsao_atual']])
    total_ano_anterior_agregado = sum([item['demanda_ano_anterior'] for item in comparacao_yoy if item['demanda_ano_anterior']])

    # Calcular variação percentual SIMPLES (igual à coluna Total da tabela)
    if total_ano_anterior_agregado > 0:
        taxa_variabilidade_media_yoy = round(((total_previsao_agregada - total_ano_anterior_agregado) / total_ano_anterior_agregado) * 100, 1)
        print(f"   [OK] YoY Total (soma das previsões individuais): {taxa_variabilidade_media_yoy:+.1f}%")
        print(f"       Total previsto (soma individual): {total_previsao_agregada:,.0f}")
        print(f"       Total ano anterior: {total_ano_anterior_agregado:,.0f}")
    else:
        taxa_variabilidade_media_yoy = 0
        print(f"   [AVISO] Total período anterior é zero")

    # Extrair período da comparação YoY
    if comparacao_yoy and len(comparacao_yoy) > 0:
        primeiro_mes_previsao = pd.to_datetime(comparacao_yoy[0]['mes'])
        ultimo_mes_previsao = pd.to_datetime(comparacao_yoy[-1]['mes'])

        # Mesmo período no ano anterior (subtrair 12 meses de cada data)
        primeiro_mes_anterior = primeiro_mes_previsao - pd.DateOffset(months=12)
        ultimo_mes_anterior = ultimo_mes_previsao - pd.DateOffset(months=12)

        # Formato: Mês/Ano
        mes_inicio = primeiro_mes_previsao.strftime('%b/%y')
        mes_fim = ultimo_mes_previsao.strftime('%b/%y')
        mes_inicio_anterior = primeiro_mes_anterior.strftime('%b/%y')
        mes_fim_anterior = ultimo_mes_anterior.strftime('%b/%y')

        # Se for apenas um mês
        if len(comparacao_yoy) == 1:
            anos_yoy = f"{mes_inicio_anterior} vs {mes_inicio}"
        else:
            anos_yoy = f"{mes_inicio_anterior}-{mes_fim_anterior} vs {mes_inicio}-{mes_fim}"
    else:
        anos_yoy = "N/A"

    # Calcular taxa de vendas perdidas (%)
    vendas_perdidas = resumo['vendas_perdidas']
    vendas_totais = df['Vendas_Corrigidas'].sum()
    vendas_potenciais = vendas_totais + vendas_perdidas
    taxa_vendas_perdidas = round((vendas_perdidas / vendas_potenciais * 100), 1) if vendas_potenciais > 0 else 0

    # Extrair período histórico (apenas ano inicial e final)
    ano_min = df['Mes'].min().year
    ano_max = df['Mes'].max().year
    periodo_vendas_perdidas = f"{ano_min}-{ano_max}" if ano_min != ano_max else str(ano_min)

    # Atualizar resumo com novas métricas
    resumo['taxa_variabilidade_media_yoy'] = taxa_variabilidade_media_yoy
    resumo['taxa_vendas_perdidas'] = taxa_vendas_perdidas
    resumo['anos_yoy'] = anos_yoy
    resumo['periodo_vendas_perdidas'] = periodo_vendas_perdidas

    print("Processamento concluído!")

    # Calcular resumo de alertas inteligentes
    from core.smart_alerts import SmartAlertGenerator
    generator = SmartAlertGenerator()
    generator.alertas = todos_alertas
    resumo_alertas = generator.get_summary()

    # Retornar resultado com metadados
    return {
        'success': True,
        'arquivo_saida': nome_arquivo,
        'resumo': resumo,
        'alertas': alertas,
        'grafico_data': grafico_data,
        'smart_alerts': todos_alertas,
        'smart_alerts_summary': resumo_alertas,
        'metadados': metadados  # Informações sobre filiais, produtos, estatísticas
    }


@app.route('/menu')
def menu():
    """Menu principal - Capa do sistema"""
    return render_template('menu.html')


@app.route('/')
def index():
    """Página de previsão de demanda"""
    return render_template('index.html')


# Endpoint /upload removido - agora apenas consulta ao banco de dados via /api/previsao/banco


@app.route('/download/<filename>')
def download_file(filename):
    """Download do arquivo Excel gerado ou exemplo"""
    # Primeiro tenta na pasta de outputs
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)

    # Se não encontrar, tenta na raiz (arquivos de exemplo)
    if not os.path.exists(filepath):
        filepath = filename

    if not os.path.exists(filepath):
        return jsonify({'erro': 'Arquivo não encontrado'}), 404

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ===== SIMULADOR DE CENÁRIOS =====

# Variável global para armazenar dados da última previsão (sessão simples)
ultima_previsao_data = None

@app.route('/simulador')
def simulador():
    """Página do simulador de cenários"""
    return render_template('simulador.html')


@app.route('/simulador/dados', methods=['GET'])
def get_dados_simulador():
    """
    Retorna dados da última previsão para o simulador
    """
    global ultima_previsao_data

    if ultima_previsao_data is None:
        return jsonify({
            'success': False,
            'erro': 'Nenhuma previsão disponível. Processe uma previsão primeiro na página principal.'
        }), 404

    # Transformar dados para formato esperado pelo simulador
    # JavaScript espera: { previsoes: [...], lojas: [...], skus: [...] }
    dados_formatados = {
        'previsoes': ultima_previsao_data.get('previsoes_lojas', []),
        'lojas': ultima_previsao_data.get('lojas', []),
        'skus': ultima_previsao_data.get('skus', [])
    }

    # Normalizar nomes de campos para minúsculas (JavaScript espera lowercase)
    for prev in dados_formatados['previsoes']:
        # Criar campos com nomes lowercase mantendo originais para compatibilidade
        if 'Loja' in prev:
            prev['loja'] = prev['Loja']
        if 'SKU' in prev:
            prev['sku'] = prev['SKU']
        if 'Mes_Previsao' in prev:
            prev['mes'] = prev['Mes_Previsao']
        if 'Previsao' in prev:
            prev['previsao'] = prev['Previsao']

    return jsonify({
        'success': True,
        'dados': dados_formatados
    })


@app.route('/simulador/aplicar', methods=['POST'])
def aplicar_cenario():
    """
    Aplica um cenário de simulação
    """
    global ultima_previsao_data

    if ultima_previsao_data is None:
        return jsonify({
            'success': False,
            'erro': 'Nenhuma previsão disponível'
        }), 404

    try:
        from core.scenario_simulator import criar_simulador_de_dados

        # Criar simulador
        simulador = criar_simulador_de_dados(ultima_previsao_data['previsoes_lojas'])

        # Obter tipo de ajuste
        dados_request = request.get_json()
        tipo_ajuste = dados_request.get('tipo')

        # Aplicar cenário
        if tipo_ajuste == 'predefinido':
            cenario = dados_request.get('cenario')
            simulador.aplicar_cenario_predefinido(cenario)

        elif tipo_ajuste == 'global':
            percentual = float(dados_request.get('percentual', 0))
            simulador.aplicar_ajuste_global(percentual)

        elif tipo_ajuste == 'por_loja':
            loja = dados_request.get('loja')
            percentual = float(dados_request.get('percentual', 0))
            simulador.aplicar_ajuste_por_loja(loja, percentual)

        elif tipo_ajuste == 'por_sku':
            sku = dados_request.get('sku')
            percentual = float(dados_request.get('percentual', 0))
            simulador.aplicar_ajuste_por_sku(sku, percentual)

        elif tipo_ajuste == 'por_mes':
            mes = int(dados_request.get('mes'))
            percentual = float(dados_request.get('percentual', 0))
            simulador.aplicar_ajuste_por_mes(mes, percentual)

        else:
            return jsonify({'success': False, 'erro': 'Tipo de ajuste inválido'}), 400

        # Calcular impacto
        impacto = simulador.calcular_impacto()

        # Obter comparação
        df_comp = simulador.get_comparacao()

        # Converter para lista de dicts
        previsoes_comparacao = df_comp.to_dict('records')

        # Serializar datas
        for prev in previsoes_comparacao:
            if 'Mes_Previsao' in prev and hasattr(prev['Mes_Previsao'], 'strftime'):
                prev['Mes_Previsao'] = prev['Mes_Previsao'].strftime('%Y-%m-%d')

        return jsonify({
            'success': True,
            'cenario_nome': simulador.cenario_nome,
            'previsoes_comparacao': previsoes_comparacao,
            'impacto': impacto,
            'ajustes_aplicados': simulador.ajustes_aplicados
        })

    except Exception as e:
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/simulador/exportar', methods=['POST'])
def exportar_cenario():
    """
    Exporta cenário simulado para Excel
    """
    global ultima_previsao_data

    if ultima_previsao_data is None:
        return jsonify({'success': False, 'erro': 'Nenhuma previsão disponível'}), 404

    try:
        from core.scenario_simulator import ScenarioSimulator
        import pandas as pd
        from io import BytesIO

        # Receber previsões simuladas do frontend
        dados_request = request.get_json()
        previsoes_simuladas = dados_request.get('previsoes_simuladas', [])

        # Converter para DataFrame
        df_base = pd.DataFrame(ultima_previsao_data['previsoes_lojas'])
        df_simulado = pd.DataFrame(previsoes_simuladas)

        # Renomear colunas para maiúsculas (padrão esperado)
        df_simulado.columns = [col.capitalize() if col.lower() in ['loja', 'sku', 'mes', 'previsao'] else col
                               for col in df_simulado.columns]

        # Criar simulador e substituir dados simulados
        simulador = ScenarioSimulator(df_base)
        simulador.previsoes_simuladas = df_simulado

        # Exportar para Excel em memória
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f"Cenario_Simulado_{timestamp}.xlsx"
        filepath = os.path.join(app.config['OUTPUT_FOLDER'], nome_arquivo)

        simulador.exportar_para_excel(filepath)

        # Enviar arquivo para download
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        print(f"Erro ao exportar cenário: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/eventos')
def eventos():
    """Página de gerenciamento de eventos"""
    return render_template('eventos.html')


@app.route('/eventos_simples')
def eventos_simples():
    """Página simplificada de gerenciamento de eventos"""
    return render_template('eventos_simples.html')


@app.route('/eventos/cadastrar', methods=['POST'])
def cadastrar_evento():
    """Cadastra um novo evento"""
    try:
        from core.event_manager import EventManager

        dados = request.get_json()

        # Validação de dados
        if not dados:
            return jsonify({'success': False, 'erro': 'Nenhum dado recebido'}), 400

        if 'tipo' not in dados or 'data_evento' not in dados:
            return jsonify({'success': False, 'erro': 'Campos obrigatórios faltando'}), 400

        # Log de debug
        print(f"[EVENTOS] Cadastrando evento:")
        print(f"   Tipo: {dados.get('tipo')}")
        print(f"   Data: {dados.get('data_evento')}")
        print(f"   Multiplicador: {dados.get('multiplicador_manual')}")
        print(f"   Duracao: {dados.get('duracao_dias')}")

        manager = EventManager()
        evento_id = manager.cadastrar_evento(
            tipo=dados['tipo'],
            data_evento=dados['data_evento'],
            multiplicador_manual=dados.get('multiplicador_manual'),
            nome_custom=dados.get('nome_custom'),
            duracao_dias=dados.get('duracao_dias'),
            observacoes=dados.get('observacoes')
        )

        print(f"   [OK] Evento cadastrado com ID: {evento_id}")

        return jsonify({
            'success': True,
            'evento_id': evento_id,
            'mensagem': 'Evento cadastrado com sucesso'
        })

    except Exception as e:
        print(f"[ERRO] Erro ao cadastrar evento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/eventos/listar', methods=['GET'])
def listar_eventos():
    """Lista todos os eventos cadastrados"""
    try:
        from core.event_manager import EventManager

        manager = EventManager()
        eventos = manager.listar_eventos(apenas_ativos=False, apenas_futuros=False)

        # Converter para JSON manualmente para garantir compatibilidade
        import json
        eventos_serializaveis = []

        for evento in eventos:
            try:
                # Criar dict limpo apenas com campos necessários (SEM timestamps)
                # Função auxiliar para converter número de forma segura
                def safe_float(valor):
                    if valor is None or valor == '':
                        return None
                    try:
                        num = float(valor)
                        # Verificar se é NaN ou infinito
                        import math
                        if math.isnan(num) or math.isinf(num):
                            return None
                        return num
                    except (ValueError, TypeError):
                        return None

                evento_limpo = {
                    'id': int(evento.get('id', 0)),
                    'tipo': str(evento.get('tipo', '')),
                    'nome': str(evento.get('nome', '')),
                    'data_evento': str(evento.get('data_evento', '')),
                    'duracao_dias': int(evento.get('duracao_dias', 1)),
                    'multiplicador_manual': safe_float(evento.get('multiplicador_manual')),
                    'multiplicador_calculado': safe_float(evento.get('multiplicador_calculado')),
                    'usar_multiplicador_manual': bool(int(evento.get('usar_multiplicador_manual', 0))),
                    'ativo': bool(int(evento.get('ativo', 1))),
                    'observacoes': str(evento['observacoes']) if evento.get('observacoes') else None
                }
                eventos_serializaveis.append(evento_limpo)
            except Exception as e:
                print(f"[AVISO] Erro ao processar evento {evento.get('id')}: {e}")
                continue

        print(f"[EVENTOS] Listando {len(eventos_serializaveis)} eventos")

        # Testar serialização
        try:
            json.dumps(eventos_serializaveis)
        except Exception as e:
            print(f"[ERRO] Falha ao serializar: {e}")
            return jsonify({'success': False, 'erro': 'Erro ao serializar eventos'}), 500

        return jsonify({
            'success': True,
            'eventos': eventos_serializaveis
        })

    except Exception as e:
        print(f"[ERRO] Erro ao listar eventos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/eventos/teste', methods=['GET'])
def teste_eventos():
    """Endpoint de teste simples"""
    return jsonify({
        'success': True,
        'mensagem': 'Teste OK',
        'eventos': [
            {'id': 1, 'nome': 'Teste', 'tipo': 'BLACK_FRIDAY', 'data_evento': '2026-11-01'}
        ]
    })


@app.route('/eventos/atualizar', methods=['POST'])
def atualizar_evento():
    """Atualiza multiplicador de um evento"""
    try:
        from core.event_manager import EventManager

        dados = request.get_json()
        evento_id = dados['evento_id']
        multiplicador = dados['multiplicador']

        manager = EventManager()
        manager.atualizar_multiplicador(evento_id, multiplicador)

        return jsonify({
            'success': True,
            'mensagem': 'Evento atualizado com sucesso'
        })

    except Exception as e:
        print(f"Erro ao atualizar evento: {e}")
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/eventos/desativar', methods=['POST'])
def desativar_evento():
    """Desativa um evento"""
    try:
        from core.event_manager import EventManager

        dados = request.get_json()
        evento_id = dados['evento_id']

        manager = EventManager()
        manager.desativar_evento(evento_id)

        return jsonify({
            'success': True,
            'mensagem': 'Evento desativado'
        })

    except Exception as e:
        print(f"Erro ao desativar evento: {e}")
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/eventos/reativar', methods=['POST'])
def reativar_evento():
    """Reativa um evento"""
    try:
        from core.event_manager import EventManager
        import sqlite3

        dados = request.get_json()
        evento_id = dados['evento_id']

        manager = EventManager()

        # Reativar manualmente via SQL
        conn = sqlite3.connect(manager.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE eventos
            SET ativo = 1, atualizado_em = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (evento_id,))
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'mensagem': 'Evento reativado'
        })

    except Exception as e:
        print(f"Erro ao reativar evento: {e}")
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/eventos/excluir', methods=['POST'])
def excluir_evento():
    """Exclui permanentemente um evento"""
    try:
        from core.event_manager import EventManager

        dados = request.get_json()
        evento_id = dados['evento_id']

        manager = EventManager()
        manager.excluir_evento(evento_id)

        return jsonify({
            'success': True,
            'mensagem': 'Evento excluído permanentemente'
        })

    except Exception as e:
        print(f"[ERRO] ao excluir evento: {e}")
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/pedido_fornecedor_integrado')
def pedido_fornecedor_integrado_page():
    """Página de pedidos ao fornecedor integrado com previsão V2"""
    return render_template('pedido_fornecedor_integrado.html')


@app.route('/api/pedido_fornecedor_integrado', methods=['POST'])
def api_pedido_fornecedor_integrado():
    """
    API para gerar pedidos ao fornecedor integrado com previsão V2.

    Características:
    - Usa previsão V2 (Bottom-Up) em tempo real
    - Cobertura baseada em ABC: Lead Time + Ciclo + Segurança_ABC
    - Curva ABC do cadastro de produtos para nível de serviço
    - Arredondamento para múltiplo de caixa
    - Destino configurável (Lojas ou CD)

    Request JSON:
    {
        "fornecedor": "TODOS" ou código específico,
        "categoria": "TODAS" ou categoria específica,
        "destino_tipo": "LOJA" ou "CD",
        "cod_empresa": código da loja/CD de destino (ou "TODAS"),
        "cobertura_dias": null (automático) ou número específico
    }

    Returns:
        JSON com pedidos calculados e agregação por fornecedor
    """
    try:
        from core.pedido_fornecedor_integrado import (
            PedidoFornecedorIntegrado,
            agregar_por_fornecedor,
            calcular_cobertura_abc
        )

        dados = request.get_json()

        # Parâmetros de filtro
        fornecedor_filtro = dados.get('fornecedor', 'TODOS')
        linha1_filtro = dados.get('linha1', 'TODAS')  # Linha 1 (categoria)
        linha3_filtro = dados.get('linha3', 'TODAS')  # Linha 3 (sublinha/codigo_linha)
        # Compatibilidade com versao anterior (categoria -> linha1)
        if linha1_filtro == 'TODAS' and 'categoria' in dados:
            linha1_filtro = dados.get('categoria', 'TODAS')
        destino_tipo = dados.get('destino_tipo', 'LOJA')  # 'LOJA' ou 'CD'
        cod_empresa = dados.get('cod_empresa', 'TODAS')
        cobertura_dias = dados.get('cobertura_dias')  # None = automático

        print(f"\n{'='*60}")
        print("PEDIDO FORNECEDOR INTEGRADO")
        print(f"{'='*60}")
        print(f"  Fornecedor: {fornecedor_filtro} (tipo: {type(fornecedor_filtro).__name__})")
        print(f"  Linha 1: {linha1_filtro} (tipo: {type(linha1_filtro).__name__})")
        print(f"  Linha 3: {linha3_filtro} (tipo: {type(linha3_filtro).__name__})")
        print(f"  Destino: {destino_tipo}")
        print(f"  Empresa: {cod_empresa}")
        print(f"  Cobertura: {'Automática (ABC)' if cobertura_dias is None else f'{cobertura_dias} dias'}")

        # Conectar ao banco usando configuração centralizada
        conn = get_db_connection()

        # Normalizar cod_empresa: pode ser string, lista ou 'TODAS'
        lojas_selecionadas = []
        if isinstance(cod_empresa, list):
            lojas_selecionadas = [int(x) for x in cod_empresa]
        elif cod_empresa != 'TODAS':
            lojas_selecionadas = [int(cod_empresa)]

        # Identificar se o destino é CD (centralizado)
        # CDs são lojas com cod_empresa >= 80
        if lojas_selecionadas:
            is_destino_cd = destino_tipo == 'CD' or any(loja >= 80 for loja in lojas_selecionadas)
            cod_destino = lojas_selecionadas[0]  # Usar a primeira loja para parâmetros
        else:
            is_destino_cd = destino_tipo == 'CD'
            cod_destino = 80  # Default para CD

        # =====================================================================
        # 1. DETERMINAR LISTA DE FORNECEDORES A PROCESSAR
        # =====================================================================
        # Suporta: string 'TODOS', string com nome, ou array de nomes
        fornecedores_a_processar = []

        # Normalizar: se for lista com 'TODOS', tratar como TODOS
        if isinstance(fornecedor_filtro, list):
            # Verificar se 'TODOS' está na lista
            if 'TODOS' in fornecedor_filtro:
                fornecedor_filtro = 'TODOS'  # Converter para string para cair no bloco seguinte
            else:
                # Multi-select: lista de fornecedores específicos
                fornecedores_a_processar = fornecedor_filtro
                print(f"  Modo: Multi-select ({len(fornecedores_a_processar)} fornecedores)")

        # Se ainda não preencheu (não era multi-select)
        if not fornecedores_a_processar:
            if fornecedor_filtro == 'TODOS':
                # Buscar todos os fornecedores que tem produtos ativos
                query_fornecedores = """
                    SELECT DISTINCT nome_fornecedor
                    FROM cadastro_produtos_completo
                    WHERE ativo = TRUE AND nome_fornecedor IS NOT NULL
                """
                params_forn = []

                # Aplicar filtro de Linha 1 (categoria) se especificado
                if linha1_filtro != 'TODAS':
                    if isinstance(linha1_filtro, list):
                        placeholders = ','.join(['%s'] * len(linha1_filtro))
                        query_fornecedores += f" AND categoria IN ({placeholders})"
                        params_forn.extend(linha1_filtro)
                    else:
                        query_fornecedores += " AND categoria = %s"
                        params_forn.append(linha1_filtro)

                # Aplicar filtro de Linha 3 (codigo_linha) se especificado
                if linha3_filtro != 'TODAS':
                    if isinstance(linha3_filtro, list):
                        placeholders = ','.join(['%s'] * len(linha3_filtro))
                        query_fornecedores += f" AND codigo_linha IN ({placeholders})"
                        params_forn.extend(linha3_filtro)
                    else:
                        query_fornecedores += " AND codigo_linha = %s"
                        params_forn.append(linha3_filtro)

                query_fornecedores += " ORDER BY nome_fornecedor"

                df_fornecedores = pd.read_sql(query_fornecedores, conn, params=params_forn if params_forn else None)
                fornecedores_a_processar = df_fornecedores['nome_fornecedor'].tolist()
                print(f"  Modo: TODOS ({len(fornecedores_a_processar)} fornecedores encontrados)")
            else:
                # Fornecedor unico (string)
                fornecedores_a_processar = [fornecedor_filtro]
                print(f"  Modo: Fornecedor unico ({fornecedor_filtro})")

        if not fornecedores_a_processar:
            conn.close()
            return jsonify({
                'success': False,
                'erro': 'Nenhum fornecedor encontrado com os filtros selecionados.'
            }), 400

        # =====================================================================
        # 2. PROCESSAR PRODUTOS POR FORNECEDOR (sem LIMIT global)
        # =====================================================================
        print(f"  Destino CD: {is_destino_cd}")

        # Verificar se tabela parametros_fornecedor existe (graceful degradation)
        cursor_check = conn.cursor()
        cursor_check.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'parametros_fornecedor'
            )
        """)
        tabela_params_existe = cursor_check.fetchone()[0]
        cursor_check.close()

        df_produtos_total = pd.DataFrame()
        total_produtos_por_fornecedor = {}

        for idx_forn, nome_fornecedor in enumerate(fornecedores_a_processar):
            # DEBUG: Log do nome do fornecedor sendo processado
            if 'FAME' in str(nome_fornecedor).upper():
                print(f"    [DEBUG-FAME] Processando fornecedor: '{nome_fornecedor}' (repr: {repr(nome_fornecedor)})")

            # Query para buscar produtos deste fornecedor
            # Se tabela parametros_fornecedor existe, faz JOIN para buscar parametros
            # Senao, usa valores padrao
            if tabela_params_existe:
                query_produtos = """
                    SELECT DISTINCT
                        p.cod_produto as codigo,
                        p.descricao,
                        p.categoria as linha1,
                        p.codigo_linha as linha3,
                        p.descricao_linha as linha3_descricao,
                        'B' as curva_abc,
                        p.cnpj_fornecedor as codigo_fornecedor,
                        p.nome_fornecedor,
                        COALESCE(pf.lead_time_dias, 15) as lead_time_dias,
                        COALESCE(pf.ciclo_pedido_dias, 7) as ciclo_pedido_dias,
                        COALESCE(pf.pedido_minimo_valor, 0) as pedido_minimo_valor,
                        CASE WHEN pf.id IS NOT NULL THEN TRUE ELSE FALSE END as fornecedor_cadastrado
                    FROM cadastro_produtos_completo p
                    LEFT JOIN parametros_fornecedor pf ON p.cnpj_fornecedor = pf.cnpj_fornecedor
                        AND pf.cod_empresa = %s AND pf.ativo = TRUE
                    WHERE p.ativo = TRUE
                        AND p.nome_fornecedor = %s
                """
                params = [cod_destino, nome_fornecedor]
            else:
                # Fallback: sem tabela de parametros, usa valores padrao
                query_produtos = """
                    SELECT DISTINCT
                        p.cod_produto as codigo,
                        p.descricao,
                        p.categoria as linha1,
                        p.codigo_linha as linha3,
                        p.descricao_linha as linha3_descricao,
                        'B' as curva_abc,
                        p.cnpj_fornecedor as codigo_fornecedor,
                        p.nome_fornecedor,
                        15 as lead_time_dias,
                        7 as ciclo_pedido_dias,
                        0 as pedido_minimo_valor,
                        FALSE as fornecedor_cadastrado
                    FROM cadastro_produtos_completo p
                    WHERE p.ativo = TRUE
                        AND p.nome_fornecedor = %s
                """
                params = [nome_fornecedor]

            # Filtro de Linha 1 (categoria) se especificado
            if linha1_filtro != 'TODAS':
                if isinstance(linha1_filtro, list):
                    placeholders = ','.join(['%s'] * len(linha1_filtro))
                    query_produtos += f" AND p.categoria IN ({placeholders})"
                    params.extend(linha1_filtro)
                else:
                    query_produtos += " AND p.categoria = %s"
                    params.append(linha1_filtro)

            # Filtro de Linha 3 (codigo_linha) se especificado
            if linha3_filtro != 'TODAS':
                if isinstance(linha3_filtro, list):
                    placeholders = ','.join(['%s'] * len(linha3_filtro))
                    query_produtos += f" AND p.codigo_linha IN ({placeholders})"
                    params.extend(linha3_filtro)
                else:
                    query_produtos += " AND p.codigo_linha = %s"
                    params.append(linha3_filtro)

            query_produtos += " ORDER BY p.cod_produto"

            df_forn = pd.read_sql(query_produtos, conn, params=params)

            if not df_forn.empty:
                total_produtos_por_fornecedor[nome_fornecedor] = len(df_forn)
                df_produtos_total = pd.concat([df_produtos_total, df_forn], ignore_index=True)

                # DEBUG: Log específico para Fame
                if 'FAME' in str(nome_fornecedor).upper():
                    print(f"    [DEBUG-FAME] Produtos encontrados: {len(df_forn)}")
            else:
                # DEBUG: Log quando não encontra produtos
                if 'FAME' in str(nome_fornecedor).upper():
                    print(f"    [DEBUG-FAME] NENHUM produto encontrado para '{nome_fornecedor}'")

            # Log de progresso a cada 10 fornecedores
            if (idx_forn + 1) % 10 == 0:
                print(f"    Fornecedores processados: {idx_forn + 1}/{len(fornecedores_a_processar)}")

        df_produtos = df_produtos_total

        if df_produtos.empty:
            conn.close()
            return jsonify({
                'success': False,
                'erro': f'Nenhum item encontrado com os filtros selecionados.'
            }), 400

        # Verificar duplicatas por código de produto (apenas debug, não remove)
        total_produtos = len(df_produtos)
        produtos_unicos = df_produtos['codigo'].nunique()
        duplicatas = total_produtos - produtos_unicos

        print(f"  Total de produtos a processar: {total_produtos} (de {len(fornecedores_a_processar)} fornecedores)")
        if duplicatas > 0:
            print(f"  [AVISO] Encontradas {duplicatas} duplicatas de código de produto!")
            # Mostrar quais produtos estão duplicados
            duplicados = df_produtos[df_produtos.duplicated(subset=['codigo'], keep=False)]
            print(f"    Produtos duplicados: {duplicados['codigo'].unique().tolist()[:10]}...")  # Mostrar primeiros 10
            # Remover duplicatas mantendo o primeiro (preserva fornecedor original do produto)
            df_produtos = df_produtos.drop_duplicates(subset=['codigo'], keep='first')
            print(f"    Após remoção: {len(df_produtos)} produtos")

        # DEBUG: Mostrar quantos produtos por fornecedor foram carregados
        for forn_nome, qtd in total_produtos_por_fornecedor.items():
            print(f"    [DEBUG] {forn_nome}: {qtd} produtos")

        # DEBUG: Contar produtos por fornecedor no DataFrame final
        contagem_final = df_produtos.groupby('nome_fornecedor').size().to_dict()
        print(f"  [DEBUG] Produtos por fornecedor no DataFrame final:")
        for forn, qtd in sorted(contagem_final.items()):
            if 'FAME' in str(forn).upper():
                print(f"    >>> {forn}: {qtd} produtos <<<")  # Destacar Fame
            else:
                print(f"    {forn}: {qtd} produtos")

        # Buscar lojas para somar demanda (lojas que vendem, cod_empresa < 80)
        # Flag para identificar pedido direto multi-loja (layout segregado por loja)
        is_pedido_multiloja = False
        mapa_nomes_lojas = {}  # {cod_empresa: nome_loja}

        if is_destino_cd:
            # Para CD: somar demanda de todas as lojas (cod_empresa 1-9)
            lojas_demanda = pd.read_sql(
                "SELECT cod_empresa FROM cadastro_lojas WHERE ativo = TRUE AND cod_empresa < 80",
                conn
            )['cod_empresa'].tolist()
            nome_destino = f"CD {cod_destino}"
        else:
            # Para loja(s) específica(s): usar as lojas selecionadas
            if lojas_selecionadas:
                lojas_demanda = lojas_selecionadas
                if len(lojas_selecionadas) == 1:
                    nome_destino = f"Loja {lojas_selecionadas[0]}"
                else:
                    # Pedido direto multi-loja: processar cada loja separadamente
                    is_pedido_multiloja = True
                    nome_destino = f"Lojas {', '.join(str(l) for l in lojas_selecionadas)}"
                    # Buscar nomes das lojas
                    placeholders = ','.join(['%s'] * len(lojas_selecionadas))
                    df_nomes_lojas = pd.read_sql(
                        f"SELECT cod_empresa, nome_loja FROM cadastro_lojas WHERE cod_empresa IN ({placeholders})",
                        conn,
                        params=lojas_selecionadas
                    )
                    mapa_nomes_lojas = dict(zip(df_nomes_lojas['cod_empresa'], df_nomes_lojas['nome_loja']))
                    print(f"  [MULTILOJA] Modo pedido direto multi-loja ativado")
                    print(f"  [MULTILOJA] Lojas: {mapa_nomes_lojas}")
            else:
                # Nenhuma loja selecionada: usar todas
                lojas_demanda = pd.read_sql(
                    "SELECT cod_empresa FROM cadastro_lojas WHERE ativo = TRUE AND cod_empresa < 80",
                    conn
                )['cod_empresa'].tolist()
                nome_destino = "Todas as Lojas"

        print(f"  Lojas para calcular demanda: {lojas_demanda}")
        print(f"  Modo multiloja: {is_pedido_multiloja}")

        # 2. Processar cada produto
        from core.demand_calculator import DemandCalculator

        processador = PedidoFornecedorIntegrado(conn)
        resultados = []

        itens_sem_historico = 0
        itens_sem_demanda = 0

        # DEBUG: Rastreamento por fornecedor
        debug_por_fornecedor = {}

        for idx, row in df_produtos.iterrows():
            try:
                codigo = row['codigo']
                nome_forn_atual = row.get('nome_fornecedor', 'DESCONHECIDO')

                # Inicializar contador para este fornecedor
                if nome_forn_atual not in debug_por_fornecedor:
                    debug_por_fornecedor[nome_forn_atual] = {
                        'total': 0, 'sem_historico': 0, 'sem_demanda': 0, 'processados': 0
                    }
                debug_por_fornecedor[nome_forn_atual]['total'] += 1

                # Obter parametros do fornecedor do cadastro
                lead_time_forn = int(row.get('lead_time_dias', 15))
                ciclo_pedido_forn = int(row.get('ciclo_pedido_dias', 7))
                pedido_min_forn = float(row.get('pedido_minimo_valor', 0))
                fornecedor_cadastrado = row.get('fornecedor_cadastrado', False)

                # =====================================================================
                # MODO MULTI-LOJA: Processar cada loja separadamente
                # IMPORTANTE: Todos os itens devem aparecer em todas as lojas para transparencia
                # =====================================================================
                if is_pedido_multiloja:
                    for loja_cod in lojas_demanda:
                        loja_nome = mapa_nomes_lojas.get(loja_cod, f"Loja {loja_cod}")

                        # Buscar histórico apenas desta loja
                        historico_loja = processador.buscar_historico_vendas(codigo, loja_cod, dias=365)

                        # Calcular demanda (pode ser 0 se não houver histórico)
                        demanda_diaria = 0
                        desvio_padrao = 0
                        metadata = {'metodo_usado': 'sem_historico'}

                        if historico_loja and len(historico_loja) >= 7:
                            demanda_media, desvio_padrao, metadata = DemandCalculator.calcular_demanda_inteligente(historico_loja, metodo='auto')
                            demanda_diaria = demanda_media if demanda_media > 0 else 0

                        # Processar item para esta loja (mesmo sem demanda, para transparencia)
                        resultado = processador.processar_item(
                            codigo=codigo,
                            cod_empresa=loja_cod,
                            previsao_diaria=demanda_diaria,
                            desvio_padrao=desvio_padrao,
                            cobertura_dias=cobertura_dias,
                            lead_time_dias=lead_time_forn,
                            ciclo_pedido_dias=ciclo_pedido_forn,
                            pedido_minimo_valor=pedido_min_forn
                        )

                        # IMPORTANTE: Incluir TODOS os itens, mesmo com erro
                        # Se houver erro, criar resultado basico para transparencia
                        if 'erro' in resultado:
                            resultado = {
                                'codigo': codigo,
                                'descricao': row.get('descricao', ''),
                                'nome_fornecedor': nome_forn_atual,
                                'curva_abc': row.get('curva_abc', 'B'),
                                'estoque_atual': 0,
                                'estoque_transito': 0,
                                'demanda_prevista': 0,
                                'demanda_prevista_diaria': 0,
                                'cobertura_atual_dias': 999,
                                'cobertura_pos_pedido_dias': 999,
                                'quantidade_pedido': 0,
                                'valor_pedido': 0,
                                'preco_custo': 0,
                                'cue': 0,
                                'deve_pedir': False,
                                'bloqueado': False,
                                'erro_processamento': resultado.get('erro', 'Erro desconhecido')
                            }

                        resultado['cod_loja'] = loja_cod
                        resultado['nome_loja'] = loja_nome
                        resultado['codigo_fornecedor'] = row.get('codigo_fornecedor', '')  # CNPJ
                        resultado['metodo_previsao'] = metadata.get('metodo_usado', 'auto')
                        resultado['lojas_consideradas'] = 1
                        resultado['fornecedor_cadastrado'] = fornecedor_cadastrado
                        resultado['lead_time_usado'] = lead_time_forn
                        resultado['ciclo_pedido_usado'] = ciclo_pedido_forn
                        resultado['pedido_minimo_fornecedor'] = pedido_min_forn
                        # Marcar se não tem demanda (para classificar como OK no frontend)
                        resultado['sem_demanda_loja'] = demanda_diaria <= 0
                        resultados.append(resultado)
                        debug_por_fornecedor[nome_forn_atual]['processados'] += 1

                else:
                    # =====================================================================
                    # MODO NORMAL: Agregar demanda de todas as lojas
                    # =====================================================================
                    # Buscar e somar histórico de todas as lojas relevantes
                    historico_total = []
                    for loja in lojas_demanda:
                        historico_loja = processador.buscar_historico_vendas(codigo, loja, dias=365)
                        if historico_loja:
                            if not historico_total:
                                historico_total = historico_loja.copy()
                            else:
                                # Somar históricos (dia a dia)
                                for i in range(min(len(historico_total), len(historico_loja))):
                                    historico_total[i] += historico_loja[i]

                    if not historico_total or len(historico_total) < 7:
                        itens_sem_historico += 1
                        debug_por_fornecedor[nome_forn_atual]['sem_historico'] += 1
                        if idx < 5:
                            print(f"    [DEBUG] Item {codigo}: sem historico suficiente")
                        continue

                    # Usar DemandCalculator para calcular demanda de forma inteligente
                    demanda_media, desvio_padrao, metadata = DemandCalculator.calcular_demanda_inteligente(historico_total, metodo='auto')

                    demanda_diaria = demanda_media

                    if demanda_diaria <= 0:
                        itens_sem_demanda += 1
                        debug_por_fornecedor[nome_forn_atual]['sem_demanda'] += 1
                        continue

                    resultado = processador.processar_item(
                        codigo=codigo,
                        cod_empresa=cod_destino,
                        previsao_diaria=demanda_diaria,
                        desvio_padrao=desvio_padrao,
                        cobertura_dias=cobertura_dias,
                        lead_time_dias=lead_time_forn,
                        ciclo_pedido_dias=ciclo_pedido_forn,
                        pedido_minimo_valor=pedido_min_forn
                    )

                    if 'erro' not in resultado:
                        resultado['nome_loja'] = nome_destino
                        resultado['cod_loja'] = cod_destino
                        resultado['codigo_fornecedor'] = row.get('codigo_fornecedor', '')  # CNPJ
                        resultado['metodo_previsao'] = metadata.get('metodo_usado', 'auto')
                        resultado['lojas_consideradas'] = len(lojas_demanda)
                        resultado['fornecedor_cadastrado'] = fornecedor_cadastrado
                        resultado['lead_time_usado'] = lead_time_forn
                        resultado['ciclo_pedido_usado'] = ciclo_pedido_forn
                        resultado['pedido_minimo_fornecedor'] = pedido_min_forn
                        resultados.append(resultado)
                        debug_por_fornecedor[nome_forn_atual]['processados'] += 1

            except Exception as e:
                print(f"  [AVISO] Erro ao processar item {row.get('codigo')}: {e}")
                import traceback
                traceback.print_exc()
                continue

            # Progresso
            if (idx + 1) % 50 == 0:
                print(f"  Processados {idx + 1}/{len(df_produtos)} produtos...")

        # =====================================================================
        # CALCULO DE OPORTUNIDADES DE TRANSFERENCIA
        # Para CD: usa logica de grupo regional
        # Para Multi-Loja: verifica excesso entre lojas selecionadas
        # =====================================================================
        oportunidades_transferencia = []
        sessao_pedido = None

        # =====================================================================
        # TRANSFERENCIAS PARA PEDIDO MULTI-LOJA (direto entre lojas selecionadas)
        # =====================================================================
        if is_pedido_multiloja and resultados and len(lojas_selecionadas) > 1:
            try:
                print(f"\n  [TRANSFERENCIAS MULTILOJA] Analisando oportunidades entre {len(lojas_selecionadas)} lojas...")

                import uuid
                sessao_pedido = f"ML_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"

                # Agrupar resultados por produto
                resultados_por_produto = {}
                for r in resultados:
                    cod = r['codigo']
                    if cod not in resultados_por_produto:
                        resultados_por_produto[cod] = {}
                    loja_cod = r.get('cod_loja')
                    if loja_cod:
                        resultados_por_produto[cod][loja_cod] = r

                total_oportunidades = 0
                cobertura_alvo = cobertura_dias or 21

                for cod_produto, lojas_data in resultados_por_produto.items():
                    if len(lojas_data) < 2:
                        continue

                    # Identificar lojas com excesso e lojas com necessidade
                    # IMPORTANTE: Uma loja NAO pode ser origem E destino ao mesmo tempo!
                    lojas_excesso = []  # (loja_cod, excesso_unidades, dados)
                    lojas_necessidade = []  # (loja_cod, necessidade_unidades, dados)

                    for loja_cod, dados in lojas_data.items():
                        demanda_diaria = dados.get('demanda_prevista_diaria', 0) or 0
                        estoque = dados.get('estoque_atual', 0) or 0
                        transito = dados.get('estoque_transito', 0) or 0
                        cobertura_atual = dados.get('cobertura_atual_dias', 0) or 0
                        deve_pedir = dados.get('deve_pedir', False)
                        qtd_pedido = dados.get('quantidade_pedido', 0) or 0

                        # ========================================================
                        # REGRAS CONFORME DOCUMENTACAO (TRANSFERENCIAS_ENTRE_LOJAS.md)
                        # ========================================================
                        # Parametros:
                        #   COBERTURA_MINIMA_DOADOR = 10 dias (minimo que doador mantem)
                        #   MARGEM_EXCESSO_DIAS = 7 dias (dias acima do alvo para ser doador)
                        #   COBERTURA_ALVO_PADRAO = 21 dias
                        #
                        # Doadora: Cobertura > alvo + 7 dias (ou seja, > 28 dias)
                        # Receptora: Cobertura < alvo (ou seja, < 21 dias)
                        #
                        # Formulas:
                        #   excesso_doador = estoque - (demanda × 10 dias)
                        #   necessidade_receptor = (demanda × 21 dias) - estoque
                        #   qtd_transferir = min(excesso, necessidade)
                        # ========================================================

                        COBERTURA_MINIMA_DOADOR = 10  # dias
                        MARGEM_EXCESSO_DIAS = 7       # dias acima do alvo

                        # REGRA PRINCIPAL: Se a loja PRECISA PEDIR, ela e RECEPTORA (nao pode ser doadora)
                        if deve_pedir and qtd_pedido > 0:
                            # Loja receptora: cobertura < alvo
                            # Necessidade = quanto falta para atingir cobertura alvo
                            estoque_efetivo = estoque + transito
                            if demanda_diaria > 0:
                                # Necessidade conforme documentacao:
                                # necessidade_receptor = (demanda × COBERTURA_ALVO) - estoque
                                necessidade_transf = int((demanda_diaria * cobertura_alvo) - estoque_efetivo)
                                necessidade_transf = max(0, necessidade_transf)
                            else:
                                necessidade_transf = 0

                            # Guardar ambos: necessidade para transferencia e qtd_pedido original
                            lojas_necessidade.append((loja_cod, necessidade_transf, dados, qtd_pedido))
                            continue  # NAO avaliar como possivel doadora

                        # Se NAO precisa pedir, verificar se tem excesso para ser DOADORA
                        # Doadora: cobertura > alvo + MARGEM_EXCESSO_DIAS
                        if demanda_diaria <= 0:
                            # Sem demanda = potencial doador se tem estoque significativo
                            if estoque > 5:
                                lojas_excesso.append((loja_cod, estoque, dados))
                        elif cobertura_atual > (cobertura_alvo + MARGEM_EXCESSO_DIAS):
                            # Tem excesso conforme documentacao: cobertura > alvo + 7 dias
                            # Excesso = estoque - (demanda × COBERTURA_MINIMA_DOADOR)
                            # Doador deve manter no minimo 10 dias de cobertura
                            estoque_disponivel = estoque + transito
                            estoque_minimo_doador = demanda_diaria * COBERTURA_MINIMA_DOADOR
                            excesso = int(estoque_disponivel - estoque_minimo_doador)
                            if excesso > 0:
                                lojas_excesso.append((loja_cod, excesso, dados))

                    # Se tem lojas com excesso e lojas com necessidade, sugerir transferencias
                    if lojas_excesso and lojas_necessidade:
                        # DEBUG: Log para verificar classificacao
                        print(f"    [TRANSF DEBUG] Produto {cod_produto}:")
                        print(f"      Lojas EXCESSO (podem enviar): {[(l[0], l[1]) for l in lojas_excesso]}")
                        print(f"      Lojas NECESSIDADE (precisam receber - transf): {[(l[0], l[1]) for l in lojas_necessidade]}")
                        print(f"      Lojas NECESSIDADE (pedido fornecedor): {[(l[0], l[3]) for l in lojas_necessidade]}")

                        # Ordenar por prioridade (maior excesso primeiro, maior necessidade primeiro)
                        lojas_excesso.sort(key=lambda x: x[1], reverse=True)
                        lojas_necessidade.sort(key=lambda x: x[1], reverse=True)

                        # lojas_necessidade agora tem 4 elementos: (loja_cod, necessidade_transf, dados, qtd_pedido)
                        for loja_dest, necessidade_transf, dados_dest, qtd_pedido_dest in lojas_necessidade:
                            qtd_restante = necessidade_transf

                            for i, (loja_orig, excesso, dados_orig) in enumerate(lojas_excesso):
                                if qtd_restante <= 0 or excesso <= 0:
                                    continue

                                # Calcular quantidade a transferir
                                qtd_transferir = min(qtd_restante, excesso)

                                if qtd_transferir > 0:
                                    # Criar oportunidade de transferencia
                                    cue = dados_orig.get('cue', 0) or dados_dest.get('cue', 0) or 0
                                    valor_estimado = qtd_transferir * cue

                                    transferencia = {
                                        'cod_produto': cod_produto,
                                        'descricao_produto': dados_orig.get('descricao', ''),
                                        'curva_abc': dados_orig.get('curva_abc', 'B'),
                                        'loja_origem': loja_orig,
                                        'nome_loja_origem': dados_orig.get('nome_loja', f'Loja {loja_orig}'),
                                        'estoque_origem': dados_orig.get('estoque_atual', 0),
                                        'cobertura_origem_dias': dados_orig.get('cobertura_atual_dias', 0),
                                        'loja_destino': loja_dest,
                                        'nome_loja_destino': dados_dest.get('nome_loja', f'Loja {loja_dest}'),
                                        'estoque_destino': dados_dest.get('estoque_atual', 0),
                                        'cobertura_destino_dias': dados_dest.get('cobertura_atual_dias', 0),
                                        'qtd_sugerida': qtd_transferir,
                                        'valor_estimado': valor_estimado,
                                        'cue': cue,
                                        'urgencia': 'ALTA' if dados_dest.get('ruptura_iminente') else 'MEDIA',
                                        'tipo': 'MULTILOJA'
                                    }

                                    oportunidades_transferencia.append(transferencia)
                                    total_oportunidades += 1

                                    # Atualizar excesso restante
                                    lojas_excesso[i] = (loja_orig, excesso - qtd_transferir, dados_orig)
                                    qtd_restante -= qtd_transferir

                print(f"  [TRANSFERENCIAS MULTILOJA] {total_oportunidades} oportunidades identificadas")

                # Salvar transferencias MULTILOJA no banco de dados
                if oportunidades_transferencia:
                    try:
                        cur = conn.cursor()
                        salvos_multiloja = 0

                        for t in oportunidades_transferencia:
                            try:
                                # Verificar se tabela existe
                                cur.execute("""
                                    SELECT EXISTS (
                                        SELECT FROM information_schema.tables
                                        WHERE table_name = 'oportunidades_transferencia'
                                    )
                                """)
                                if not cur.fetchone()[0]:
                                    print("  [TRANSFERENCIAS MULTILOJA] Tabela oportunidades_transferencia nao existe")
                                    break

                                cur.execute("""
                                    INSERT INTO oportunidades_transferencia (
                                        sessao_pedido, grupo_id,
                                        cod_produto, descricao_produto, curva_abc,
                                        loja_origem, nome_loja_origem, estoque_origem, transito_origem,
                                        demanda_diaria_origem, cobertura_origem_dias, excesso_unidades,
                                        loja_destino, nome_loja_destino, estoque_destino, transito_destino,
                                        demanda_diaria_destino, cobertura_destino_dias, necessidade_unidades,
                                        qtd_sugerida, valor_estimado, cue, urgencia
                                    ) VALUES (
                                        %s, %s, %s, %s, %s,
                                        %s, %s, %s, %s, %s, %s, %s,
                                        %s, %s, %s, %s, %s, %s, %s,
                                        %s, %s, %s, %s
                                    )
                                    ON CONFLICT (sessao_pedido, cod_produto, loja_origem, loja_destino)
                                    DO UPDATE SET
                                        qtd_sugerida = EXCLUDED.qtd_sugerida,
                                        valor_estimado = EXCLUDED.valor_estimado,
                                        urgencia = EXCLUDED.urgencia,
                                        data_calculo = NOW()
                                """, (
                                    sessao_pedido, None,  # grupo_id = None para MULTILOJA
                                    t.get('cod_produto'), t.get('descricao_produto', ''), t.get('curva_abc', 'B'),
                                    t.get('loja_origem'), t.get('nome_loja_origem', ''), t.get('estoque_origem', 0), 0,
                                    0, t.get('cobertura_origem_dias', 0), 0,
                                    t.get('loja_destino'), t.get('nome_loja_destino', ''), t.get('estoque_destino', 0), 0,
                                    0, t.get('cobertura_destino_dias', 0), 0,
                                    t.get('qtd_sugerida', 0), t.get('valor_estimado', 0), t.get('cue', 0), t.get('urgencia', 'MEDIA')
                                ))
                                salvos_multiloja += 1
                            except Exception as e_item:
                                print(f"  [TRANSFERENCIAS MULTILOJA] Erro ao salvar item: {e_item}")

                        conn.commit()
                        print(f"  [TRANSFERENCIAS MULTILOJA] {salvos_multiloja} oportunidades salvas no banco")
                        cur.close()
                    except Exception as e_save:
                        print(f"  [TRANSFERENCIAS MULTILOJA] Erro ao salvar no banco: {e_save}")

            except Exception as e:
                print(f"  [TRANSFERENCIAS MULTILOJA] Erro: {e}")
                import traceback
                traceback.print_exc()

        # =====================================================================
        # TRANSFERENCIAS PARA PEDIDO CD (logica original com grupo regional)
        # =====================================================================
        elif is_destino_cd and resultados:
            try:
                from core.transferencia_regional import TransferenciaRegional

                print(f"\n  [TRANSFERENCIAS] Calculando oportunidades de transferencia...")

                calculador = TransferenciaRegional(conn)
                sessao_pedido = calculador.gerar_sessao_pedido()

                # Identificar grupo regional pelo CD
                grupo = calculador._get_grupo_por_cd(cod_destino)

                if grupo:
                    grupo_id = grupo['id']
                    total_oportunidades = 0

                    # Para cada item processado, verificar desbalanceamento
                    for resultado in resultados:
                        codigo = resultado['codigo']
                        descricao = resultado.get('descricao', '')
                        curva_abc = resultado.get('curva_abc', 'B')

                        # Consolidar posicao do item
                        posicao = calculador.consolidar_posicao_item(
                            cod_produto=int(codigo),
                            grupo_id=grupo_id,
                            cd_principal=cod_destino
                        )

                        # Montar demanda por loja
                        demanda_por_loja = {}
                        demanda_diaria_total = resultado.get('demanda_prevista_diaria', 0)

                        if lojas_demanda and demanda_diaria_total > 0:
                            # Distribuir demanda proporcionalmente (simplificado)
                            demanda_por_loja_unit = demanda_diaria_total / len(lojas_demanda)
                            for loja in lojas_demanda:
                                demanda_por_loja[loja] = demanda_por_loja_unit

                        if demanda_por_loja:
                            # Analisar desbalanceamento
                            analise = calculador.analisar_desbalanceamento(
                                posicao=posicao,
                                demanda_por_loja=demanda_por_loja,
                                cobertura_alvo=cobertura_dias or 21
                            )

                            # Calcular transferencias se houver desbalanceamento
                            if analise['tem_desbalanceamento']:
                                transferencias = calculador.calcular_transferencias(
                                    posicao=posicao,
                                    analise=analise,
                                    cobertura_alvo=cobertura_dias or 21
                                )

                                if transferencias:
                                    # Salvar oportunidades
                                    salvos = calculador.salvar_oportunidades(
                                        transferencias=transferencias,
                                        grupo_id=grupo_id,
                                        sessao_pedido=sessao_pedido,
                                        descricao_produto=descricao,
                                        curva_abc=curva_abc
                                    )
                                    total_oportunidades += salvos
                                    oportunidades_transferencia.extend(transferencias)

                    print(f"  [TRANSFERENCIAS] {total_oportunidades} oportunidades identificadas e salvas")
                else:
                    print(f"  [TRANSFERENCIAS] CD {cod_destino} nao pertence a nenhum grupo de transferencia")

            except Exception as e:
                print(f"  [TRANSFERENCIAS] Erro ao calcular transferencias: {e}")
                import traceback
                traceback.print_exc()

        conn.close()

        print(f"  [DEBUG] Itens sem histórico suficiente: {itens_sem_historico}")
        print(f"  [DEBUG] Itens sem demanda: {itens_sem_demanda}")
        print(f"  [DEBUG] Itens processados com sucesso: {len(resultados)}")

        # DEBUG: Detalhamento por fornecedor
        print(f"\n  [DEBUG] ====== DETALHAMENTO POR FORNECEDOR ======")
        for forn_nome, stats in sorted(debug_por_fornecedor.items()):
            if stats['total'] > 0:  # Mostrar apenas fornecedores com produtos
                print(f"    {forn_nome}: {stats['total']} produtos -> "
                      f"{stats['processados']} processados, "
                      f"{stats['sem_historico']} sem histórico, "
                      f"{stats['sem_demanda']} sem demanda")
        print(f"  [DEBUG] ================================================")

        if not resultados:
            return jsonify({
                'success': False,
                'erro': f'Nenhum item com dados suficientes para gerar pedido. Itens analisados: {len(df_produtos)}, sem histórico: {itens_sem_historico}, sem demanda: {itens_sem_demanda}'
            }), 400

        print(f"  Total processado: {len(resultados)} itens")

        # Converter tipos NumPy para tipos Python nativos (para serialização JSON)
        def converter_tipos_json(obj):
            """Converte tipos NumPy para tipos Python nativos e trata NaN/Infinito/None"""
            import math

            # Tratar None primeiro
            if obj is None:
                return None

            # Tratar dicionários
            if isinstance(obj, dict):
                return {k: converter_tipos_json(v) for k, v in obj.items()}

            # Tratar listas
            elif isinstance(obj, list):
                return [converter_tipos_json(item) for item in obj]

            # Tratar tipos numpy
            elif isinstance(obj, (np.bool_, np.generic)):
                valor = obj.item()
                # Tratar NaN e infinito
                if isinstance(valor, float):
                    if math.isnan(valor):
                        return 0
                    if math.isinf(valor):
                        return 999 if valor > 0 else -999
                return valor

            # Tratar arrays numpy
            elif isinstance(obj, np.ndarray):
                return converter_tipos_json(obj.tolist())

            # Tratar floats Python (incluindo pandas NaN)
            elif isinstance(obj, float):
                if math.isnan(obj):
                    return 0
                if math.isinf(obj):
                    return 999 if obj > 0 else -999
                return obj

            # Tratar pandas NA/NaT (com try/except para tipos não suportados)
            else:
                try:
                    if pd.isna(obj):
                        return None
                except (TypeError, ValueError):
                    pass

            return obj

        resultados = converter_tipos_json(resultados)

        # 3. Agregar por fornecedor
        agregacao = agregar_por_fornecedor(resultados)

        # 4. Separar itens que devem ser pedidos, bloqueados e ok
        itens_pedido = [r for r in resultados if r.get('deve_pedir') and not r.get('bloqueado')]
        itens_bloqueados = [r for r in resultados if r.get('bloqueado')]
        itens_ok = [r for r in resultados if not r.get('deve_pedir') and not r.get('bloqueado')]

        # =====================================================================
        # VALIDACAO DE PEDIDO MINIMO POR FORNECEDOR/LOJA
        # =====================================================================
        # Agregar valor total por fornecedor e loja para validar pedido minimo
        validacao_pedido_minimo = {}  # {(cod_fornecedor, cod_loja): {'valor_total': X, 'pedido_minimo': Y, 'abaixo_minimo': bool}}

        for item in itens_pedido:
            cod_fornecedor = item.get('codigo_fornecedor', '')
            cod_loja = item.get('cod_loja', item.get('cod_empresa', 0))
            nome_fornecedor = item.get('nome_fornecedor', '')
            nome_loja = item.get('nome_loja', f'Loja {cod_loja}')
            valor_pedido = item.get('valor_pedido', 0) or 0
            pedido_minimo = item.get('pedido_minimo_fornecedor', 0) or 0

            chave = (cod_fornecedor, cod_loja)

            if chave not in validacao_pedido_minimo:
                validacao_pedido_minimo[chave] = {
                    'cod_fornecedor': cod_fornecedor,
                    'nome_fornecedor': nome_fornecedor,
                    'cod_loja': cod_loja,
                    'nome_loja': nome_loja,
                    'valor_total': 0,
                    'pedido_minimo': pedido_minimo,
                    'abaixo_minimo': False,
                    'diferenca': 0
                }

            validacao_pedido_minimo[chave]['valor_total'] += valor_pedido
            # Atualizar pedido_minimo se ainda nao foi definido
            if validacao_pedido_minimo[chave]['pedido_minimo'] == 0 and pedido_minimo > 0:
                validacao_pedido_minimo[chave]['pedido_minimo'] = pedido_minimo

        # Verificar quais estao abaixo do minimo
        alertas_pedido_minimo = []
        for chave, info in validacao_pedido_minimo.items():
            pedido_min = info['pedido_minimo']
            valor_total = info['valor_total']

            if pedido_min > 0 and valor_total < pedido_min:
                info['abaixo_minimo'] = True
                info['diferenca'] = pedido_min - valor_total
                alertas_pedido_minimo.append(info)
                print(f"  [PEDIDO MINIMO] Alerta: {info['nome_fornecedor']} / {info['nome_loja']} - "
                      f"Valor R$ {valor_total:,.2f} < Minimo R$ {pedido_min:,.2f}")

        # Marcar itens que estao em pedidos abaixo do minimo
        for item in itens_pedido:
            cod_fornecedor = item.get('codigo_fornecedor', '')
            cod_loja = item.get('cod_loja', item.get('cod_empresa', 0))
            chave = (cod_fornecedor, cod_loja)

            if chave in validacao_pedido_minimo:
                info = validacao_pedido_minimo[chave]
                item['pedido_abaixo_minimo'] = info['abaixo_minimo']
                item['valor_total_pedido_loja'] = info['valor_total']
                if info['abaixo_minimo']:
                    item['critica_pedido_minimo'] = f"Abaixo do pedido mínimo (R$ {info['valor_total']:,.2f} < R$ {info['pedido_minimo']:,.2f})"

        if alertas_pedido_minimo:
            print(f"  [PEDIDO MINIMO] {len(alertas_pedido_minimo)} fornecedor/loja(s) abaixo do minimo")

        # DEBUG: Contar por fornecedor nos resultados finais (apenas FAME)
        print(f"\n  [DEBUG] ====== FAME - DISTRIBUIÇÃO POR CATEGORIA ======")
        fame_pedido = len([r for r in itens_pedido if 'FAME' in str(r.get('nome_fornecedor', '')).upper()])
        fame_bloq = len([r for r in itens_bloqueados if 'FAME' in str(r.get('nome_fornecedor', '')).upper()])
        fame_ok = len([r for r in itens_ok if 'FAME' in str(r.get('nome_fornecedor', '')).upper()])
        print(f"    FAME - Pedido: {fame_pedido} itens")
        print(f"    FAME - Bloqueados: {fame_bloq} itens")
        print(f"    FAME - OK (sem necessidade): {fame_ok} itens")
        print(f"    FAME - Total na tabela (pedido+bloq): {fame_pedido + fame_bloq} itens")
        print(f"    FAME - Total geral: {fame_pedido + fame_bloq + fame_ok} itens")
        print(f"  [DEBUG] ================================================")

        # 5. Estatísticas
        # Filtrar itens não bloqueados para cálculos
        itens_nao_bloqueados = [r for r in resultados if not r.get('bloqueado')]

        # Verificar quantos itens usaram previsão semanal
        itens_com_previsao_semanal = len([r for r in itens_pedido if r.get('previsao_semanal', {}).get('usado', False)])

        # Calcular coberturas medias de forma segura (evitando NaN)
        coberturas_atuais = [r.get('cobertura_atual_dias', 0) for r in itens_nao_bloqueados if r.get('cobertura_atual_dias', 0) < 900]
        coberturas_pos = [r.get('cobertura_pos_pedido_dias', 0) for r in itens_pedido if r.get('cobertura_pos_pedido_dias', 0) < 900]

        cobertura_media_atual = round(np.mean(coberturas_atuais), 1) if coberturas_atuais else 0
        cobertura_media_pos = round(np.mean(coberturas_pos), 1) if coberturas_pos else 0

        # Garantir que nao ha NaN
        import math
        if math.isnan(cobertura_media_atual) or math.isinf(cobertura_media_atual):
            cobertura_media_atual = 0
        if math.isnan(cobertura_media_pos) or math.isinf(cobertura_media_pos):
            cobertura_media_pos = 0

        estatisticas = {
            'total_itens_analisados': len(resultados),
            'total_itens_pedido': len(itens_pedido),
            'total_itens_bloqueados': len(itens_bloqueados),
            'total_itens_ok': len(itens_ok),
            'valor_total_pedido': round(sum(r.get('valor_pedido', 0) for r in itens_pedido), 2),
            'itens_ruptura_iminente': len([r for r in itens_nao_bloqueados if r.get('ruptura_iminente')]),
            'itens_risco_ruptura': len([r for r in itens_nao_bloqueados if r.get('risco_ruptura')]),
            'cobertura_media_atual': cobertura_media_atual,
            'cobertura_media_pos_pedido': cobertura_media_pos,
            'itens_previsao_semanal': itens_com_previsao_semanal,
            'usa_previsao_semanal': itens_com_previsao_semanal > 0,
            # Debug info - ajuda a identificar problemas
            'total_produtos_consultados': len(df_produtos),
            'total_fornecedores_processados': len(fornecedores_a_processar),
            'itens_sem_historico': itens_sem_historico,
            'itens_sem_demanda': itens_sem_demanda
        }

        print(f"\n{'='*60}")
        print("RESULTADO")
        print(f"{'='*60}")
        print(f"  Itens a pedir: {estatisticas['total_itens_pedido']}")
        print(f"  Itens bloqueados: {estatisticas['total_itens_bloqueados']}")
        print(f"  Valor total: R$ {estatisticas['valor_total_pedido']:,.2f}")
        print(f"  Fornecedores: {agregacao['total_fornecedores']}")
        if oportunidades_transferencia:
            print(f"  Oportunidades de transferencia: {len(oportunidades_transferencia)}")
        print(f"{'='*60}\n")

        # Adicionar info de transferencias nas estatisticas
        estatisticas['total_oportunidades_transferencia'] = len(oportunidades_transferencia)
        estatisticas['sessao_pedido'] = sessao_pedido

        # Aplicar conversao JSON a todo o resultado (para garantir que nao ha NaN)
        resposta = converter_tipos_json({
            'success': True,
            'estatisticas': estatisticas,
            'agregacao_fornecedor': agregacao,
            'itens_pedido': itens_pedido,
            'itens_bloqueados': itens_bloqueados,
            'itens_ok': itens_ok,  # Todos os itens OK (sem necessidade de pedido)
            # Informacoes para layout multi-loja
            'is_pedido_multiloja': is_pedido_multiloja,
            'lojas_info': [
                {'cod_loja': cod, 'nome_loja': nome}
                for cod, nome in mapa_nomes_lojas.items()
            ] if is_pedido_multiloja else [],
            'transferencias': {
                'total': len(oportunidades_transferencia),
                'sessao_pedido': sessao_pedido,
                'tem_oportunidades': len(oportunidades_transferencia) > 0,
                'produtos_com_transferencia': list(set(t.get('cod_produto') for t in oportunidades_transferencia)),
                'detalhes': [
                    {
                        'cod_produto': t.get('cod_produto'),
                        'qtd_sugerida': t.get('qtd_sugerida'),
                        'loja_origem': t.get('nome_loja_origem'),
                        'loja_destino': t.get('nome_loja_destino'),
                        'cod_loja_origem': t.get('loja_origem'),  # Codigo numerico
                        'cod_loja_destino': t.get('loja_destino'),  # Codigo numerico
                        'urgencia': t.get('urgencia'),
                        'tipo': t.get('tipo', 'REGIONAL')  # MULTILOJA ou REGIONAL
                    }
                    for t in oportunidades_transferencia
                ]
            },
            # Alertas de pedido minimo por fornecedor/loja
            'alertas_pedido_minimo': [
                {
                    'cod_fornecedor': a['cod_fornecedor'],
                    'nome_fornecedor': a['nome_fornecedor'],
                    'cod_loja': a['cod_loja'],
                    'nome_loja': a['nome_loja'],
                    'valor_total': round(a['valor_total'], 2),
                    'pedido_minimo': round(a['pedido_minimo'], 2),
                    'diferenca': round(a['diferenca'], 2)
                }
                for a in alertas_pedido_minimo
            ],
            'tem_alertas_pedido_minimo': len(alertas_pedido_minimo) > 0
        })

        return jsonify(resposta)

    except Exception as e:
        import traceback
        print(f"[ERRO] {e}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'erro': str(e)
        }), 500


@app.route('/api/pedido_fornecedor_integrado/exportar', methods=['POST'])
def api_pedido_fornecedor_integrado_exportar():
    """Exporta o pedido ao fornecedor para Excel."""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        dados = request.get_json()

        if not dados:
            return jsonify({'success': False, 'erro': 'Dados não fornecidos'}), 400

        itens_pedido = dados.get('itens_pedido', [])
        estatisticas = dados.get('estatisticas', {})
        agregacao = dados.get('agregacao_fornecedor', {})

        if not itens_pedido:
            return jsonify({'success': False, 'erro': 'Nenhum item para exportar'}), 400

        # Criar workbook
        wb = Workbook()

        # =====================================================================
        # ABA 1: RESUMO
        # =====================================================================
        ws_resumo = wb.active
        ws_resumo.title = 'Resumo'

        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Titulo
        ws_resumo['A1'] = 'PEDIDO AO FORNECEDOR - RESUMO'
        ws_resumo['A1'].font = Font(bold=True, size=14)
        ws_resumo.merge_cells('A1:D1')

        ws_resumo['A2'] = f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}'

        # Estatisticas
        ws_resumo['A4'] = 'Estatisticas'
        ws_resumo['A4'].font = Font(bold=True)

        stats_data = [
            ('Total de Itens a Pedir', estatisticas.get('total_itens_pedido', 0)),
            ('Valor Total do Pedido', f"R$ {estatisticas.get('valor_total_pedido', 0):,.2f}"),
            ('Itens Analisados', estatisticas.get('total_itens_analisados', 0)),
            ('Itens em Ruptura Iminente', estatisticas.get('itens_ruptura_iminente', 0)),
            ('Total de Fornecedores', agregacao.get('total_fornecedores', 0)),
        ]

        for i, (label, valor) in enumerate(stats_data, start=5):
            ws_resumo[f'A{i}'] = label
            ws_resumo[f'B{i}'] = valor

        # Resumo por fornecedor
        ws_resumo['A12'] = 'Resumo por Fornecedor'
        ws_resumo['A12'].font = Font(bold=True)

        fornecedores = agregacao.get('fornecedores', [])
        if fornecedores:
            headers_forn = ['Fornecedor', 'Itens', 'Unidades', 'Valor']
            for col, header in enumerate(headers_forn, start=1):
                cell = ws_resumo.cell(row=13, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            for i, f in enumerate(fornecedores, start=14):
                ws_resumo.cell(row=i, column=1, value=f.get('nome_fornecedor', '')).border = border
                ws_resumo.cell(row=i, column=2, value=f.get('total_itens', 0)).border = border
                ws_resumo.cell(row=i, column=3, value=f.get('total_unidades', 0)).border = border
                ws_resumo.cell(row=i, column=4, value=f.get('valor_total', 0)).border = border
                ws_resumo.cell(row=i, column=4).number_format = '#,##0.00'

        # Ajustar largura das colunas
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20
        ws_resumo.column_dimensions['C'].width = 15
        ws_resumo.column_dimensions['D'].width = 15

        # =====================================================================
        # ABA 2: ITENS DO PEDIDO (estrutura por filial)
        # Estrutura: Cod Filial, Nome Filial, Cod Item, Descricao, CNPJ Forn, Nome Forn, Qtd, CUE
        # =====================================================================
        ws_itens = wb.create_sheet('Itens Pedido')

        # Headers na estrutura solicitada (com coluna Critica)
        headers = [
            'Cod Filial', 'Nome Filial', 'Cod Item', 'Descricao Item',
            'CNPJ Fornecedor', 'Nome Fornecedor', 'Qtd Pedido', 'CUE', 'Critica'
        ]

        # Estilo para celulas com critica (amarelo)
        critica_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        critica_font = Font(color='B45309')

        for col, header in enumerate(headers, start=1):
            cell = ws_itens.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Ordenar itens por Codigo Filial e CNPJ Fornecedor
        itens_ordenados = sorted(itens_pedido, key=lambda x: (
            x.get('cod_loja', x.get('cod_empresa', 0)) or 0,
            x.get('cnpj_fornecedor', x.get('codigo_fornecedor', '')) or ''
        ))

        # Dados na estrutura solicitada (ordenados por filial e CNPJ)
        for row_idx, item in enumerate(itens_ordenados, start=2):
            # Codigo da filial (cod_loja ou cod_empresa)
            cod_filial = item.get('cod_loja', item.get('cod_empresa', ''))
            ws_itens.cell(row=row_idx, column=1, value=cod_filial).border = border

            # Nome da filial
            nome_filial = item.get('nome_loja', item.get('nome_empresa', f'Filial {cod_filial}'))
            ws_itens.cell(row=row_idx, column=2, value=nome_filial).border = border

            # Codigo do item
            ws_itens.cell(row=row_idx, column=3, value=item.get('codigo', '')).border = border

            # Descricao do item
            ws_itens.cell(row=row_idx, column=4, value=item.get('descricao', '')[:60]).border = border

            # CNPJ do fornecedor
            cnpj_forn = item.get('cnpj_fornecedor', item.get('codigo_fornecedor', ''))
            ws_itens.cell(row=row_idx, column=5, value=cnpj_forn).border = border

            # Nome do fornecedor
            ws_itens.cell(row=row_idx, column=6, value=item.get('nome_fornecedor', '')).border = border

            # Quantidade do pedido
            ws_itens.cell(row=row_idx, column=7, value=item.get('quantidade_pedido', 0)).border = border

            # CUE (Custo Unitario Efetivo)
            cue = item.get('cue', item.get('preco_custo', 0))
            ws_itens.cell(row=row_idx, column=8, value=cue).border = border
            ws_itens.cell(row=row_idx, column=8).number_format = '#,##0.00'

            # Critica - Pedido abaixo do minimo
            critica = item.get('critica_pedido_minimo', '')
            cell_critica = ws_itens.cell(row=row_idx, column=9, value=critica)
            cell_critica.border = border
            if critica:
                cell_critica.fill = critica_fill
                cell_critica.font = critica_font

        # Ajustar largura das colunas
        col_widths = [12, 25, 12, 50, 18, 30, 12, 12, 45]
        for i, width in enumerate(col_widths, start=1):
            ws_itens.column_dimensions[chr(64 + i)].width = width

        # Congelar primeira linha
        ws_itens.freeze_panes = 'A2'

        # =====================================================================
        # ABA 3: POR FORNECEDOR (agrupado)
        # =====================================================================
        ws_forn = wb.create_sheet('Por Fornecedor')

        # Agrupar itens por fornecedor
        itens_por_fornecedor = {}
        for item in itens_pedido:
            forn = item.get('nome_fornecedor', 'Sem Fornecedor')
            if forn not in itens_por_fornecedor:
                itens_por_fornecedor[forn] = []
            itens_por_fornecedor[forn].append(item)

        row = 1
        for fornecedor, itens in sorted(itens_por_fornecedor.items()):
            # Header do fornecedor
            ws_forn.cell(row=row, column=1, value=fornecedor)
            ws_forn.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws_forn.merge_cells(f'A{row}:F{row}')

            valor_forn = sum(i.get('valor_pedido', 0) for i in itens)
            ws_forn.cell(row=row, column=7, value=f'Total: R$ {valor_forn:,.2f}')
            ws_forn.cell(row=row, column=7).font = Font(bold=True)

            row += 1

            # Headers da tabela
            headers_forn = ['Codigo', 'Descricao', 'Qtd Pedido', 'CUE', 'Valor']
            for col, header in enumerate(headers_forn, start=1):
                cell = ws_forn.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

            row += 1

            # Itens do fornecedor
            for item in itens:
                ws_forn.cell(row=row, column=1, value=item.get('codigo', ''))
                ws_forn.cell(row=row, column=2, value=item.get('descricao', '')[:40])
                ws_forn.cell(row=row, column=3, value=item.get('quantidade_pedido', 0))
                ws_forn.cell(row=row, column=4, value=item.get('preco_custo', 0))
                ws_forn.cell(row=row, column=4).number_format = '#,##0.00'
                ws_forn.cell(row=row, column=5, value=item.get('valor_pedido', 0))
                ws_forn.cell(row=row, column=5).number_format = '#,##0.00'
                row += 1

            row += 1  # Linha em branco entre fornecedores

        # Ajustar largura
        ws_forn.column_dimensions['A'].width = 12
        ws_forn.column_dimensions['B'].width = 40
        ws_forn.column_dimensions['C'].width = 12
        ws_forn.column_dimensions['D'].width = 12
        ws_forn.column_dimensions['E'].width = 14

        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pedido_fornecedor_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/transferencias')
def transferencias():
    """Página de transferências entre lojas - Oportunidades calculadas automaticamente"""
    return render_template('transferencias.html')


@app.route('/api/transferencias/oportunidades', methods=['GET'])
def api_transferencias_oportunidades():
    """
    Retorna oportunidades de transferência calculadas.
    Query params:
        - grupo_id: Filtrar por grupo (opcional)
        - horas: Limitar a oportunidades das últimas X horas (default: 24)
    """
    try:
        from core.transferencia_regional import TransferenciaRegional

        grupo_id = request.args.get('grupo_id', type=int)
        horas = request.args.get('horas', default=24, type=int)

        conn = get_db_connection()

        # Verificar se tabelas de transferencia existem (graceful degradation)
        cursor_check = conn.cursor()
        cursor_check.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'oportunidades_transferencia'
            )
        """)
        tabela_existe = cursor_check.fetchone()[0]
        cursor_check.close()

        if not tabela_existe:
            conn.close()
            return jsonify({
                'success': True,
                'oportunidades': [],
                'resumo': {'total': 0, 'criticas': 0, 'altas': 0, 'valor_total': 0, 'produtos_unicos': 0},
                'aviso': 'Tabela de transferencias ainda nao foi criada. Execute o script de migracao (migration_v5.sql).'
            })

        calculador = TransferenciaRegional(conn)

        oportunidades = calculador.buscar_oportunidades_recentes(
            grupo_id=grupo_id,
            limite_horas=horas
        )
        resumo = calculador.resumo_oportunidades(oportunidades)

        conn.close()

        return jsonify({
            'success': True,
            'oportunidades': oportunidades,
            'resumo': resumo
        })

    except Exception as e:
        print(f"Erro ao buscar oportunidades: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/api/transferencias/grupos', methods=['GET'])
def api_transferencias_grupos():
    """Retorna grupos de transferência disponíveis."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Verificar se tabela existe (graceful degradation)
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'grupos_transferencia'
            )
        """)
        tabela_existe = cursor.fetchone()[0]

        if not tabela_existe:
            cursor.close()
            conn.close()
            return jsonify({
                'success': True,
                'grupos': [],
                'aviso': 'Tabelas de transferencia ainda nao foram criadas. Execute o script de migracao (migration_v5.sql).'
            })

        cursor.close()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT
                gt.id,
                gt.nome,
                gt.descricao,
                gt.cd_principal,
                COUNT(lgt.id) as total_lojas
            FROM grupos_transferencia gt
            LEFT JOIN lojas_grupo_transferencia lgt ON gt.id = lgt.grupo_id AND lgt.ativo = TRUE
            WHERE gt.ativo = TRUE
            GROUP BY gt.id, gt.nome, gt.descricao, gt.cd_principal
            ORDER BY gt.nome
        """)

        grupos = cursor.fetchall()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'grupos': grupos
        })

    except Exception as e:
        print(f"Erro ao buscar grupos: {e}")
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/api/transferencias/exportar', methods=['GET'])
def api_transferencias_exportar():
    """Exporta oportunidades de transferência para Excel."""
    try:
        from core.transferencia_regional import TransferenciaRegional
        import io

        grupo_id = request.args.get('grupo_id', type=int)
        horas = request.args.get('horas', default=24, type=int)

        conn = get_db_connection()

        # Verificar se tabela existe (graceful degradation)
        cursor_check = conn.cursor()
        cursor_check.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'oportunidades_transferencia'
            )
        """)
        tabela_existe = cursor_check.fetchone()[0]
        cursor_check.close()

        if not tabela_existe:
            conn.close()
            return jsonify({
                'success': False,
                'erro': 'Tabela de transferencias ainda nao foi criada. Execute o script de migracao (migration_v5.sql).'
            }), 404

        calculador = TransferenciaRegional(conn)

        oportunidades = calculador.buscar_oportunidades_recentes(
            grupo_id=grupo_id,
            limite_horas=horas
        )
        conn.close()

        if not oportunidades:
            return jsonify({'success': False, 'erro': 'Nenhuma oportunidade encontrada'}), 404

        # Criar DataFrame
        df = pd.DataFrame(oportunidades)

        # Selecionar e renomear colunas
        colunas = {
            'cod_produto': 'Codigo',
            'descricao_produto': 'Descricao',
            'curva_abc': 'ABC',
            'nome_loja_origem': 'Loja Origem',
            'estoque_origem': 'Estoque Origem',
            'cobertura_origem_dias': 'Cobertura Origem (dias)',
            'nome_loja_destino': 'Loja Destino',
            'estoque_destino': 'Estoque Destino',
            'cobertura_destino_dias': 'Cobertura Destino (dias)',
            'qtd_sugerida': 'Qtd Transferir',
            'valor_estimado': 'Valor Estimado',
            'urgencia': 'Urgencia',
            'grupo_nome': 'Grupo Regional'
        }

        df_export = df[[c for c in colunas.keys() if c in df.columns]].copy()
        df_export.columns = [colunas[c] for c in df_export.columns]

        # Gerar Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Transferencias', index=False)

            # Ajustar largura das colunas
            worksheet = writer.sheets['Transferencias']
            for idx, col in enumerate(df_export.columns):
                max_len = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)

        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'transferencias_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/api/transferencias/limpar', methods=['DELETE'])
def api_transferencias_limpar():
    """Limpa todo o histórico de oportunidades de transferência."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Verificar se tabela existe (graceful degradation)
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'oportunidades_transferencia'
            )
        """)
        tabela_existe = cursor.fetchone()[0]

        if not tabela_existe:
            cursor.close()
            conn.close()
            return jsonify({
                'success': True,
                'registros_removidos': 0,
                'mensagem': 'Tabela de transferencias ainda nao foi criada. Execute o script de migracao.'
            })

        # Contar registros antes de deletar
        cursor.execute("SELECT COUNT(*) FROM oportunidades_transferencia")
        total_antes = cursor.fetchone()[0]

        # Deletar todos os registros
        cursor.execute("DELETE FROM oportunidades_transferencia")
        conn.commit()

        cursor.close()
        conn.close()

        print(f"[TRANSFERENCIAS] Historico limpo: {total_antes} registros removidos")

        return jsonify({
            'success': True,
            'registros_removidos': total_antes,
            'mensagem': f'{total_antes} registros removidos com sucesso'
        })

    except Exception as e:
        print(f"Erro ao limpar historico: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/pedido_manual')
def pedido_manual():
    """Página de pedido manual simplificado"""
    return render_template('pedido_manual.html')


@app.route('/pedido_quantidade')
def pedido_quantidade():
    """Página de pedido por quantidade"""
    return render_template('pedido_quantidade.html')


@app.route('/pedido_cobertura')
def pedido_cobertura():
    """Página de pedido por cobertura"""
    return render_template('pedido_cobertura.html')


@app.route('/processar_pedido_quantidade', methods=['POST'])
def processar_pedido_quantidade():
    """
    Processa pedido por quantidade informada pelo usuário
    """
    try:
        from core.order_processor import OrderProcessor, gerar_relatorio_pedido
        from datetime import datetime

        if 'file' not in request.files:
            return jsonify({'success': False, 'erro': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'success': False, 'erro': 'Nenhum arquivo selecionado'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'erro': 'Arquivo deve ser Excel'}), 400

        # Salvar arquivo
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Carregar dados
        df_entrada = pd.read_excel(filepath)

        # Processar pedido
        df_resultado = OrderProcessor.processar_pedido_por_quantidade(
            df_entrada,
            validar_embalagem=True
        )

        # Gerar resumo
        resumo = gerar_relatorio_pedido(df_resultado, 'quantidade')

        # Salvar resultado
        nome_arquivo_saida = f"pedido_quantidade_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_saida = os.path.join(app.config['OUTPUT_FOLDER'], nome_arquivo_saida)

        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df_resultado.to_excel(writer, sheet_name='Pedido', index=False)

            # Criar aba de resumo
            itens_ajustados = df_resultado[df_resultado['Foi_Ajustado'] == True]
            if len(itens_ajustados) > 0:
                itens_ajustados.to_excel(writer, sheet_name='Itens_Ajustados', index=False)

        # Limpar arquivo de upload
        try:
            os.remove(filepath)
        except:
            pass

        return jsonify({
            'success': True,
            'arquivo_saida': nome_arquivo_saida,
            'resumo': resumo,
            'dados_tabela': df_resultado.to_dict('records')
        })

    except ValueError as e:
        return jsonify({'success': False, 'erro': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'erro': f'Erro: {str(e)}'}), 500


@app.route('/processar_pedido_cobertura', methods=['POST'])
def processar_pedido_cobertura():
    """
    Processa pedido por cobertura desejada informada pelo usuário
    """
    try:
        from core.order_processor import OrderProcessor, gerar_relatorio_pedido
        from datetime import datetime

        if 'file' not in request.files:
            return jsonify({'success': False, 'erro': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'success': False, 'erro': 'Nenhum arquivo selecionado'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'erro': 'Arquivo deve ser Excel'}), 400

        # Salvar arquivo
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Carregar dados
        df_entrada = pd.read_excel(filepath)

        # Processar pedido
        df_resultado = OrderProcessor.processar_pedido_por_cobertura(df_entrada)

        # Gerar resumo
        resumo = gerar_relatorio_pedido(df_resultado, 'cobertura')

        # Salvar resultado
        nome_arquivo_saida = f"pedido_cobertura_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_saida = os.path.join(app.config['OUTPUT_FOLDER'], nome_arquivo_saida)

        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df_resultado.to_excel(writer, sheet_name='Pedido', index=False)

            # Criar aba de itens sem necessidade
            itens_sem_necessidade = df_resultado[df_resultado['Quantidade_Pedido'] == 0]
            if len(itens_sem_necessidade) > 0:
                itens_sem_necessidade.to_excel(writer, sheet_name='Sem_Necessidade', index=False)

            # Criar aba de itens ajustados
            itens_ajustados = df_resultado[df_resultado['Foi_Ajustado'] == True]
            if len(itens_ajustados) > 0:
                itens_ajustados.to_excel(writer, sheet_name='Itens_Ajustados', index=False)

        # Limpar arquivo de upload
        try:
            os.remove(filepath)
        except:
            pass

        return jsonify({
            'success': True,
            'arquivo_saida': nome_arquivo_saida,
            'resumo': resumo,
            'dados_tabela': df_resultado.to_dict('records')
        })

    except ValueError as e:
        return jsonify({'success': False, 'erro': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'erro': f'Erro: {str(e)}'}), 500


@app.route('/api/carregar_dados_pedido', methods=['POST'])
def carregar_dados_pedido():
    """
    Carrega dados de reabastecimento para uso no simulador
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'erro': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'success': False, 'erro': 'Nenhum arquivo selecionado'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'erro': 'Arquivo deve ser Excel'}), 400

        # Salvar arquivo temporário
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{filename}")
        file.save(filepath)

        # Carregar Excel
        df = pd.read_excel(filepath)

        # Validar colunas mínimas necessárias
        colunas_necessarias = [
            'Loja', 'SKU', 'Demanda_Media_Mensal', 'Lead_Time_Dias',
            'Estoque_Disponivel', 'Ponto_Pedido'
        ]

        colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]

        if colunas_faltantes:
            os.remove(filepath)
            return jsonify({
                'success': False,
                'erro': f'Colunas faltando: {", ".join(colunas_faltantes)}'
            }), 400

        # Calcular demanda diária
        if 'Demanda_Diaria' not in df.columns:
            df['Demanda_Diaria'] = df['Demanda_Media_Mensal'] / 30

        # Garantir que quantidade de pedido existe
        if 'Quantidade_Pedido' not in df.columns:
            df['Quantidade_Pedido'] = 0

        if 'Lote_Minimo' not in df.columns:
            df['Lote_Minimo'] = 1

        # Selecionar colunas relevantes
        colunas_retornar = [
            'Loja', 'SKU', 'Demanda_Diaria', 'Lead_Time_Dias',
            'Estoque_Disponivel', 'Ponto_Pedido', 'Quantidade_Pedido',
            'Lote_Minimo'
        ]

        # Adicionar colunas opcionais se existirem
        for col in ['Estoque_Transito', 'Pedidos_Abertos', 'Estoque_Seguranca']:
            if col in df.columns:
                colunas_retornar.append(col)

        df_resultado = df[colunas_retornar].copy()

        # Limpar arquivo temporário
        try:
            os.remove(filepath)
        except:
            pass

        return jsonify({
            'success': True,
            'dados': df_resultado.to_dict('records'),
            'total_itens': len(df_resultado)
        })

    except Exception as e:
        return jsonify({'success': False, 'erro': f'Erro ao carregar: {str(e)}'}), 500


@app.route('/api/buscar_item', methods=['POST'])
def buscar_item():
    """
    Busca item específico no último resultado processado (placeholder)
    """
    try:
        data = request.get_json()
        loja = data.get('loja')
        sku = data.get('sku')

        if not loja or not sku:
            return jsonify({'success': False, 'erro': 'Loja e SKU são obrigatórios'}), 400

        # TODO: Implementar busca em cache ou banco de dados
        # Por enquanto, retorna erro informativo
        return jsonify({
            'success': False,
            'erro': 'Carregue um arquivo Excel primeiro para buscar itens'
        }), 404

    except Exception as e:
        return jsonify({'success': False, 'erro': str(e)}), 500


# ===== ROTAS DE KPIs =====
@app.route('/kpis')
def kpis():
    """Página de KPIs"""
    return render_template('kpis.html')


@app.route('/api/kpis/filtros', methods=['GET'])
def kpis_filtros():
    """Retorna opções para filtros de KPIs"""
    try:
        # TODO: Buscar do banco de dados ou cache
        # Por enquanto, retornar dados mockados
        filtros = {
            'lojas': [
                {'id': 1, 'nome': 'Loja 1'},
                {'id': 2, 'nome': 'Loja 2'},
                {'id': 3, 'nome': 'Loja 3'}
            ],
            'produtos': [
                {'id': 1001, 'nome': 'Produto A'},
                {'id': 1002, 'nome': 'Produto B'},
                {'id': 1003, 'nome': 'Produto C'}
            ],
            'categorias': [
                'Alimentos',
                'Bebidas',
                'Limpeza',
                'Higiene'
            ],
            'fornecedores': [
                'Fornecedor 1',
                'Fornecedor 2',
                'Fornecedor 3'
            ]
        }
        return jsonify(filtros)
    except Exception as e:
        print(f"Erro ao buscar filtros: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/kpis/dados', methods=['GET'])
def kpis_dados():
    """Retorna dados de KPIs com base nos filtros"""
    try:
        visao = request.args.get('visao', 'mensal')
        loja = request.args.get('loja', '')
        produto = request.args.get('produto', '')
        categoria = request.args.get('categoria', '')
        fornecedor = request.args.get('fornecedor', '')

        # TODO: Buscar dados reais do banco/cache baseado nos filtros
        # Por enquanto, retornar dados mockados

        import random
        from datetime import datetime, timedelta

        # Gerar dados mockados para demonstração
        if visao == 'mensal':
            # Últimos 12 meses
            meses = []
            data_base = datetime.now()
            for i in range(12, 0, -1):
                mes = data_base - timedelta(days=i*30)
                meses.append(mes.strftime('%b/%y'))

            wmape_mensal = [{'mes': m, 'wmape': random.uniform(5, 15)} for m in meses]
            bias_mensal = [{'mes': m, 'bias': random.uniform(-2, 2)} for m in meses]
            ruptura_mensal = [{'mes': m, 'taxa_ruptura': random.uniform(2, 8)} for m in meses]
            cobertura_mensal = [{'mes': m, 'cobertura_media': random.uniform(15, 25)} for m in meses]
            servico_mensal = [{'mes': m, 'nivel_servico': random.uniform(92, 98)} for m in meses]

            classificacao = []
            for m in meses:
                classificacao.append({
                    'mes': m,
                    'excelente': random.randint(50, 80),
                    'bom': random.randint(30, 50),
                    'aceitavel': random.randint(10, 30),
                    'fraca': random.randint(5, 15)
                })

        else:  # semanal
            # Últimas 26 semanas
            semanas = []
            data_base = datetime.now()
            for i in range(26, 0, -1):
                semana = data_base - timedelta(days=i*7)
                semanas.append(semana.strftime('S%V/%y'))

            wmape_mensal = [{'semana': s, 'wmape': random.uniform(5, 15)} for s in semanas]
            bias_mensal = [{'semana': s, 'bias': random.uniform(-2, 2)} for s in semanas]
            ruptura_mensal = [{'semana': s, 'taxa_ruptura': random.uniform(2, 8)} for s in semanas]
            cobertura_mensal = [{'semana': s, 'cobertura_media': random.uniform(15, 25)} for s in semanas]
            classificacao = []
            servico_mensal = []

        # Métricas atuais (últimos 30 dias)
        metricas_atuais = {
            'wmape': random.uniform(8, 12),
            'bias': random.uniform(-1.5, 1.5),
            'previsoes_excelentes': random.randint(120, 180),
            'total_skus': 245,
            'taxa_ruptura': random.uniform(3, 7),
            'cobertura_media': random.uniform(18, 22),
            'nivel_servico': random.uniform(93, 97),
            'skus_criticos': random.randint(5, 15),

            # Tendências
            'wmape_tendencia': {'tipo': 'down', 'valor': '2.3%'},
            'bias_tendencia': {'tipo': 'stable', 'valor': '0.1%'},
            'excelentes_tendencia': {'tipo': 'up', 'valor': '5.2%'},
            'ruptura_tendencia': {'tipo': 'down', 'valor': '1.2%'},
            'cobertura_tendencia': {'tipo': 'up', 'valor': '0.5 dias'},
            'servico_tendencia': {'tipo': 'up', 'valor': '2.1%'},
            'criticos_tendencia': {'tipo': 'down', 'valor': '3 SKUs'}
        }

        # Top/Bottom Performers
        performers = [
            {
                'sku': '1001',
                'descricao': 'Produto A - Amostra',
                'loja': 'Loja 1',
                'wmape': 5.2,
                'bias': 0.3,
                'taxa_ruptura': 1.5,
                'cobertura': 22.5
            },
            {
                'sku': '1002',
                'descricao': 'Produto B - Amostra',
                'loja': 'Loja 2',
                'wmape': 8.7,
                'bias': -0.8,
                'taxa_ruptura': 3.2,
                'cobertura': 19.3
            },
            {
                'sku': '1003',
                'descricao': 'Produto C - Amostra',
                'loja': 'Loja 1',
                'wmape': 25.4,
                'bias': 5.2,
                'taxa_ruptura': 18.7,
                'cobertura': 8.2
            },
            {
                'sku': '1004',
                'descricao': 'Produto D - Amostra',
                'loja': 'Loja 3',
                'wmape': 12.3,
                'bias': -2.1,
                'taxa_ruptura': 5.8,
                'cobertura': 15.7
            }
        ]

        resultado = {
            'metricas_atuais': metricas_atuais,
            'series_temporais': {
                'wmape_mensal' if visao == 'mensal' else 'wmape_semanal': wmape_mensal,
                'bias_mensal' if visao == 'mensal' else 'bias_semanal': bias_mensal,
                'classificacao': classificacao if visao == 'mensal' else [],
                'ruptura_mensal' if visao == 'mensal' else 'ruptura_semanal': ruptura_mensal,
                'cobertura_mensal' if visao == 'mensal' else 'cobertura_semanal': cobertura_mensal,
                'servico_mensal': servico_mensal if visao == 'mensal' else []
            },
            'performers': performers
        }

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar dados de KPIs: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'erro': str(e)}), 500


# ==============================================================================
# ROTAS DE API - INTEGRAÇÃO COM BANCO DE DADOS
# ==============================================================================

import psycopg2
from psycopg2.extras import RealDictCursor

# Configuração do banco de dados via variáveis de ambiente
# Para usar localmente, defina as variáveis ou crie um arquivo .env
DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'localhost'),
    'database': os.environ.get('DB_NAME', 'previsao_demanda'),
    'user': os.environ.get('DB_USER', 'postgres'),
    'password': os.environ.get('DB_PASSWORD', 'FerreiraCost@01'),
    'port': int(os.environ.get('DB_PORT', 5432))
}

def get_db_connection():
    """Cria conexão com o banco PostgreSQL"""
    conn = psycopg2.connect(**DB_CONFIG)
    conn.set_client_encoding('LATIN1')
    return conn


@app.route('/visualizacao')
def visualizacao_demanda():
    """Página de visualização de demanda com dados reais"""
    return render_template('visualizacao_demanda.html')


@app.route('/api/lojas', methods=['GET'])
def api_lojas():
    """Retorna lista de lojas do banco"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT cod_empresa, nome_loja, regiao, tipo
            FROM cadastro_lojas
            WHERE ativo = TRUE
            ORDER BY nome_loja
        """)

        lojas = cursor.fetchall()
        cursor.close()
        conn.close()

        # Adicionar opção "Todas"
        resultado = [{'cod_empresa': 'TODAS', 'nome_loja': 'TODAS - Todas as Lojas (Agregado)', 'regiao': None, 'tipo': None}]
        resultado.extend([dict(row) for row in lojas])

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar lojas: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/categorias', methods=['GET'])
def api_categorias():
    """Retorna lista de categorias do banco"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT DISTINCT categoria
            FROM cadastro_produtos
            WHERE categoria IS NOT NULL AND ativo = TRUE
            ORDER BY categoria
        """)

        categorias = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()

        # Adicionar opção "Todas"
        resultado = ['TODAS - Todas as Categorias']
        resultado.extend(categorias)

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar categorias: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/produtos', methods=['GET'])
def api_produtos():
    """Retorna lista de produtos do banco (com filtros opcionais)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Parâmetros de filtro
        loja = request.args.get('loja', 'TODAS')
        categoria = request.args.get('categoria', 'TODAS')

        # Query base
        query = """
            SELECT DISTINCT p.codigo, p.descricao, p.categoria, p.subcategoria, p.curva_abc
            FROM cadastro_produtos p
            WHERE p.ativo = TRUE
        """

        params = []

        # Filtrar por categoria
        if categoria and categoria != 'TODAS' and not categoria.startswith('TODAS -'):
            query += " AND p.categoria = %s"
            params.append(categoria)

        # Se filtrar por loja específica, mostrar só produtos com vendas nessa loja
        if loja and loja != 'TODAS':
            query += """
                AND EXISTS (
                    SELECT 1 FROM historico_vendas_diario h
                    WHERE h.codigo = p.codigo AND h.cod_empresa = %s
                )
            """
            params.append(int(loja))

        query += " ORDER BY p.descricao"

        cursor.execute(query, params)
        produtos = cursor.fetchall()
        cursor.close()
        conn.close()

        # Adicionar opção "Todos"
        resultado = [{'codigo': 'TODOS', 'descricao': 'TODOS - Todos os Produtos (Agregado)', 'categoria': None, 'subcategoria': None, 'curva_abc': None}]
        resultado.extend(produtos)

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar produtos: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/fornecedores', methods=['GET'])
def api_fornecedores():
    """Retorna lista de fornecedores do banco"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT DISTINCT nome_fornecedor, cnpj_fornecedor
            FROM cadastro_produtos_completo
            WHERE ativo = TRUE AND nome_fornecedor IS NOT NULL
            ORDER BY nome_fornecedor
        """)

        fornecedores = cursor.fetchall()
        cursor.close()
        conn.close()

        # Adicionar opcao "Todos"
        resultado = [{'nome_fornecedor': 'TODOS', 'cnpj_fornecedor': None}]
        resultado.extend(fornecedores)

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar fornecedores: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/linhas', methods=['GET'])
def api_linhas():
    """Retorna lista de linhas (categorias nivel 1) do banco"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT DISTINCT categoria as linha
            FROM cadastro_produtos_completo
            WHERE ativo = TRUE AND categoria IS NOT NULL
            ORDER BY categoria
        """)

        linhas = cursor.fetchall()
        cursor.close()
        conn.close()

        # Adicionar opcao "Todas"
        resultado = [{'linha': 'TODAS'}]
        resultado.extend(linhas)

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar linhas: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/sublinhas', methods=['GET'])
def api_sublinhas():
    """Retorna lista de sublinhas (categorias nivel 3) do banco"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Parametro de filtro por linha
        linha = request.args.get('linha', 'TODAS')

        query = """
            SELECT DISTINCT codigo_linha, descricao_linha
            FROM cadastro_produtos_completo
            WHERE ativo = TRUE AND codigo_linha IS NOT NULL
        """
        params = []

        if linha and linha != 'TODAS':
            query += " AND categoria = %s"
            params.append(linha)

        query += " ORDER BY descricao_linha"

        cursor.execute(query, params)
        sublinhas = cursor.fetchall()
        cursor.close()
        conn.close()

        # Adicionar opcao "Todas"
        resultado = [{'codigo_linha': 'TODAS', 'descricao_linha': 'TODAS - Todas as Sublinhas'}]
        resultado.extend(sublinhas)

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar sublinhas: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/produtos_completo', methods=['GET'])
def api_produtos_completo():
    """Retorna lista de produtos com dados completos (fornecedor, linha, sublinha)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Parametros de filtro
        loja = request.args.get('loja', 'TODAS')
        fornecedor = request.args.get('fornecedor', 'TODOS')
        linha = request.args.get('linha', 'TODAS')
        sublinha = request.args.get('sublinha', 'TODAS')

        query = """
            SELECT DISTINCT
                p.cod_produto as codigo,
                p.descricao,
                p.nome_fornecedor,
                p.cnpj_fornecedor,
                p.categoria as linha,
                p.codigo_linha,
                p.descricao_linha as sublinha
            FROM cadastro_produtos_completo p
            WHERE p.ativo = TRUE
        """
        params = []

        # Filtrar por fornecedor
        if fornecedor and fornecedor != 'TODOS':
            query += " AND p.nome_fornecedor = %s"
            params.append(fornecedor)

        # Filtrar por linha
        if linha and linha != 'TODAS':
            query += " AND p.categoria = %s"
            params.append(linha)

        # Filtrar por sublinha
        if sublinha and sublinha != 'TODAS':
            query += " AND p.codigo_linha = %s"
            params.append(sublinha)

        # Se filtrar por loja especifica, mostrar so produtos com vendas nessa loja
        if loja and loja != 'TODAS':
            query += """
                AND EXISTS (
                    SELECT 1 FROM historico_vendas_diario h
                    WHERE h.codigo = p.cod_produto AND h.cod_empresa = %s
                )
            """
            params.append(int(loja))

        query += " ORDER BY p.descricao"

        cursor.execute(query, params)
        produtos = cursor.fetchall()
        cursor.close()
        conn.close()

        # Adicionar opcao "Todos"
        resultado = [{
            'codigo': 'TODOS',
            'descricao': 'TODOS - Todos os Produtos (Agregado)',
            'nome_fornecedor': None,
            'cnpj_fornecedor': None,
            'linha': None,
            'codigo_linha': None,
            'sublinha': None
        }]
        resultado.extend(produtos)

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar produtos completo: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/vendas', methods=['GET'])
def api_vendas():
    """Retorna dados de vendas históricos"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Parâmetros
        loja = request.args.get('loja', 'TODAS')
        produto = request.args.get('produto', 'TODOS')
        categoria = request.args.get('categoria', 'TODAS')
        meses = int(request.args.get('meses', 12))

        # Query base
        query = """
            SELECT
                DATE_TRUNC('day', h.data) as data,
                SUM(h.qtd_venda) as qtd_venda,
                SUM(h.valor_venda) as valor_venda,
                COUNT(DISTINCT h.cod_empresa) as num_lojas,
                COUNT(DISTINCT h.codigo) as num_produtos
            FROM historico_vendas_diario h
        """

        # Joins se necessário
        if categoria and categoria != 'TODAS' and not categoria.startswith('TODAS -'):
            query += " JOIN cadastro_produtos p ON h.codigo = p.codigo"

        query += " WHERE h.data >= CURRENT_DATE - INTERVAL '%s months'"
        params = [meses]

        # Filtros
        if loja and loja != 'TODAS':
            query += " AND h.cod_empresa = %s"
            params.append(int(loja))

        if produto and produto != 'TODOS':
            query += " AND h.codigo = %s"
            params.append(int(produto))

        if categoria and categoria != 'TODAS' and not categoria.startswith('TODAS -'):
            query += " AND p.categoria = %s"
            params.append(categoria)

        query += """
            GROUP BY DATE_TRUNC('day', h.data)
            ORDER BY data
        """

        cursor.execute(query, params)
        vendas = cursor.fetchall()
        cursor.close()
        conn.close()

        # Converter para formato JSON-friendly
        resultado = []
        for row in vendas:
            resultado.append({
                'data': row['data'].strftime('%Y-%m-%d'),
                'qtd_venda': float(row['qtd_venda']),
                'valor_venda': float(row['valor_venda']) if row['valor_venda'] else 0,
                'num_lojas': row['num_lojas'],
                'num_produtos': row['num_produtos']
            })

        return jsonify(resultado)

    except Exception as e:
        print(f"Erro ao buscar vendas: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'erro': str(e)}), 500


@app.route('/api/gerar_previsao_banco', methods=['POST'])
def api_gerar_previsao_banco():
    """
    Gera previsão de demanda usando dados do banco PostgreSQL

    Esta função agora redireciona para a versão otimizada (Bottom-Up),
    que calcula cada item individualmente e depois agrega.

    Benefícios:
    - Sazonalidade por dia da semana (diário) ou semana do ano (semanal)
    - Tendência com limitadores (queda >40% ignorada, alta >50% limitada)
    - Soma dos itens = Total (consistência matemática)
    """
    # Redirecionar para a função otimizada (Bottom-Up)
    return api_gerar_previsao_banco_v2_interno()


# Função legada mantida para referência (não utilizada)
def _api_gerar_previsao_banco_legado():
    """
    [LEGADO] Versão anterior da previsão - mantida apenas para referência.
    A versão atual usa Bottom-Up (api_gerar_previsao_banco_v2_interno).
    """
    try:
        # Inicializar auto-logger para auditoria
        logger = get_auto_logger()

        dados = request.get_json()

        loja = dados.get('loja', '')
        categoria = dados.get('categoria', '')
        produto = dados.get('produto', '')
        granularidade = dados.get('granularidade', 'mensal')

        # Extrair parâmetros de período customizado
        tipo_periodo = dados.get('tipo_periodo', None)

        # Variáveis para armazenar período de previsão (calculado depois)
        periodo_inicio_customizado = None
        periodo_fim_customizado = None

        if tipo_periodo == 'mensal':
            mes_inicio = int(dados.get('mes_inicio', 1))
            ano_inicio = int(dados.get('ano_inicio', 2026))
            mes_fim = int(dados.get('mes_fim', 3))
            ano_fim = int(dados.get('ano_fim', 2026))
            periodo_inicio_customizado = pd.Timestamp(year=ano_inicio, month=mes_inicio, day=1)
            from calendar import monthrange
            ultimo_dia = monthrange(ano_fim, mes_fim)[1]
            periodo_fim_customizado = pd.Timestamp(year=ano_fim, month=mes_fim, day=ultimo_dia)
            print(f"\nPeríodo customizado MENSAL: {mes_inicio:02d}/{ano_inicio} até {mes_fim:02d}/{ano_fim}", flush=True)

        elif tipo_periodo == 'semanal':
            semana_inicio = int(dados.get('semana_inicio', 1))
            ano_semana_inicio = int(dados.get('ano_semana_inicio', 2026))
            semana_fim = int(dados.get('semana_fim', 12))
            ano_semana_fim = int(dados.get('ano_semana_fim', 2026))
            # Converter semana/ano para data
            from datetime import datetime, timedelta
            # Semana 1 = primeira segunda-feira do ano ou 1 de janeiro
            periodo_inicio_customizado = pd.Timestamp(datetime.strptime(f'{ano_semana_inicio}-W{semana_inicio:02d}-1', '%G-W%V-%u'))
            periodo_fim_customizado = pd.Timestamp(datetime.strptime(f'{ano_semana_fim}-W{semana_fim:02d}-7', '%G-W%V-%u'))
            print(f"\nPeríodo customizado SEMANAL: Semana {semana_inicio}/{ano_semana_inicio} até Semana {semana_fim}/{ano_semana_fim}", flush=True)

        elif tipo_periodo == 'diario':
            data_inicio_str = dados.get('data_inicio', '')
            data_fim_str = dados.get('data_fim', '')
            if data_inicio_str and data_fim_str:
                periodo_inicio_customizado = pd.Timestamp(data_inicio_str)
                periodo_fim_customizado = pd.Timestamp(data_fim_str)
                print(f"\nPeríodo customizado DIÁRIO: {data_inicio_str} até {data_fim_str}", flush=True)

        # Fallback para meses_previsao se não houver período customizado
        meses_previsao = int(dados.get('meses_previsao', 3))

        # Normalizar filtros ANTES de fazer prints (remover emojis e prefixos)
        # Se categoria começa com "TODAS -", considerar como TODAS
        if categoria and categoria.startswith('TODAS'):
            categoria = 'TODAS'

        # Se produto começa com "TODOS" ou tem emoji, considerar como TODOS
        if produto and (produto.startswith('TODOS') or '📦' in produto):
            produto = 'TODOS'

        # Se loja começa com "TODAS" ou tem emoji, considerar como TODAS
        if loja and (loja.startswith('TODAS') or '📊' in loja):
            loja = 'TODAS'

        print(f"\n=== Gerando Previsao do Banco ===")
        print(f"Loja: {loja}")
        print(f"Categoria: {categoria}")
        print(f"Produto: {produto}")
        print(f"Granularidade: {granularidade}")
        print(f"Tipo período recebido: {tipo_periodo}")
        if periodo_inicio_customizado and periodo_fim_customizado:
            print(f"Período CUSTOMIZADO: {periodo_inicio_customizado.strftime('%Y-%m-%d')} até {periodo_fim_customizado.strftime('%Y-%m-%d')}")
        else:
            print(f"Meses Previsao (fallback): {meses_previsao}")

        # NOTA: O cálculo exato de períodos será feito APÓS obter a última data do histórico
        # para garantir que o período de previsão corresponda exatamente aos meses solicitados
        # A variável periodos_previsao será calculada mais adiante no código
        periodos_previsao_aproximado = meses_previsao  # valor temporário, será recalculado

        print(f"Período de previsao solicitado", flush=True)

        # Conectar ao banco
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Construir query baseada nos filtros
        where_clauses = []
        params = []

        # Filtro de loja (se não for TODAS)
        if loja and loja != 'TODAS' and not loja.startswith('TODAS'):
            where_clauses.append("h.cod_empresa = %s")
            params.append(loja)

        # Filtro de categoria (se não for TODAS)
        if categoria and categoria != 'TODAS' and not categoria.startswith('TODAS'):
            where_clauses.append("p.categoria = %s")
            params.append(categoria)

        # Filtro de produto (se não for TODOS)
        if produto and produto != 'TODOS' and not produto.startswith('TODOS'):
            where_clauses.append("h.codigo = %s")
            params.append(produto)

        where_sql = " AND " + " AND ".join(where_clauses) if where_clauses else ""

        # Query para buscar dados históricos baseado na granularidade
        # IMPORTANTE: Usar DATE_TRUNC('month', ...) para garantir que começamos no primeiro dia do mês
        # Isso garante meses completos e proporção correta entre meses e semanas
        # Exemplo: Se hoje é 2026-01-09, o boundary será 2024-01-01 (não 2024-01-09)
        if granularidade == 'diario':
            query = f"""
                SELECT
                    h.data,
                    SUM(h.qtd_venda) as qtd_venda,
                    SUM(h.valor_venda) as valor_venda
                FROM historico_vendas_diario h
                JOIN cadastro_produtos p ON h.codigo = p.codigo
                WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                {where_sql}
                GROUP BY h.data
                ORDER BY h.data
            """
        elif granularidade == 'semanal':
            # Para semanal, agrupar por semana usando EXATAMENTE os mesmos dias do período
            # IMPORTANTE: Não filtrar semanas, apenas agrupar os dias que já foram filtrados
            # Isso garante que mensal e semanal somem EXATAMENTE os mesmos dias
            query = f"""
                WITH dados_diarios AS (
                    SELECT
                        h.data,
                        SUM(h.qtd_venda) as qtd_venda,
                        SUM(h.valor_venda) as valor_venda
                    FROM historico_vendas_diario h
                    JOIN cadastro_produtos p ON h.codigo = p.codigo
                    WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                    {where_sql}
                    GROUP BY h.data
                )
                SELECT
                    DATE_TRUNC('week', data)::date as data,
                    SUM(qtd_venda) as qtd_venda,
                    SUM(valor_venda) as valor_venda
                FROM dados_diarios
                GROUP BY DATE_TRUNC('week', data)
                ORDER BY DATE_TRUNC('week', data)
            """
        else:  # mensal
            query = f"""
                SELECT
                    DATE_TRUNC('month', h.data) as data,
                    SUM(h.qtd_venda) as qtd_venda,
                    SUM(h.valor_venda) as valor_venda
                FROM historico_vendas_diario h
                JOIN cadastro_produtos p ON h.codigo = p.codigo
                WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                {where_sql}
                GROUP BY DATE_TRUNC('month', h.data)
                ORDER BY DATE_TRUNC('month', h.data)
            """

        cursor.execute(query, params)
        dados_historicos = cursor.fetchall()

        # Log do total de dados históricos base
        if dados_historicos:
            total_historico_base = sum(float(d['qtd_venda']) for d in dados_historicos)
            print(f"Total dados históricos (últimos 2 anos): {total_historico_base:,.2f} em {len(dados_historicos)} períodos", flush=True)

        # Buscar dados do mesmo período do ano anterior para comparação
        print(f"Buscando dados do ano anterior para comparação...", flush=True)

        # Primeiro, precisamos processar dados históricos para calcular o período
        if dados_historicos and len(dados_historicos) >= 3:
            # Criar DataFrame temporário para obter datas
            df_temp = pd.DataFrame(dados_historicos)
            df_temp['data'] = pd.to_datetime(df_temp['data'])
            df_temp.set_index('data', inplace=True)
            df_temp.sort_index(inplace=True)

            # Calcular período correspondente do ano anterior
            # IMPORTANTE: Se período customizado foi fornecido, usar ele para buscar ano anterior
            # Isso garante que a comparação na tabela seja coerente com o período de previsão
            if periodo_inicio_customizado and periodo_fim_customizado:
                # Usar período de previsão customizado para buscar ano anterior correspondente
                inicio_periodo_previsao = periodo_inicio_customizado
                fim_periodo_previsao = periodo_fim_customizado
                print(f"  Usando período customizado para ano anterior: {inicio_periodo_previsao.strftime('%Y-%m-%d')} até {fim_periodo_previsao.strftime('%Y-%m-%d')}", flush=True)

                # CORREÇÃO PARA SEMANAL: Usar EXATAMENTE as mesmas semanas ISO do ano anterior
                # Exemplo: Semana 1-13/2026 -> Semana 1-13/2025
                # Isso garante que o número de períodos seja igual e a comparação seja justa
                if granularidade == 'semanal' and tipo_periodo == 'semanal':
                    from datetime import datetime as dt

                    # Obter os parâmetros originais de semana
                    semana_inicio_param = int(dados.get('semana_inicio', 1))
                    ano_semana_inicio_param = int(dados.get('ano_semana_inicio', 2026))
                    semana_fim_param = int(dados.get('semana_fim', 13))
                    ano_semana_fim_param = int(dados.get('ano_semana_fim', 2026))

                    # Ano anterior
                    ano_inicio_ant = ano_semana_inicio_param - 1
                    ano_fim_ant = ano_semana_fim_param - 1

                    # Converter semana ISO para data: mesmas semanas, ano anterior
                    # Semana X do ano Y-1
                    inicio_ano_anterior = pd.Timestamp(dt.strptime(f'{ano_inicio_ant}-W{semana_inicio_param:02d}-1', '%G-W%V-%u'))
                    fim_ano_anterior = pd.Timestamp(dt.strptime(f'{ano_fim_ant}-W{semana_fim_param:02d}-7', '%G-W%V-%u'))
                    print(f"  Semanal: Semanas {semana_inicio_param}-{semana_fim_param}/{ano_semana_inicio_param} -> Semanas {semana_inicio_param}-{semana_fim_param}/{ano_inicio_ant}", flush=True)
                else:
                    # Para mensal e diário, subtrair 1 ano normalmente
                    inicio_ano_anterior = inicio_periodo_previsao - pd.DateOffset(years=1)
                    fim_ano_anterior = fim_periodo_previsao - pd.DateOffset(years=1)
            else:
                # CORREÇÃO: Calcular o período de PREVISÃO (não histórico) para buscar ano anterior
                # O ano anterior deve corresponder aos mesmos períodos que serão previstos
                ultima_data_hist = df_temp.index[-1]

                if granularidade == 'semanal':
                    # Para semanal sem período customizado:
                    # A previsão começa na semana seguinte ao último histórico
                    # Buscar ano anterior: mesmas semanas do ano Y-1
                    from datetime import datetime as dt

                    # Data de início da previsão: semana seguinte ao último histórico
                    inicio_previsao = ultima_data_hist + pd.Timedelta(weeks=1)
                    # Data fim: calcular baseado em meses_previsao
                    mes_destino_calc = ultima_data_hist.month + meses_previsao
                    ano_destino_calc = ultima_data_hist.year
                    while mes_destino_calc > 12:
                        mes_destino_calc -= 12
                        ano_destino_calc += 1
                    from calendar import monthrange
                    ultimo_dia = monthrange(ano_destino_calc, mes_destino_calc)[1]
                    fim_previsao = pd.Timestamp(year=ano_destino_calc, month=mes_destino_calc, day=ultimo_dia)

                    # Semanas do período de previsão
                    semana_inicio = inicio_previsao.isocalendar()[1]
                    ano_semana_inicio = inicio_previsao.isocalendar()[0]
                    semana_fim = fim_previsao.isocalendar()[1]
                    ano_semana_fim = fim_previsao.isocalendar()[0]

                    # Ano anterior: mesmas semanas ISO, ano -1
                    ano_inicio_ant = ano_semana_inicio - 1
                    ano_fim_ant = ano_semana_fim - 1

                    # Tratar caso especial: semana 53 pode não existir em alguns anos
                    # Nesse caso, limitar à semana 52 ou usar a última disponível
                    semana_fim_ajustada = semana_fim
                    if semana_fim == 53:
                        # Verificar se o ano anterior tem semana 53
                        try:
                            dt.strptime(f'{ano_fim_ant}-W53-1', '%G-W%V-%u')
                        except ValueError:
                            # Ano anterior não tem semana 53, usar semana 52
                            semana_fim_ajustada = 52
                            print(f"  Nota: Semana 53 não existe em {ano_fim_ant}, usando S52", flush=True)

                    try:
                        # IMPORTANTE: Começar 1 semana ANTES para garantir que temos todas as semanas necessárias
                        # Isso cobre casos onde a semana incompleta no histórico desloca o início da previsão
                        semana_busca_inicio = semana_inicio - 1 if semana_inicio > 1 else 52
                        ano_busca_inicio = ano_inicio_ant if semana_inicio > 1 else ano_inicio_ant - 1

                        inicio_ano_anterior = pd.Timestamp(dt.strptime(f'{ano_busca_inicio}-W{semana_busca_inicio:02d}-1', '%G-W%V-%u'))
                        fim_ano_anterior = pd.Timestamp(dt.strptime(f'{ano_fim_ant}-W{semana_fim_ajustada:02d}-7', '%G-W%V-%u'))
                        print(f"  Semanal (auto): Previsão S{semana_inicio}-S{semana_fim}/{ano_semana_inicio} -> Busca S{semana_busca_inicio}-S{semana_fim_ajustada}/{ano_busca_inicio}-{ano_fim_ant}", flush=True)
                    except ValueError as e:
                        # Se ainda assim não funcionar, usar DateOffset
                        inicio_ano_anterior = inicio_previsao - pd.DateOffset(years=1)
                        fim_ano_anterior = fim_previsao - pd.DateOffset(years=1)
                        print(f"  Semanal (fallback): usando DateOffset ({e})", flush=True)
                else:
                    # Para mensal e diário: usar o PERÍODO DE PREVISÃO, não o histórico
                    # Calcular período de previsão baseado em ultima_data_hist e meses_previsao
                    inicio_previsao = ultima_data_hist + pd.Timedelta(days=1)

                    mes_destino_calc = ultima_data_hist.month + meses_previsao
                    ano_destino_calc = ultima_data_hist.year
                    while mes_destino_calc > 12:
                        mes_destino_calc -= 12
                        ano_destino_calc += 1
                    from calendar import monthrange
                    ultimo_dia = monthrange(ano_destino_calc, mes_destino_calc)[1]
                    fim_previsao = pd.Timestamp(year=ano_destino_calc, month=mes_destino_calc, day=ultimo_dia)

                    inicio_ano_anterior = inicio_previsao - pd.DateOffset(years=1)
                    fim_ano_anterior = fim_previsao - pd.DateOffset(years=1)
                    print(f"  {granularidade.capitalize()} (auto): {inicio_previsao.strftime('%Y-%m-%d')} a {fim_previsao.strftime('%Y-%m-%d')} -> {inicio_ano_anterior.strftime('%Y-%m-%d')} a {fim_ano_anterior.strftime('%Y-%m-%d')}", flush=True)

            print(f"  Buscando ano anterior: {inicio_ano_anterior.strftime('%Y-%m-%d')} até {fim_ano_anterior.strftime('%Y-%m-%d')}", flush=True)

            # Buscar dados do ano anterior (mesmos filtros, exceto as datas)
            # params contém os filtros, mas não inclui as datas inicialmente
            # Todas as granularidades: [inicio, fim, ...filtros]
            params_ano_anterior = [
                inicio_ano_anterior.strftime('%Y-%m-%d'),
                fim_ano_anterior.strftime('%Y-%m-%d')
            ] + params

            if granularidade == 'mensal':
                query_ano_anterior = f"""
                    SELECT
                        DATE_TRUNC('month', h.data) as data,
                        SUM(h.qtd_venda) as qtd_venda
                    FROM historico_vendas_diario h
                    JOIN cadastro_produtos p ON h.codigo = p.codigo
                    WHERE h.data >= %s
                      AND h.data <= %s
                      {where_sql}
                    GROUP BY DATE_TRUNC('month', h.data)
                    ORDER BY DATE_TRUNC('month', h.data)
                """
            elif granularidade == 'semanal':
                # Mesma lógica do query principal: agrupar EXATAMENTE os mesmos dias
                query_ano_anterior = f"""
                    WITH dados_diarios AS (
                        SELECT
                            h.data,
                            SUM(h.qtd_venda) as qtd_venda
                        FROM historico_vendas_diario h
                        JOIN cadastro_produtos p ON h.codigo = p.codigo
                        WHERE h.data >= %s
                          AND h.data <= %s
                          {where_sql}
                        GROUP BY h.data
                    )
                    SELECT
                        DATE_TRUNC('week', data)::date as data,
                        SUM(qtd_venda) as qtd_venda
                    FROM dados_diarios
                    GROUP BY DATE_TRUNC('week', data)
                    ORDER BY DATE_TRUNC('week', data)
                """
            else:  # diario
                query_ano_anterior = f"""
                    SELECT
                        h.data,
                        SUM(h.qtd_venda) as qtd_venda
                    FROM historico_vendas_diario h
                    JOIN cadastro_produtos p ON h.codigo = p.codigo
                    WHERE h.data >= %s
                      AND h.data <= %s
                      {where_sql}
                    GROUP BY h.data
                    ORDER BY h.data
                """

            cursor.execute(query_ano_anterior, params_ano_anterior)
            dados_ano_anterior = cursor.fetchall()

            # Processar dados do ano anterior - INCLUIR DATAS
            ano_anterior_datas = [d['data'].strftime('%Y-%m-%d') for d in dados_ano_anterior] if dados_ano_anterior else []
            ano_anterior_valores = [float(d['qtd_venda']) for d in dados_ano_anterior] if dados_ano_anterior else []
            ano_anterior_total = sum(ano_anterior_valores) if ano_anterior_valores else 0

            print(f"Dados ano anterior: {len(ano_anterior_valores)} períodos, total: {ano_anterior_total}", flush=True)
            if ano_anterior_datas:
                print(f"  Período ano anterior: {ano_anterior_datas[0]} até {ano_anterior_datas[-1]}", flush=True)
        else:
            ano_anterior_datas = []
            ano_anterior_valores = []
            ano_anterior_total = 0

        # Fechar conexão
        cursor.close()
        conn.close()

        # Verificar se há dados suficientes (mínimo 3 períodos)
        if not dados_historicos or len(dados_historicos) < 3:
            return jsonify({
                'erro': f'Dados insuficientes para gerar previsão. Encontrados {len(dados_historicos) if dados_historicos else 0} períodos, necessário pelo menos 3 períodos de histórico.',
                'detalhes': {
                    'periodos_encontrados': len(dados_historicos) if dados_historicos else 0,
                    'filtros': {
                        'loja': loja,
                        'categoria': categoria,
                        'produto': produto,
                        'granularidade': granularidade
                    }
                }
            }), 400

        # Converter para DataFrame do pandas
        df = pd.DataFrame(dados_historicos)
        df['data'] = pd.to_datetime(df['data'])
        # Converter qtd_venda para float (pode vir como Decimal do PostgreSQL)
        df['qtd_venda'] = df['qtd_venda'].astype(float)
        df = df.sort_values('data')

        # Preparar série temporal
        df.set_index('data', inplace=True)
        serie_temporal_completa = df['qtd_venda'].tolist()
        datas_completas = df.index.tolist()

        # CALCULAR PERÍODOS DE PREVISÃO
        # Agora suporta período CUSTOMIZADO informado pelo usuário
        from dateutil.relativedelta import relativedelta
        from calendar import monthrange

        ultima_data_historico = datas_completas[-1]

        # Se período customizado foi fornecido, usar diretamente
        if periodo_inicio_customizado and periodo_fim_customizado:
            data_inicio_previsao = periodo_inicio_customizado
            data_fim_previsao = periodo_fim_customizado

            # Para mensal, extrair mês/ano início e destino
            primeiro_mes_previsao = data_inicio_previsao.month
            ano_primeiro_mes = data_inicio_previsao.year
            mes_destino = data_fim_previsao.month
            ano_destino = data_fim_previsao.year

            print(f"Periodo de previsao CUSTOMIZADO:", flush=True)
            print(f"  Data inicio: {data_inicio_previsao.strftime('%Y-%m-%d')}", flush=True)
            print(f"  Data fim: {data_fim_previsao.strftime('%Y-%m-%d')}", flush=True)
        else:
            # LÓGICA ORIGINAL baseada em meses_previsao
            # Primeiro mês de previsão = mês onde a previsão começa
            primeiro_mes_previsao = ultima_data_historico.month
            ano_primeiro_mes = ultima_data_historico.year

            # O MÊS DESTINO é o primeiro mês + (meses_previsao - 1)
            mes_destino = primeiro_mes_previsao + (meses_previsao - 1)
            ano_destino = ano_primeiro_mes
            while mes_destino > 12:
                mes_destino -= 12
                ano_destino += 1

            # Data fim é o ÚLTIMO DIA do mês de destino
            ultimo_dia_mes_destino = monthrange(ano_destino, mes_destino)[1]
            data_fim_previsao = pd.Timestamp(year=ano_destino, month=mes_destino, day=ultimo_dia_mes_destino)

            print(f"Periodo de previsao alvo:", flush=True)
            print(f"  Ultima data historico: {ultima_data_historico.strftime('%Y-%m-%d')}", flush=True)
            print(f"  Primeiro mes previsao: {primeiro_mes_previsao:02d}/{ano_primeiro_mes}", flush=True)
            print(f"  Mes alvo (destino): {mes_destino:02d}/{ano_destino}", flush=True)
            print(f"  Data fim previsao: {data_fim_previsao.strftime('%Y-%m-%d')}", flush=True)

        # Usar pd.date_range para calcular o número EXATO de períodos
        # Isso garante consistência com a geração de datas futuras
        if granularidade == 'diario':
            # Para diário: usar data_inicio_previsao se customizado, senão dia seguinte ao histórico
            if periodo_inicio_customizado:
                start_date = periodo_inicio_customizado
            else:
                start_date = ultima_data_historico + pd.Timedelta(days=1)

            datas_previsao_temp = pd.date_range(
                start=start_date,
                end=data_fim_previsao,
                freq='D'
            )
            periodos_previsao = len(datas_previsao_temp)

        elif granularidade == 'semanal':
            # Para semanal: usar data_inicio_previsao se customizado
            if periodo_inicio_customizado:
                start_date = periodo_inicio_customizado
            else:
                start_date = ultima_data_historico + pd.Timedelta(weeks=1)

            datas_previsao_temp = pd.date_range(
                start=start_date,
                end=data_fim_previsao,
                freq='W-MON'
            )
            periodos_previsao = len(datas_previsao_temp)

        else:  # mensal
            # Para mensal: usar período customizado ou calcular baseado em meses_previsao
            if periodo_inicio_customizado:
                # Período customizado: usar EXATAMENTE o mês/ano que o usuário solicitou
                # primeiro_mes_previsao = mês de início informado pelo usuário
                # ano_primeiro_mes = ano de início informado pelo usuário
                data_inicio_temp = pd.Timestamp(year=ano_primeiro_mes, month=primeiro_mes_previsao, day=1)
                data_fim_temp = pd.Timestamp(year=ano_destino, month=mes_destino, day=1)
                print(f"  Mensal customizado: {primeiro_mes_previsao:02d}/{ano_primeiro_mes} até {mes_destino:02d}/{ano_destino}", flush=True)
            else:
                # Original: primeiro período é o mês SEGUINTE ao último histórico
                mes_inicio_mensal = primeiro_mes_previsao + 1
                ano_inicio_mensal = ano_primeiro_mes
                if mes_inicio_mensal > 12:
                    mes_inicio_mensal = 1
                    ano_inicio_mensal += 1

                data_inicio_temp = pd.Timestamp(year=ano_inicio_mensal, month=mes_inicio_mensal, day=1)
                data_fim_temp = pd.Timestamp(year=ano_destino, month=mes_destino, day=1)

            datas_previsao_temp = pd.date_range(
                start=data_inicio_temp,
                end=data_fim_temp,
                freq='MS'
            )
            periodos_previsao = len(datas_previsao_temp)

        print(f"  Periodos de previsao: {periodos_previsao} ({granularidade})", flush=True)

        # Validar série temporal antes do processamento
        print(f"\nValidando serie temporal...", flush=True)
        validacao = validate_series(
            serie_temporal_completa,
            min_length=3,
            allow_zeros=True,
            check_outliers=False  # Outliers serão tratados separadamente
        )

        if not validacao['valid']:
            return jsonify({
                'erro': 'Dados inválidos para previsão',
                'codigo_erro': validacao.get('errors', [{}])[0].get('code', 'ERR000'),
                'mensagem': validacao.get('errors', [{}])[0].get('message', 'Erro desconhecido'),
                'detalhes': validacao.get('statistics', {})
            }), 400

        print(f"  Validacao OK - {len(serie_temporal_completa)} periodos", flush=True)

        # Log detalhado da série temporal completa
        total_serie_completa = sum(serie_temporal_completa)
        media_serie_completa = total_serie_completa / len(serie_temporal_completa)
        print(f"  Total da série completa: {total_serie_completa:,.2f}", flush=True)
        print(f"  Média por período: {media_serie_completa:,.2f}", flush=True)

        # Dividir dados: 50% base histórica, 25% teste, 25% previsão
        total_periodos = len(serie_temporal_completa)
        print(f"\nTotal de periodos historicos: {total_periodos}", flush=True)

        # Calcular tamanhos - Divisão 50/25/25
        tamanho_base = int(total_periodos * 0.50)
        tamanho_teste = int(total_periodos * 0.25)
        # Usar os períodos restantes para validação futura (comparação YoY)
        # O meses_previsao será usado para prever ALÉM dos dados históricos
        tamanho_validacao_futura = total_periodos - tamanho_base - tamanho_teste

        # Ajustar se necessário para garantir que base + teste <= total
        if tamanho_base + tamanho_teste > total_periodos:
            tamanho_teste = total_periodos - tamanho_base

        print(f"Divisao dos dados (50/25/25):", flush=True)
        print(f"  - Base historica: {tamanho_base} periodos (50%)", flush=True)
        print(f"  - Teste/validacao: {tamanho_teste} periodos (25%)", flush=True)
        print(f"  - Validacao futura (YoY): {tamanho_validacao_futura} periodos (25%)", flush=True)
        print(f"  - Previsao para alem dos dados: {periodos_previsao} periodos", flush=True)

        # IMPORTANTE: Guardar série ORIGINAL para mostrar histórico real ao usuário
        # Outliers só devem ser removidos para CÁLCULO da previsão, não para mostrar dados históricos
        serie_temporal_original = serie_temporal_completa.copy()
        datas_completas_original = datas_completas.copy()

        # Dividir os dados ORIGINAIS (para mostrar ao usuário no gráfico)
        serie_base_original = serie_temporal_original[:tamanho_base]
        serie_teste_original = serie_temporal_original[tamanho_base:tamanho_base + tamanho_teste]

        # IMPORTANTE: Preservar datas ORIGINAIS para exibição ao usuário
        # Estas NÃO devem ser afetadas pela remoção de outliers
        datas_base_original = datas_completas_original[:tamanho_base]
        datas_teste_original = datas_completas_original[tamanho_base:tamanho_base + tamanho_teste]

        # Datas para cálculo (podem ser alteradas se outliers forem removidos)
        datas_base = datas_completas_original[:tamanho_base]
        datas_teste = datas_completas_original[tamanho_base:tamanho_base + tamanho_teste]

        # Calcular índices sazonais para melhorar previsão
        print(f"\nCalculando indices sazonais...", flush=True)
        serie_completa_array = np.array(serie_temporal_completa)

        # Usar módulo robusto de detecção de outliers (IQR ao invés de Z-Score)
        resultado_outliers = auto_clean_outliers(serie_completa_array)
        serie_temporal_completa_limpa = resultado_outliers['cleaned_data']  # Série LIMPA para previsão

        if resultado_outliers['outliers_count'] > 0:
            print(f"  Outliers detectados: {resultado_outliers['outliers_count']}", flush=True)
            print(f"  Método usado: {resultado_outliers['method_used']}", flush=True)
            print(f"  Tratamento: {resultado_outliers['treatment']}", flush=True)
            print(f"  Razão: {resultado_outliers['reason']}", flush=True)
            print(f"  Confiança: {resultado_outliers['confidence']:.2f}", flush=True)

        # Usar série LIMPA para cálculo de previsão
        # NOTA: Se outliers foram REMOVIDOS, a série limpa é mais curta!
        # Precisamos recalcular os tamanhos de base/teste E ajustar as datas
        if resultado_outliers['treatment'] == 'REMOVE' and resultado_outliers['outliers_count'] > 0:
            total_periodos_limpo = len(serie_temporal_completa_limpa)
            tamanho_base_limpo = int(total_periodos_limpo * 0.50)
            tamanho_teste_limpo = int(total_periodos_limpo * 0.25)

            serie_base = serie_temporal_completa_limpa[:tamanho_base_limpo]
            serie_teste = serie_temporal_completa_limpa[tamanho_base_limpo:tamanho_base_limpo + tamanho_teste_limpo]

            # CRÍTICO: Atualizar tamanho_teste para usar o valor limpo
            # Isso garante que modelo.predict() gere o mesmo número de previsões que serie_teste
            tamanho_teste = tamanho_teste_limpo

            # Usar série limpa para cálculos
            serie_temporal_completa = serie_temporal_completa_limpa

            # CRÍTICO: Remover as datas correspondentes aos outliers removidos
            # para manter alinhamento entre datas e valores
            outlier_indices_set = set(resultado_outliers['outliers_detected'])  # Converter para set para lookup rápido
            datas_completas_limpas = [datas_completas_original[i] for i in range(len(datas_completas_original)) if i not in outlier_indices_set]
            datas_completas = datas_completas_limpas

            # Também atualizar datas_base para o loop de ajuste sazonal
            datas_base = datas_completas[:tamanho_base_limpo]

            print(f"  Série ajustada após remoção: {total_periodos_limpo} períodos (era {total_periodos})", flush=True)
            print(f"  Datas ajustadas: {len(datas_completas)} datas (removidos {len(outlier_indices_set)} outliers)", flush=True)
        else:
            # Se substituiu ou não removeu, série tem mesmo tamanho
            serie_temporal_completa = serie_temporal_completa_limpa
            serie_base = serie_temporal_completa[:tamanho_base]
            serie_teste = serie_temporal_completa[tamanho_base:tamanho_base + tamanho_teste]

        # Detectar sazonalidade automaticamente com validação estatística
        print(f"\nDetectando sazonalidade automaticamente...", flush=True)
        resultado_sazonalidade = detect_seasonality(serie_temporal_completa)

        tem_sazonalidade = resultado_sazonalidade['has_seasonality']
        periodo_sazonal = resultado_sazonalidade['seasonal_period']
        forca_sazonal = resultado_sazonalidade['strength']

        print(f"  Sazonalidade detectada: {tem_sazonalidade}", flush=True)
        if tem_sazonalidade:
            print(f"  Período sazonal: {periodo_sazonal}", flush=True)
            print(f"  Força: {forca_sazonal:.2f}", flush=True)
            print(f"  Razão: {resultado_sazonalidade['reason']}", flush=True)
        else:
            print(f"  Razão: {resultado_sazonalidade['reason']}", flush=True)

        # Calcular fatores sazonais com base na GRANULARIDADE, não apenas na detecção automática
        # Isso garante que previsões mensais tenham variação mês a mês mesmo quando a detecção
        # estatística não encontra padrão significativo
        fatores_sazonais = {}

        if granularidade == 'mensal' and len(serie_temporal_completa) >= 12:
            # Calcular fatores sazonais mensais usando TODOS os dados históricos limpos
            # Usar o tamanho mínimo entre datas e série para evitar index out of range
            indices_sazonais = {}
            tamanho_para_sazonalidade = min(len(datas_completas), len(serie_temporal_completa))
            for i in range(tamanho_para_sazonalidade):
                mes = datas_completas[i].month
                if mes not in indices_sazonais:
                    indices_sazonais[mes] = []
                indices_sazonais[mes].append(serie_temporal_completa[i])

            media_geral = np.mean(serie_temporal_completa)
            for mes, valores in indices_sazonais.items():
                fatores_sazonais[mes] = np.mean(valores) / media_geral

            print(f"  Fatores sazonais mensais calculados: {len(fatores_sazonais)} meses", flush=True)
            print(f"  Valores dos fatores mensais: {dict(sorted(fatores_sazonais.items()))}", flush=True)

        elif granularidade == 'semanal' and len(serie_temporal_completa) >= 52:
            # Calcular fatores sazonais semanais usando SEMANA DO ANO (1-52)
            # Assim como mensal usa 12 meses, semanal deve usar 52 semanas para capturar sazonalidade anual
            # Exemplo: Semana 50 da previsão é influenciada pela média histórica das semanas 50
            indices_sazonais = {}
            tamanho_para_sazonalidade = min(len(datas_completas), len(serie_temporal_completa))

            for i in range(tamanho_para_sazonalidade):
                # Usar número da semana no ano (1-52/53)
                semana_ano = datas_completas[i].isocalendar()[1]  # Semana ISO do ano (1-52)

                if semana_ano not in indices_sazonais:
                    indices_sazonais[semana_ano] = []
                indices_sazonais[semana_ano].append(serie_temporal_completa[i])

            media_geral = np.mean(serie_temporal_completa)
            for semana, valores in indices_sazonais.items():
                fatores_sazonais[semana] = np.mean(valores) / media_geral

            print(f"  Fatores sazonais semanais calculados: {len(fatores_sazonais)} semanas do ano", flush=True)
            print(f"  Valores dos fatores semanais: Min={min(fatores_sazonais.values()):.3f}, Max={max(fatores_sazonais.values()):.3f}, Amplitude={(max(fatores_sazonais.values())-min(fatores_sazonais.values()))*100:.2f}%", flush=True)

        elif granularidade == 'diario' and len(serie_temporal_completa) >= 7:
            # Calcular fatores sazonais diários (dia da semana)
            indices_sazonais = {}
            tamanho_para_sazonalidade = min(len(datas_completas), len(serie_temporal_completa))
            for i in range(tamanho_para_sazonalidade):
                dia_semana = datas_completas[i].weekday()  # 0=segunda, 6=domingo
                if dia_semana not in indices_sazonais:
                    indices_sazonais[dia_semana] = []
                indices_sazonais[dia_semana].append(serie_temporal_completa[i])

            media_geral = np.mean(serie_temporal_completa)
            for dia, valores in indices_sazonais.items():
                fatores_sazonais[dia] = np.mean(valores) / media_geral

            print(f"  Fatores sazonais diários calculados: {len(fatores_sazonais)} dias da semana", flush=True)
        else:
            print(f"  Sem ajuste sazonal (dados insuficientes para granularidade {granularidade})", flush=True)

        # Aplicar todos os 6 modelos principais de previsão da ferramenta
        modelos_testar = [
            'Média Móvel Simples',
            'Média Móvel Ponderada',
            'Suavização Exponencial Simples',
            'Holt',
            'Holt-Winters',
            'Regressão Linear'
        ]

        resultados_modelos = {}
        metricas = {}
        melhor_modelo = None
        melhor_wmape = float('inf')

        print(f"\nTreinando e testando modelos...", flush=True)

        for nome_modelo in modelos_testar:
            try:
                print(f"Aplicando modelo: {nome_modelo}", flush=True)
                modelo = get_modelo(nome_modelo)

                # Treinar com dados base
                modelo.fit(serie_base)

                # Prever para o período de teste
                previsoes_teste_base = modelo.predict(tamanho_teste)

                # APLICAR AJUSTE SAZONAL nas previsões de teste baseado na GRANULARIDADE
                if fatores_sazonais:
                    previsoes_teste = []
                    # A última data da base é datas_base[-1]
                    ultima_data_base = datas_base[-1]

                    for i in range(tamanho_teste):
                        # Calcular a chave sazonal com base na granularidade
                        if granularidade == 'mensal':
                            # Para mensal, usar o número do mês (1-12)
                            # CORREÇÃO: O primeiro período de teste é o mês SEGUINTE ao último da base
                            mes_previsao = ((ultima_data_base.month + i) % 12) + 1
                            chave_sazonal = mes_previsao
                        elif granularidade == 'semanal':
                            # Para semanal, usar SEMANA DO ANO (1-52)
                            from datetime import timedelta
                            data_previsao = ultima_data_base + timedelta(weeks=i)
                            semana_ano = data_previsao.isocalendar()[1]  # 1-52/53
                            chave_sazonal = semana_ano  # Usar semana do ano diretamente
                        elif granularidade == 'diario':
                            # Para diário, usar dia da semana (0-6)
                            # CORREÇÃO: Usar i+1 porque o primeiro período de teste é o dia SEGUINTE à última data base
                            from datetime import timedelta
                            data_previsao = ultima_data_base + timedelta(days=i+1)
                            chave_sazonal = data_previsao.weekday()
                        else:
                            chave_sazonal = None

                        # Aplicar fator sazonal
                        fator = fatores_sazonais.get(chave_sazonal, 1.0)
                        valor_ajustado = previsoes_teste_base[i] * fator
                        previsoes_teste.append(valor_ajustado)
                else:
                    previsoes_teste = previsoes_teste_base

                # Calcular métricas no período de teste
                if len(previsoes_teste) == len(serie_teste):
                    erros_absolutos = []
                    erros_percentuais = []
                    erros_reais = []

                    for i in range(len(serie_teste)):
                        real = serie_teste[i]
                        previsto = previsoes_teste[i]

                        if real > 0:
                            erro_abs = abs(previsto - real)
                            erro_perc = (erro_abs / real) * 100
                            erro_real = previsto - real

                            erros_absolutos.append(erro_abs)
                            erros_percentuais.append(erro_perc)
                            erros_reais.append(erro_real)

                    if len(erros_percentuais) > 0:
                        wmape = sum(erros_percentuais) / len(erros_percentuais)
                        bias = (sum(erros_reais) / sum(serie_teste)) * 100 if sum(serie_teste) > 0 else 0
                        mae = sum(erros_absolutos) / len(erros_absolutos)

                        metricas[nome_modelo] = {
                            'wmape': wmape,
                            'bias': bias,
                            'mae': mae
                        }

                        print(f"  OK {nome_modelo}: WMAPE={wmape:.2f}%, BIAS={bias:.2f}%, MAE={mae:.2f}", flush=True)

                        if wmape < melhor_wmape:
                            melhor_wmape = wmape
                            melhor_modelo = nome_modelo
                    else:
                        metricas[nome_modelo] = {'wmape': 0.0, 'bias': 0.0, 'mae': 0.0}
                        print(f"  ! {nome_modelo}: Sem erros calculados", flush=True)
                else:
                    metricas[nome_modelo] = {'wmape': 0.0, 'bias': 0.0, 'mae': 0.0}
                    print(f"  ! {nome_modelo}: Tamanho incompatível (previsto={len(previsoes_teste)}, esperado={len(serie_teste)})", flush=True)

                # Agora treinar com TODOS os dados (base + teste) para fazer previsão futura
                modelo.fit(serie_temporal_completa)
                # Prever para o número de períodos solicitado pelo usuário
                previsoes_futuro_base = modelo.predict(periodos_previsao)

                # Log da previsão base (antes do ajuste sazonal)
                total_base = sum(previsoes_futuro_base)
                media_base = total_base / len(previsoes_futuro_base)
                print(f"    Previsão base (sem ajuste): Total={total_base:,.2f}, Média={media_base:,.2f}", flush=True)

                # APLICAR AJUSTE SAZONAL nas previsões futuras baseado na GRANULARIDADE
                if fatores_sazonais:
                    # Aplicar fatores sazonais nas previsões futuras
                    previsoes_futuro = []
                    # IMPORTANTE: Usar ultima_data_historico (data original antes de remoção de outliers)
                    # para garantir alinhamento com as datas de previsão geradas
                    ultima_data = ultima_data_historico

                    for i in range(periodos_previsao):
                        # Calcular a chave sazonal com base na granularidade
                        if granularidade == 'mensal':
                            # Para mensal, usar o número do mês (1-12)
                            # CORREÇÃO: O primeiro período de previsão é o mês SEGUINTE ao último histórico
                            # Por isso usamos (ultima_data.month + i + 1) ao invés de (ultima_data.month + i)
                            mes_previsao = ((ultima_data.month + i) % 12) + 1
                            chave_sazonal = mes_previsao
                        elif granularidade == 'semanal':
                            # Para semanal, usar SEMANA DO ANO (1-52)
                            from datetime import timedelta
                            data_previsao = ultima_data + timedelta(weeks=i)
                            semana_ano = data_previsao.isocalendar()[1]  # 1-52/53
                            chave_sazonal = semana_ano  # Usar semana do ano diretamente
                        elif granularidade == 'diario':
                            # Para diário, usar dia da semana (0-6)
                            # CORREÇÃO: Usar i+1 porque o primeiro período de previsão é o dia SEGUINTE ao último histórico
                            from datetime import timedelta
                            data_previsao = ultima_data + timedelta(days=i+1)
                            chave_sazonal = data_previsao.weekday()
                        else:
                            chave_sazonal = None

                        # Aplicar fator sazonal
                        fator = fatores_sazonais.get(chave_sazonal, 1.0)
                        valor_ajustado = previsoes_futuro_base[i] * fator
                        previsoes_futuro.append(valor_ajustado)

                    if tem_sazonalidade:
                        print(f"    Ajuste sazonal aplicado (força: {forca_sazonal:.2f}, var: {min(fatores_sazonais.values()):.3f} - {max(fatores_sazonais.values()):.3f})", flush=True)
                    else:
                        print(f"    Ajuste sazonal aplicado baseado em padrões históricos (var: {min(fatores_sazonais.values()):.3f} - {max(fatores_sazonais.values()):.3f})", flush=True)
                else:
                    previsoes_futuro = previsoes_futuro_base
                    print(f"    Sem ajuste sazonal (dados insuficientes)", flush=True)

                resultados_modelos[nome_modelo] = {
                    'previsao_teste': previsoes_teste[:len(serie_teste)],
                    'previsao_futuro': previsoes_futuro
                }

                # Log detalhado do total previsto
                total_previsto = sum(previsoes_futuro)
                print(f"    Total previsto para {periodos_previsao} períodos: {total_previsto:,.2f}", flush=True)

            except Exception as e:
                print(f"  ERRO ao aplicar modelo {nome_modelo}: {e}", flush=True)
                import traceback
                print(traceback.format_exc(), flush=True)
                continue

        # Se nenhum modelo funcionou, usar fallback
        if not resultados_modelos or melhor_modelo is None:
            print("Nenhum modelo funcionou, usando fallback...", flush=True)
            modelo = get_modelo('Média Móvel Simples')
            modelo.fit(serie_base)
            previsoes_teste_base = modelo.predict(tamanho_teste)
            modelo.fit(serie_temporal_completa)
            previsoes_futuro_base = modelo.predict(periodos_previsao)

            # IMPORTANTE: Aplicar fatores sazonais também no fallback!
            if fatores_sazonais:
                # Aplicar sazonalidade nas previsões de TESTE
                previsoes_teste = []
                # Data inicial do período de teste
                data_inicio_teste = datas_base[-1] if len(datas_base) > 0 else datas_completas[tamanho_base - 1]

                for i in range(len(previsoes_teste_base)):
                    if granularidade == 'mensal':
                        # CORREÇÃO: O primeiro período de teste é o mês SEGUINTE ao último da base
                        mes_teste = ((data_inicio_teste.month + i) % 12) + 1
                        chave_sazonal = mes_teste
                    elif granularidade == 'semanal':
                        semana_base = data_inicio_teste.isocalendar()[1]
                        semana_teste = ((semana_base + i - 1) % 52) + 1
                        chave_sazonal = semana_teste
                    elif granularidade == 'diario':
                        # CORREÇÃO: Usar i+1 porque o primeiro período de teste é o dia SEGUINTE à data_inicio_teste
                        from datetime import timedelta
                        data_previsao = data_inicio_teste + timedelta(days=i+1)
                        chave_sazonal = data_previsao.weekday()
                    else:
                        chave_sazonal = None

                    fator = fatores_sazonais.get(chave_sazonal, 1.0)
                    valor_ajustado = previsoes_teste_base[i] * fator
                    previsoes_teste.append(valor_ajustado)

                # Aplicar sazonalidade nas previsões FUTURAS
                previsoes_futuro = []
                # IMPORTANTE: Usar ultima_data_historico (data original antes de remoção de outliers)
                # para garantir alinhamento com as datas de previsão geradas
                ultima_data = ultima_data_historico

                for i in range(len(previsoes_futuro_base)):
                    if granularidade == 'mensal':
                        # CORREÇÃO: O primeiro período de previsão é o mês SEGUINTE ao último histórico
                        mes_futuro = ((ultima_data.month + i) % 12) + 1
                        chave_sazonal = mes_futuro
                    elif granularidade == 'semanal':
                        semana_base = ultima_data.isocalendar()[1]
                        semana_futuro = ((semana_base + i - 1) % 52) + 1
                        chave_sazonal = semana_futuro
                    elif granularidade == 'diario':
                        # CORREÇÃO: Usar i+1 porque o primeiro período de previsão é o dia SEGUINTE ao último histórico
                        from datetime import timedelta
                        data_previsao = ultima_data + timedelta(days=i+1)
                        chave_sazonal = data_previsao.weekday()
                    else:
                        chave_sazonal = None

                    fator = fatores_sazonais.get(chave_sazonal, 1.0)
                    valor_ajustado = previsoes_futuro_base[i] * fator
                    previsoes_futuro.append(valor_ajustado)

                print(f"  Fallback com ajuste sazonal aplicado (var: {min(fatores_sazonais.values()):.3f} - {max(fatores_sazonais.values()):.3f})", flush=True)
            else:
                previsoes_teste = previsoes_teste_base
                previsoes_futuro = previsoes_futuro_base
                print(f"  Fallback sem ajuste sazonal (dados insuficientes)", flush=True)

            resultados_modelos = {
                'Média Móvel Simples': {
                    'previsao_teste': previsoes_teste,
                    'previsao_futuro': previsoes_futuro
                }
            }

            # Calcular métricas WMAPE e BIAS para o fallback
            wmape_fallback = 0
            bias_fallback = 0
            mae_fallback = 0

            if len(previsoes_teste) > 0 and len(serie_teste) > 0:
                erros_absolutos = []
                erros_percentuais = []
                erros_reais = []

                tamanho_comparacao = min(len(previsoes_teste), len(serie_teste))
                for i in range(tamanho_comparacao):
                    real = serie_teste[i]
                    previsto = previsoes_teste[i]

                    if real > 0:
                        erro_abs = abs(previsto - real)
                        erro_perc = (erro_abs / real) * 100
                        erro_real = previsto - real

                        erros_absolutos.append(erro_abs)
                        erros_percentuais.append(erro_perc)
                        erros_reais.append(erro_real)

                if len(erros_percentuais) > 0:
                    wmape_fallback = sum(erros_percentuais) / len(erros_percentuais)
                    soma_reais = sum(serie_teste[:tamanho_comparacao])
                    bias_fallback = (sum(erros_reais) / soma_reais) * 100 if soma_reais > 0 else 0
                    mae_fallback = sum(erros_absolutos) / len(erros_absolutos)

                print(f"  Fallback métricas: WMAPE={wmape_fallback:.2f}%, BIAS={bias_fallback:.2f}%", flush=True)

            metricas = {'Média Móvel Simples': {'wmape': wmape_fallback, 'bias': bias_fallback, 'mae': mae_fallback}}
            melhor_modelo = 'Média Móvel Simples'
            melhor_wmape = wmape_fallback

        print(f"\nMelhor modelo: {melhor_modelo}", flush=True)

        # Registrar seleção do modelo no auto-logger para auditoria
        if melhor_modelo and melhor_modelo in metricas:
            # Extrair características da série temporal
            caracteristicas = {
                'tamanho': total_periodos,
                'media': float(np.mean(serie_temporal_completa)),
                'desvio_padrao': float(np.std(serie_temporal_completa)),
                'coef_variacao': float(np.std(serie_temporal_completa) / np.mean(serie_temporal_completa)),
                'sazonalidade': tem_sazonalidade,
                'periodo_sazonal': periodo_sazonal if tem_sazonalidade else None,
                'forca_sazonal': forca_sazonal if tem_sazonalidade else 0.0,
                'outliers_detectados': resultado_outliers.get('outliers_count', 0),
                'metodo_outlier': resultado_outliers.get('method_used', 'none')
            }

            logger.log_selection(
                metodo_selecionado=melhor_modelo,
                confianca=1.0 - (melhor_wmape / 100.0) if melhor_wmape < 100 else 0.0,  # Converter WMAPE em confiança
                razao=f"Menor WMAPE: {melhor_wmape:.2f}%",
                caracteristicas=caracteristicas,
                sku=produto if produto != 'TODOS' else None,
                loja=loja if loja != 'TODAS' else None,
                horizonte=meses_previsao,
                sucesso=True
            )
            print(f"  Seleção registrada no auto-logger", flush=True)

        # Preparar resposta com 3 séries de dados
        # IMPORTANTE: Usar série ORIGINAL para mostrar histórico ao usuário
        # Outliers só devem ser removidos para CÁLCULO da previsão, não para exibição
        resposta = {
            'sucesso': True,
            'melhor_modelo': melhor_modelo,
            'metricas': metricas,
            'modelos': {},
            'granularidade': granularidade,  # Adicionar granularidade para formatação de labels
            'historico_base': {
                'datas': [d.strftime('%Y-%m-%d') for d in datas_base_original],  # Usar datas ORIGINAIS
                'valores': serie_base_original  # Usar série ORIGINAL (sem remoção de outliers)
            },
            'historico_teste': {
                'datas': [d.strftime('%Y-%m-%d') for d in datas_teste_original],  # Usar datas ORIGINAIS
                'valores': serie_teste_original  # Usar série ORIGINAL (sem remoção de outliers)
            },
            'ano_anterior': {
                'datas': ano_anterior_datas,
                'valores': ano_anterior_valores,
                'total': ano_anterior_total
            }
        }

        # IMPORTANTE: Para granularidade semanal, realinhar dados do ano anterior
        # para corresponder às mesmas semanas ISO das previsões
        # Isso será feito após gerar as datas futuras da primeira previsão

        # Adicionar previsões de cada modelo
        metricas_futuro = {}

        # Verificar se há dados para gerar datas futuras
        if len(datas_completas) == 0:
            return jsonify({'erro': 'Não há dados históricos suficientes para gerar previsão'}), 400

        for nome_modelo, resultado in resultados_modelos.items():
            # Gerar datas futuras baseado na granularidade e no período solicitado
            # Suporta período customizado quando fornecido pelo usuário
            ultima_data = datas_completas[-1]

            if granularidade == 'diario':
                # Para diário: usar período customizado ou calcular baseado no histórico
                if periodo_inicio_customizado:
                    start_date = periodo_inicio_customizado
                else:
                    start_date = ultima_data + pd.Timedelta(days=1)

                datas_futuras = pd.date_range(
                    start=start_date,
                    end=data_fim_previsao,
                    freq='D'
                )
            elif granularidade == 'semanal':
                # Para semanal: usar período customizado ou calcular baseado no histórico
                if periodo_inicio_customizado:
                    start_date = periodo_inicio_customizado
                else:
                    start_date = ultima_data + pd.Timedelta(weeks=1)

                datas_futuras = pd.date_range(
                    start=start_date,
                    end=data_fim_previsao,
                    freq='W-MON'  # Semanas começando na segunda-feira
                )
            else:  # mensal
                # Para mensal: usar período customizado ou calcular baseado no histórico
                if periodo_inicio_customizado:
                    data_inicio_mensal = pd.Timestamp(year=ano_primeiro_mes, month=primeiro_mes_previsao, day=1)
                    data_fim_mensal = pd.Timestamp(year=ano_destino, month=mes_destino, day=1)
                else:
                    # Original: primeiro período é o mês SEGUINTE ao último histórico
                    mes_inicio_mensal = primeiro_mes_previsao + 1
                    ano_inicio_mensal = ano_primeiro_mes
                    if mes_inicio_mensal > 12:
                        mes_inicio_mensal = 1
                        ano_inicio_mensal += 1

                    data_inicio_mensal = pd.Timestamp(year=ano_inicio_mensal, month=mes_inicio_mensal, day=1)
                    data_fim_mensal = pd.Timestamp(year=ano_destino, month=mes_destino, day=1)

                datas_futuras = pd.date_range(
                    start=data_inicio_mensal,
                    end=data_fim_mensal,
                    freq='MS'
                )

            # Ajustar periodos_previsao para corresponder às datas geradas
            periodos_previsao_efetivo = len(datas_futuras)

            # REALINHAMENTO DO ANO ANTERIOR para corresponder às datas futuras
            # Isso é especialmente importante para granularidade semanal onde
            # a mesma semana ISO pode cair em datas diferentes em anos diferentes
            if granularidade == 'semanal' and len(ano_anterior_datas) > 0 and len(datas_futuras) > 0:
                # Criar dicionário de valores do ano anterior indexado por (ano_iso, semana_iso)
                # Isso evita que dados de anos diferentes se sobrescrevam
                ano_anterior_por_semana = {}
                for i, data_str in enumerate(ano_anterior_datas):
                    if data_str is None:
                        continue
                    data_ant = pd.Timestamp(data_str)
                    if pd.isna(data_ant):
                        continue
                    ano_iso = data_ant.isocalendar()[0]
                    semana_iso = data_ant.isocalendar()[1]
                    # Usar tupla (ano, semana) como chave
                    chave = (ano_iso, semana_iso)
                    ano_anterior_por_semana[chave] = {
                        'data': data_str,
                        'valor': ano_anterior_valores[i]
                    }

                # Realinhar para corresponder às datas futuras
                ano_anterior_datas_realinhado = []
                ano_anterior_valores_realinhado = []

                for data_futura in datas_futuras[:periodos_previsao_efetivo]:
                    ano_futura = data_futura.isocalendar()[0]
                    semana_futura = data_futura.isocalendar()[1]
                    # Buscar pelo ano anterior (ano da previsão - 1) e mesma semana
                    ano_busca = ano_futura - 1
                    chave_busca = (ano_busca, semana_futura)

                    if chave_busca in ano_anterior_por_semana:
                        ano_anterior_datas_realinhado.append(ano_anterior_por_semana[chave_busca]['data'])
                        ano_anterior_valores_realinhado.append(ano_anterior_por_semana[chave_busca]['valor'])
                    else:
                        # Se não encontrar a semana (raro), usar None/0
                        # Isso pode acontecer com semana 53 que não existe em todos os anos
                        ano_anterior_datas_realinhado.append(None)
                        ano_anterior_valores_realinhado.append(0)

                # Atualizar resposta com dados realinhados
                resposta['ano_anterior']['datas'] = ano_anterior_datas_realinhado
                resposta['ano_anterior']['valores'] = ano_anterior_valores_realinhado
                # Total permanece o mesmo (soma de todos os valores do período)
                print(f"  Ano anterior realinhado: {len(ano_anterior_valores_realinhado)} semanas correspondentes", flush=True)

                # Atualizar variáveis locais também para cálculo de métricas
                ano_anterior_valores = ano_anterior_valores_realinhado
                ano_anterior_datas = ano_anterior_datas_realinhado

            elif granularidade == 'mensal' and len(ano_anterior_datas) > 0 and len(datas_futuras) > 0:
                # Para mensal, realinhar por mês
                ano_anterior_por_mes = {}
                for i, data_str in enumerate(ano_anterior_datas):
                    if data_str is None:
                        continue
                    data_ant = pd.Timestamp(data_str)
                    if pd.isna(data_ant):
                        continue
                    mes = data_ant.month
                    ano_anterior_por_mes[mes] = {
                        'data': data_str,
                        'valor': ano_anterior_valores[i]
                    }

                ano_anterior_datas_realinhado = []
                ano_anterior_valores_realinhado = []

                for data_futura in datas_futuras[:periodos_previsao_efetivo]:
                    mes_futuro = data_futura.month
                    if mes_futuro in ano_anterior_por_mes:
                        ano_anterior_datas_realinhado.append(ano_anterior_por_mes[mes_futuro]['data'])
                        ano_anterior_valores_realinhado.append(ano_anterior_por_mes[mes_futuro]['valor'])
                    else:
                        ano_anterior_datas_realinhado.append(None)
                        ano_anterior_valores_realinhado.append(0)

                resposta['ano_anterior']['datas'] = ano_anterior_datas_realinhado
                resposta['ano_anterior']['valores'] = ano_anterior_valores_realinhado
                print(f"  Ano anterior realinhado: {len(ano_anterior_valores_realinhado)} meses correspondentes", flush=True)

                ano_anterior_valores = ano_anterior_valores_realinhado
                ano_anterior_datas = ano_anterior_datas_realinhado

            # Estruturar dados do modelo com teste e previsão futura
            modelo_data = {}

            # Adicionar previsões do período de teste
            if 'previsao_teste' in resultado and len(resultado['previsao_teste']) > 0:
                modelo_data['teste'] = {
                    'datas': [d.strftime('%Y-%m-%d') for d in datas_teste],
                    'valores': [float(v) for v in resultado['previsao_teste']]
                }

            # Adicionar previsões futuras
            if 'previsao_futuro' in resultado and len(resultado['previsao_futuro']) > 0:
                modelo_data['futuro'] = {
                    'datas': [d.strftime('%Y-%m-%d') for d in datas_futuras[:len(resultado['previsao_futuro'])]],
                    'valores': [float(v) for v in resultado['previsao_futuro']]
                }

                # Calcular WMAPE e BIAS para previsão futura comparando com ano anterior
                # Usar apenas os primeiros períodos correspondentes à validação futura
                previsao_futuro = resultado['previsao_futuro']
                if len(ano_anterior_valores) > 0 and len(previsao_futuro) > 0:
                    # Comparar apenas os períodos que temos no ano anterior (tamanho_validacao_futura)
                    tamanho_comparacao = min(tamanho_validacao_futura, len(ano_anterior_valores), len(previsao_futuro))

                    erros_absolutos_futuro = []
                    erros_percentuais_futuro = []
                    erros_reais_futuro = []

                    for i in range(tamanho_comparacao):
                        real = ano_anterior_valores[i]
                        previsto = previsao_futuro[i]

                        if real > 0:
                            erro_abs = abs(previsto - real)
                            erro_perc = (erro_abs / real) * 100
                            erro_real = previsto - real

                            erros_absolutos_futuro.append(erro_abs)
                            erros_percentuais_futuro.append(erro_perc)
                            erros_reais_futuro.append(erro_real)

                    if len(erros_percentuais_futuro) > 0:
                        wmape_futuro = sum(erros_percentuais_futuro) / len(erros_percentuais_futuro)
                        bias_futuro = (sum(erros_reais_futuro) / sum(ano_anterior_valores[:tamanho_comparacao])) * 100
                        mae_futuro = sum(erros_absolutos_futuro) / len(erros_absolutos_futuro)

                        metricas_futuro[nome_modelo] = {
                            'wmape': wmape_futuro,
                            'bias': bias_futuro,
                            'mae': mae_futuro
                        }

                        print(f"  Métricas futuras {nome_modelo}: WMAPE={wmape_futuro:.2f}%, BIAS={bias_futuro:.2f}%", flush=True)

            resposta['modelos'][nome_modelo] = modelo_data

        # Adicionar métricas futuras à resposta (comparação com ano anterior)
        resposta['metricas_futuro'] = metricas_futuro

        # Gerar alertas inteligentes baseado na previsão
        print(f"\nGerando alertas inteligentes...", flush=True)
        if melhor_modelo and melhor_modelo in resultados_modelos:
            modelo_info = {
                'modelo': melhor_modelo,
                'wmape': metricas.get(melhor_modelo, {}).get('wmape', 0),
                'bias': metricas.get(melhor_modelo, {}).get('bias', 0),
                'mae': metricas.get(melhor_modelo, {}).get('mae', 0),
                'seasonality_detected': {
                    'has_seasonality': tem_sazonalidade,
                    'detected_period': periodo_sazonal if tem_sazonalidade else None,
                    'strength': forca_sazonal if tem_sazonalidade else 0.0
                },
                'outliers_count': resultado_outliers.get('outliers_count', 0)
            }

            resultado_alertas = generate_alerts_for_forecast(
                sku=produto if produto != 'TODOS' else 'AGREGADO',
                loja=loja if loja != 'TODAS' else 'TODAS',
                historico=serie_temporal_completa,
                previsao=resultados_modelos[melhor_modelo].get('previsao_futuro', []),
                modelo_params=modelo_info,
                estoque_atual=None,  # Não disponível neste endpoint
                lead_time_dias=7,
                custo_unitario=None  # Não disponível neste endpoint
            )

            alertas = resultado_alertas['alertas']
            resumo_alertas = resultado_alertas['resumo']

            resposta['alertas'] = alertas
            resposta['resumo_alertas'] = resumo_alertas
            print(f"  {len(alertas)} alertas gerados", flush=True)

            # Resumo por prioridade
            criticos = sum(1 for a in alertas if a.get('tipo') == 'CRITICAL')
            avisos = sum(1 for a in alertas if a.get('tipo') == 'WARNING')
            info = sum(1 for a in alertas if a.get('tipo') == 'INFO')

            if criticos > 0:
                print(f"  [CRITICO] {criticos} alertas criticos", flush=True)
            if avisos > 0:
                print(f"  [AVISO] {avisos} avisos", flush=True)
            if info > 0:
                print(f"  [INFO] {info} informativos", flush=True)
        else:
            resposta['alertas'] = []

        # =====================================================
        # RELATÓRIO DETALHADO POR ITEM/FORNECEDOR
        # =====================================================
        print(f"\nGerando relatório detalhado por item/fornecedor...", flush=True)

        try:
            # Criar nova conexão para o relatório (a anterior foi fechada)
            conn_rel = psycopg2.connect(**DB_CONFIG)
            cursor_rel = conn_rel.cursor(cursor_factory=RealDictCursor)

            # Buscar lista de itens com seus fornecedores
            # Nota: h.codigo é integer, p.cod_produto é varchar - precisa cast
            query_itens = f"""
                SELECT DISTINCT
                    h.codigo as cod_produto,
                    p.descricao,
                    p.nome_fornecedor,
                    p.cnpj_fornecedor,
                    p.categoria,
                    p.descricao_linha
                FROM historico_vendas_diario h
                JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
                WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                {where_sql}
                ORDER BY p.nome_fornecedor, p.descricao
            """
            cursor_rel.execute(query_itens, params)
            lista_itens = cursor_rel.fetchall()

            if lista_itens and len(lista_itens) > 0:
                print(f"  Encontrados {len(lista_itens)} itens para o relatório", flush=True)

                # Buscar vendas por item e período
                if granularidade == 'mensal':
                    query_vendas_item = f"""
                        SELECT
                            h.codigo as cod_produto,
                            DATE_TRUNC('month', h.data) as periodo,
                            SUM(h.qtd_venda) as qtd_venda
                        FROM historico_vendas_diario h
                        JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
                        WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                        {where_sql}
                        GROUP BY h.codigo, DATE_TRUNC('month', h.data)
                        ORDER BY h.codigo, DATE_TRUNC('month', h.data)
                    """
                elif granularidade == 'semanal':
                    query_vendas_item = f"""
                        SELECT
                            h.codigo as cod_produto,
                            DATE_TRUNC('week', h.data) as periodo,
                            SUM(h.qtd_venda) as qtd_venda
                        FROM historico_vendas_diario h
                        JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
                        WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                        {where_sql}
                        GROUP BY h.codigo, DATE_TRUNC('week', h.data)
                        ORDER BY h.codigo, DATE_TRUNC('week', h.data)
                    """
                else:  # diario
                    query_vendas_item = f"""
                        SELECT
                            h.codigo as cod_produto,
                            h.data as periodo,
                            SUM(h.qtd_venda) as qtd_venda
                        FROM historico_vendas_diario h
                        JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
                        WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                        {where_sql}
                        GROUP BY h.codigo, h.data
                        ORDER BY h.codigo, h.data
                    """

                cursor_rel.execute(query_vendas_item, params)
                vendas_por_item = cursor_rel.fetchall()

                # Organizar vendas por item
                vendas_item_dict = {}
                for venda in vendas_por_item:
                    cod = venda['cod_produto']
                    if cod not in vendas_item_dict:
                        vendas_item_dict[cod] = []
                    vendas_item_dict[cod].append({
                        'periodo': venda['periodo'].strftime('%Y-%m-%d') if venda['periodo'] else None,
                        'qtd_venda': float(venda['qtd_venda']) if venda['qtd_venda'] else 0
                    })

                # Calcular previsão para cada item com seleção INDIVIDUAL de modelo
                relatorio_itens = []
                datas_previsao = resposta.get('modelos', {}).get(melhor_modelo, {}).get('futuro', {}).get('datas', [])

                # Modelos disponíveis para teste individual
                modelos_disponiveis = [
                    'Média Móvel Simples',
                    'Média Móvel Ponderada',
                    'Suavização Exponencial Simples',
                    'Holt',
                    'Holt-Winters',
                    'Regressão Linear'
                ]

                print(f"  Selecionando melhor modelo para cada item individualmente...", flush=True)

                for idx, item in enumerate(lista_itens):
                    cod_produto = item['cod_produto']
                    vendas_hist = vendas_item_dict.get(cod_produto, [])

                    # Calcular total histórico do item
                    total_item_hist = sum(v['qtd_venda'] for v in vendas_hist)

                    # Extrair série temporal do item (ordenada por período)
                    vendas_hist_sorted = sorted(vendas_hist, key=lambda x: x['periodo'] if x['periodo'] else '')
                    serie_item = [v['qtd_venda'] for v in vendas_hist_sorted]

                    # Selecionar melhor modelo para ESTE item
                    melhor_modelo_item = melhor_modelo  # Fallback para modelo agregado
                    previsao_item_periodos = []

                    if len(serie_item) >= 6:  # Mínimo de 6 períodos para testar modelos
                        # Dividir série do item: 70% treino, 30% teste
                        tamanho_treino = int(len(serie_item) * 0.7)
                        if tamanho_treino < 3:
                            tamanho_treino = 3
                        tamanho_teste_item = len(serie_item) - tamanho_treino

                        if tamanho_teste_item >= 2:
                            serie_treino_item = serie_item[:tamanho_treino]
                            serie_teste_item = serie_item[tamanho_treino:]

                            melhor_wmape_item = float('inf')

                            for nome_modelo in modelos_disponiveis:
                                try:
                                    modelo_item = get_modelo(nome_modelo)
                                    modelo_item.fit(serie_treino_item)
                                    previsoes_teste_item = modelo_item.predict(tamanho_teste_item)

                                    # Calcular WMAPE para este modelo neste item
                                    if len(previsoes_teste_item) == len(serie_teste_item):
                                        erros_pct = []
                                        for i in range(len(serie_teste_item)):
                                            if serie_teste_item[i] > 0:
                                                erro = abs(previsoes_teste_item[i] - serie_teste_item[i]) / serie_teste_item[i] * 100
                                                erros_pct.append(erro)

                                        if erros_pct:
                                            wmape_item = sum(erros_pct) / len(erros_pct)
                                            if wmape_item < melhor_wmape_item:
                                                melhor_wmape_item = wmape_item
                                                melhor_modelo_item = nome_modelo
                                except:
                                    pass  # Modelo falhou para este item, tentar próximo

                            # Gerar previsão com o melhor modelo selecionado para este item
                            try:
                                modelo_final = get_modelo(melhor_modelo_item)
                                modelo_final.fit(serie_item)  # Treinar com toda a série do item
                                previsao_item_base = modelo_final.predict(periodos_previsao)

                                # APLICAR AJUSTE SAZONAL nas previsões do item
                                if fatores_sazonais and len(previsao_item_base) > 0:
                                    previsao_item_periodos = []
                                    for i in range(len(previsao_item_base)):
                                        # Calcular chave sazonal baseada na granularidade
                                        if granularidade == 'mensal':
                                            mes_previsao = ((ultima_data_historico.month + i) % 12) + 1
                                            chave_sazonal = mes_previsao
                                        elif granularidade == 'semanal':
                                            from datetime import timedelta
                                            data_prev = ultima_data_historico + timedelta(weeks=i)
                                            chave_sazonal = data_prev.isocalendar()[1]
                                        elif granularidade == 'diario':
                                            from datetime import timedelta
                                            data_prev = ultima_data_historico + timedelta(days=i+1)
                                            chave_sazonal = data_prev.weekday()
                                        else:
                                            chave_sazonal = None

                                        # Aplicar fator sazonal
                                        fator = fatores_sazonais.get(chave_sazonal, 1.0)
                                        valor_ajustado = previsao_item_base[i] * fator
                                        previsao_item_periodos.append(valor_ajustado)
                                else:
                                    previsao_item_periodos = previsao_item_base
                            except:
                                previsao_item_periodos = []

                    # Se não conseguiu gerar previsão individual, usar proporção do agregado
                    # (que já tem sazonalidade aplicada)
                    if not previsao_item_periodos or len(previsao_item_periodos) == 0:
                        total_historico = sum(serie_temporal_original) if serie_temporal_original else 1
                        proporcao_item = total_item_hist / total_historico if total_historico > 0 else 0
                        previsoes_modelo = resultados_modelos.get(melhor_modelo, {}).get('previsao_futuro', [])
                        previsao_item_periodos = [p * proporcao_item for p in previsoes_modelo]
                        melhor_modelo_item = melhor_modelo + " (proporcional)"

                    # Montar previsão por período
                    previsao_por_periodo = []
                    total_previsao_item = 0
                    for i, data_prev in enumerate(datas_previsao):
                        if i < len(previsao_item_periodos):
                            prev_periodo = round(previsao_item_periodos[i], 1)
                            previsao_por_periodo.append({
                                'periodo': data_prev,
                                'previsao': prev_periodo
                            })
                            total_previsao_item += prev_periodo

                    # Buscar demanda REAL do ano anterior para este item
                    # Filtrar vendas do item que estão dentro do período do ano anterior
                    demanda_ano_anterior_item = 0

                    if ano_anterior_datas and len(ano_anterior_datas) > 0:
                        # Determinar período do ano anterior (datas já estão no formato 'YYYY-MM-DD')
                        data_inicio_aa = ano_anterior_datas[0]   # Primeira data do ano anterior
                        data_fim_aa = ano_anterior_datas[-1]     # Última data do ano anterior

                        # Filtrar vendas do item que caem dentro desse período
                        for venda in vendas_hist_sorted:
                            data_venda = venda['periodo']
                            if data_venda and data_inicio_aa <= data_venda <= data_fim_aa:
                                demanda_ano_anterior_item += venda['qtd_venda']

                    demanda_ano_anterior_item = round(demanda_ano_anterior_item, 1)

                    # Calcular variação
                    if demanda_ano_anterior_item > 0:
                        variacao_pct = round(((total_previsao_item - demanda_ano_anterior_item) / demanda_ano_anterior_item) * 100, 1)
                    else:
                        variacao_pct = None

                    # Determinar sinal de alerta baseado na variação
                    if variacao_pct is None:
                        sinal_alerta = 'cinza'
                        sinal_emoji = '⚪'
                    elif abs(variacao_pct) > 50:
                        sinal_alerta = 'vermelho' if variacao_pct < -50 else 'amarelo'
                        sinal_emoji = '🔴' if variacao_pct < -50 else '🟡'
                    elif abs(variacao_pct) > 20:
                        sinal_alerta = 'azul'
                        sinal_emoji = '🔵'
                    else:
                        sinal_alerta = 'verde'
                        sinal_emoji = '🟢'

                    relatorio_itens.append({
                        'cod_produto': cod_produto,
                        'descricao': item['descricao'],
                        'nome_fornecedor': item['nome_fornecedor'] or 'SEM FORNECEDOR',
                        'cnpj_fornecedor': item['cnpj_fornecedor'],
                        'categoria': item['categoria'],
                        'linha': item['descricao_linha'],
                        'demanda_prevista_total': round(total_previsao_item, 1),
                        'demanda_ano_anterior': demanda_ano_anterior_item,
                        'variacao_percentual': variacao_pct,
                        'sinal_alerta': sinal_alerta,
                        'sinal_emoji': sinal_emoji,
                        'metodo_estatistico': melhor_modelo_item,
                        'previsao_por_periodo': previsao_por_periodo,
                        'historico_total': round(total_item_hist, 1)
                    })

                    # Log de progresso a cada 50 itens
                    if (idx + 1) % 50 == 0:
                        print(f"    Processados {idx + 1}/{len(lista_itens)} itens...", flush=True)

                # Formatar períodos para o frontend
                periodos_formatados = []
                for data_str in datas_previsao:
                    if data_str:
                        try:
                            from datetime import datetime as dt_parse
                            data_obj = dt_parse.strptime(data_str, '%Y-%m-%d')
                            if granularidade == 'mensal':
                                periodos_formatados.append({
                                    'mes': data_obj.month,
                                    'ano': data_obj.year
                                })
                            elif granularidade == 'semanal':
                                # Calcular número da semana ISO
                                semana_iso = data_obj.isocalendar()[1]
                                ano_iso = data_obj.isocalendar()[0]
                                periodos_formatados.append({
                                    'semana': semana_iso,
                                    'ano': ano_iso
                                })
                            else:  # diario
                                periodos_formatados.append(data_str)
                        except:
                            periodos_formatados.append(data_str)

                # Adicionar à resposta
                resposta['relatorio_detalhado'] = {
                    'itens': relatorio_itens,
                    'periodos_previsao': periodos_formatados,
                    'total_itens': len(relatorio_itens),
                    'granularidade': granularidade
                }

                print(f"  Relatório gerado com {len(relatorio_itens)} itens", flush=True)
            else:
                resposta['relatorio_detalhado'] = {
                    'itens': [],
                    'periodos_previsao': [],
                    'total_itens': 0,
                    'granularidade': granularidade
                }
                print(f"  Nenhum item encontrado para o relatório", flush=True)

            # Fechar conexão do relatório
            cursor_rel.close()
            conn_rel.close()

        except Exception as e_relatorio:
            print(f"  Erro ao gerar relatório detalhado: {e_relatorio}", flush=True)
            resposta['relatorio_detalhado'] = {
                'itens': [],
                'periodos_previsao': [],
                'total_itens': 0,
                'erro': str(e_relatorio)
            }
            # Tentar fechar conexão em caso de erro
            try:
                cursor_rel.close()
                conn_rel.close()
            except:
                pass

        print(f"\nPrevisão gerada com sucesso! Melhor modelo: {melhor_modelo}")

        return jsonify(resposta)

    except Exception as e:
        import traceback
        erro_msg = traceback.format_exc()
        # Gravar erro em arquivo para debug
        with open('erro_previsao.log', 'w', encoding='utf-8') as f:
            f.write(erro_msg)
        try:
            print(f"Erro ao gerar previsao: {e}", flush=True)
        except:
            pass
        return jsonify({'erro': str(e)}), 500


# ==============================================================================
# API V2 - BOTTOM-UP FORECASTING (Calcula itens primeiro, depois agrega)
# ==============================================================================

@app.route('/api/gerar_previsao_banco_v2', methods=['POST'])
def api_gerar_previsao_banco_v2():
    """
    [COMPATIBILIDADE] Endpoint V2 mantido para compatibilidade.
    Ambos os endpoints agora usam a mesma lógica otimizada.
    """
    return api_gerar_previsao_banco_v2_interno()


def api_gerar_previsao_banco_v2_interno():
    """
    BOTTOM-UP FORECASTING (Versão Otimizada)
    Calcula previsão para cada item individualmente, depois agrega para o total.
    Isso garante que: Soma dos itens = Total do gráfico/tabela

    Características:
    - Sazonalidade por dia da semana (diário) ou semana do ano (semanal)
    - Tendência com limitadores (queda >40% ignorada, alta >50% limitada)
    - Comparação mesmo dia da semana do ano anterior (granularidade diária)
    """
    try:
        dados = request.get_json()

        # Extrair parâmetros
        loja = dados.get('loja', 'TODAS')
        categoria = dados.get('categoria', 'TODAS')
        produto = dados.get('produto', 'TODOS')
        granularidade = dados.get('granularidade', 'mensal')

        # Parâmetros adicionais de filtro
        fornecedor = dados.get('fornecedor', 'TODOS')
        linha = dados.get('linha', 'TODAS')
        sublinha = dados.get('sublinha', 'TODAS')

        # Parâmetros de período
        mes_inicio = dados.get('mes_inicio')
        ano_inicio = dados.get('ano_inicio')
        mes_fim = dados.get('mes_fim')
        ano_fim = dados.get('ano_fim')
        semana_inicio = dados.get('semana_inicio')
        semana_fim = dados.get('semana_fim')
        ano_semana_inicio = dados.get('ano_semana_inicio')
        ano_semana_fim = dados.get('ano_semana_fim')
        data_inicio = dados.get('data_inicio')
        data_fim = dados.get('data_fim')

        # Calcular número de períodos de previsão baseado no período selecionado
        meses_previsao = 6  # Default
        if granularidade == 'mensal' and mes_inicio and ano_inicio and mes_fim and ano_fim:
            # Calcular meses entre as datas
            meses_previsao = (int(ano_fim) - int(ano_inicio)) * 12 + (int(mes_fim) - int(mes_inicio)) + 1
        elif granularidade == 'semanal' and semana_inicio and ano_semana_inicio and semana_fim and ano_semana_fim:
            # Calcular semanas entre as datas
            # S27 até S29 = 3 semanas (27, 28, 29)
            semanas = (int(ano_semana_fim) - int(ano_semana_inicio)) * 52 + (int(semana_fim) - int(semana_inicio)) + 1
            meses_previsao = semanas
            print(f"DEBUG SEMANAL: semana_inicio={semana_inicio}, semana_fim={semana_fim}, ano_inicio={ano_semana_inicio}, ano_fim={ano_semana_fim}", flush=True)
            print(f"DEBUG SEMANAL: semanas calculadas = {semanas}", flush=True)
        elif granularidade == 'diario' and data_inicio and data_fim:
            # Calcular dias entre as datas
            from datetime import datetime
            d1 = datetime.strptime(data_inicio, '%Y-%m-%d')
            d2 = datetime.strptime(data_fim, '%Y-%m-%d')
            dias = (d2 - d1).days + 1
            meses_previsao = dias  # Será usado diretamente

        print(f"\n{'='*60}", flush=True)
        print(f"=== PREVISÃO V2 - BOTTOM-UP FORECASTING ===", flush=True)
        print(f"{'='*60}", flush=True)
        print(f"Loja: {loja}", flush=True)
        print(f"Fornecedor: {fornecedor}", flush=True)
        print(f"Linha: {linha}", flush=True)
        print(f"Granularidade: {granularidade}", flush=True)
        print(f"Períodos de previsão calculados: {meses_previsao}", flush=True)

        # Para granularidade mensal, meses_previsao já é o número correto de períodos
        # Para semanal e diário, já foi calculado corretamente acima
        periodos_previsao = meses_previsao

        # Conectar ao banco
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Construir filtros SQL
        where_conditions = []
        params = []

        if loja and loja != 'TODAS':
            where_conditions.append("h.cod_empresa = %s")
            params.append(int(loja))

        if fornecedor and fornecedor != 'TODOS':
            where_conditions.append("p.nome_fornecedor = %s")
            params.append(fornecedor)

        if linha and linha != 'TODAS':
            where_conditions.append("p.categoria = %s")
            params.append(linha)

        if sublinha and sublinha != 'TODAS':
            where_conditions.append("p.codigo_linha = %s")
            params.append(sublinha)

        if produto and produto != 'TODOS':
            where_conditions.append("h.codigo = %s")
            params.append(int(produto))

        where_sql = ""
        if where_conditions:
            where_sql = "AND " + " AND ".join(where_conditions)

        # =====================================================
        # PASSO 1: BUSCAR LISTA DE ITENS (COM SITUAÇÃO DE COMPRA)
        # =====================================================
        print(f"\n[PASSO 1] Buscando lista de itens...", flush=True)

        # Verificar se tabelas de situação de compra existem
        tabela_sit_compra_existe = False
        try:
            cursor.execute("SELECT to_regclass('public.situacao_compra_itens') IS NOT NULL as existe")
            result = cursor.fetchone()
            tabela_sit_compra_existe = result['existe'] if result else False
        except Exception:
            tabela_sit_compra_existe = False

        if tabela_sit_compra_existe:
            print(f"  Tabela situacao_compra_itens encontrada - filtro SIT_COMPRA ATIVO", flush=True)
            # Query com LEFT JOIN para trazer situação de compra
            # Busca a loja específica ou qualquer loja se filtro for TODAS
            sit_compra_join = """
                LEFT JOIN situacao_compra_itens sci ON h.codigo = sci.codigo
                    AND (sci.data_fim IS NULL OR sci.data_fim >= CURRENT_DATE)
                LEFT JOIN situacao_compra_regras scr ON sci.sit_compra = scr.codigo_situacao
                    AND scr.ativo = TRUE
            """
            sit_compra_select = ", sci.sit_compra, scr.descricao as sit_compra_descricao"
        else:
            print(f"  Tabela situacao_compra_itens NÃO encontrada - filtro SIT_COMPRA DESATIVADO", flush=True)
            sit_compra_join = ""
            sit_compra_select = ", NULL as sit_compra, NULL as sit_compra_descricao"

        query_itens = f"""
            SELECT DISTINCT
                h.codigo as cod_produto,
                p.descricao,
                p.nome_fornecedor,
                p.cnpj_fornecedor,
                p.categoria,
                p.descricao_linha
                {sit_compra_select}
            FROM historico_vendas_diario h
            JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
            {sit_compra_join}
            WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
            {where_sql}
            ORDER BY p.nome_fornecedor, p.descricao
        """
        cursor.execute(query_itens, params)
        lista_itens = cursor.fetchall()

        if not lista_itens or len(lista_itens) == 0:
            cursor.close()
            conn.close()
            return jsonify({'erro': 'Nenhum item encontrado com os filtros selecionados'}), 400

        print(f"  Encontrados {len(lista_itens)} itens", flush=True)

        # =====================================================
        # PASSO 2: BUSCAR VENDAS HISTÓRICAS POR ITEM (COM ESTOQUE PARA SANEAMENTO)
        # =====================================================
        print(f"\n[PASSO 2] Buscando vendas históricas por item (com dados de estoque)...", flush=True)

        # Verificar se a tabela de estoque existe (para graceful degradation)
        tabela_estoque_existe = False
        try:
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_name = 'historico_estoque_diario'
                )
            """)
            tabela_estoque_existe = cursor.fetchone()[0] if cursor.fetchone else False
            # Verificação alternativa para RealDictCursor
            cursor.execute("SELECT to_regclass('public.historico_estoque_diario') IS NOT NULL as existe")
            result = cursor.fetchone()
            tabela_estoque_existe = result['existe'] if result else False
        except Exception as e:
            print(f"  AVISO: Não foi possível verificar tabela de estoque: {e}", flush=True)
            tabela_estoque_existe = False

        if tabela_estoque_existe:
            print(f"  Tabela historico_estoque_diario encontrada - saneamento de rupturas ATIVO", flush=True)
        else:
            print(f"  Tabela historico_estoque_diario NÃO encontrada - saneamento de rupturas DESATIVADO", flush=True)

        # Query com LEFT JOIN para trazer dados de estoque (para detecção de rupturas)
        # Se a tabela não existir, usa query sem o JOIN (graceful degradation)
        if tabela_estoque_existe:
            # Query COM dados de estoque
            join_estoque = """
                LEFT JOIN historico_estoque_diario e
                    ON h.data = e.data
                    AND h.codigo = e.codigo
                    AND h.cod_empresa = e.cod_empresa
            """
            select_estoque = "AVG(COALESCE(e.estoque_diario, -1)) as estoque_medio"
        else:
            # Query SEM dados de estoque (fallback)
            join_estoque = ""
            select_estoque = "-1 as estoque_medio"

        if granularidade == 'mensal':
            query_vendas = f"""
                SELECT
                    h.codigo as cod_produto,
                    DATE_TRUNC('month', h.data) as periodo,
                    SUM(h.qtd_venda) as qtd_venda,
                    {select_estoque}
                FROM historico_vendas_diario h
                JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
                {join_estoque}
                WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                {where_sql}
                GROUP BY h.codigo, DATE_TRUNC('month', h.data)
                ORDER BY h.codigo, DATE_TRUNC('month', h.data)
            """
        elif granularidade == 'semanal':
            query_vendas = f"""
                SELECT
                    h.codigo as cod_produto,
                    DATE_TRUNC('week', h.data) as periodo,
                    SUM(h.qtd_venda) as qtd_venda,
                    {select_estoque}
                FROM historico_vendas_diario h
                JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
                {join_estoque}
                WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                {where_sql}
                GROUP BY h.codigo, DATE_TRUNC('week', h.data)
                ORDER BY h.codigo, DATE_TRUNC('week', h.data)
            """
        else:  # diario
            query_vendas = f"""
                SELECT
                    h.codigo as cod_produto,
                    h.data as periodo,
                    SUM(h.qtd_venda) as qtd_venda,
                    {select_estoque}
                FROM historico_vendas_diario h
                JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
                {join_estoque}
                WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
                {where_sql}
                GROUP BY h.codigo, h.data
                ORDER BY h.codigo, h.data
            """

        cursor.execute(query_vendas, params)
        vendas_por_item_raw = cursor.fetchall()

        # =====================================================
        # PASSO 2.1: SANEAR RUPTURAS NO HISTÓRICO
        # =====================================================
        print(f"\n[PASSO 2.1] Saneando rupturas no histórico...", flush=True)

        # Organizar dados em DataFrame para saneamento
        dados_para_sanear = []
        for venda in vendas_por_item_raw:
            estoque = float(venda['estoque_medio']) if venda['estoque_medio'] is not None else -1
            dados_para_sanear.append({
                'cod_produto': venda['cod_produto'],
                'periodo': venda['periodo'].strftime('%Y-%m-%d') if venda['periodo'] else None,
                'qtd_venda': float(venda['qtd_venda']) if venda['qtd_venda'] else 0,
                'estoque': estoque if estoque >= 0 else None  # -1 significa sem dados
            })

        df_historico = pd.DataFrame(dados_para_sanear)

        # Aplicar saneamento de rupturas
        sanitizer = RupturaSanitizer()
        estatisticas_saneamento = {'total_rupturas': 0, 'rupturas_saneadas': 0, 'metodos': {}}

        if len(df_historico) > 0 and 'estoque' in df_historico.columns:
            # Adicionar coluna dummy de loja para o sanitizer
            df_historico['cod_empresa'] = 1

            df_saneado, stats = sanitizer.sanear_serie_temporal(
                df_historico,
                col_data='periodo',
                col_venda='qtd_venda',
                col_estoque='estoque',
                col_produto='cod_produto',
                col_loja='cod_empresa',
                granularidade=granularidade
            )

            estatisticas_saneamento = {
                'total_rupturas': stats.get('total_rupturas_detectadas', 0),
                'rupturas_saneadas': stats.get('rupturas_saneadas', 0),
                'metodos': stats.get('metodos_usados', {})
            }

            print(f"  Rupturas detectadas: {estatisticas_saneamento['total_rupturas']}", flush=True)
            print(f"  Rupturas saneadas: {estatisticas_saneamento['rupturas_saneadas']}", flush=True)
            if estatisticas_saneamento['metodos']:
                for metodo, qtd in estatisticas_saneamento['metodos'].items():
                    print(f"    - {metodo}: {qtd}", flush=True)
        else:
            df_saneado = df_historico.copy()
            df_saneado['qtd_venda_saneada'] = df_saneado['qtd_venda']
            df_saneado['foi_saneado'] = False
            print(f"  Sem dados de estoque - saneamento não aplicado", flush=True)

        # Organizar vendas por item (usando valores saneados)
        vendas_item_dict = {}
        todas_datas = set()
        for _, row in df_saneado.iterrows():
            cod = row['cod_produto']
            if cod not in vendas_item_dict:
                vendas_item_dict[cod] = []
            data_str = row['periodo']
            if data_str:
                todas_datas.add(data_str)
            # Usar qtd_venda_saneada se disponível, senão qtd_venda original
            qtd = row.get('qtd_venda_saneada', row['qtd_venda'])
            vendas_item_dict[cod].append({
                'periodo': data_str,
                'qtd_venda': float(qtd) if qtd else 0,
                'qtd_venda_original': float(row['qtd_venda']) if row['qtd_venda'] else 0,
                'foi_saneado': row.get('foi_saneado', False)
            })

        # Ordenar datas
        datas_ordenadas = sorted(list(todas_datas))
        print(f"  Período histórico: {datas_ordenadas[0] if datas_ordenadas else 'N/A'} até {datas_ordenadas[-1] if datas_ordenadas else 'N/A'}", flush=True)

        # =====================================================
        # PASSO 3: CALCULAR FATORES SAZONAIS (do agregado)
        # =====================================================
        print(f"\n[PASSO 3] Calculando fatores sazonais...", flush=True)

        # Agregar vendas para calcular sazonalidade
        vendas_agregadas = {}
        for cod, vendas in vendas_item_dict.items():
            for v in vendas:
                periodo = v['periodo']
                if periodo:
                    if periodo not in vendas_agregadas:
                        vendas_agregadas[periodo] = 0
                    vendas_agregadas[periodo] += v['qtd_venda']

        # Criar série temporal agregada ordenada
        serie_agregada = [vendas_agregadas.get(d, 0) for d in datas_ordenadas]

        # Calcular fatores sazonais
        fatores_sazonais = {}
        if len(serie_agregada) >= 12:
            if granularidade == 'mensal':
                # Calcular média por mês
                vendas_por_mes = {}
                for i, data_str in enumerate(datas_ordenadas):
                    mes = int(data_str[5:7])
                    if mes not in vendas_por_mes:
                        vendas_por_mes[mes] = []
                    vendas_por_mes[mes].append(serie_agregada[i])

                media_geral = sum(serie_agregada) / len(serie_agregada) if serie_agregada else 1

                for mes in range(1, 13):
                    if mes in vendas_por_mes and len(vendas_por_mes[mes]) > 0:
                        media_mes = sum(vendas_por_mes[mes]) / len(vendas_por_mes[mes])
                        fatores_sazonais[mes] = media_mes / media_geral if media_geral > 0 else 1.0
                    else:
                        fatores_sazonais[mes] = 1.0

                print(f"  Fatores sazonais mensais calculados: {len(fatores_sazonais)} meses", flush=True)

            elif granularidade == 'semanal':
                # Calcular média por semana do ano
                vendas_por_semana = {}
                for i, data_str in enumerate(datas_ordenadas):
                    from datetime import datetime as dt
                    data_obj = dt.strptime(data_str, '%Y-%m-%d')
                    semana = data_obj.isocalendar()[1]
                    if semana not in vendas_por_semana:
                        vendas_por_semana[semana] = []
                    vendas_por_semana[semana].append(serie_agregada[i])

                media_geral = sum(serie_agregada) / len(serie_agregada) if serie_agregada else 1

                for semana in range(1, 53):
                    if semana in vendas_por_semana and len(vendas_por_semana[semana]) > 0:
                        media_semana = sum(vendas_por_semana[semana]) / len(vendas_por_semana[semana])
                        fatores_sazonais[semana] = media_semana / media_geral if media_geral > 0 else 1.0
                    else:
                        fatores_sazonais[semana] = 1.0

                print(f"  Fatores sazonais semanais calculados: {len(fatores_sazonais)} semanas", flush=True)

            else:  # diario
                # Calcular média por dia da semana
                vendas_por_dia = {}
                for i, data_str in enumerate(datas_ordenadas):
                    from datetime import datetime as dt
                    data_obj = dt.strptime(data_str, '%Y-%m-%d')
                    dia_semana = data_obj.weekday()
                    if dia_semana not in vendas_por_dia:
                        vendas_por_dia[dia_semana] = []
                    vendas_por_dia[dia_semana].append(serie_agregada[i])

                media_geral = sum(serie_agregada) / len(serie_agregada) if serie_agregada else 1

                for dia in range(7):
                    if dia in vendas_por_dia and len(vendas_por_dia[dia]) > 0:
                        media_dia = sum(vendas_por_dia[dia]) / len(vendas_por_dia[dia])
                        fatores_sazonais[dia] = media_dia / media_geral if media_geral > 0 else 1.0
                    else:
                        fatores_sazonais[dia] = 1.0

                print(f"  Fatores sazonais diários calculados: {len(fatores_sazonais)} dias", flush=True)

        # Determinar última data do histórico
        from datetime import datetime as dt
        ultima_data_historico = dt.strptime(datas_ordenadas[-1], '%Y-%m-%d') if datas_ordenadas else dt.now()

        # =====================================================
        # PASSO 4: CALCULAR PREVISÃO PARA CADA ITEM (BOTTOM-UP)
        # =====================================================
        print(f"\n[PASSO 4] Calculando previsão para cada item (Bottom-Up)...", flush=True)

        modelos_disponiveis = [
            'Média Móvel Simples',
            'Média Móvel Ponderada',
            'Suavização Exponencial Simples',
            'Holt',
            'Holt-Winters',
            'Regressão Linear'
        ]

        # Gerar datas de previsão BASEADO NO PERÍODO SELECIONADO PELO USUÁRIO
        datas_previsao = []
        from datetime import timedelta
        # pd já importado globalmente na linha 9

        if granularidade == 'mensal' and mes_inicio and ano_inicio:
            # Usar as datas que o usuário selecionou
            for i in range(periodos_previsao):
                mes_atual = int(mes_inicio) + i
                ano_atual = int(ano_inicio)
                # Ajustar se passar de dezembro
                while mes_atual > 12:
                    mes_atual -= 12
                    ano_atual += 1
                data_prev = f"{ano_atual:04d}-{mes_atual:02d}-01"
                datas_previsao.append(data_prev)
        elif granularidade == 'semanal' and semana_inicio and ano_semana_inicio:
            # Calcular data inicial da semana usando formato ISO correto
            from datetime import datetime
            ano_base = int(ano_semana_inicio)
            semana_base = int(semana_inicio)

            # Usar formato ISO para obter a segunda-feira da semana correta
            # %G = ano ISO, %V = semana ISO, %u = dia da semana (1=segunda)
            try:
                data_base = datetime.strptime(f'{ano_base}-W{semana_base:02d}-1', '%G-W%V-%u')
            except ValueError:
                # Fallback para semanas que podem não existir em alguns anos
                print(f"  AVISO: Semana {semana_base}/{ano_base} pode ser inválida, ajustando...", flush=True)
                data_base = datetime(ano_base, 1, 1)
                data_base = data_base + timedelta(weeks=semana_base - 1)

            print(f"  Semana inicial: S{semana_base}/{ano_base} = {data_base.strftime('%Y-%m-%d')}", flush=True)
            print(f"  Períodos a gerar: {periodos_previsao}", flush=True)

            for i in range(periodos_previsao):
                data_prev = data_base + timedelta(weeks=i)
                # Usar formato YYYY-Www para melhor identificação de semanas
                semana_atual = data_prev.isocalendar()[1]
                ano_atual = data_prev.isocalendar()[0]
                # Armazenar como string no formato que será exibido
                data_formatada = f"{ano_atual}-S{semana_atual:02d}"
                datas_previsao.append(data_formatada)
                print(f"    Período {i+1}: {data_formatada} (data: {data_prev.strftime('%Y-%m-%d')})", flush=True)
        elif granularidade == 'diario' and data_inicio:
            from datetime import datetime
            data_base = datetime.strptime(data_inicio, '%Y-%m-%d')
            for i in range(periodos_previsao):
                data_prev = data_base + timedelta(days=i)
                datas_previsao.append(data_prev.strftime('%Y-%m-%d'))
        else:
            # Fallback: usar última data histórica
            for i in range(periodos_previsao):
                if granularidade == 'mensal':
                    data_prev = ultima_data_historico + pd.DateOffset(months=i+1)
                    data_prev = data_prev.replace(day=1)
                    datas_previsao.append(data_prev.strftime('%Y-%m-%d'))
                elif granularidade == 'semanal':
                    data_prev = ultima_data_historico + timedelta(weeks=i+1)
                    semana_atual = data_prev.isocalendar()[1]
                    ano_atual = data_prev.isocalendar()[0]
                    datas_previsao.append(f"{ano_atual}-S{semana_atual:02d}")
                else:  # diario
                    data_prev = ultima_data_historico + timedelta(days=i+1)
                    datas_previsao.append(data_prev.strftime('%Y-%m-%d'))

        print(f"  Datas de previsão geradas: {datas_previsao[:3]}...{datas_previsao[-1] if len(datas_previsao) > 3 else ''}", flush=True)

        # Estruturas para armazenar resultados
        relatorio_itens = []
        previsoes_agregadas_por_periodo = {d: 0 for d in datas_previsao}
        contagem_modelos = {}

        # Buscar período do ano anterior para comparação
        if datas_previsao:
            from datetime import datetime as dt
            # Parsear primeira e última data (pode ser formato YYYY-MM-DD ou YYYY-SWW)
            primeira_data = datas_previsao[0]
            ultima_data = datas_previsao[-1]

            if '-S' in primeira_data:
                # Formato semanal: YYYY-SWW
                ano_ini, sem_ini = primeira_data.split('-S')
                ano_fim, sem_fim = ultima_data.split('-S')
                data_inicio_prev = dt.strptime(f'{ano_ini}-W{sem_ini}-1', '%G-W%V-%u')
                data_fim_prev = dt.strptime(f'{ano_fim}-W{sem_fim}-7', '%G-W%V-%u')
            else:
                # Formato normal: YYYY-MM-DD
                data_inicio_prev = dt.strptime(primeira_data, '%Y-%m-%d')
                data_fim_prev = dt.strptime(ultima_data, '%Y-%m-%d')
            data_inicio_aa = (data_inicio_prev - pd.DateOffset(years=1)).strftime('%Y-%m-%d')
            data_fim_aa = (data_fim_prev - pd.DateOffset(years=1)).strftime('%Y-%m-%d')
        else:
            data_inicio_aa = None
            data_fim_aa = None

        for idx, item in enumerate(lista_itens):
            cod_produto = item['cod_produto']

            # =====================================================
            # VERIFICAR SITUAÇÃO DE COMPRA (FL, EN = sem cálculo de demanda)
            # =====================================================
            sit_compra = item.get('sit_compra')
            sit_compra_descricao = item.get('sit_compra_descricao')

            # Itens Fora de Linha (FL) ou Encomenda (EN) não têm previsão calculada
            pular_calculo_demanda = sit_compra in ('FL', 'EN')

            if pular_calculo_demanda:
                # Adicionar item ao relatório com demanda zerada
                relatorio_itens.append({
                    'cod_produto': cod_produto,
                    'descricao': item['descricao'],
                    'nome_fornecedor': item['nome_fornecedor'] or 'SEM FORNECEDOR',
                    'cnpj_fornecedor': item['cnpj_fornecedor'],
                    'categoria': item['categoria'],
                    'linha': item['descricao_linha'],
                    'demanda_prevista_total': 0,
                    'demanda_ano_anterior': 0,
                    'variacao_percentual': None,
                    'sinal_alerta': 'cinza',
                    'sinal_emoji': '⚪',
                    'metodo_estatistico': f'Sem cálculo ({sit_compra_descricao or sit_compra})',
                    'previsao_por_periodo': [{'periodo': d, 'previsao': 0} for d in datas_previsao],
                    'historico_total': 0,
                    'sit_compra': sit_compra,
                    'sit_compra_descricao': sit_compra_descricao
                })
                continue  # Pular para o próximo item

            vendas_hist = vendas_item_dict.get(cod_produto, [])

            # Ordenar vendas por período
            vendas_hist_sorted = sorted(vendas_hist, key=lambda x: x['periodo'] if x['periodo'] else '')
            serie_item = [v['qtd_venda'] for v in vendas_hist_sorted]
            total_item_hist = sum(serie_item)

            # =====================================================
            # DETECTAR SAZONALIDADE INDIVIDUAL DO ITEM
            # =====================================================
            tem_sazonalidade_item = False
            fatores_sazonais_item = {}
            forca_sazonalidade = 0
            tendencia_por_mes_item = {}  # NOVO: tendência de crescimento por mês
            tendencia_anomala = False  # NOVO: flag para queda extrema (< -40%)
            tendencia_limitada = False  # NOVO: flag para alta extrema (> +50%)

            if granularidade == 'mensal' and len(serie_item) >= 12:
                # Calcular média por mês para ESTE item
                vendas_por_mes_item = {}
                # NOVO: Armazenar vendas por ano/mês para calcular tendência
                vendas_por_ano_mes_item = {}

                for i, v in enumerate(vendas_hist_sorted):
                    if v['periodo']:
                        ano = int(v['periodo'][:4])
                        mes = int(v['periodo'][5:7])
                        if mes not in vendas_por_mes_item:
                            vendas_por_mes_item[mes] = []
                        vendas_por_mes_item[mes].append(v['qtd_venda'])

                        # Armazenar por ano/mês
                        if mes not in vendas_por_ano_mes_item:
                            vendas_por_ano_mes_item[mes] = {}
                        if ano not in vendas_por_ano_mes_item[mes]:
                            vendas_por_ano_mes_item[mes][ano] = 0
                        vendas_por_ano_mes_item[mes][ano] += v['qtd_venda']

                # Calcular média geral do item
                media_geral_item = sum(serie_item) / len(serie_item) if serie_item else 1

                # Calcular fatores sazonais individuais
                for mes in range(1, 13):
                    if mes in vendas_por_mes_item and len(vendas_por_mes_item[mes]) > 0:
                        media_mes = sum(vendas_por_mes_item[mes]) / len(vendas_por_mes_item[mes])
                        fatores_sazonais_item[mes] = media_mes / media_geral_item if media_geral_item > 0 else 1.0
                    else:
                        fatores_sazonais_item[mes] = 1.0

                # =====================================================
                # NOVO: CALCULAR TENDÊNCIA DE CRESCIMENTO POR MÊS
                # =====================================================
                tendencia_por_mes_item = {}

                # Flag para detectar se houve queda extrema (possível anomalia)
                tendencia_anomala = False
                tendencias_brutas = []

                for mes in range(1, 13):
                    if mes in vendas_por_ano_mes_item:
                        anos_ordenados = sorted(vendas_por_ano_mes_item[mes].keys())
                        valores_por_ano = [vendas_por_ano_mes_item[mes][ano] for ano in anos_ordenados]

                        if len(anos_ordenados) >= 3:
                            # Com 3+ anos: usar regressão linear para projetar tendência
                            n = len(anos_ordenados)
                            x_mean = sum(range(n)) / n
                            y_mean = sum(valores_por_ano) / n

                            numerador = sum((i - x_mean) * (valores_por_ano[i] - y_mean) for i in range(n))
                            denominador = sum((i - x_mean) ** 2 for i in range(n))

                            if denominador > 0 and y_mean > 0:
                                slope = numerador / denominador
                                # Tendência = crescimento relativo por ano
                                tendencia_bruta = slope / y_mean
                                tendencias_brutas.append(tendencia_bruta)
                                tendencia_por_mes_item[mes] = tendencia_bruta
                            else:
                                tendencia_por_mes_item[mes] = 0

                        elif len(anos_ordenados) == 2:
                            # Com 2 anos: usar crescimento simples YoY
                            v_antigo = valores_por_ano[0]
                            v_recente = valores_por_ano[1]

                            if v_antigo > 0:
                                tendencia_bruta = (v_recente - v_antigo) / v_antigo
                                tendencias_brutas.append(tendencia_bruta)
                                tendencia_por_mes_item[mes] = tendencia_bruta
                            else:
                                tendencia_por_mes_item[mes] = 0
                        else:
                            tendencia_por_mes_item[mes] = 0
                    else:
                        tendencia_por_mes_item[mes] = 0

                # =====================================================
                # DETECTAR ANOMALIA E LIMITAR TENDÊNCIAS EXTREMAS
                # - Queda extrema (< -40%): Ignora tendência (possível anomalia)
                # - Alta extrema (> +50%): Limita a +50% (evita projeções irrealistas)
                # =====================================================
                # CORREÇÃO: Aplicar AMBOS os limites independentemente
                if tendencias_brutas:
                    media_tendencia = sum(tendencias_brutas) / len(tendencias_brutas)

                    # Aplicar limites individuais a cada período (independente da média)
                    for mes in tendencia_por_mes_item:
                        # Limitar tendências muito negativas a -40%
                        if tendencia_por_mes_item[mes] < -0.40:
                            tendencia_por_mes_item[mes] = -0.40
                            tendencia_anomala = True
                        # Limitar tendências muito positivas a +50%
                        elif tendencia_por_mes_item[mes] > 0.50:
                            tendencia_por_mes_item[mes] = 0.50
                            tendencia_limitada = True

                # Detectar força da sazonalidade (coeficiente de variação dos fatores)
                if fatores_sazonais_item:
                    valores_fatores = list(fatores_sazonais_item.values())
                    media_fatores = sum(valores_fatores) / len(valores_fatores)
                    variancia = sum((f - media_fatores) ** 2 for f in valores_fatores) / len(valores_fatores)
                    desvio = variancia ** 0.5
                    forca_sazonalidade = desvio / media_fatores if media_fatores > 0 else 0

                    # Sazonalidade forte se CV dos fatores > 0.3 (30% de variação)
                    tem_sazonalidade_item = forca_sazonalidade > 0.3

            elif granularidade == 'semanal' and len(serie_item) >= 52:
                # Calcular média por semana do ano
                vendas_por_semana_item = {}
                # NOVO: Armazenar vendas por ano/semana para calcular tendência
                vendas_por_ano_semana_item = {}

                for v in vendas_hist_sorted:
                    if v['periodo']:
                        data_obj = dt.strptime(v['periodo'], '%Y-%m-%d')
                        ano = data_obj.year
                        semana = data_obj.isocalendar()[1]
                        if semana not in vendas_por_semana_item:
                            vendas_por_semana_item[semana] = []
                        vendas_por_semana_item[semana].append(v['qtd_venda'])

                        # Armazenar por ano/semana
                        if semana not in vendas_por_ano_semana_item:
                            vendas_por_ano_semana_item[semana] = {}
                        if ano not in vendas_por_ano_semana_item[semana]:
                            vendas_por_ano_semana_item[semana][ano] = 0
                        vendas_por_ano_semana_item[semana][ano] += v['qtd_venda']

                media_geral_item = sum(serie_item) / len(serie_item) if serie_item else 1

                for semana in range(1, 53):
                    if semana in vendas_por_semana_item and len(vendas_por_semana_item[semana]) > 0:
                        media_semana = sum(vendas_por_semana_item[semana]) / len(vendas_por_semana_item[semana])
                        fatores_sazonais_item[semana] = media_semana / media_geral_item if media_geral_item > 0 else 1.0
                    else:
                        fatores_sazonais_item[semana] = 1.0

                # NOVO: Calcular tendência por semana
                tendencias_brutas_semana = []
                for semana in range(1, 53):
                    if semana in vendas_por_ano_semana_item:
                        anos_ordenados = sorted(vendas_por_ano_semana_item[semana].keys())
                        valores_por_ano = [vendas_por_ano_semana_item[semana][ano] for ano in anos_ordenados]

                        if len(anos_ordenados) >= 2:
                            v_antigo = valores_por_ano[0]
                            v_recente = valores_por_ano[-1]
                            if v_antigo > 0:
                                tendencia_bruta = (v_recente - v_antigo) / v_antigo
                                tendencias_brutas_semana.append(tendencia_bruta)
                                tendencia_por_mes_item[semana] = tendencia_bruta
                            else:
                                tendencia_por_mes_item[semana] = 0
                        else:
                            tendencia_por_mes_item[semana] = 0
                    else:
                        tendencia_por_mes_item[semana] = 0

                # Aplicar limitadores de tendência para semanal
                # CORREÇÃO: Aplicar AMBOS os limites independentemente
                if tendencias_brutas_semana:
                    # Aplicar limites individuais a cada semana (independente da média)
                    for semana in tendencia_por_mes_item:
                        # Limitar tendências muito negativas a -40%
                        if tendencia_por_mes_item[semana] < -0.40:
                            tendencia_por_mes_item[semana] = -0.40
                            tendencia_anomala = True
                        # Limitar tendências muito positivas a +50%
                        elif tendencia_por_mes_item[semana] > 0.50:
                            tendencia_por_mes_item[semana] = 0.50
                            tendencia_limitada = True

                if fatores_sazonais_item:
                    valores_fatores = list(fatores_sazonais_item.values())
                    media_fatores = sum(valores_fatores) / len(valores_fatores)
                    variancia = sum((f - media_fatores) ** 2 for f in valores_fatores) / len(valores_fatores)
                    desvio = variancia ** 0.5
                    forca_sazonalidade = desvio / media_fatores if media_fatores > 0 else 0
                    tem_sazonalidade_item = forca_sazonalidade > 0.3

            elif granularidade == 'diario' and len(serie_item) >= 14:
                # Calcular média por dia da semana (0=segunda, 6=domingo)
                vendas_por_dia_item = {}
                # Armazenar vendas por semana/dia para calcular tendência
                vendas_por_semana_dia_item = {}

                for v in vendas_hist_sorted:
                    if v['periodo']:
                        data_obj = dt.strptime(v['periodo'], '%Y-%m-%d')
                        dia_semana = data_obj.weekday()
                        semana_ano = data_obj.isocalendar()[1]

                        if dia_semana not in vendas_por_dia_item:
                            vendas_por_dia_item[dia_semana] = []
                        vendas_por_dia_item[dia_semana].append(v['qtd_venda'])

                        # Armazenar por semana/dia para tendência
                        if dia_semana not in vendas_por_semana_dia_item:
                            vendas_por_semana_dia_item[dia_semana] = {}
                        if semana_ano not in vendas_por_semana_dia_item[dia_semana]:
                            vendas_por_semana_dia_item[dia_semana][semana_ano] = 0
                        vendas_por_semana_dia_item[dia_semana][semana_ano] += v['qtd_venda']

                media_geral_item = sum(serie_item) / len(serie_item) if serie_item else 1

                # Calcular fatores sazonais por dia da semana
                for dia in range(7):
                    if dia in vendas_por_dia_item and len(vendas_por_dia_item[dia]) > 0:
                        media_dia = sum(vendas_por_dia_item[dia]) / len(vendas_por_dia_item[dia])
                        fatores_sazonais_item[dia] = media_dia / media_geral_item if media_geral_item > 0 else 1.0
                    else:
                        fatores_sazonais_item[dia] = 1.0

                # Calcular tendência por dia da semana
                tendencias_brutas_dia = []
                for dia in range(7):
                    if dia in vendas_por_semana_dia_item:
                        semanas_ordenadas = sorted(vendas_por_semana_dia_item[dia].keys())
                        if len(semanas_ordenadas) >= 4:  # Pelo menos 4 semanas de dados
                            # Comparar primeira metade com segunda metade
                            meio = len(semanas_ordenadas) // 2
                            valores_primeira = [vendas_por_semana_dia_item[dia][s] for s in semanas_ordenadas[:meio]]
                            valores_segunda = [vendas_por_semana_dia_item[dia][s] for s in semanas_ordenadas[meio:]]
                            media_primeira = sum(valores_primeira) / len(valores_primeira) if valores_primeira else 0
                            media_segunda = sum(valores_segunda) / len(valores_segunda) if valores_segunda else 0

                            if media_primeira > 0:
                                tendencia_bruta = (media_segunda - media_primeira) / media_primeira
                                tendencias_brutas_dia.append(tendencia_bruta)
                                tendencia_por_mes_item[dia] = tendencia_bruta
                            else:
                                tendencia_por_mes_item[dia] = 0
                        else:
                            tendencia_por_mes_item[dia] = 0
                    else:
                        tendencia_por_mes_item[dia] = 0

                # Aplicar limitadores de tendência para diário
                # CORREÇÃO: Aplicar AMBOS os limites independentemente
                if tendencias_brutas_dia:
                    # Aplicar limites individuais a cada dia (independente da média)
                    for dia in tendencia_por_mes_item:
                        # Limitar tendências muito negativas a -40%
                        if tendencia_por_mes_item[dia] < -0.40:
                            tendencia_por_mes_item[dia] = -0.40
                            tendencia_anomala = True
                        # Limitar tendências muito positivas a +50%
                        elif tendencia_por_mes_item[dia] > 0.50:
                            tendencia_por_mes_item[dia] = 0.50
                            tendencia_limitada = True

                if fatores_sazonais_item:
                    valores_fatores = list(fatores_sazonais_item.values())
                    media_fatores = sum(valores_fatores) / len(valores_fatores)
                    variancia = sum((f - media_fatores) ** 2 for f in valores_fatores) / len(valores_fatores)
                    desvio = variancia ** 0.5
                    forca_sazonalidade = desvio / media_fatores if media_fatores > 0 else 0
                    tem_sazonalidade_item = forca_sazonalidade > 0.3

            # =====================================================
            # SELECIONAR ESTRATÉGIA DE PREVISÃO
            # =====================================================
            melhor_modelo_item = 'Média Móvel Simples'
            previsao_item_periodos = []
            usou_sazonalidade_individual = False

            if len(serie_item) >= 6:
                # ESTRATÉGIA 1: Sazonalidade forte → usar média histórica por período COM TENDÊNCIA
                if tem_sazonalidade_item and fatores_sazonais_item:
                    # Usar previsão baseada na média histórica do período (mais preciso para itens sazonais)
                    media_geral_item = sum(serie_item) / len(serie_item) if serie_item else 0

                    # Calcular ano de previsão para aplicar tendência
                    ano_previsao = ultima_data_historico.year + 1  # Próximo ano

                    previsao_item_periodos = []
                    tendencias_aplicadas = []

                    for i in range(periodos_previsao):
                        # Usar a data de previsão já calculada (baseada no período selecionado pelo usuário)
                        data_prev_str = datas_previsao[i] if i < len(datas_previsao) else None

                        if granularidade == 'mensal':
                            # Extrair mês da data de previsão real
                            if data_prev_str:
                                mes_prev = int(data_prev_str[5:7])  # Formato: YYYY-MM-DD
                                ano_prev = int(data_prev_str[:4])
                            else:
                                mes_prev = ((ultima_data_historico.month + i) % 12) + 1
                                ano_prev = ultima_data_historico.year + 1
                            chave = mes_prev

                            # Calcular quantos anos à frente estamos prevendo
                            anos_a_frente = ano_prev - ultima_data_historico.year

                            # Obter tendência do mês (se disponível)
                            tendencia_mes = tendencia_por_mes_item.get(mes_prev, 0) if tendencia_por_mes_item else 0

                            # Limitar tendência para evitar valores extremos (-50% a +100%)
                            tendencia_mes = max(-0.5, min(1.0, tendencia_mes))

                            # Aplicar tendência proporcional ao número de anos à frente
                            fator_tendencia = 1 + (tendencia_mes * anos_a_frente)
                            tendencias_aplicadas.append(tendencia_mes)

                        elif granularidade == 'semanal':
                            if data_prev_str:
                                from datetime import datetime as dt_parse
                                # Suportar formato YYYY-SWW (ex: 2026-S27) ou YYYY-MM-DD
                                if '-S' in data_prev_str:
                                    ano_str, sem_str = data_prev_str.split('-S')
                                    chave = int(sem_str)
                                    ano_prev = int(ano_str)
                                else:
                                    data_prev_obj = dt_parse.strptime(data_prev_str, '%Y-%m-%d')
                                    chave = data_prev_obj.isocalendar()[1]
                                    ano_prev = data_prev_obj.year
                            else:
                                data_prev_obj = ultima_data_historico + timedelta(weeks=i+1)
                                chave = data_prev_obj.isocalendar()[1]
                                ano_prev = data_prev_obj.year

                            # Calcular anos à frente para semanal
                            anos_a_frente = ano_prev - ultima_data_historico.year

                            # Obter tendência da semana (se disponível)
                            tendencia_semana = tendencia_por_mes_item.get(chave, 0) if tendencia_por_mes_item else 0
                            tendencia_semana = max(-0.5, min(1.0, tendencia_semana))
                            fator_tendencia = 1 + (tendencia_semana * max(1, anos_a_frente))
                            tendencias_aplicadas.append(tendencia_semana)

                        else:  # diario
                            if data_prev_str:
                                from datetime import datetime as dt_parse
                                data_prev_obj = dt_parse.strptime(data_prev_str, '%Y-%m-%d')
                                chave = data_prev_obj.weekday()
                            else:
                                data_prev_obj = ultima_data_historico + timedelta(days=i+1)
                                chave = data_prev_obj.weekday()

                            # Obter tendência do dia da semana (se disponível)
                            tendencia_dia = tendencia_por_mes_item.get(chave, 0) if tendencia_por_mes_item else 0
                            tendencia_dia = max(-0.5, min(1.0, tendencia_dia))
                            fator_tendencia = 1 + tendencia_dia
                            tendencias_aplicadas.append(tendencia_dia)

                        # Usar fator sazonal individual para calcular previsão
                        fator_sazonal = fatores_sazonais_item.get(chave, 1.0)
                        previsao_periodo = media_geral_item * fator_sazonal * fator_tendencia
                        previsao_item_periodos.append(previsao_periodo)

                    # Montar descrição do modelo com info de tendência
                    if tendencia_anomala and tendencias_aplicadas:
                        # Tendência foi limitada a -40% por ser queda extrema
                        tendencia_media = sum(tendencias_aplicadas) / len(tendencias_aplicadas)
                        melhor_modelo_item = f"Sazonal+Tendência (força={forca_sazonalidade:.2f}, tend={tendencia_media:+.1%}) [limitada:-40%]"
                    elif tendencia_limitada and tendencias_aplicadas and any(t != 0 for t in tendencias_aplicadas):
                        # Tendência foi limitada (alta > +50%)
                        tendencia_media = sum(tendencias_aplicadas) / len(tendencias_aplicadas)
                        melhor_modelo_item = f"Sazonal+Tendência (força={forca_sazonalidade:.2f}, tend={tendencia_media:+.1%}) [limitada:+50%]"
                    elif tendencias_aplicadas and any(t != 0 for t in tendencias_aplicadas):
                        tendencia_media = sum(tendencias_aplicadas) / len(tendencias_aplicadas)
                        melhor_modelo_item = f"Sazonal+Tendência (força={forca_sazonalidade:.2f}, tend={tendencia_media:+.1%})"
                    else:
                        melhor_modelo_item = f"Sazonal Individual (força={forca_sazonalidade:.2f})"
                    usou_sazonalidade_individual = True

                # ESTRATÉGIA 2: Sem sazonalidade forte → testar modelos e aplicar fator sazonal individual
                else:
                    # Dividir série: 70% treino, 30% teste para selecionar melhor modelo
                    tamanho_treino = max(3, int(len(serie_item) * 0.7))
                    tamanho_teste = len(serie_item) - tamanho_treino

                    if tamanho_teste >= 2:
                        serie_treino = serie_item[:tamanho_treino]
                        serie_teste = serie_item[tamanho_treino:]

                        melhor_wmape = float('inf')

                        for nome_modelo in modelos_disponiveis:
                            try:
                                modelo = get_modelo(nome_modelo)
                                modelo.fit(serie_treino)
                                previsoes_teste = modelo.predict(tamanho_teste)

                                if len(previsoes_teste) == len(serie_teste):
                                    erros = []
                                    for i in range(len(serie_teste)):
                                        if serie_teste[i] > 0:
                                            erro = abs(previsoes_teste[i] - serie_teste[i]) / serie_teste[i] * 100
                                            erros.append(erro)

                                    if erros:
                                        wmape = sum(erros) / len(erros)
                                        if wmape < melhor_wmape:
                                            melhor_wmape = wmape
                                            melhor_modelo_item = nome_modelo
                            except:
                                pass

                    # Gerar previsão com o melhor modelo
                    try:
                        modelo_final = get_modelo(melhor_modelo_item)
                        modelo_final.fit(serie_item)
                        previsao_base = modelo_final.predict(periodos_previsao)

                        # Aplicar sazonalidade INDIVIDUAL do item (se disponível) ou do agregado
                        fatores_usar = fatores_sazonais_item if fatores_sazonais_item else fatores_sazonais
                        if fatores_usar:
                            previsao_item_periodos = []
                            for i in range(len(previsao_base)):
                                # Usar a data de previsão já calculada (baseada no período selecionado)
                                data_prev_str = datas_previsao[i] if i < len(datas_previsao) else None

                                if granularidade == 'mensal':
                                    if data_prev_str:
                                        mes_prev = int(data_prev_str[5:7])
                                    else:
                                        mes_prev = ((ultima_data_historico.month + i) % 12) + 1
                                    chave = mes_prev
                                elif granularidade == 'semanal':
                                    if data_prev_str:
                                        from datetime import datetime as dt_parse
                                        # Suportar formato YYYY-SWW (ex: 2026-S27) ou YYYY-MM-DD
                                        if '-S' in data_prev_str:
                                            ano_str, sem_str = data_prev_str.split('-S')
                                            chave = int(sem_str)
                                        else:
                                            data_prev_obj = dt_parse.strptime(data_prev_str, '%Y-%m-%d')
                                            chave = data_prev_obj.isocalendar()[1]
                                    else:
                                        data_prev_obj = ultima_data_historico + timedelta(weeks=i+1)
                                        chave = data_prev_obj.isocalendar()[1]
                                else:
                                    if data_prev_str:
                                        from datetime import datetime as dt_parse
                                        data_prev_obj = dt_parse.strptime(data_prev_str, '%Y-%m-%d')
                                        chave = data_prev_obj.weekday()
                                    else:
                                        data_prev_obj = ultima_data_historico + timedelta(days=i+1)
                                        chave = data_prev_obj.weekday()

                                fator = fatores_usar.get(chave, 1.0)
                                previsao_item_periodos.append(previsao_base[i] * fator)
                        else:
                            previsao_item_periodos = list(previsao_base)
                    except:
                        previsao_item_periodos = []

            # Fallback: média simples se não conseguiu prever
            if not previsao_item_periodos or len(previsao_item_periodos) < periodos_previsao:
                if len(serie_item) > 0:
                    media = sum(serie_item) / len(serie_item)
                    previsao_item_periodos = [media] * periodos_previsao
                    melhor_modelo_item = "Média Simples (fallback)"
                else:
                    previsao_item_periodos = [0] * periodos_previsao
                    melhor_modelo_item = "Sem histórico"

            # Contar modelos utilizados
            if melhor_modelo_item not in contagem_modelos:
                contagem_modelos[melhor_modelo_item] = 0
            contagem_modelos[melhor_modelo_item] += 1

            # Montar previsão por período e AGREGAR
            previsao_por_periodo = []
            total_previsao_item = 0

            for i, data_prev in enumerate(datas_previsao):
                if i < len(previsao_item_periodos):
                    valor = round(previsao_item_periodos[i], 1)
                    previsao_por_periodo.append({
                        'periodo': data_prev,
                        'previsao': valor
                    })
                    total_previsao_item += valor
                    # AGREGAR para o total
                    previsoes_agregadas_por_periodo[data_prev] += valor

            # Buscar demanda real do ano anterior para este item
            demanda_ano_anterior_item = 0
            if data_inicio_aa and data_fim_aa:
                for venda in vendas_hist_sorted:
                    data_venda = venda['periodo']
                    if data_venda and data_inicio_aa <= data_venda <= data_fim_aa:
                        demanda_ano_anterior_item += venda['qtd_venda']
            demanda_ano_anterior_item = round(demanda_ano_anterior_item, 1)

            # Calcular variação
            if demanda_ano_anterior_item > 0:
                variacao_pct = round(((total_previsao_item - demanda_ano_anterior_item) / demanda_ano_anterior_item) * 100, 1)
            else:
                variacao_pct = None

            # Determinar sinal de alerta
            if variacao_pct is None:
                sinal_alerta = 'cinza'
                sinal_emoji = '⚪'
            elif abs(variacao_pct) > 50:
                sinal_alerta = 'vermelho' if variacao_pct < -50 else 'amarelo'
                sinal_emoji = '🔴' if variacao_pct < -50 else '🟡'
            elif abs(variacao_pct) > 20:
                sinal_alerta = 'azul'
                sinal_emoji = '🔵'
            else:
                sinal_alerta = 'verde'
                sinal_emoji = '🟢'

            relatorio_itens.append({
                'cod_produto': cod_produto,
                'descricao': item['descricao'],
                'nome_fornecedor': item['nome_fornecedor'] or 'SEM FORNECEDOR',
                'cnpj_fornecedor': item['cnpj_fornecedor'],
                'categoria': item['categoria'],
                'linha': item['descricao_linha'],
                'demanda_prevista_total': round(total_previsao_item, 1),
                'demanda_ano_anterior': demanda_ano_anterior_item,
                'variacao_percentual': variacao_pct,
                'sinal_alerta': sinal_alerta,
                'sinal_emoji': sinal_emoji,
                'metodo_estatistico': melhor_modelo_item,
                'previsao_por_periodo': previsao_por_periodo,
                'historico_total': round(total_item_hist, 1),
                'sit_compra': sit_compra,
                'sit_compra_descricao': sit_compra_descricao
            })

            # Log de progresso
            if (idx + 1) % 50 == 0:
                print(f"    Processados {idx + 1}/{len(lista_itens)} itens...", flush=True)

        print(f"  Todos os {len(lista_itens)} itens processados!", flush=True)
        print(f"  Modelos utilizados: {contagem_modelos}", flush=True)

        # =====================================================
        # PASSO 5: CONSTRUIR SÉRIE AGREGADA (soma dos itens)
        # =====================================================
        print(f"\n[PASSO 5] Construindo série agregada (Bottom-Up)...", flush=True)

        # Série histórica agregada
        serie_historica_agregada = serie_agregada
        datas_historicas = datas_ordenadas

        # Dividir série histórica em BASE (75%) e TESTE (25%) para exibição no gráfico
        split_index = int(len(datas_historicas) * 0.75)
        datas_base = datas_historicas[:split_index]
        valores_base = serie_historica_agregada[:split_index]
        datas_teste = datas_historicas[split_index:]
        valores_teste = serie_historica_agregada[split_index:]

        print(f"  Série dividida: Base={len(datas_base)} períodos, Teste={len(datas_teste)} períodos", flush=True)

        # =====================================================
        # CÁLCULO DO BACKTEST - USAR MESMO MÉTODO DO HISTÓRICO
        # =====================================================
        # O backtest mais preciso é usar os valores do mesmo período do ano anterior
        # ajustado pela tendência observada. Isso é mais realista do que recalcular
        # os modelos porque mostra como o histórico se comportou.
        #
        # Abordagem: Para cada mês de teste, buscar o mesmo mês do ano anterior
        # na base e aplicar fator de crescimento/tendência observado.

        previsao_teste = []

        if len(datas_teste) > 0 and len(valores_base) >= 3:
            print(f"  Calculando backtest para {len(datas_teste)} períodos...", flush=True)

            # Criar índice de vendas agregadas por período para lookup rápido
            vendas_por_data = {d: v for d, v in zip(datas_historicas, serie_historica_agregada)}

            # Calcular tendência geral (primeira metade vs segunda metade da base)
            meio_base = len(valores_base) // 2
            if meio_base > 0:
                media_primeira_metade = sum(valores_base[:meio_base]) / meio_base
                media_segunda_metade = sum(valores_base[meio_base:]) / (len(valores_base) - meio_base)
                if media_primeira_metade > 0:
                    fator_tendencia_geral = media_segunda_metade / media_primeira_metade
                    # Limitar entre 0.7 e 1.5
                    fator_tendencia_geral = max(0.7, min(1.5, fator_tendencia_geral))
                else:
                    fator_tendencia_geral = 1.0
            else:
                fator_tendencia_geral = 1.0

            print(f"    Fator de tendência geral: {fator_tendencia_geral:.2f}", flush=True)

            for i, data_teste_str in enumerate(datas_teste):
                previsao_periodo = 0

                try:
                    from datetime import datetime as dt_parse
                    from dateutil.relativedelta import relativedelta

                    # Determinar chave sazonal e buscar período correspondente na base
                    if '-S' in data_teste_str:
                        ano_str, sem_str = data_teste_str.split('-S')
                        chave_teste = int(sem_str)
                        ano_teste = int(ano_str)
                        # Para semanal: buscar mesma semana do ano anterior
                        ano_anterior = ano_teste - 1
                        data_aa_str = f"{ano_anterior}-S{chave_teste:02d}"
                    else:
                        data_obj = dt_parse.strptime(data_teste_str, '%Y-%m-%d')
                        if granularidade == 'mensal':
                            chave_teste = data_obj.month
                            # Buscar mesmo mês do ano anterior
                            data_aa = data_obj - relativedelta(years=1)
                            data_aa_str = data_aa.strftime('%Y-%m-%d')
                        elif granularidade == 'semanal':
                            chave_teste = data_obj.isocalendar()[1]
                            # Buscar mesma semana do ano anterior
                            data_aa = data_obj - relativedelta(years=1)
                            data_aa_str = data_aa.strftime('%Y-%m-%d')
                        else:
                            chave_teste = data_obj.weekday()
                            data_aa = data_obj - relativedelta(years=1)
                            data_aa_str = data_aa.strftime('%Y-%m-%d')

                    # Tentar encontrar valor do ano anterior na base
                    valor_aa = vendas_por_data.get(data_aa_str, 0)

                    if valor_aa > 0:
                        # Usar valor do ano anterior com ajuste de tendência
                        previsao_periodo = valor_aa * fator_tendencia_geral
                    else:
                        # Fallback: usar fator sazonal agregado
                        media_base = sum(valores_base) / len(valores_base)
                        fator = fatores_sazonais.get(chave_teste, 1.0)
                        previsao_periodo = media_base * fator

                except Exception as e:
                    # Fallback simples
                    media_base = sum(valores_base) / len(valores_base)
                    previsao_periodo = media_base

                previsao_teste.append(round(previsao_periodo, 2))

            print(f"  Backtest calculado: {len(previsao_teste)} períodos", flush=True)

        # Calcular erro do backtest para diagnóstico (WMAPE e BIAS)
        # Inicializar fora do bloco para garantir que existam
        wmape_backtest = 0
        bias_backtest = 0
        if previsao_teste and valores_teste:
            erros_bt = []
            erros_reais_bt = []
            for i in range(min(len(previsao_teste), len(valores_teste))):
                if valores_teste[i] > 0:
                    erro_pct = abs(previsao_teste[i] - valores_teste[i]) / valores_teste[i] * 100
                    erro_real = previsao_teste[i] - valores_teste[i]
                    erros_bt.append(erro_pct)
                    erros_reais_bt.append(erro_real)
            if erros_bt:
                wmape_backtest = sum(erros_bt) / len(erros_bt)
                soma_reais = sum(valores_teste[:len(erros_reais_bt)])
                bias_backtest = (sum(erros_reais_bt) / soma_reais) * 100 if soma_reais > 0 else 0
                print(f"  WMAPE do backtest: {wmape_backtest:.1f}%", flush=True)
                print(f"  BIAS do backtest: {bias_backtest:.1f}%", flush=True)
                # Mostrar comparação período a período
                print(f"    Comparação período a período:", flush=True)
                for i in range(min(3, len(previsao_teste), len(valores_teste))):
                    print(f"      {datas_teste[i]}: Real={valores_teste[i]:,.0f}, Prev={previsao_teste[i]:,.0f}", flush=True)

        # Fallback: usar média simples se não conseguiu calcular
        if not previsao_teste and len(valores_base) >= 3 and len(datas_teste) > 0:
            media_base = sum(valores_base) / len(valores_base)
            previsao_teste = [round(media_base, 2)] * len(datas_teste)
            print(f"  Previsão backtest (fallback média): {len(previsao_teste)} períodos", flush=True)

        # Série de previsão agregada (soma dos itens)
        previsao_agregada = [previsoes_agregadas_por_periodo[d] for d in datas_previsao]

        total_previsao = sum(previsao_agregada)
        total_historico = sum(serie_historica_agregada)

        print(f"  Total histórico (2 anos): {total_historico:,.2f}", flush=True)
        print(f"  Total previsão (soma dos itens): {total_previsao:,.2f}", flush=True)

        # =====================================================
        # PASSO 6: BUSCAR ANO ANTERIOR AGREGADO
        # =====================================================
        print(f"\n[PASSO 6] Buscando ano anterior agregado...", flush=True)

        ano_anterior_valores = []
        ano_anterior_datas = []

        if granularidade == 'diario':
            # Para diário: buscar o MESMO DIA DA SEMANA do ano anterior
            # Criar índice de vendas por dia da semana e semana ISO
            vendas_por_dia_semana = {}  # {(ano, semana_iso, dia_semana): valor}
            for data_str in datas_ordenadas:
                data_obj = dt.strptime(data_str, '%Y-%m-%d')
                ano_iso, semana_iso, dia_iso = data_obj.isocalendar()
                dia_semana = data_obj.weekday()
                chave = (ano_iso, semana_iso, dia_semana)
                vendas_por_dia_semana[chave] = vendas_agregadas.get(data_str, 0)

            # DEBUG: Mostrar anos disponíveis nos dados históricos
            anos_disponiveis = set()
            for chave in vendas_por_dia_semana.keys():
                anos_disponiveis.add(chave[0])
            print(f"  DEBUG: Anos ISO disponíveis no histórico: {sorted(anos_disponiveis)}", flush=True)

            # Para cada data de previsão, buscar o mesmo dia da semana na mesma semana ISO do ano anterior
            dias_semana_nomes = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
            debug_count = 0
            for data_prev_str in datas_previsao:
                data_prev = dt.strptime(data_prev_str, '%Y-%m-%d')
                ano_iso_prev, semana_iso_prev, _ = data_prev.isocalendar()
                dia_semana_prev = data_prev.weekday()

                # Buscar ano anterior: mesma semana ISO, mesmo dia da semana
                ano_anterior_iso = ano_iso_prev - 1
                chave_aa = (ano_anterior_iso, semana_iso_prev, dia_semana_prev)

                if chave_aa in vendas_por_dia_semana:
                    valor_aa = vendas_por_dia_semana[chave_aa]
                    ano_anterior_valores.append(valor_aa)
                    # Reconstruir a data do ano anterior
                    data_aa = dt.strptime(f'{ano_anterior_iso}-W{semana_iso_prev:02d}-{dia_semana_prev + 1}', '%G-W%V-%u')
                    ano_anterior_datas.append(data_aa.strftime('%Y-%m-%d'))
                    if debug_count < 5:
                        print(f"    DEBUG: Previsão {data_prev_str} ({dias_semana_nomes[dia_semana_prev]}) -> Ano anterior {data_aa.strftime('%Y-%m-%d')} ({dias_semana_nomes[dia_semana_prev]}) = {valor_aa:,.0f}", flush=True)
                        debug_count += 1
                else:
                    # Se não encontrar, tentar semana anterior ou posterior
                    encontrado = False
                    for offset in [1, -1, 2, -2]:
                        chave_alt = (ano_anterior_iso, semana_iso_prev + offset, dia_semana_prev)
                        if chave_alt in vendas_por_dia_semana:
                            valor_aa = vendas_por_dia_semana[chave_alt]
                            ano_anterior_valores.append(valor_aa)
                            data_aa = dt.strptime(f'{ano_anterior_iso}-W{(semana_iso_prev + offset):02d}-{dia_semana_prev + 1}', '%G-W%V-%u')
                            ano_anterior_datas.append(data_aa.strftime('%Y-%m-%d'))
                            if debug_count < 5:
                                print(f"    DEBUG: Previsão {data_prev_str} ({dias_semana_nomes[dia_semana_prev]}) -> Ano anterior (offset {offset}) {data_aa.strftime('%Y-%m-%d')} = {valor_aa:,.0f}", flush=True)
                                debug_count += 1
                            encontrado = True
                            break
                    if not encontrado:
                        ano_anterior_valores.append(0)
                        ano_anterior_datas.append('')
                        if debug_count < 5:
                            print(f"    DEBUG: Previsão {data_prev_str} ({dias_semana_nomes[dia_semana_prev]}) -> NÃO ENCONTRADO (procurou ano {ano_anterior_iso})", flush=True)
                            debug_count += 1

            print(f"  Ano anterior (mesmo dia da semana): {len(ano_anterior_valores)} períodos", flush=True)
            if ano_anterior_valores:
                print(f"  DEBUG: Primeiros 5 valores do ano anterior: {ano_anterior_valores[:5]}", flush=True)
                print(f"  DEBUG: Primeiras 5 datas do ano anterior: {ano_anterior_datas[:5]}", flush=True)

        elif data_inicio_aa and data_fim_aa:
            # Para mensal e semanal: buscar pelo período de datas
            for data_str in datas_ordenadas:
                if data_inicio_aa <= data_str <= data_fim_aa:
                    ano_anterior_datas.append(data_str)
                    ano_anterior_valores.append(vendas_agregadas.get(data_str, 0))

        ano_anterior_total = sum(ano_anterior_valores)
        print(f"  Ano anterior: {len(ano_anterior_valores)} períodos, total: {ano_anterior_total:,.2f}", flush=True)

        # =====================================================
        # PASSO 7: MONTAR RESPOSTA
        # =====================================================
        print(f"\n[PASSO 7] Montando resposta...", flush=True)

        # Formatar períodos para frontend
        periodos_formatados = []
        for data_str in datas_previsao:
            try:
                if '-S' in data_str:
                    # Formato semanal: YYYY-SWW
                    ano_str, sem_str = data_str.split('-S')
                    periodos_formatados.append({
                        'semana': int(sem_str),
                        'ano': int(ano_str)
                    })
                else:
                    # Formato data: YYYY-MM-DD
                    data_obj = dt.strptime(data_str, '%Y-%m-%d')
                    if granularidade == 'mensal':
                        periodos_formatados.append({
                            'mes': data_obj.month,
                            'ano': data_obj.year
                        })
                    elif granularidade == 'semanal':
                        periodos_formatados.append({
                            'semana': data_obj.isocalendar()[1],
                            'ano': data_obj.isocalendar()[0]
                        })
                    else:
                        periodos_formatados.append(data_str)
            except Exception as e:
                print(f"  AVISO: Erro ao formatar período {data_str}: {e}", flush=True)
                periodos_formatados.append(data_str)

        # Usar métricas do backtest (comparação previsão vs dados reais do período de teste)
        # wmape_backtest e bias_backtest já foram calculados acima no bloco de backtest

        # Montar resposta no mesmo formato da v1
        resposta = {
            'sucesso': True,
            'versao': 'v2_bottom_up',
            'granularidade': granularidade,
            'filtros': {
                'loja': loja,
                'fornecedor': fornecedor,
                'linha': linha,
                'sublinha': sublinha,
                'produto': produto
            },
            'serie_temporal': {
                'datas': datas_historicas,
                'valores': serie_historica_agregada
            },
            # Dados divididos para gráfico (compatibilidade com frontend)
            'historico_base': {
                'datas': datas_base,
                'valores': valores_base
            },
            'historico_teste': {
                'datas': datas_teste,
                'valores': valores_teste
            },
            'melhor_modelo': 'Bottom-Up (Individual por Item)',
            'modelos': {
                'Bottom-Up (Individual por Item)': {
                    'metricas': {
                        'wmape': round(wmape_backtest, 2),
                        'bias': round(bias_backtest, 2)
                    },
                    'teste': {
                        'datas': datas_teste,
                        'valores': previsao_teste  # Previsão do modelo para o período de teste (linha verde tracejada)
                    },
                    'futuro': {
                        'datas': datas_previsao,
                        'valores': [round(v, 2) for v in previsao_agregada]
                    }
                }
            },
            'ano_anterior': {
                'datas': ano_anterior_datas,
                'valores': ano_anterior_valores,
                'total': ano_anterior_total
            },
            'relatorio_detalhado': {
                'itens': relatorio_itens,
                'periodos_previsao': periodos_formatados,
                'total_itens': len(relatorio_itens),
                'granularidade': granularidade
            },
            'estatisticas_modelos': contagem_modelos,
            'saneamento_rupturas': {
                'total_rupturas_detectadas': estatisticas_saneamento.get('total_rupturas', 0),
                'rupturas_saneadas': estatisticas_saneamento.get('rupturas_saneadas', 0),
                'metodos_usados': estatisticas_saneamento.get('metodos', {}),
                'descricao': 'Rupturas (venda=0 E estoque=0) foram substituídas por valores estimados usando interpolação sazonal'
            }
        }

        cursor.close()
        conn.close()

        print(f"\n{'='*60}", flush=True)
        print(f"PREVISÃO V2 CONCLUÍDA COM SUCESSO!", flush=True)
        print(f"  Total itens: {len(lista_itens)}", flush=True)
        print(f"  Total previsão: {total_previsao:,.2f}", flush=True)
        print(f"{'='*60}", flush=True)

        return jsonify(resposta)

    except Exception as e:
        import traceback
        erro_msg = traceback.format_exc()
        print(f"ERRO na previsão V2: {erro_msg}", flush=True)
        with open('erro_previsao_v2.log', 'w', encoding='utf-8') as f:
            f.write(erro_msg)
        return jsonify({'erro': str(e)}), 500


# ==============================================================================
# PARAMETROS DE FORNECEDOR - Importacao e Consulta
# ==============================================================================

@app.route('/importar-parametros-fornecedor')
def pagina_importar_parametros_fornecedor():
    """Pagina de importacao de parametros de fornecedor."""
    return render_template('importar_parametros_fornecedor.html')


@app.route('/api/parametros-fornecedor', methods=['GET'])
def api_listar_parametros_fornecedor():
    """Lista parametros de fornecedor cadastrados."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Verificar se tabela existe
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'parametros_fornecedor'
            )
        """)
        if not cursor.fetchone()['exists']:
            conn.close()
            return jsonify([])

        # Filtros opcionais
        cnpj = request.args.get('cnpj')
        cod_empresa = request.args.get('cod_empresa')

        query = """
            SELECT cnpj_fornecedor, nome_fornecedor, cod_empresa, tipo_destino,
                   lead_time_dias, ciclo_pedido_dias, pedido_minimo_valor, ativo
            FROM parametros_fornecedor
            WHERE ativo = TRUE
        """
        params = []

        if cnpj:
            query += " AND cnpj_fornecedor = %s"
            params.append(cnpj)

        if cod_empresa:
            query += " AND cod_empresa = %s"
            params.append(int(cod_empresa))

        query += " ORDER BY nome_fornecedor, cod_empresa"

        cursor.execute(query, params)
        resultados = cursor.fetchall()
        conn.close()

        return jsonify(resultados)

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/api/importar-parametros-fornecedor', methods=['POST'])
def api_importar_parametros_fornecedor():
    """Importa parametros de fornecedor do arquivo Excel."""
    try:
        dados = request.get_json()
        registros = dados.get('dados', [])
        modo = dados.get('modo', 'atualizar')

        if not registros:
            return jsonify({
                'sucesso': False,
                'mensagem': 'Nenhum registro para importar'
            })

        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Criar tabela se nao existir
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS parametros_fornecedor (
                id SERIAL PRIMARY KEY,
                cnpj_fornecedor VARCHAR(20) NOT NULL,
                nome_fornecedor VARCHAR(200),
                cod_empresa INTEGER NOT NULL,
                tipo_destino VARCHAR(20) DEFAULT 'LOJA',
                lead_time_dias INTEGER DEFAULT 15,
                ciclo_pedido_dias INTEGER DEFAULT 7,
                pedido_minimo_valor DECIMAL(12,2) DEFAULT 0,
                data_importacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                ativo BOOLEAN DEFAULT TRUE,
                UNIQUE(cnpj_fornecedor, cod_empresa)
            )
        """)
        conn.commit()

        # Criar indices
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_param_forn_cnpj ON parametros_fornecedor(cnpj_fornecedor)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_param_forn_empresa ON parametros_fornecedor(cod_empresa)")
        conn.commit()

        # Se modo substituir, desativar todos
        if modo == 'substituir':
            cursor.execute("UPDATE parametros_fornecedor SET ativo = FALSE")
            conn.commit()

        # Inserir/atualizar registros
        from psycopg2.extras import execute_values
        valores = []
        for r in registros:
            # Garantir CNPJ com 14 digitos (adicionar zeros a esquerda se necessario)
            cnpj_raw = str(r.get('cnpj_fornecedor', '')).strip()
            cnpj_formatado = cnpj_raw.zfill(14) if cnpj_raw and len(cnpj_raw) < 14 else cnpj_raw
            valores.append((
                cnpj_formatado,
                str(r.get('nome_fornecedor', '')).strip(),
                int(r.get('cod_empresa', 0)),
                str(r.get('tipo_destino', 'LOJA')).upper(),
                int(r.get('lead_time_dias', 15)),
                int(r.get('ciclo_pedido_dias', 7)),
                float(r.get('pedido_minimo_valor', 0))
            ))

        execute_values(cursor, """
            INSERT INTO parametros_fornecedor
                (cnpj_fornecedor, nome_fornecedor, cod_empresa, tipo_destino,
                 lead_time_dias, ciclo_pedido_dias, pedido_minimo_valor)
            VALUES %s
            ON CONFLICT (cnpj_fornecedor, cod_empresa) DO UPDATE SET
                nome_fornecedor = EXCLUDED.nome_fornecedor,
                tipo_destino = EXCLUDED.tipo_destino,
                lead_time_dias = EXCLUDED.lead_time_dias,
                ciclo_pedido_dias = EXCLUDED.ciclo_pedido_dias,
                pedido_minimo_valor = EXCLUDED.pedido_minimo_valor,
                data_importacao = CURRENT_TIMESTAMP,
                ativo = TRUE
        """, valores)

        conn.commit()
        conn.close()

        return jsonify({
            'sucesso': True,
            'mensagem': f'Importacao concluida com sucesso',
            'total': len(valores)
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'mensagem': f'Erro na importacao: {str(e)}'
        }), 500


@app.route('/api/parametros-fornecedor-produto', methods=['GET'])
def api_parametros_fornecedor_por_produto():
    """
    Busca parametros do fornecedor de um produto especifico.
    Params: codigo (codigo do produto), cod_empresa (loja destino)
    """
    try:
        codigo = request.args.get('codigo')
        cod_empresa = request.args.get('cod_empresa')

        if not codigo or not cod_empresa:
            return jsonify({'erro': 'Parametros codigo e cod_empresa sao obrigatorios'}), 400

        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Verificar se tabela existe
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'parametros_fornecedor'
            )
        """)
        if not cursor.fetchone()['exists']:
            conn.close()
            return jsonify({
                'encontrado': False,
                'mensagem': 'Tabela de parametros de fornecedor nao existe. Execute a importacao primeiro.'
            })

        # Buscar CNPJ do fornecedor do produto
        cursor.execute("""
            SELECT cnpj_fornecedor, nome_fornecedor
            FROM cadastro_produtos_completo
            WHERE cod_produto = %s
            LIMIT 1
        """, [str(codigo)])

        produto = cursor.fetchone()
        if not produto or not produto['cnpj_fornecedor']:
            conn.close()
            return jsonify({
                'encontrado': False,
                'mensagem': 'Produto nao encontrado ou sem fornecedor cadastrado'
            })

        cnpj = str(produto['cnpj_fornecedor']).strip()

        # Buscar parametros do fornecedor para a loja
        cursor.execute("""
            SELECT cnpj_fornecedor, nome_fornecedor, cod_empresa, tipo_destino,
                   lead_time_dias, ciclo_pedido_dias, pedido_minimo_valor
            FROM parametros_fornecedor
            WHERE cnpj_fornecedor = %s
              AND cod_empresa = %s
              AND ativo = TRUE
            LIMIT 1
        """, [cnpj, int(cod_empresa)])

        params = cursor.fetchone()
        conn.close()

        if not params:
            return jsonify({
                'encontrado': False,
                'cnpj_fornecedor': cnpj,
                'nome_fornecedor': produto.get('nome_fornecedor', ''),
                'mensagem': f'Fornecedor nao cadastrado para loja {cod_empresa}'
            })

        return jsonify({
            'encontrado': True,
            'cnpj_fornecedor': params['cnpj_fornecedor'],
            'nome_fornecedor': params['nome_fornecedor'],
            'cod_empresa': params['cod_empresa'],
            'tipo_destino': params['tipo_destino'],
            'lead_time_dias': params['lead_time_dias'],
            'ciclo_pedido_dias': params['ciclo_pedido_dias'],
            'pedido_minimo_valor': float(params['pedido_minimo_valor']) if params['pedido_minimo_valor'] else 0
        })

    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ============================================================================
# CENTRAL DE PARAMETROS DE FORNECEDOR - Consulta e Edicao
# ============================================================================

@app.route('/central-parametros-fornecedor')
def pagina_central_parametros_fornecedor():
    """Pagina da Central de Parametros de Fornecedor."""
    return render_template('central_parametros_fornecedor.html')


@app.route('/api/central-parametros/buscar', methods=['GET'])
def api_central_parametros_buscar():
    """
    Busca fornecedor e seus produtos para consulta de parametros.
    Params: codigo (produto), cnpj, nome (fornecedor), cod_empresa (loja)
    """
    conn = None
    try:
        codigo = request.args.get('codigo', '').strip()
        cnpj = request.args.get('cnpj', '').strip()
        nome = request.args.get('nome', '').strip()
        cod_empresa = request.args.get('cod_empresa', '').strip()

        conn = psycopg2.connect(**DB_CONFIG)
        conn.autocommit = True  # Evita problemas de transacao abortada
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cnpj_fornecedor = None
        nome_fornecedor = None

        # Se busca por codigo de produto, primeiro encontrar o fornecedor
        if codigo:
            cursor.execute("""
                SELECT cnpj_fornecedor, nome_fornecedor
                FROM cadastro_produtos_completo
                WHERE cod_produto = %s
                LIMIT 1
            """, [codigo])
            produto = cursor.fetchone()

            # Se nao encontrou por codigo exato, buscar por LIKE
            if not produto:
                cursor.execute("""
                    SELECT cnpj_fornecedor, nome_fornecedor
                    FROM cadastro_produtos_completo
                    WHERE cod_produto LIKE %s
                    LIMIT 1
                """, [f"%{codigo}%"])
                produto = cursor.fetchone()

            if produto:
                cnpj_fornecedor = produto['cnpj_fornecedor']
                nome_fornecedor = produto['nome_fornecedor']

        # Se busca por CNPJ
        elif cnpj:
            # Normalizar CNPJ (remover formatacao e adicionar zeros)
            cnpj_limpo = cnpj.replace('.', '').replace('/', '').replace('-', '').strip()
            if len(cnpj_limpo) < 14:
                cnpj_limpo = cnpj_limpo.zfill(14)
            cnpj_fornecedor = cnpj_limpo

            # Buscar nome do fornecedor
            cursor.execute("""
                SELECT DISTINCT nome_fornecedor
                FROM cadastro_produtos_completo
                WHERE cnpj_fornecedor = %s
                LIMIT 1
            """, [cnpj_fornecedor])
            forn = cursor.fetchone()
            if forn:
                nome_fornecedor = forn['nome_fornecedor']

        # Se busca por nome
        elif nome:
            cursor.execute("""
                SELECT DISTINCT cnpj_fornecedor, nome_fornecedor
                FROM cadastro_produtos_completo
                WHERE UPPER(nome_fornecedor) LIKE UPPER(%s)
                AND cnpj_fornecedor IS NOT NULL
                LIMIT 1
            """, [f"%{nome}%"])
            forn = cursor.fetchone()
            if forn:
                cnpj_fornecedor = forn['cnpj_fornecedor']
                nome_fornecedor = forn['nome_fornecedor']

        if not cnpj_fornecedor:
            if conn:
                conn.close()
            return jsonify({
                'success': False,
                'mensagem': 'Fornecedor nao encontrado com os filtros informados'
            })

        # Buscar parametros do fornecedor
        fornecedor_params = None
        parametros_cadastrados = False

        # Verificar se tabela de parametros existe
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'parametros_fornecedor'
            )
        """)
        tabela_params_existe = cursor.fetchone()['exists']

        if tabela_params_existe:
            query_params = """
                SELECT cnpj_fornecedor, nome_fornecedor, cod_empresa, tipo_destino,
                       lead_time_dias, ciclo_pedido_dias, pedido_minimo_valor
                FROM parametros_fornecedor
                WHERE cnpj_fornecedor = %s AND ativo = TRUE
            """
            params_list = [cnpj_fornecedor]

            if cod_empresa:
                query_params += " AND cod_empresa = %s"
                params_list.append(int(cod_empresa))

            query_params += " ORDER BY cod_empresa LIMIT 1"
            cursor.execute(query_params, params_list)
            fornecedor_params = cursor.fetchone()
            parametros_cadastrados = fornecedor_params is not None

        # Montar dados do fornecedor
        fornecedor = {
            'cnpj_fornecedor': cnpj_fornecedor,
            'nome_fornecedor': nome_fornecedor,
            'cod_empresa': int(cod_empresa) if cod_empresa else (fornecedor_params['cod_empresa'] if fornecedor_params else None),
            'tipo_destino': fornecedor_params['tipo_destino'] if fornecedor_params else 'LOJA',
            'lead_time_dias': fornecedor_params['lead_time_dias'] if fornecedor_params else 15,
            'ciclo_pedido_dias': fornecedor_params['ciclo_pedido_dias'] if fornecedor_params else 7,
            'pedido_minimo_valor': float(fornecedor_params['pedido_minimo_valor']) if fornecedor_params and fornecedor_params['pedido_minimo_valor'] else 0,
            'parametros_cadastrados': parametros_cadastrados
        }

        # Buscar produtos do fornecedor - query simplificada e robusta
        cod_empresa_filter = int(cod_empresa) if cod_empresa else None

        # Verificar quais tabelas existem para montar query dinamicamente
        cursor.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_name IN ('cadastro_produtos', 'estoque_posicao_atual', 'situacao_compra_itens')
        """)
        tabelas_existentes = [row['table_name'] for row in cursor.fetchall()]

        # Query base - sempre funciona
        produtos = []
        cursor.execute("""
            SELECT cod_produto, descricao, categoria, codigo_linha, descricao_linha
            FROM cadastro_produtos_completo
            WHERE cnpj_fornecedor = %s AND ativo = TRUE
            ORDER BY descricao
            LIMIT 500
        """, [cnpj_fornecedor])
        produtos_base = cursor.fetchall()

        # Enriquecer com dados adicionais se as tabelas existirem
        for p in produtos_base:
            produto = dict(p)
            produto['curva_abc'] = None
            produto['preco_custo'] = None
            produto['sit_compra'] = None

            # Buscar curva ABC
            if 'cadastro_produtos' in tabelas_existentes:
                try:
                    cursor.execute("""
                        SELECT curva_abc FROM cadastro_produtos
                        WHERE codigo::text = %s
                        LIMIT 1
                    """, [produto['cod_produto']])
                    abc = cursor.fetchone()
                    if abc:
                        produto['curva_abc'] = abc['curva_abc']
                except:
                    pass

            # Buscar CUE (Custo Unitario de Entrada) da tabela estoque_posicao_atual
            # Mesma fonte usada no calculo do pedido ao fornecedor
            if 'estoque_posicao_atual' in tabelas_existentes:
                try:
                    if cod_empresa_filter:
                        cursor.execute("""
                            SELECT cue FROM estoque_posicao_atual
                            WHERE codigo::text = %s AND cod_empresa = %s
                            LIMIT 1
                        """, [produto['cod_produto'], cod_empresa_filter])
                    else:
                        cursor.execute("""
                            SELECT cue FROM estoque_posicao_atual
                            WHERE codigo::text = %s
                            ORDER BY cod_empresa
                            LIMIT 1
                        """, [produto['cod_produto']])
                    cue = cursor.fetchone()
                    if cue and cue['cue']:
                        produto['preco_custo'] = float(cue['cue'])
                except:
                    pass

            # Buscar situacao compra
            if 'situacao_compra_itens' in tabelas_existentes:
                try:
                    if cod_empresa_filter:
                        cursor.execute("""
                            SELECT sit_compra FROM situacao_compra_itens
                            WHERE codigo::text = %s AND cod_empresa = %s
                            LIMIT 1
                        """, [produto['cod_produto'], cod_empresa_filter])
                    else:
                        cursor.execute("""
                            SELECT sit_compra FROM situacao_compra_itens
                            WHERE codigo::text = %s
                            LIMIT 1
                        """, [produto['cod_produto']])
                    sit = cursor.fetchone()
                    if sit:
                        produto['sit_compra'] = sit['sit_compra']
                except:
                    pass

            produtos.append(produto)

        conn.close()

        return jsonify({
            'success': True,
            'fornecedor': fornecedor,
            'produtos': produtos
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        if conn:
            try:
                conn.close()
            except:
                pass
        return jsonify({
            'success': False,
            'mensagem': f'Erro ao buscar: {str(e)}'
        }), 500


@app.route('/api/central-parametros/salvar', methods=['POST'])
def api_central_parametros_salvar():
    """
    Salva ou atualiza parametros de um fornecedor.
    """
    try:
        dados = request.get_json()

        cnpj = dados.get('cnpj_fornecedor')
        nome = dados.get('nome_fornecedor', '')
        cod_empresa = dados.get('cod_empresa')
        lead_time = dados.get('lead_time_dias', 15)
        ciclo = dados.get('ciclo_pedido_dias', 7)
        fat_minimo = dados.get('pedido_minimo_valor', 0)
        tipo_destino = dados.get('tipo_destino', 'LOJA')

        if not cnpj:
            return jsonify({'success': False, 'mensagem': 'CNPJ e obrigatorio'})

        if not cod_empresa:
            return jsonify({'success': False, 'mensagem': 'Codigo da empresa e obrigatorio'})

        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Criar tabela se nao existir
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS parametros_fornecedor (
                id SERIAL PRIMARY KEY,
                cnpj_fornecedor VARCHAR(20) NOT NULL,
                nome_fornecedor VARCHAR(200),
                cod_empresa INTEGER NOT NULL,
                tipo_destino VARCHAR(20) DEFAULT 'LOJA',
                lead_time_dias INTEGER DEFAULT 15,
                ciclo_pedido_dias INTEGER DEFAULT 7,
                pedido_minimo_valor DECIMAL(12,2) DEFAULT 0,
                data_importacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                ativo BOOLEAN DEFAULT TRUE,
                UNIQUE(cnpj_fornecedor, cod_empresa)
            )
        """)

        # Inserir ou atualizar
        cursor.execute("""
            INSERT INTO parametros_fornecedor
                (cnpj_fornecedor, nome_fornecedor, cod_empresa, tipo_destino,
                 lead_time_dias, ciclo_pedido_dias, pedido_minimo_valor, ativo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
            ON CONFLICT (cnpj_fornecedor, cod_empresa) DO UPDATE SET
                nome_fornecedor = EXCLUDED.nome_fornecedor,
                tipo_destino = EXCLUDED.tipo_destino,
                lead_time_dias = EXCLUDED.lead_time_dias,
                ciclo_pedido_dias = EXCLUDED.ciclo_pedido_dias,
                pedido_minimo_valor = EXCLUDED.pedido_minimo_valor,
                data_importacao = CURRENT_TIMESTAMP,
                ativo = TRUE
        """, [cnpj, nome, cod_empresa, tipo_destino, lead_time, ciclo, fat_minimo])

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'mensagem': 'Parametros salvos com sucesso'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'mensagem': f'Erro ao salvar: {str(e)}'
        }), 500


# ============================================================================
# APIS DE DEMANDA VALIDADA E PEDIDO PLANEJADO (v6.0)
# ============================================================================

@app.route('/api/demanda_validada/salvar', methods=['POST'])
def api_salvar_demanda_validada():
    """
    Salva demanda validada pelo usuario para um periodo especifico.

    Request JSON:
    {
        "itens": [
            {
                "cod_produto": "123456",
                "cod_loja": 1,
                "cod_fornecedor": 100,
                "data_inicio": "2026-07-01",
                "data_fim": "2026-07-31",
                "demanda_diaria": 10.5,
                "demanda_total_periodo": 325.5
            }
        ],
        "usuario": "nome_usuario",
        "observacao": "Validacao para julho 2026"
    }
    """
    try:
        dados = request.get_json()

        itens = dados.get('itens', [])
        usuario = dados.get('usuario', 'sistema')
        observacao = dados.get('observacao', '')

        if not itens:
            return jsonify({'success': False, 'erro': 'Nenhum item para salvar'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # Verificar se tabela existe, criar se necessario
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'demanda_validada'
            )
        """)
        if not cursor.fetchone()[0]:
            # Criar tabela se nao existir
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS demanda_validada (
                    id SERIAL PRIMARY KEY,
                    cod_produto VARCHAR(20) NOT NULL,
                    cod_loja INTEGER NOT NULL,
                    cod_fornecedor INTEGER,
                    data_inicio DATE NOT NULL,
                    data_fim DATE NOT NULL,
                    semana_ano VARCHAR(10),
                    demanda_diaria NUMERIC(12,4),
                    demanda_total_periodo NUMERIC(12,2),
                    data_validacao TIMESTAMP DEFAULT NOW(),
                    usuario_validacao VARCHAR(100),
                    versao INTEGER DEFAULT 1,
                    status VARCHAR(20) DEFAULT 'validado',
                    ativo BOOLEAN DEFAULT TRUE,
                    observacao TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()

        itens_salvos = 0
        for item in itens:
            cod_produto = str(item.get('cod_produto'))
            # cod_loja pode ser None quando selecionado "TODAS" as lojas
            cod_loja_raw = item.get('cod_loja')
            cod_loja = int(cod_loja_raw) if cod_loja_raw is not None else 0  # 0 = todas as lojas
            cod_fornecedor = item.get('cod_fornecedor')
            data_inicio = item.get('data_inicio')
            data_fim = item.get('data_fim')
            demanda_diaria = float(item.get('demanda_diaria') or 0)
            demanda_total = float(item.get('demanda_total_periodo') or 0)

            # Calcular semana_ano (formato: YYYY-SWW)
            from datetime import datetime as dt
            data_ini = dt.strptime(data_inicio, '%Y-%m-%d')
            semana_ano = f"{data_ini.year}-S{data_ini.isocalendar()[1]:02d}"

            # Desativar validacoes anteriores que se sobrepoem
            cursor.execute("""
                UPDATE demanda_validada
                SET ativo = FALSE, updated_at = NOW()
                WHERE cod_produto = %s
                  AND cod_loja = %s
                  AND ativo = TRUE
                  AND (
                      (data_inicio <= %s AND data_fim >= %s)
                      OR (data_inicio >= %s AND data_inicio <= %s)
                      OR (data_fim >= %s AND data_fim <= %s)
                  )
            """, [cod_produto, cod_loja, data_inicio, data_fim,
                  data_inicio, data_fim, data_inicio, data_fim])

            # Inserir nova validacao
            cursor.execute("""
                INSERT INTO demanda_validada
                (cod_produto, cod_loja, cod_fornecedor, data_inicio, data_fim,
                 semana_ano, demanda_diaria, demanda_total_periodo,
                 usuario_validacao, observacao)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [cod_produto, cod_loja, cod_fornecedor, data_inicio, data_fim,
                  semana_ano, demanda_diaria, demanda_total, usuario, observacao])

            itens_salvos += 1

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'itens_salvos': itens_salvos,
            'mensagem': f'{itens_salvos} itens validados com sucesso'
        })

    except Exception as e:
        print(f"[ERRO] api_salvar_demanda_validada: {e}")
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/api/demanda_validada/listar', methods=['GET'])
def api_listar_demanda_validada():
    """
    Lista demandas validadas com filtros opcionais.

    Query params:
        cod_produto: filtrar por produto
        cod_loja: filtrar por loja
        cod_fornecedor: filtrar por fornecedor
        data_inicio: filtrar por periodo (inicio)
        data_fim: filtrar por periodo (fim)
        apenas_ativos: true/false (default: true)
    """
    try:
        cod_produto = request.args.get('cod_produto')
        cod_loja = request.args.get('cod_loja')
        cod_fornecedor = request.args.get('cod_fornecedor')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Verificar se tabela existe
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'demanda_validada'
            )
        """)
        if not cursor.fetchone()['exists']:
            conn.close()
            return jsonify([])

        # Construir query com filtros
        where_conditions = []
        params = []

        if apenas_ativos:
            where_conditions.append("ativo = TRUE")

        if cod_produto:
            where_conditions.append("cod_produto = %s")
            params.append(str(cod_produto))

        if cod_loja:
            where_conditions.append("cod_loja = %s")
            params.append(int(cod_loja))

        if cod_fornecedor:
            where_conditions.append("cod_fornecedor = %s")
            params.append(int(cod_fornecedor))

        if data_inicio:
            where_conditions.append("data_fim >= %s")
            params.append(data_inicio)

        if data_fim:
            where_conditions.append("data_inicio <= %s")
            params.append(data_fim)

        where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"

        cursor.execute(f"""
            SELECT id, cod_produto, cod_loja, cod_fornecedor,
                   data_inicio, data_fim, semana_ano,
                   demanda_diaria, demanda_total_periodo,
                   data_validacao, usuario_validacao,
                   versao, status, observacao
            FROM demanda_validada
            WHERE {where_clause}
            ORDER BY data_validacao DESC
            LIMIT 1000
        """, params)

        resultados = cursor.fetchall()
        conn.close()

        # Converter datas para string
        for r in resultados:
            if r.get('data_inicio'):
                r['data_inicio'] = str(r['data_inicio'])
            if r.get('data_fim'):
                r['data_fim'] = str(r['data_fim'])
            if r.get('data_validacao'):
                r['data_validacao'] = str(r['data_validacao'])

        return jsonify(resultados)

    except Exception as e:
        print(f"[ERRO] api_listar_demanda_validada: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/demanda_validada/buscar_para_periodo', methods=['GET'])
def api_buscar_demanda_validada_periodo():
    """
    Busca demanda validada para um periodo especifico (usado no pedido planejado).

    Query params:
        cod_loja: codigo da loja (obrigatorio)
        data_inicio: inicio do periodo (obrigatorio)
        data_fim: fim do periodo (obrigatorio)
        cod_fornecedor: filtrar por fornecedor (opcional)
    """
    try:
        cod_loja = request.args.get('cod_loja')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        cod_fornecedor = request.args.get('cod_fornecedor')

        if not cod_loja or not data_inicio or not data_fim:
            return jsonify({'erro': 'Parametros obrigatorios: cod_loja, data_inicio, data_fim'}), 400

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Verificar se tabela existe
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'demanda_validada'
            )
        """)
        if not cursor.fetchone()['exists']:
            conn.close()
            return jsonify({'itens': [], 'tem_validacao': False})

        # Buscar validacoes que cobrem o periodo
        params = [int(cod_loja), data_inicio, data_fim]
        fornecedor_filter = ""
        if cod_fornecedor:
            fornecedor_filter = " AND cod_fornecedor = %s"
            params.append(int(cod_fornecedor))

        cursor.execute(f"""
            SELECT cod_produto, cod_loja, cod_fornecedor,
                   data_inicio, data_fim,
                   demanda_diaria, demanda_total_periodo,
                   data_validacao, usuario_validacao
            FROM demanda_validada
            WHERE cod_loja = %s
              AND ativo = TRUE
              AND status = 'validado'
              AND data_inicio <= %s
              AND data_fim >= %s
              {fornecedor_filter}
            ORDER BY data_validacao DESC
        """, params)

        resultados = cursor.fetchall()
        conn.close()

        # Converter para dict indexado por produto
        itens_dict = {}
        for r in resultados:
            cod = str(r['cod_produto'])
            if cod not in itens_dict:  # Manter apenas a mais recente
                itens_dict[cod] = {
                    'cod_produto': cod,
                    'demanda_diaria': float(r['demanda_diaria']) if r['demanda_diaria'] else 0,
                    'demanda_total_periodo': float(r['demanda_total_periodo']) if r['demanda_total_periodo'] else 0,
                    'data_validacao': str(r['data_validacao']),
                    'usuario_validacao': r['usuario_validacao']
                }

        return jsonify({
            'itens': list(itens_dict.values()),
            'tem_validacao': len(itens_dict) > 0,
            'total_itens': len(itens_dict)
        })

    except Exception as e:
        print(f"[ERRO] api_buscar_demanda_validada_periodo: {e}")
        return jsonify({'erro': str(e)}), 500


@app.route('/api/pedido_planejado', methods=['POST'])
def api_pedido_planejado():
    """
    Gera pedido planejado para um periodo futuro.

    Diferenca do pedido normal (reabastecimento):
    - Usa demanda validada para o periodo especificado
    - Calcula data de emissao com offset de lead time
    - Nao considera estoque atual (assume que sera consumido ate la)

    Request JSON:
    {
        "destino_tipo": "Loja" ou "CD",
        "cod_empresa": codigo ou lista de codigos ou "TODAS",
        "fornecedor": nome ou lista de nomes ou "TODOS",
        "linha1": categoria ou lista ou "TODAS",
        "linha3": sublinha ou lista ou "TODAS",
        "periodo_inicio": "2026-07-01",
        "periodo_fim": "2026-07-31",
        "usar_demanda_validada": true/false (default: true)
    }
    """
    try:
        from core.pedido_fornecedor_integrado import (
            PedidoFornecedorIntegrado,
            agregar_por_fornecedor
        )
        from datetime import datetime as dt, timedelta
        import math

        dados = request.get_json()

        # Parametros de filtro (igual ao pedido normal)
        fornecedor_filtro = dados.get('fornecedor', 'TODOS')
        linha1_filtro = dados.get('linha1', 'TODAS')
        linha3_filtro = dados.get('linha3', 'TODAS')
        destino_tipo = dados.get('destino_tipo', 'LOJA')
        cod_empresa = dados.get('cod_empresa', 'TODAS')

        # Parametros especificos do pedido planejado
        periodo_inicio = dados.get('periodo_inicio')
        periodo_fim = dados.get('periodo_fim')
        usar_demanda_validada = dados.get('usar_demanda_validada', True)

        if not periodo_inicio or not periodo_fim:
            return jsonify({'erro': 'Parametros obrigatorios: periodo_inicio, periodo_fim'}), 400

        print(f"\n{'='*60}")
        print("PEDIDO PLANEJADO")
        print(f"{'='*60}")
        print(f"  Periodo: {periodo_inicio} ate {periodo_fim}")
        print(f"  Fornecedor: {fornecedor_filtro}")
        print(f"  Destino: {destino_tipo}")
        print(f"  Empresa: {cod_empresa}")
        print(f"  Usar demanda validada: {usar_demanda_validada}")

        conn = get_db_connection()

        # Normalizar cod_empresa
        lojas_selecionadas = []
        if isinstance(cod_empresa, list):
            lojas_selecionadas = [int(x) for x in cod_empresa]
        elif cod_empresa != 'TODAS':
            lojas_selecionadas = [int(cod_empresa)]

        # Buscar lojas se nao especificadas
        if not lojas_selecionadas:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            tipo_filtro = 'Loja' if destino_tipo.upper() == 'LOJA' else 'CD'
            cursor.execute("""
                SELECT cod_empresa FROM cadastro_lojas
                WHERE tipo = %s AND ativo = TRUE
            """, [tipo_filtro])
            lojas_selecionadas = [r['cod_empresa'] for r in cursor.fetchall()]

        # Calcular dias do periodo
        data_ini = dt.strptime(periodo_inicio, '%Y-%m-%d')
        data_fim_dt = dt.strptime(periodo_fim, '%Y-%m-%d')
        dias_periodo = (data_fim_dt - data_ini).days + 1

        print(f"  Dias no periodo: {dias_periodo}")
        print(f"  Lojas: {lojas_selecionadas}")

        # Buscar demanda validada se solicitado
        demanda_validada_dict = {}
        demanda_agregada_dict = {}  # Demanda com cod_loja=0 (agregada)
        proporcao_vendas_loja = {}  # Proporção histórica de vendas por loja/produto

        if usar_demanda_validada:
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            # 1. Buscar demanda específica por loja
            for loja in lojas_selecionadas:
                cursor.execute("""
                    SELECT cod_produto, demanda_diaria
                    FROM demanda_validada
                    WHERE cod_loja = %s
                      AND ativo = TRUE
                      AND status = 'validado'
                      AND data_inicio <= %s
                      AND data_fim >= %s
                """, [loja, periodo_inicio, periodo_fim])

                for r in cursor.fetchall():
                    key = (str(r['cod_produto']), loja)
                    demanda_validada_dict[key] = float(r['demanda_diaria']) if r['demanda_diaria'] else 0

            print(f"  Demanda validada por loja: {len(demanda_validada_dict)} itens")

            # 2. Buscar demanda agregada (cod_loja=0) para fallback
            cursor.execute("""
                SELECT cod_produto, demanda_diaria
                FROM demanda_validada
                WHERE cod_loja = 0
                  AND ativo = TRUE
                  AND status = 'validado'
                  AND data_inicio <= %s
                  AND data_fim >= %s
            """, [periodo_inicio, periodo_fim])

            for r in cursor.fetchall():
                cod_produto = str(r['cod_produto'])
                demanda_agregada_dict[cod_produto] = float(r['demanda_diaria']) if r['demanda_diaria'] else 0

            print(f"  Demanda validada agregada (fallback): {len(demanda_agregada_dict)} itens")

            # 3. Calcular proporção histórica de vendas por loja para cada produto agregado
            if demanda_agregada_dict and len(lojas_selecionadas) > 1:
                produtos_agregados = list(demanda_agregada_dict.keys())
                placeholders = ','.join(['%s'] * len(produtos_agregados))
                lojas_placeholders = ','.join(['%s'] * len(lojas_selecionadas))

                # Buscar vendas históricas por loja para calcular proporção
                cursor.execute(f"""
                    SELECT
                        codigo::text as cod_produto,
                        cod_empresa as cod_loja,
                        COALESCE(SUM(qtd_venda), 0) as total_vendas
                    FROM historico_vendas_diario
                    WHERE codigo::text IN ({placeholders})
                      AND cod_empresa IN ({lojas_placeholders})
                      AND data >= CURRENT_DATE - INTERVAL '365 days'
                    GROUP BY codigo, cod_empresa
                """, produtos_agregados + lojas_selecionadas)

                # Calcular total por produto e proporção por loja
                vendas_por_produto = {}
                for r in cursor.fetchall():
                    cod = str(r['cod_produto'])
                    loja = r['cod_loja']
                    vendas = float(r['total_vendas']) if r['total_vendas'] else 0

                    if cod not in vendas_por_produto:
                        vendas_por_produto[cod] = {}
                    vendas_por_produto[cod][loja] = vendas

                # Calcular proporção para cada produto/loja
                for cod_produto, vendas_lojas in vendas_por_produto.items():
                    total_vendas = sum(vendas_lojas.values())
                    if total_vendas > 0:
                        # Calcular proporção para lojas com vendas
                        for loja_v, vendas in vendas_lojas.items():
                            proporcao = vendas / total_vendas
                            proporcao_vendas_loja[(cod_produto, loja_v)] = proporcao

                        # Lojas sem vendas deste produto recebem proporção 0
                        # (a demanda será distribuída apenas para quem vende)
                        for loja_s in lojas_selecionadas:
                            if loja_s not in vendas_lojas:
                                proporcao_vendas_loja[(cod_produto, loja_s)] = 0.0
                    else:
                        # Se não houver histórico, distribuir igualmente
                        proporcao_igual = 1.0 / len(lojas_selecionadas)
                        for loja_s in lojas_selecionadas:
                            proporcao_vendas_loja[(cod_produto, loja_s)] = proporcao_igual

                # Garantir que produtos sem histórico em alguma loja tenham proporção
                for cod_produto in produtos_agregados:
                    if cod_produto not in vendas_por_produto:
                        # Produto sem histórico em nenhuma loja - distribuir igualmente
                        proporcao_igual = 1.0 / len(lojas_selecionadas)
                        for loja_s in lojas_selecionadas:
                            proporcao_vendas_loja[(cod_produto, loja_s)] = proporcao_igual

                print(f"  Proporções calculadas para distribuição: {len(proporcao_vendas_loja)} combinações produto/loja")

        # Instanciar processador de pedidos
        processador = PedidoFornecedorIntegrado(conn)
        processador.carregar_regras_sit_compra()

        # Normalizar filtros (mesmo codigo do pedido normal)
        fornecedores_lista = None
        if isinstance(fornecedor_filtro, list):
            fornecedores_lista = fornecedor_filtro
        elif fornecedor_filtro and fornecedor_filtro != 'TODOS':
            fornecedores_lista = [fornecedor_filtro]

        linhas1_lista = None
        if isinstance(linha1_filtro, list):
            linhas1_lista = linha1_filtro
        elif linha1_filtro and linha1_filtro != 'TODAS':
            linhas1_lista = [linha1_filtro]

        linhas3_lista = None
        if isinstance(linha3_filtro, list):
            linhas3_lista = linha3_filtro
        elif linha3_filtro and linha3_filtro != 'TODAS':
            linhas3_lista = [linha3_filtro]

        # Buscar produtos
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        where_conditions = ["1=1"]
        params = []

        if fornecedores_lista:
            placeholders = ','.join(['%s'] * len(fornecedores_lista))
            where_conditions.append(f"p.nome_fornecedor IN ({placeholders})")
            params.extend(fornecedores_lista)

        if linhas1_lista:
            placeholders = ','.join(['%s'] * len(linhas1_lista))
            # categoria corresponde a Linha 1 (DESCR_LINHA1)
            where_conditions.append(f"p.categoria IN ({placeholders})")
            params.extend(linhas1_lista)

        if linhas3_lista:
            placeholders = ','.join(['%s'] * len(linhas3_lista))
            where_conditions.append(f"p.codigo_linha IN ({placeholders})")
            params.extend(linhas3_lista)

        where_clause = " AND ".join(where_conditions)

        cursor.execute(f"""
            SELECT DISTINCT
                p.cod_produto as codigo,
                p.descricao,
                p.nome_fornecedor,
                p.cnpj_fornecedor,
                p.categoria as linha,
                p.codigo_linha,
                'B' as curva_abc,
                0 as custo_unitario
            FROM cadastro_produtos_completo p
            WHERE {where_clause}
        """, params)

        produtos = cursor.fetchall()
        print(f"  Produtos encontrados: {len(produtos)}")

        # Processar itens
        itens_pedido = []
        itens_bloqueados = []

        for loja in lojas_selecionadas:
            # Buscar nome da loja
            cursor.execute("""
                SELECT nome_loja FROM cadastro_lojas WHERE cod_empresa = %s
            """, [loja])
            nome_loja_row = cursor.fetchone()
            nome_loja = nome_loja_row['nome_loja'] if nome_loja_row else f'Loja {loja}'

            # Buscar lead time e pedido minimo do fornecedor para esta loja
            lead_times = {}
            pedidos_minimos = {}
            cursor.execute("""
                SELECT cnpj_fornecedor, lead_time_dias, COALESCE(pedido_minimo_valor, 0) as pedido_minimo_valor
                FROM parametros_fornecedor
                WHERE cod_empresa = %s AND ativo = TRUE
            """, [loja])
            for r in cursor.fetchall():
                lead_times[r['cnpj_fornecedor']] = r['lead_time_dias']
                pedidos_minimos[r['cnpj_fornecedor']] = float(r['pedido_minimo_valor'])

            for produto in produtos:
                codigo = str(produto['codigo'])

                # Verificar situacao de compra
                # verificar_situacao_compra retorna None se liberado, ou dict com info do bloqueio
                situacao = processador.verificar_situacao_compra(codigo, loja)
                if situacao is not None:
                    itens_bloqueados.append({
                        'codigo': codigo,
                        'descricao': produto['descricao'],
                        'nome_fornecedor': produto['nome_fornecedor'],
                        'cod_loja': loja,
                        'nome_loja': nome_loja,
                        'motivo_bloqueio': situacao.get('descricao', 'Bloqueado'),
                        'curva_abc': produto.get('curva_abc', 'B')
                    })
                    continue

                # Buscar demanda (com fallback para demanda agregada)
                key = (codigo, loja)
                usa_demanda_validada_loja = False
                usa_demanda_agregada = False

                if key in demanda_validada_dict:
                    # 1. Prioridade: demanda validada específica da loja
                    demanda_diaria = demanda_validada_dict[key]
                    usa_demanda_validada_loja = True
                elif codigo in demanda_agregada_dict:
                    # 2. Fallback: demanda agregada distribuída proporcionalmente
                    demanda_agregada_total = demanda_agregada_dict[codigo]
                    proporcao_key = (codigo, loja)

                    if proporcao_key in proporcao_vendas_loja:
                        proporcao = proporcao_vendas_loja[proporcao_key]
                        if proporcao > 0:
                            # Distribuir proporcionalmente ao histórico de vendas da loja
                            demanda_diaria = demanda_agregada_total * proporcao
                            usa_demanda_agregada = True
                        else:
                            # Loja não vende este produto - usar histórico próprio
                            demanda_diaria = 0  # Será calculado abaixo
                    else:
                        # Se não tiver proporção calculada, distribuir igualmente
                        demanda_diaria = demanda_agregada_total / len(lojas_selecionadas)
                        usa_demanda_agregada = True

                if not usa_demanda_validada_loja and not usa_demanda_agregada:
                    # 3. Último recurso: calcular demanda com base no histórico (média últimos 90 dias)
                    cursor.execute("""
                        SELECT AVG(qtd_venda) as media_diaria
                        FROM historico_vendas_diario
                        WHERE codigo = %s::integer
                          AND cod_empresa = %s
                          AND data >= CURRENT_DATE - INTERVAL '90 days'
                    """, [codigo, loja])
                    row = cursor.fetchone()
                    demanda_diaria = float(row['media_diaria']) if row and row['media_diaria'] else 0

                if demanda_diaria <= 0:
                    continue  # Sem demanda, pular

                demanda_periodo = demanda_diaria * dias_periodo

                # Buscar estoque atual e custo da tabela estoque_posicao_atual
                cursor.execute("""
                    SELECT
                        COALESCE(estoque, 0) as estoque_disponivel,
                        COALESCE(qtd_pendente, 0) as qtd_pendente,
                        COALESCE(qtd_pend_transf, 0) as qtd_pend_transf,
                        COALESCE(cue, 0) as custo_unitario
                    FROM estoque_posicao_atual
                    WHERE codigo = %s::integer AND cod_empresa = %s
                """, [codigo, loja])
                estoque_row = cursor.fetchone()

                if estoque_row:
                    estoque_disponivel = float(estoque_row['estoque_disponivel'])
                    qtd_pendente = float(estoque_row['qtd_pendente'])
                    qtd_pend_transf = float(estoque_row['qtd_pend_transf'])
                    custo_unitario = float(estoque_row['custo_unitario'])
                    estoque_transito = qtd_pendente + qtd_pend_transf
                else:
                    estoque_disponivel = 0
                    estoque_transito = 0
                    custo_unitario = 0

                # Calcular cobertura atual (em dias)
                if demanda_diaria > 0:
                    cobertura_atual = estoque_disponivel / demanda_diaria
                    cobertura_com_transito = (estoque_disponivel + estoque_transito) / demanda_diaria
                else:
                    cobertura_atual = 999
                    cobertura_com_transito = 999

                # Calcular lead time e obter pedido minimo
                cnpj = produto.get('cnpj_fornecedor')
                lead_time = lead_times.get(cnpj, 15)  # Default 15 dias
                pedido_minimo = pedidos_minimos.get(cnpj, 0)  # Default 0

                # Calcular data de emissao (periodo_inicio - lead_time)
                data_emissao = data_ini - timedelta(days=lead_time)

                # Calcular valor do pedido
                valor_pedido = demanda_periodo * custo_unitario

                # Calcular cobertura pós-pedido
                estoque_pos_pedido = estoque_disponivel + estoque_transito + demanda_periodo
                if demanda_diaria > 0:
                    cobertura_pos_pedido = estoque_pos_pedido / demanda_diaria
                else:
                    cobertura_pos_pedido = 999

                # Calcular numero de caixas (assumindo multiplo 1 se nao tiver parametro)
                # Usar ceil para garantir que demandas pequenas não fiquem zeradas
                qtd_pedido = max(1, math.ceil(demanda_periodo)) if demanda_periodo > 0 else 0
                numero_caixas = int(qtd_pedido)  # 1 unidade = 1 caixa por padrao

                # Campos compatíveis com lógica de transferências do pedido CD
                itens_pedido.append({
                    'codigo': codigo,
                    'descricao': produto['descricao'],
                    'codigo_fornecedor': cnpj,  # Usado para agregacao
                    'nome_fornecedor': produto['nome_fornecedor'],
                    'cnpj_fornecedor': cnpj,
                    'cod_loja': loja,
                    'nome_loja': nome_loja,
                    'curva_abc': produto.get('curva_abc', 'B'),
                    # Campos de demanda (compatíveis com pedido CD)
                    'demanda_diaria': round(demanda_diaria, 2),
                    'demanda_prevista_diaria': round(demanda_diaria, 2),  # Alias para compatibilidade
                    'demanda_periodo': math.ceil(demanda_periodo) if demanda_periodo > 0 else 0,
                    'quantidade_pedido': qtd_pedido,
                    'numero_caixas': numero_caixas,
                    'valor_pedido': round(valor_pedido, 2),
                    'custo_unitario': round(custo_unitario, 2),
                    'cue': round(custo_unitario, 2),  # Alias para compatibilidade
                    'estoque_atual': round(estoque_disponivel, 0),
                    'estoque_transito': round(estoque_transito, 0),
                    'cobertura_atual_dias': round(cobertura_atual, 1),
                    'cobertura_com_transito_dias': round(cobertura_com_transito, 1),
                    'cobertura_pos_pedido_dias': round(cobertura_pos_pedido, 1),
                    'lead_time': lead_time,
                    'pedido_minimo_fornecedor': pedido_minimo,
                    'data_emissao_sugerida': data_emissao.strftime('%Y-%m-%d'),
                    'periodo_inicio': periodo_inicio,
                    'periodo_fim': periodo_fim,
                    'dias_periodo': dias_periodo,
                    'usa_demanda_validada': usa_demanda_validada_loja,
                    'usa_demanda_agregada': usa_demanda_agregada,
                    'origem_demanda': 'validada_loja' if usa_demanda_validada_loja else ('agregada_proporcional' if usa_demanda_agregada else 'historico'),
                    'deve_pedir': True,  # Pedido planejado sempre deve pedir
                    'ruptura_iminente': cobertura_atual <= 3,  # Para cálculo de urgência
                    'risco_ruptura': cobertura_atual <= 7,
                    'bloqueado': False
                })

        conn.close()

        # =====================================================================
        # VALIDACAO DE PEDIDO MINIMO POR FORNECEDOR/LOJA
        # =====================================================================
        validacao_pedido_minimo = {}

        for item in itens_pedido:
            cod_fornecedor = item.get('codigo_fornecedor', '')
            cod_loja = item.get('cod_loja', 0)
            nome_fornecedor = item.get('nome_fornecedor', '')
            nome_loja = item.get('nome_loja', f'Loja {cod_loja}')
            valor_pedido = item.get('valor_pedido', 0) or 0
            pedido_min = item.get('pedido_minimo_fornecedor', 0) or 0

            chave = (cod_fornecedor, cod_loja)

            if chave not in validacao_pedido_minimo:
                validacao_pedido_minimo[chave] = {
                    'cod_fornecedor': cod_fornecedor,
                    'nome_fornecedor': nome_fornecedor,
                    'cod_loja': cod_loja,
                    'nome_loja': nome_loja,
                    'valor_total': 0,
                    'pedido_minimo': pedido_min,
                    'abaixo_minimo': False,
                    'diferenca': 0
                }

            validacao_pedido_minimo[chave]['valor_total'] += valor_pedido
            if validacao_pedido_minimo[chave]['pedido_minimo'] == 0 and pedido_min > 0:
                validacao_pedido_minimo[chave]['pedido_minimo'] = pedido_min

        # Verificar quais estao abaixo do minimo
        alertas_pedido_minimo = []
        for chave, info in validacao_pedido_minimo.items():
            pedido_min = info['pedido_minimo']
            valor_total = info['valor_total']

            if pedido_min > 0 and valor_total < pedido_min:
                info['abaixo_minimo'] = True
                info['diferenca'] = pedido_min - valor_total
                alertas_pedido_minimo.append(info)
                print(f"  [PEDIDO MINIMO] Alerta: {info['nome_fornecedor']} / {info['nome_loja']} - "
                      f"Valor R$ {valor_total:,.2f} < Minimo R$ {pedido_min:,.2f}")

        # Adicionar flag nos itens que estao abaixo do minimo
        for item in itens_pedido:
            cod_fornecedor = item.get('codigo_fornecedor', '')
            cod_loja = item.get('cod_loja', 0)
            chave = (cod_fornecedor, cod_loja)

            if chave in validacao_pedido_minimo:
                info = validacao_pedido_minimo[chave]
                item['pedido_abaixo_minimo'] = info['abaixo_minimo']
                item['valor_total_pedido_loja'] = info['valor_total']
                if info['abaixo_minimo']:
                    item['critica_pedido_minimo'] = f"Abaixo do pedido mínimo (R$ {info['valor_total']:,.2f} < R$ {info['pedido_minimo']:,.2f})"

        if alertas_pedido_minimo:
            print(f"  [PEDIDO MINIMO] {len(alertas_pedido_minimo)} fornecedor/loja(s) abaixo do minimo")

        # =====================================================================
        # CALCULO DE TRANSFERENCIAS ENTRE LOJAS (respeitando grupos regionais)
        # =====================================================================
        oportunidades_transferencia = []
        lojas_unicas = set(i.get('cod_loja') for i in itens_pedido)
        is_multiloja = len(lojas_unicas) > 1

        if is_multiloja and itens_pedido:
            try:
                print(f"\n  [TRANSFERENCIAS] Analisando oportunidades entre {len(lojas_unicas)} lojas...")

                # Reconectar ao banco para buscar grupos de transferencia
                conn_transf = get_db_connection()
                cursor_transf = conn_transf.cursor(cursor_factory=RealDictCursor)

                # Buscar grupos de transferencia ativos
                cursor_transf.execute("""
                    SELECT id, nome, cd_principal
                    FROM grupos_transferencia
                    WHERE ativo = TRUE
                """)
                grupos = cursor_transf.fetchall()

                # Buscar lojas por grupo
                grupos_lojas = {}  # {grupo_id: {cod_loja: {pode_doar, pode_receber, nome_loja}}}
                loja_para_grupo = {}  # {cod_loja: grupo_id}

                for grupo in grupos:
                    grupo_id = grupo['id']
                    cursor_transf.execute("""
                        SELECT cod_empresa, nome_loja, pode_doar, pode_receber, prioridade_recebimento
                        FROM lojas_grupo_transferencia
                        WHERE grupo_id = %s AND ativo = TRUE
                    """, [grupo_id])
                    lojas_grupo = cursor_transf.fetchall()

                    grupos_lojas[grupo_id] = {}
                    for loja in lojas_grupo:
                        cod = loja['cod_empresa']
                        grupos_lojas[grupo_id][cod] = {
                            'pode_doar': loja['pode_doar'],
                            'pode_receber': loja['pode_receber'],
                            'nome_loja': loja['nome_loja'],
                            'prioridade': loja['prioridade_recebimento']
                        }
                        loja_para_grupo[cod] = grupo_id

                conn_transf.close()

                if not grupos_lojas:
                    print(f"  [TRANSFERENCIAS] Nenhum grupo de transferencia configurado")
                else:
                    print(f"  [TRANSFERENCIAS] {len(grupos_lojas)} grupos de transferencia encontrados")

                    import uuid
                    sessao_pedido = f"PL_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"

                    # Agrupar itens por produto
                    itens_por_produto = {}
                    for item in itens_pedido:
                        cod = item['codigo']
                        if cod not in itens_por_produto:
                            itens_por_produto[cod] = {}
                        loja_cod = item.get('cod_loja')
                        if loja_cod:
                            itens_por_produto[cod][loja_cod] = item

                    cobertura_alvo = 21  # dias
                    COBERTURA_MINIMA_DOADOR = 10  # dias
                    MARGEM_EXCESSO_DIAS = 7  # dias acima do alvo

                    for cod_produto, lojas_data in itens_por_produto.items():
                        if len(lojas_data) < 2:
                            continue

                        # Processar por grupo regional (apenas lojas do mesmo grupo podem transferir)
                        for grupo_id, lojas_config in grupos_lojas.items():
                            # Filtrar apenas lojas deste grupo que estao no pedido
                            lojas_no_grupo = {
                                loja_cod: dados
                                for loja_cod, dados in lojas_data.items()
                                if loja_cod in lojas_config
                            }

                            if len(lojas_no_grupo) < 2:
                                continue

                            lojas_excesso = []
                            lojas_necessidade = []

                            for loja_cod, dados in lojas_no_grupo.items():
                                config_loja = lojas_config.get(loja_cod, {})
                                pode_doar = config_loja.get('pode_doar', False)
                                pode_receber = config_loja.get('pode_receber', False)

                                # Usar mesmos campos do pedido CD
                                demanda_diaria = dados.get('demanda_prevista_diaria', 0) or dados.get('demanda_diaria', 0) or 0
                                estoque = dados.get('estoque_atual', 0) or 0
                                transito = dados.get('estoque_transito', 0) or 0
                                cobertura_atual = dados.get('cobertura_atual_dias', 0) or 0
                                deve_pedir = dados.get('deve_pedir', False)
                                qtd_pedido = dados.get('quantidade_pedido', 0) or 0

                                # REGRA PRINCIPAL: Se a loja PRECISA PEDIR E pode receber, ela e RECEPTORA
                                if pode_receber and deve_pedir and qtd_pedido > 0:
                                    estoque_efetivo = estoque + transito
                                    if demanda_diaria > 0:
                                        # Necessidade conforme documentação:
                                        # necessidade_receptor = (demanda × COBERTURA_ALVO) - estoque
                                        necessidade_transf = int((demanda_diaria * cobertura_alvo) - estoque_efetivo)
                                        necessidade_transf = max(0, necessidade_transf)
                                    else:
                                        necessidade_transf = 0
                                    prioridade = config_loja.get('prioridade', 99)
                                    # Guardar: (loja, necessidade_transf, dados, qtd_pedido_original, prioridade)
                                    lojas_necessidade.append((loja_cod, necessidade_transf, dados, qtd_pedido, prioridade))
                                    continue  # NAO avaliar como possivel doadora

                                # Se NAO precisa pedir E pode doar, verificar se tem excesso para ser DOADORA
                                if pode_doar:
                                    if demanda_diaria <= 0:
                                        # Sem demanda = potencial doador se tem estoque significativo
                                        if estoque > 5:
                                            lojas_excesso.append((loja_cod, estoque, dados))
                                    elif cobertura_atual > (cobertura_alvo + MARGEM_EXCESSO_DIAS):
                                        # Tem excesso conforme documentação: cobertura > alvo + 7 dias
                                        # Excesso = estoque - (demanda × COBERTURA_MINIMA_DOADOR)
                                        estoque_disponivel = estoque + transito
                                        estoque_minimo_doador = demanda_diaria * COBERTURA_MINIMA_DOADOR
                                        excesso = int(estoque_disponivel - estoque_minimo_doador)
                                        if excesso > 0:
                                            lojas_excesso.append((loja_cod, excesso, dados))

                            # Se tem lojas com excesso e necessidade no MESMO GRUPO, criar transferencias
                            if lojas_excesso and lojas_necessidade:
                                lojas_excesso.sort(key=lambda x: x[1], reverse=True)
                                # Ordenar por prioridade de recebimento (menor = mais prioritario)
                                lojas_necessidade.sort(key=lambda x: (x[4], -x[1]))

                                for loja_dest, necessidade_transf, dados_dest, qtd_pedido_dest, _ in lojas_necessidade:
                                    qtd_restante = necessidade_transf

                                    for i, (loja_orig, excesso, dados_orig) in enumerate(lojas_excesso):
                                        if qtd_restante <= 0 or excesso <= 0:
                                            continue

                                        qtd_transferir = min(qtd_restante, excesso)

                                        if qtd_transferir > 0:
                                            # Criar oportunidade de transferência (mesmos campos do pedido CD)
                                            cue = dados_orig.get('cue', 0) or dados_dest.get('cue', 0) or dados_orig.get('custo_unitario', 0) or 0
                                            valor_estimado = qtd_transferir * cue

                                            transferencia = {
                                                'cod_produto': cod_produto,
                                                'descricao_produto': dados_orig.get('descricao', ''),
                                                'curva_abc': dados_orig.get('curva_abc', 'B'),
                                                # Origem
                                                'loja_origem': loja_orig,
                                                'nome_loja_origem': dados_orig.get('nome_loja', f'Loja {loja_orig}'),
                                                'estoque_origem': dados_orig.get('estoque_atual', 0),
                                                'cobertura_origem_dias': dados_orig.get('cobertura_atual_dias', 0),
                                                # Destino
                                                'loja_destino': loja_dest,
                                                'nome_loja_destino': dados_dest.get('nome_loja', f'Loja {loja_dest}'),
                                                'estoque_destino': dados_dest.get('estoque_atual', 0),
                                                'cobertura_destino_dias': dados_dest.get('cobertura_atual_dias', 0),
                                                # Valores
                                                'qtd_sugerida': qtd_transferir,
                                                'valor_estimado': round(valor_estimado, 2),
                                                'cue': cue,
                                                'urgencia': 'ALTA' if dados_dest.get('ruptura_iminente') else 'MEDIA',
                                                'tipo': 'MULTILOJA',
                                                'grupo_id': grupo_id,
                                                'sessao_pedido': sessao_pedido
                                            }
                                            oportunidades_transferencia.append(transferencia)

                                            # Atualizar excesso restante
                                            lojas_excesso[i] = (loja_orig, excesso - qtd_transferir, dados_orig)
                                            qtd_restante -= qtd_transferir

                print(f"  [TRANSFERENCIAS] {len(oportunidades_transferencia)} oportunidades identificadas")

            except Exception as e:
                print(f"  [TRANSFERENCIAS] Erro: {e}")
                import traceback
                traceback.print_exc()

        # Agregar por fornecedor
        agregado = agregar_por_fornecedor(itens_pedido)

        # Calcular resumo
        total_itens = len(itens_pedido)
        total_valor = sum(i.get('valor_pedido', 0) for i in itens_pedido)
        total_fornecedores = len(agregado)

        return jsonify({
            'success': True,
            'tipo_pedido': 'planejado',
            'periodo': {
                'inicio': periodo_inicio,
                'fim': periodo_fim,
                'dias': dias_periodo
            },
            'resumo': {
                'total_itens': total_itens,
                'total_valor': round(total_valor, 2),
                'total_fornecedores': total_fornecedores,
                'itens_demanda_validada_loja': sum(1 for i in itens_pedido if i.get('usa_demanda_validada')),
                'itens_demanda_agregada_proporcional': sum(1 for i in itens_pedido if i.get('usa_demanda_agregada')),
                'itens_demanda_historico': sum(1 for i in itens_pedido if i.get('origem_demanda') == 'historico'),
                'itens_com_demanda_validada': sum(1 for i in itens_pedido if i.get('usa_demanda_validada') or i.get('usa_demanda_agregada')),
                'itens_bloqueados': len(itens_bloqueados)
            },
            'itens_pedido': itens_pedido,
            'itens_bloqueados': itens_bloqueados,
            'agregado_fornecedor': agregado,
            'is_pedido_multiloja': is_multiloja,
            # Alertas de pedido minimo por fornecedor/loja
            'alertas_pedido_minimo': [
                {
                    'cod_fornecedor': a['cod_fornecedor'],
                    'nome_fornecedor': a['nome_fornecedor'],
                    'cod_loja': a['cod_loja'],
                    'nome_loja': a['nome_loja'],
                    'valor_total': round(a['valor_total'], 2),
                    'pedido_minimo': round(a['pedido_minimo'], 2),
                    'diferenca': round(a['diferenca'], 2)
                }
                for a in alertas_pedido_minimo
            ],
            'tem_alertas_pedido_minimo': len(alertas_pedido_minimo) > 0,
            # Transferencias entre lojas (mesma estrutura do pedido CD)
            'transferencias': {
                'total': len(oportunidades_transferencia),
                'sessao_pedido': oportunidades_transferencia[0].get('sessao_pedido') if oportunidades_transferencia else None,
                'tem_oportunidades': len(oportunidades_transferencia) > 0,
                'produtos_com_transferencia': list(set(t.get('cod_produto') for t in oportunidades_transferencia)),
                'detalhes': [
                    {
                        'cod_produto': t.get('cod_produto'),
                        'qtd_sugerida': t.get('qtd_sugerida'),
                        'loja_origem': t.get('nome_loja_origem'),
                        'loja_destino': t.get('nome_loja_destino'),
                        'cod_loja_origem': t.get('loja_origem'),  # Codigo numerico
                        'cod_loja_destino': t.get('loja_destino'),  # Codigo numerico
                        'urgencia': t.get('urgencia'),
                        'tipo': t.get('tipo', 'MULTILOJA'),
                        'grupo_id': t.get('grupo_id')
                    }
                    for t in oportunidades_transferencia
                ]
            }
        })

    except Exception as e:
        import traceback
        print(f"[ERRO] api_pedido_planejado: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@app.route('/api/pedido_planejado/exportar', methods=['POST'])
def api_pedido_planejado_exportar():
    """Exporta o pedido planejado para Excel."""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        dados = request.get_json()

        if not dados:
            return jsonify({'success': False, 'erro': 'Dados não fornecidos'}), 400

        itens_pedido = dados.get('itens_pedido', [])
        estatisticas = dados.get('estatisticas', {})
        agregacao = dados.get('agregado_fornecedor', {})
        alertas_pedido_minimo = dados.get('alertas_pedido_minimo', [])
        transferencias = dados.get('transferencias', {})

        if not itens_pedido:
            return jsonify({'success': False, 'erro': 'Nenhum item para exportar'}), 400

        # Criar workbook
        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        critica_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        critica_font = Font(color='B45309')

        # =====================================================================
        # ABA 1: RESUMO
        # =====================================================================
        ws_resumo = wb.active
        ws_resumo.title = 'Resumo'

        ws_resumo['A1'] = 'PEDIDO PLANEJADO - RESUMO'
        ws_resumo['A1'].font = Font(bold=True, size=14)
        ws_resumo.merge_cells('A1:D1')

        ws_resumo['A2'] = f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}'

        # Periodo do pedido
        if itens_pedido:
            periodo_inicio = itens_pedido[0].get('periodo_inicio', '')
            periodo_fim = itens_pedido[0].get('periodo_fim', '')
            ws_resumo['A3'] = f'Período: {periodo_inicio} até {periodo_fim}'

        ws_resumo['A5'] = 'Estatísticas'
        ws_resumo['A5'].font = Font(bold=True)

        stats_data = [
            ('Total de Itens', estatisticas.get('total_itens', len(itens_pedido))),
            ('Valor Total do Pedido', f"R$ {estatisticas.get('valor_total', sum(i.get('valor_pedido', 0) for i in itens_pedido)):,.2f}"),
            ('Lojas Processadas', estatisticas.get('lojas_processadas', len(set(i.get('cod_loja') for i in itens_pedido)))),
            ('Fornecedores', estatisticas.get('fornecedores_unicos', len(set(i.get('nome_fornecedor') for i in itens_pedido)))),
            ('Itens com Demanda Validada', estatisticas.get('itens_com_demanda_validada', 0)),
        ]

        for i, (label, valor) in enumerate(stats_data, start=6):
            ws_resumo[f'A{i}'] = label
            ws_resumo[f'B{i}'] = valor

        # Alertas de pedido minimo
        if alertas_pedido_minimo:
            row_alerta = 13
            ws_resumo[f'A{row_alerta}'] = 'Alertas de Pedido Mínimo'
            ws_resumo[f'A{row_alerta}'].font = Font(bold=True, color='B45309')
            row_alerta += 1

            headers_alerta = ['Fornecedor', 'Loja', 'Valor Pedido', 'Mínimo', 'Diferença']
            for col, header in enumerate(headers_alerta, start=1):
                cell = ws_resumo.cell(row=row_alerta, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
                cell.border = border
            row_alerta += 1

            for alerta in alertas_pedido_minimo:
                ws_resumo.cell(row=row_alerta, column=1, value=alerta.get('nome_fornecedor', '')).border = border
                ws_resumo.cell(row=row_alerta, column=2, value=alerta.get('nome_loja', '')).border = border
                ws_resumo.cell(row=row_alerta, column=3, value=alerta.get('valor_total', 0)).border = border
                ws_resumo.cell(row=row_alerta, column=3).number_format = '#,##0.00'
                ws_resumo.cell(row=row_alerta, column=4, value=alerta.get('pedido_minimo', 0)).border = border
                ws_resumo.cell(row=row_alerta, column=4).number_format = '#,##0.00'
                ws_resumo.cell(row=row_alerta, column=5, value=alerta.get('diferenca', 0)).border = border
                ws_resumo.cell(row=row_alerta, column=5).number_format = '#,##0.00'
                row_alerta += 1

        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20
        ws_resumo.column_dimensions['C'].width = 15
        ws_resumo.column_dimensions['D'].width = 15
        ws_resumo.column_dimensions['E'].width = 15

        # =====================================================================
        # ABA 2: ITENS DO PEDIDO
        # =====================================================================
        ws_itens = wb.create_sheet('Itens Pedido')

        headers = [
            'Cod Filial', 'Nome Filial', 'Cod Item', 'Descrição Item',
            'CNPJ Fornecedor', 'Nome Fornecedor', 'Demanda Diária', 'Demanda Período',
            'Qtd Pedido', 'CUE', 'Valor Pedido', 'Estoque Atual', 'Estoque Trânsito',
            'Cobertura Atual', 'Cobertura c/ Trânsito', 'Data Emissão', 'Origem Demanda', 'Crítica'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_itens.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Ordenar itens por Codigo Filial e CNPJ Fornecedor
        itens_ordenados = sorted(itens_pedido, key=lambda x: (
            x.get('cod_loja', 0) or 0,
            x.get('cnpj_fornecedor', '') or ''
        ))

        for row_idx, item in enumerate(itens_ordenados, start=2):
            ws_itens.cell(row=row_idx, column=1, value=item.get('cod_loja', '')).border = border
            ws_itens.cell(row=row_idx, column=2, value=item.get('nome_loja', '')).border = border
            ws_itens.cell(row=row_idx, column=3, value=item.get('codigo', '')).border = border
            ws_itens.cell(row=row_idx, column=4, value=item.get('descricao', '')[:60]).border = border
            ws_itens.cell(row=row_idx, column=5, value=item.get('cnpj_fornecedor', '')).border = border
            ws_itens.cell(row=row_idx, column=6, value=item.get('nome_fornecedor', '')).border = border

            ws_itens.cell(row=row_idx, column=7, value=item.get('demanda_diaria', 0)).border = border
            ws_itens.cell(row=row_idx, column=7).number_format = '#,##0.00'

            ws_itens.cell(row=row_idx, column=8, value=item.get('demanda_periodo', 0)).border = border
            ws_itens.cell(row=row_idx, column=8).number_format = '#,##0'

            ws_itens.cell(row=row_idx, column=9, value=item.get('quantidade_pedido', 0)).border = border
            ws_itens.cell(row=row_idx, column=9).number_format = '#,##0'

            ws_itens.cell(row=row_idx, column=10, value=item.get('cue', item.get('custo_unitario', 0))).border = border
            ws_itens.cell(row=row_idx, column=10).number_format = '#,##0.00'

            ws_itens.cell(row=row_idx, column=11, value=item.get('valor_pedido', 0)).border = border
            ws_itens.cell(row=row_idx, column=11).number_format = '#,##0.00'

            ws_itens.cell(row=row_idx, column=12, value=item.get('estoque_atual', 0)).border = border
            ws_itens.cell(row=row_idx, column=12).number_format = '#,##0'

            ws_itens.cell(row=row_idx, column=13, value=item.get('estoque_transito', 0)).border = border
            ws_itens.cell(row=row_idx, column=13).number_format = '#,##0'

            ws_itens.cell(row=row_idx, column=14, value=item.get('cobertura_atual_dias', 0)).border = border
            ws_itens.cell(row=row_idx, column=14).number_format = '#,##0.0'

            ws_itens.cell(row=row_idx, column=15, value=item.get('cobertura_com_transito_dias', 0)).border = border
            ws_itens.cell(row=row_idx, column=15).number_format = '#,##0.0'

            ws_itens.cell(row=row_idx, column=16, value=item.get('data_emissao_sugerida', '')).border = border
            ws_itens.cell(row=row_idx, column=17, value=item.get('origem_demanda', '')).border = border

            critica = item.get('critica_pedido_minimo', '')
            cell_critica = ws_itens.cell(row=row_idx, column=18, value=critica)
            cell_critica.border = border
            if critica:
                cell_critica.fill = critica_fill
                cell_critica.font = critica_font

        col_widths = [10, 20, 10, 40, 18, 25, 12, 12, 10, 10, 12, 10, 10, 10, 12, 12, 15, 40]
        for i, width in enumerate(col_widths, start=1):
            ws_itens.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

        ws_itens.freeze_panes = 'A2'

        # =====================================================================
        # ABA 3: TRANSFERÊNCIAS (se houver)
        # =====================================================================
        detalhes_transf = transferencias.get('detalhes', [])
        if detalhes_transf:
            ws_transf = wb.create_sheet('Transferências')

            headers_transf = [
                'Código Produto', 'Loja Origem', 'Loja Destino', 'Qtd Sugerida', 'Urgência'
            ]

            for col, header in enumerate(headers_transf, start=1):
                cell = ws_transf.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
                cell.border = border

            for row_idx, transf in enumerate(detalhes_transf, start=2):
                ws_transf.cell(row=row_idx, column=1, value=transf.get('cod_produto', '')).border = border
                ws_transf.cell(row=row_idx, column=2, value=transf.get('loja_origem', '')).border = border
                ws_transf.cell(row=row_idx, column=3, value=transf.get('loja_destino', '')).border = border
                ws_transf.cell(row=row_idx, column=4, value=transf.get('qtd_sugerida', 0)).border = border
                ws_transf.cell(row=row_idx, column=5, value=transf.get('urgencia', '')).border = border

            ws_transf.column_dimensions['A'].width = 15
            ws_transf.column_dimensions['B'].width = 25
            ws_transf.column_dimensions['C'].width = 25
            ws_transf.column_dimensions['D'].width = 15
            ws_transf.column_dimensions['E'].width = 12

        # Salvar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pedido_planejado_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(f"[ERRO] api_pedido_planejado_exportar: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


if __name__ == '__main__':
    # Criar pastas necessárias
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('outputs', exist_ok=True)

    print("=" * 50)
    print("  SISTEMA DE PREVISÃO DE DEMANDA")
    print("  Acesse: http://localhost:5001")
    print("=" * 50)

    # Rodar servidor
    app.run(debug=True, host='0.0.0.0', port=5001)
