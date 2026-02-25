"""
Blueprint: Pedido ao Fornecedor Integrado
Rotas: /pedido_fornecedor_integrado, /api/pedido_fornecedor_integrado, /api/pedido_fornecedor_integrado/exportar
"""

import os
import io
from datetime import datetime
import math
import pandas as pd
import numpy as np
from flask import Blueprint, render_template, request, jsonify, send_file, current_app

from app.utils.db_connection import get_db_connection
from app.utils.demanda_pre_calculada import (
    obter_demanda_diaria_efetiva,
    verificar_dados_disponiveis,
    precarregar_demanda_em_lote,
    obter_demanda_do_cache,
    calcular_proporcoes_vendas_por_loja
)

pedido_fornecedor_bp = Blueprint('pedido_fornecedor', __name__)


@pedido_fornecedor_bp.route('/pedido_fornecedor_integrado')
def pedido_fornecedor_integrado_page():
    """Pagina de pedidos ao fornecedor integrado com previsao V2"""
    return render_template('pedido_fornecedor_integrado.html')


@pedido_fornecedor_bp.route('/api/pedido_fornecedor_integrado', methods=['POST'])
def api_pedido_fornecedor_integrado():
    """
    API para gerar pedidos ao fornecedor integrado com previsao V2.

    Caracteristicas:
    - Usa previsao V2 (Bottom-Up) em tempo real
    - Cobertura baseada em ABC: Lead Time + Ciclo + Seguranca_ABC
    - Curva ABC do cadastro de produtos para nivel de servico
    - Arredondamento para multiplo de caixa
    - Destino configuravel (Lojas ou CD)

    Request JSON:
    {
        "fornecedor": "TODOS" ou codigo especifico,
        "categoria": "TODAS" ou categoria especifica,
        "destino_tipo": "LOJA" ou "CD",
        "cod_empresa": codigo da loja/CD de destino (ou "TODAS"),
        "cobertura_dias": null (automatico) ou numero especifico
    }

    Returns:
        JSON com pedidos calculados e agregacao por fornecedor
    """
    try:
        from core.pedido_fornecedor_integrado import (
            PedidoFornecedorIntegrado,
            agregar_por_fornecedor,
            calcular_cobertura_abc,
            arredondar_para_multiplo,
            calcular_es_pooling_cd
        )
        from core.demand_calculator import DemandCalculator
        from core.transferencia_regional import TransferenciaRegional

        dados = request.get_json()

        # Parametros de filtro
        fornecedor_filtro = dados.get('fornecedor', 'TODOS')
        linha1_filtro = dados.get('linha1', 'TODAS')
        linha3_filtro = dados.get('linha3', 'TODAS')
        if linha1_filtro == 'TODAS' and 'categoria' in dados:
            linha1_filtro = dados.get('categoria', 'TODAS')
        destino_tipo = dados.get('destino_tipo', 'LOJA')
        cod_empresa = dados.get('cod_empresa', 'TODAS')
        cobertura_dias = dados.get('cobertura_dias')

        # Normalizar filtros de linha - converter array de 1 elemento para string
        # IMPORTANTE: Trata casos onde JSON envia arrays onde esperamos string
        def normalizar_filtro(filtro, valor_padrao='TODAS'):
            # DEBUG: Mostrar o que está chegando
            print(f"    [NORMALIZAR] Entrada: {filtro} (tipo={type(filtro).__name__})")

            if filtro is None:
                return valor_padrao

            # Converter para string se for bytes
            if isinstance(filtro, bytes):
                filtro = filtro.decode('utf-8')

            # Se for string, retornar como está
            if isinstance(filtro, str):
                return filtro

            # Se for lista ou qualquer iterável (exceto string)
            try:
                # Forçar conversão para lista Python nativa
                filtro_lista = list(filtro)
                if len(filtro_lista) == 0:
                    return valor_padrao
                elif len(filtro_lista) == 1:
                    # GARANTIR que é string, não outro tipo
                    valor = filtro_lista[0]
                    if valor is None:
                        return valor_padrao
                    return str(valor) if not isinstance(valor, str) else valor
                else:
                    # Converter cada elemento para string nativa Python
                    return [str(v) if not isinstance(v, str) else v for v in filtro_lista]
            except (TypeError, ValueError):
                # Não é iterável, retornar como está
                pass

            return filtro

        linha1_filtro = normalizar_filtro(linha1_filtro, 'TODAS')
        linha3_filtro = normalizar_filtro(linha3_filtro, 'TODAS')

        # DEBUG EXTRA: Verificar tipos APÓS normalização
        print(f"    [POS-NORMALIZACAO] linha1_filtro: {linha1_filtro} (tipo={type(linha1_filtro).__name__})")
        print(f"    [POS-NORMALIZACAO] linha3_filtro: {linha3_filtro} (tipo={type(linha3_filtro).__name__})")

        print(f"\n{'='*60}")
        print("PEDIDO FORNECEDOR INTEGRADO")
        print(f"{'='*60}")
        print(f"  Fornecedor: {fornecedor_filtro}")
        print(f"  Linha 1: {linha1_filtro} (tipo: {type(linha1_filtro).__name__})")
        print(f"  Linha 3: {linha3_filtro} (tipo: {type(linha3_filtro).__name__})")
        print(f"  Destino: {destino_tipo}")
        print(f"  Empresa: {cod_empresa}")
        print(f"  Cobertura: {'Automatica (ABC)' if cobertura_dias is None else f'{cobertura_dias} dias'}")

        conn = get_db_connection()

        # Normalizar cod_empresa
        lojas_selecionadas = []
        if isinstance(cod_empresa, list):
            lojas_selecionadas = [int(x) for x in cod_empresa]
        elif cod_empresa != 'TODAS':
            lojas_selecionadas = [int(cod_empresa)]

        # Identificar se o destino e CD
        if lojas_selecionadas:
            is_destino_cd = destino_tipo == 'CD' or any(loja >= 80 for loja in lojas_selecionadas)
            cod_destino = lojas_selecionadas[0]
        else:
            is_destino_cd = destino_tipo == 'CD'
            cod_destino = 80

        # Determinar lista de fornecedores a processar
        fornecedores_a_processar = []

        if isinstance(fornecedor_filtro, list):
            if 'TODOS' in fornecedor_filtro:
                fornecedor_filtro = 'TODOS'
            else:
                fornecedores_a_processar = fornecedor_filtro

        if not fornecedores_a_processar:
            if fornecedor_filtro == 'TODOS':
                query_fornecedores = """
                    SELECT DISTINCT nome_fornecedor
                    FROM cadastro_produtos_completo
                    WHERE ativo = TRUE
                    AND nome_fornecedor IS NOT NULL
                    AND TRIM(nome_fornecedor) != ''
                """
                params_forn = []

                # FILTRO LINHA1 para query de fornecedores
                if linha1_filtro != 'TODAS':
                    is_lista = isinstance(linha1_filtro, (list, tuple))
                    if is_lista and len(linha1_filtro) > 1:
                        placeholders = ','.join(['%s'] * len(linha1_filtro))
                        query_fornecedores += f" AND categoria IN ({placeholders})"
                        for v in linha1_filtro:
                            params_forn.append(str(v) if not isinstance(v, str) else v)
                    else:
                        query_fornecedores += " AND categoria = %s"
                        valor = linha1_filtro[0] if is_lista else linha1_filtro
                        params_forn.append(str(valor) if not isinstance(valor, str) else valor)

                # FILTRO LINHA3 para query de fornecedores
                if linha3_filtro != 'TODAS':
                    is_lista = isinstance(linha3_filtro, (list, tuple))
                    if is_lista and len(linha3_filtro) > 1:
                        placeholders = ','.join(['%s'] * len(linha3_filtro))
                        query_fornecedores += f" AND codigo_linha IN ({placeholders})"
                        for v in linha3_filtro:
                            params_forn.append(str(v) if not isinstance(v, str) else v)
                    else:
                        query_fornecedores += " AND codigo_linha = %s"
                        valor = linha3_filtro[0] if is_lista else linha3_filtro
                        params_forn.append(str(valor) if not isinstance(valor, str) else valor)

                query_fornecedores += " ORDER BY nome_fornecedor"
                df_fornecedores = pd.read_sql(query_fornecedores, conn, params=params_forn if params_forn else None)
                fornecedores_a_processar = df_fornecedores['nome_fornecedor'].tolist()
            else:
                fornecedores_a_processar = [fornecedor_filtro]

        if not fornecedores_a_processar:
            conn.close()
            return jsonify({
                'success': False,
                'erro': 'Nenhum fornecedor encontrado com os filtros selecionados.'
            }), 400

        # Verificar se tabela parametros_fornecedor existe
        cursor_check = conn.cursor()
        cursor_check.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'parametros_fornecedor'
            )
        """)
        tabela_params_existe = cursor_check.fetchone()[0]
        cursor_check.close()

        df_produtos_total = pd.DataFrame()
        total_produtos_por_fornecedor = {}

        for idx_forn, nome_fornecedor in enumerate(fornecedores_a_processar):
            if tabela_params_existe:
                query_produtos = """
                    SELECT DISTINCT
                        p.cod_produto as codigo,
                        p.descricao,
                        p.categoria as linha1,
                        p.codigo_linha as linha3,
                        p.descricao_linha as linha3_descricao,
                        'B' as curva_abc,
                        p.cnpj_fornecedor as codigo_fornecedor,
                        p.nome_fornecedor,
                        COALESCE(pf.lead_time_dias, 15) as lead_time_dias,
                        COALESCE(pf.ciclo_pedido_dias, 7) as ciclo_pedido_dias,
                        COALESCE(pf.pedido_minimo_valor, 0) as pedido_minimo_valor,
                        CASE WHEN pf.id IS NOT NULL THEN TRUE ELSE FALSE END as fornecedor_cadastrado
                    FROM cadastro_produtos_completo p
                    LEFT JOIN parametros_fornecedor pf ON p.cnpj_fornecedor = pf.cnpj_fornecedor
                        AND pf.cod_empresa = %s AND pf.ativo = TRUE
                    WHERE p.ativo = TRUE
                        AND p.nome_fornecedor = %s
                """
                params = [cod_destino, nome_fornecedor]
            else:
                query_produtos = """
                    SELECT DISTINCT
                        p.cod_produto as codigo,
                        p.descricao,
                        p.categoria as linha1,
                        p.codigo_linha as linha3,
                        p.descricao_linha as linha3_descricao,
                        'B' as curva_abc,
                        p.cnpj_fornecedor as codigo_fornecedor,
                        p.nome_fornecedor,
                        15 as lead_time_dias,
                        7 as ciclo_pedido_dias,
                        0 as pedido_minimo_valor,
                        FALSE as fornecedor_cadastrado
                    FROM cadastro_produtos_completo p
                    WHERE p.ativo = TRUE
                        AND p.nome_fornecedor = %s
                """
                params = [nome_fornecedor]

            # FILTRO LINHA1 (categoria) - com proteção contra arrays
            if linha1_filtro != 'TODAS':
                # DEBUG: mostrar estado do filtro antes de usar na query
                is_lista = isinstance(linha1_filtro, (list, tuple))
                print(f"      [QUERY] linha1_filtro: {linha1_filtro}, is_lista={is_lista}")

                if is_lista and len(linha1_filtro) > 1:
                    # Múltiplos valores: usar IN
                    placeholders = ','.join(['%s'] * len(linha1_filtro))
                    query_produtos += f" AND p.categoria IN ({placeholders})"
                    # Garantir que cada valor é string
                    for v in linha1_filtro:
                        params.append(str(v) if not isinstance(v, str) else v)
                else:
                    # Valor único: usar = com string simples
                    query_produtos += " AND p.categoria = %s"
                    valor = linha1_filtro[0] if is_lista else linha1_filtro
                    # GARANTIR que é string Python pura
                    params.append(str(valor) if not isinstance(valor, str) else valor)
                    print(f"      [QUERY] Adicionado param categoria: {params[-1]} (tipo={type(params[-1]).__name__})")

            # FILTRO LINHA3 (codigo_linha) - com proteção contra arrays
            if linha3_filtro != 'TODAS':
                is_lista = isinstance(linha3_filtro, (list, tuple))
                print(f"      [QUERY] linha3_filtro: {linha3_filtro}, is_lista={is_lista}")

                if is_lista and len(linha3_filtro) > 1:
                    placeholders = ','.join(['%s'] * len(linha3_filtro))
                    query_produtos += f" AND p.codigo_linha IN ({placeholders})"
                    for v in linha3_filtro:
                        params.append(str(v) if not isinstance(v, str) else v)
                else:
                    query_produtos += " AND p.codigo_linha = %s"
                    valor = linha3_filtro[0] if is_lista else linha3_filtro
                    params.append(str(valor) if not isinstance(valor, str) else valor)
                    print(f"      [QUERY] Adicionado param codigo_linha: {params[-1]} (tipo={type(params[-1]).__name__})")

            query_produtos += " ORDER BY p.cod_produto"
            df_forn = pd.read_sql(query_produtos, conn, params=params)

            if not df_forn.empty:
                total_produtos_por_fornecedor[nome_fornecedor] = len(df_forn)
                df_produtos_total = pd.concat([df_produtos_total, df_forn], ignore_index=True)

        df_produtos = df_produtos_total

        if df_produtos.empty:
            conn.close()
            return jsonify({
                'success': False,
                'erro': 'Nenhum item encontrado com os filtros selecionados.'
            }), 400

        # Remover duplicatas
        df_produtos = df_produtos.drop_duplicates(subset=['codigo'], keep='first')

        # Configurar lojas para demanda
        is_pedido_multiloja = False
        mapa_nomes_lojas = {}

        if is_destino_cd:
            # Destino CD: buscar todas as lojas para calcular demanda e transferencias
            df_lojas_cd = pd.read_sql(
                "SELECT cod_empresa, nome_loja FROM cadastro_lojas WHERE ativo = TRUE AND cod_empresa < 80",
                conn
            )
            lojas_demanda = df_lojas_cd['cod_empresa'].tolist()
            mapa_nomes_lojas = dict(zip(df_lojas_cd['cod_empresa'], df_lojas_cd['nome_loja']))
            nome_destino = f"CD {cod_destino}"
            # IMPORTANTE: Habilitar multiloja para calcular transferencias entre lojas
            is_pedido_multiloja = len(lojas_demanda) > 1
        else:
            if lojas_selecionadas:
                lojas_demanda = lojas_selecionadas
                if len(lojas_selecionadas) == 1:
                    nome_destino = f"Loja {lojas_selecionadas[0]}"
                else:
                    is_pedido_multiloja = True
                    nome_destino = f"Lojas {', '.join(str(l) for l in lojas_selecionadas)}"
                    placeholders = ','.join(['%s'] * len(lojas_selecionadas))
                    df_nomes_lojas = pd.read_sql(
                        f"SELECT cod_empresa, nome_loja FROM cadastro_lojas WHERE cod_empresa IN ({placeholders})",
                        conn,
                        params=lojas_selecionadas
                    )
                    mapa_nomes_lojas = dict(zip(df_nomes_lojas['cod_empresa'], df_nomes_lojas['nome_loja']))
            else:
                # Todas as lojas selecionadas: buscar todas e habilitar multiloja
                df_todas_lojas = pd.read_sql(
                    "SELECT cod_empresa, nome_loja FROM cadastro_lojas WHERE ativo = TRUE AND cod_empresa < 80",
                    conn
                )
                lojas_demanda = df_todas_lojas['cod_empresa'].tolist()
                mapa_nomes_lojas = dict(zip(df_todas_lojas['cod_empresa'], df_todas_lojas['nome_loja']))
                nome_destino = "Todas as Lojas"
                # IMPORTANTE: Habilitar multiloja para calcular transferencias entre lojas
                is_pedido_multiloja = len(lojas_demanda) > 1

        # Processar produtos
        processador = PedidoFornecedorIntegrado(conn)

        codigos_produtos = df_produtos['codigo'].astype(int).tolist()
        processador.precarregar_situacoes_compra(codigos_produtos, lojas_demanda)
        processador.precarregar_estoque(codigos_produtos, lojas_demanda)
        processador.precarregar_historico_vendas(codigos_produtos, lojas_demanda)
        processador.precarregar_embalagens(codigos_produtos)
        processador.precarregar_produtos(codigos_produtos)

        # PRE-CARREGAR DEMANDA EM LOTE (otimizacao: 1 query em vez de N queries)
        # Agrupar produtos por CNPJ do fornecedor
        cache_demanda_global = {}
        cnpjs_unicos = df_produtos['codigo_fornecedor'].unique()
        for cnpj_forn in cnpjs_unicos:
            if cnpj_forn:
                produtos_deste_fornecedor = df_produtos[df_produtos['codigo_fornecedor'] == cnpj_forn]['codigo'].tolist()
                cache_demanda = precarregar_demanda_em_lote(
                    conn,
                    produtos_deste_fornecedor,
                    cnpj_forn,
                    cod_empresas=lojas_demanda if is_pedido_multiloja else None
                )
                cache_demanda_global.update(cache_demanda)
        print(f"  [CACHE] Demanda pre-carregada: {len(cache_demanda_global)} registros")

        # PRE-CALCULAR PROPORCOES DE VENDAS POR LOJA (para rateio proporcional)
        # Em vez de dividir uniformemente, rateia pela participacao de vendas de cada loja
        cache_proporcoes = {}
        if is_pedido_multiloja:
            cache_proporcoes = calcular_proporcoes_vendas_por_loja(
                conn,
                codigos_produtos,
                lojas_demanda,
                dias_historico=365  # Usa ultimo ano para proporcao
            )
            print(f"  [CACHE] Proporcoes por loja calculadas: {len(cache_proporcoes)} registros")

        resultados = []
        itens_sem_historico = 0
        itens_sem_demanda = 0

        for idx, row in df_produtos.iterrows():
            try:
                codigo = row['codigo']
                lead_time_forn = int(row.get('lead_time_dias', 15))
                ciclo_pedido_forn = int(row.get('ciclo_pedido_dias', 7))
                pedido_min_forn = float(row.get('pedido_minimo_valor', 0))
                fornecedor_cadastrado = row.get('fornecedor_cadastrado', False)

                # Obter CNPJ do fornecedor para busca na tabela pre-calculada
                cnpj_forn = row.get('codigo_fornecedor', '')

                if is_pedido_multiloja:
                    num_lojas = len(lojas_demanda)

                    # V26: Detectar metodo de previsao para aplicar limitador de cobertura (apenas TSB)
                    metodo_produto = None
                    for loja_pre in lojas_demanda:
                        proporcao_pre = cache_proporcoes.get((str(codigo), loja_pre))
                        _, _, meta_pre = obter_demanda_do_cache(
                            cache_demanda_global, str(codigo), cod_empresa=loja_pre,
                            num_lojas=num_lojas, proporcao_loja=proporcao_pre
                        )
                        if metodo_produto is None and meta_pre.get('metodo_usado'):
                            metodo_produto = meta_pre.get('metodo_usado')
                            break  # Metodo e o mesmo para todas as lojas

                    # V26: Limitador de cobertura 90d apenas para itens TSB (demanda intermitente)
                    is_tsb = metodo_produto and metodo_produto.lower() == 'tsb'

                    for loja_cod in lojas_demanda:
                        loja_nome = mapa_nomes_lojas.get(loja_cod, f"Loja {loja_cod}")

                        # Buscar proporcao de vendas desta loja para este produto
                        proporcao_loja = cache_proporcoes.get((str(codigo), loja_cod))

                        # PRIORIDADE 1: Usar demanda pre-calculada do CACHE (otimizado)
                        # Se nao encontrar por loja, usa consolidado rateado PROPORCIONALMENTE
                        demanda_diaria, desvio_padrao, metadata = obter_demanda_do_cache(
                            cache_demanda_global, str(codigo), cod_empresa=loja_cod,
                            num_lojas=num_lojas, proporcao_loja=proporcao_loja
                        )

                        # FALLBACK: Se nao houver no cache, calcular em tempo real
                        if metadata.get('fonte') == 'sem_dados':
                            historico_loja = processador.buscar_historico_vendas(codigo, loja_cod)
                            if historico_loja and len(historico_loja) >= 7:
                                demanda_media, desvio_padrao, metadata = DemandCalculator.calcular_demanda_inteligente(historico_loja, metodo='auto')
                                # calcular_demanda_inteligente retorna demanda MENSAL
                                demanda_diaria = (demanda_media / 30) if demanda_media > 0 else 0
                                metadata['fonte'] = 'tempo_real'
                            else:
                                demanda_diaria = 0
                                desvio_padrao = 0
                                metadata = {'metodo_usado': 'sem_historico', 'fonte': 'sem_dados'}

                        resultado = processador.processar_item(
                            codigo=codigo,
                            cod_empresa=loja_cod,
                            previsao_diaria=demanda_diaria,
                            desvio_padrao=desvio_padrao,
                            cobertura_dias=cobertura_dias,
                            lead_time_dias=lead_time_forn,
                            ciclo_pedido_dias=ciclo_pedido_forn,
                            pedido_minimo_valor=pedido_min_forn,
                            aplicar_limitador_cobertura=is_tsb
                        )

                        if 'erro' in resultado:
                            resultado = {
                                'codigo': codigo,
                                'descricao': row.get('descricao', ''),
                                'nome_fornecedor': row.get('nome_fornecedor', ''),
                                'curva_abc': row.get('curva_abc', 'B'),
                                'estoque_atual': 0,
                                'estoque_transito': 0,
                                'demanda_prevista': 0,
                                'demanda_prevista_diaria': 0,
                                'cobertura_atual_dias': 999,
                                'cobertura_pos_pedido_dias': 999,
                                'quantidade_pedido': 0,
                                'valor_pedido': 0,
                                'preco_custo': 0,
                                'cue': 0,
                                'deve_pedir': False,
                                'bloqueado': False
                            }

                        resultado['cod_loja'] = loja_cod
                        resultado['nome_loja'] = loja_nome
                        resultado['codigo_fornecedor'] = row.get('codigo_fornecedor', '')
                        resultado['metodo_previsao'] = metadata.get('metodo_usado', 'auto')
                        resultado['fonte_demanda'] = metadata.get('fonte', 'tempo_real')  # pre_calculada ou tempo_real
                        resultado['metodo_rateio'] = metadata.get('metodo_rateio')  # proporcional ou uniforme
                        resultado['proporcao_loja'] = metadata.get('proporcao_loja')  # 0.0 a 1.0
                        resultado['lojas_consideradas'] = 1
                        resultado['fornecedor_cadastrado'] = fornecedor_cadastrado
                        resultado['lead_time_usado'] = lead_time_forn
                        resultado['ciclo_pedido_usado'] = ciclo_pedido_forn
                        resultado['pedido_minimo_fornecedor'] = pedido_min_forn
                        resultado['sem_demanda_loja'] = demanda_diaria <= 0

                        # DEBUG: Qualquer item com cobertura insuficiente
                        cobertura_alvo_debug = cobertura_dias if cobertura_dias else 21
                        cobertura_pos = resultado.get('cobertura_pos_pedido_dias', 999)
                        if cobertura_pos < cobertura_alvo_debug and resultado.get('quantidade_pedido', 0) > 0:
                            print(f"  [DEBUG COBERTURA] Item {codigo} Loja {loja_cod}: demanda_dia={demanda_diaria:.3f}, estoque={resultado.get('estoque_atual',0)}, ES={resultado.get('estoque_seguranca',0)}, demanda_periodo={resultado.get('demanda_prevista',0)}, qtd_pedido={resultado.get('quantidade_pedido',0)}, cobertura_alvo={cobertura_alvo_debug}d, cobertura_pos={cobertura_pos:.1f}d, fonte={metadata.get('fonte', 'N/A')}")

                        resultados.append(resultado)
                else:
                    # PRIORIDADE 1: Usar demanda pre-calculada do CACHE (otimizado)
                    # Para mono-loja, buscar consolidado (cod_empresa=None)
                    demanda_diaria, desvio_padrao, metadata = obter_demanda_do_cache(
                        cache_demanda_global, str(codigo), cod_empresa=None
                    )

                    # FALLBACK: Se nao houver no cache, calcular em tempo real
                    if metadata.get('fonte') == 'sem_dados':
                        historico_total = []
                        for loja in lojas_demanda:
                            historico_loja = processador.buscar_historico_vendas(codigo, loja)
                            if historico_loja:
                                if not historico_total:
                                    historico_total = historico_loja.copy()
                                else:
                                    for i in range(min(len(historico_total), len(historico_loja))):
                                        historico_total[i] += historico_loja[i]

                        if not historico_total or len(historico_total) < 7:
                            itens_sem_historico += 1
                            continue

                        demanda_media, desvio_padrao, metadata = DemandCalculator.calcular_demanda_inteligente(historico_total, metodo='auto')
                        # calcular_demanda_inteligente retorna demanda MENSAL
                        demanda_diaria = demanda_media / 30 if demanda_media > 0 else 0
                        metadata['fonte'] = 'tempo_real'

                    if demanda_diaria <= 0:
                        itens_sem_demanda += 1
                        continue

                    resultado = processador.processar_item(
                        codigo=codigo,
                        cod_empresa=cod_destino,
                        previsao_diaria=demanda_diaria,
                        desvio_padrao=desvio_padrao,
                        cobertura_dias=cobertura_dias,
                        lead_time_dias=lead_time_forn,
                        ciclo_pedido_dias=ciclo_pedido_forn,
                        pedido_minimo_valor=pedido_min_forn
                    )

                    if 'erro' not in resultado:
                        resultado['nome_loja'] = nome_destino
                        resultado['cod_loja'] = cod_destino
                        resultado['codigo_fornecedor'] = row.get('codigo_fornecedor', '')
                        resultado['metodo_previsao'] = metadata.get('metodo_usado', 'auto')
                        resultado['fonte_demanda'] = metadata.get('fonte', 'tempo_real')  # pre_calculada ou tempo_real
                        resultado['lojas_consideradas'] = len(lojas_demanda)
                        resultado['fornecedor_cadastrado'] = fornecedor_cadastrado
                        resultado['lead_time_usado'] = lead_time_forn
                        resultado['ciclo_pedido_usado'] = ciclo_pedido_forn
                        resultado['pedido_minimo_fornecedor'] = pedido_min_forn
                        resultados.append(resultado)

            except Exception as e:
                print(f"  [AVISO] Erro ao processar item {row.get('codigo')}: {e}")
                continue

        conn.close()

        if not resultados:
            return jsonify({
                'success': False,
                'erro': f'Nenhum item com dados suficientes para gerar pedido.'
            }), 400

        # Converter tipos numpy para JSON
        def converter_tipos_json(obj):
            if obj is None:
                return None
            if isinstance(obj, dict):
                return {k: converter_tipos_json(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [converter_tipos_json(item) for item in obj]
            elif isinstance(obj, (np.bool_, np.generic)):
                valor = obj.item()
                if isinstance(valor, float):
                    if math.isnan(valor):
                        return 0
                    if math.isinf(valor):
                        return 999 if valor > 0 else -999
                return valor
            elif isinstance(obj, np.ndarray):
                return converter_tipos_json(obj.tolist())
            elif isinstance(obj, float):
                if math.isnan(obj):
                    return 0
                if math.isinf(obj):
                    return 999 if obj > 0 else -999
                return obj
            else:
                try:
                    if pd.isna(obj):
                        return None
                except (TypeError, ValueError):
                    pass
            return obj

        resultados = converter_tipos_json(resultados)

        # Agregar por fornecedor
        agregacao = agregar_por_fornecedor(resultados)

        # Separar itens
        itens_pedido = [r for r in resultados if r.get('deve_pedir') and not r.get('bloqueado')]
        itens_bloqueados = [r for r in resultados if r.get('bloqueado')]
        itens_ok = [r for r in resultados if not r.get('deve_pedir') and not r.get('bloqueado')]

        # ==============================================================================
        # PRE-CARREGAR ESTOQUE DO CD (para distribuicao V29)
        # ==============================================================================
        # V29: Detectar padrao de compra que direciona para CD(s), mesmo com destino "Lojas"
        # Suporta multiplos CDs (cod_empresa >= 80): cada item pode ir para um CD diferente
        estoque_cd = {}          # {codigo_item: {estoque, qtd_pendente, ...}}
        cd_por_item = {}         # {codigo_item: cod_cd}  - mapeia cada item ao seu CD
        cds_detectados = set()   # Conjunto de CDs encontrados

        if is_destino_cd:
            # Destino unico selecionado pelo usuario (ex: "CD 80")
            for cod in codigos_produtos:
                cd_por_item[cod] = cod_destino
            cds_detectados.add(cod_destino)

        if not is_destino_cd and codigos_produtos:
            # V30: Para pedidos direto loja, verificar estoque em TODOS os CDs
            # Mesmo fornecedores com padrao direto loja podem ter estoque nos CDs
            # (negociacoes comerciais, compras de oportunidade, etc.)
            try:
                conn_cd_check = get_db_connection()
                cursor_cd_check = conn_cd_check.cursor()
                cursor_cd_check.execute("""
                    SELECT cod_empresa, codigo,
                           COALESCE(estoque, 0) as estoque
                    FROM estoque_posicao_atual
                    WHERE cod_empresa >= 80
                      AND codigo = ANY(%s)
                      AND COALESCE(estoque, 0) > 0
                """, (codigos_produtos,))
                for row in cursor_cd_check.fetchall():
                    cod_cd = int(row[0])
                    cod_item = int(row[1])
                    # Mapear item ao CD que tem estoque
                    # Se multiplos CDs tem estoque do mesmo item, pegar o primeiro
                    if cod_item not in cd_por_item:
                        cd_por_item[cod_item] = cod_cd
                        cds_detectados.add(cod_cd)
                cursor_cd_check.close()
                conn_cd_check.close()
                if cds_detectados:
                    print(f"  [CD V30] Estoque CD detectado para direto-loja: {len(cd_por_item)} itens em CDs {sorted(cds_detectados)}")
            except Exception as e_cd30:
                print(f"  [CD V30] Erro ao verificar estoque CDs: {e_cd30}")

        # Carregar estoque de TODOS os CDs detectados
        if cds_detectados and codigos_produtos:
            try:
                conn_cd = get_db_connection()
                cursor_cd = conn_cd.cursor()
                cds_list = list(cds_detectados)
                cursor_cd.execute("""
                    SELECT cod_empresa, codigo,
                           COALESCE(estoque, 0) as estoque,
                           COALESCE(qtd_pendente, 0) as qtd_pendente,
                           COALESCE(qtd_pend_transf, 0) as qtd_pend_transf,
                           COALESCE(cue, 0) as cue
                    FROM estoque_posicao_atual
                    WHERE cod_empresa = ANY(%s) AND codigo = ANY(%s)
                """, (cds_list, codigos_produtos))
                for row in cursor_cd.fetchall():
                    cod_cd = int(row[0])
                    cod_item = int(row[1])
                    # Indexar por item (cada item tem 1 CD via padrao_compra)
                    if cd_por_item.get(cod_item) == cod_cd:
                        estoque_cd[cod_item] = {
                            'cod_cd': cod_cd,
                            'estoque': float(row[2]),
                            'qtd_pendente': float(row[3]),
                            'qtd_pend_transf': float(row[4]),
                            'cue': float(row[5])
                        }
                cursor_cd.close()
                conn_cd.close()
                for cd in sorted(cds_detectados):
                    itens_cd = [k for k, v in estoque_cd.items() if v.get('cod_cd') == cd]
                    print(f"  [CD V29] Estoque CD {cd} carregado: {len(itens_cd)} itens")
            except Exception as e_cd:
                print(f"  [CD V29] Erro ao carregar estoque CD: {e_cd}")

        # Variavel de compatibilidade (usado no JSON de resposta)
        cd_destino_v29 = sorted(cds_detectados)[0] if cds_detectados else None

        # ==============================================================================
        # ETAPA 2: DISTRIBUICAO DO ESTOQUE DO CD PARA LOJAS (V29/V30)
        # ==============================================================================
        # Roda ANTES das transferencias loja<->loja (V25).
        # Estoque no CD e improdutivo (CD nao vende), entao prioriza-se esvaziar o CD
        # antes de transferir estoque entre lojas que vendem.
        # CD = "super-doador" sem restricao de grupo regional.
        # Reserva ES do CD com efeito pooling (sigma_total = sqrt(sum(sigma^2)))
        transferencias_cd = []
        info_distribuicao_cd = {}

        if cds_detectados and estoque_cd:
            print(f"\n  [CD V29/V30] Iniciando distribuicao do estoque de {len(cds_detectados)} CD(s) {sorted(cds_detectados)} para lojas...")

            # Agrupar resultados por codigo de produto (normalizar para int)
            itens_por_codigo_cd = {}
            for item in resultados:
                codigo = int(item.get('codigo', 0))
                if codigo not in itens_por_codigo_cd:
                    itens_por_codigo_cd[codigo] = []
                itens_por_codigo_cd[codigo].append(item)

            # Carregar multiplos de embalagem
            embalagens_multiplo_cd = {}
            try:
                conn_emb_cd = get_db_connection()
                cursor_emb_cd = conn_emb_cd.cursor()
                cursor_emb_cd.execute("""
                    SELECT codigo, qtd_embalagem
                    FROM embalagem_arredondamento
                    WHERE qtd_embalagem > 1
                """)
                embalagens_multiplo_cd = {int(row[0]): int(row[1]) for row in cursor_emb_cd.fetchall()}
                cursor_emb_cd.close()
                conn_emb_cd.close()
            except Exception as e_emb:
                print(f"  [CD V29/V30] Erro ao carregar embalagens: {e_emb}")

            total_distribuido_cd = 0
            total_itens_distribuidos = 0
            valor_economia_cd = 0

            for codigo, itens_loja in itens_por_codigo_cd.items():
                info_cd = estoque_cd.get(codigo, {})
                estoque_cd_disp = info_cd.get('estoque', 0) + info_cd.get('qtd_pend_transf', 0)
                cd_do_item = cd_por_item.get(codigo)

                if estoque_cd_disp <= 0 or not cd_do_item:
                    continue

                # Calcular ES do CD com efeito pooling
                desvios_lojas = []
                curva_item = 'B'
                lead_time_item = 15  # default
                for item in itens_loja:
                    desvio = item.get('desvio_padrao_diario', 0) or 0
                    desvios_lojas.append(desvio)
                    curva_item = item.get('curva_abc', 'B') or 'B'
                    lead_time_item = item.get('lead_time_usado', 15) or 15

                es_cd_info = calcular_es_pooling_cd(desvios_lojas, lead_time_item, curva_item)
                es_cd_valor = es_cd_info['es_cd']

                # Estoque distribuivel = estoque CD - ES do CD (reserva minima)
                estoque_distribuivel = max(0, estoque_cd_disp - es_cd_valor)

                if estoque_distribuivel <= 0:
                    continue

                # Identificar lojas que precisam (pedido > 0)
                receptores_cd = []
                for item in itens_loja:
                    qtd_pedido = float(item.get('quantidade_pedido', 0) or 0)
                    cobertura = float(item.get('cobertura_atual_dias', 0) or 0)
                    demanda_d = float(item.get('demanda_prevista_diaria', 0) or 0)
                    estoque_loja = float(item.get('estoque_atual', 0) or 0)

                    if qtd_pedido > 0 and demanda_d > 0:
                        if estoque_loja == 0:
                            prioridade = 0
                            faixa = 'RUPTURA'
                        elif cobertura <= 30:
                            prioridade = 1
                            faixa = 'CRITICA'
                        elif cobertura <= 60:
                            prioridade = 2
                            faixa = 'ALTA'
                        elif cobertura <= 90:
                            prioridade = 3
                            faixa = 'MEDIA'
                        else:
                            continue  # > 90 dias, nao precisa

                        receptores_cd.append({
                            'item_ref': item,
                            'cod_loja': item.get('cod_loja'),
                            'nome_loja': item.get('nome_loja', f"Loja {item.get('cod_loja')}"),
                            'necessidade': qtd_pedido,
                            'necessidade_restante': qtd_pedido,
                            'prioridade': prioridade,
                            'faixa': faixa,
                            'cobertura_dias': cobertura,
                            'demanda_diaria': demanda_d
                        })

                if not receptores_cd:
                    continue

                # Ordenar por prioridade (RUPTURA primeiro), depois por cobertura
                receptores_cd.sort(key=lambda x: (x['prioridade'], x['cobertura_dias']))

                # Distribuir estoque do CD
                estoque_restante_cd = estoque_distribuivel
                multiplo_emb = embalagens_multiplo_cd.get(codigo, 1)

                for receptor in receptores_cd:
                    if estoque_restante_cd <= 0:
                        break

                    qtd_bruta = min(receptor['necessidade_restante'], estoque_restante_cd)
                    # Arredondar para BAIXO (caixa fechada)
                    if multiplo_emb > 1:
                        qtd_distribuir = (int(qtd_bruta) // multiplo_emb) * multiplo_emb
                    else:
                        qtd_distribuir = int(qtd_bruta)

                    if qtd_distribuir <= 0:
                        continue

                    # Reduzir pedido da loja
                    item_dest = receptor['item_ref']
                    qtd_antes = float(item_dest.get('quantidade_pedido', 0) or 0)
                    nova_qtd = max(0, qtd_antes - qtd_distribuir)
                    # Re-arredondar para multiplo de caixa (para cima)
                    if multiplo_emb > 1:
                        nova_qtd = arredondar_para_multiplo(nova_qtd, multiplo_emb, 'cima')
                    item_dest['quantidade_pedido'] = nova_qtd
                    cue_item = float(item_dest.get('cue', 0) or item_dest.get('preco_custo', 0) or 0)
                    item_dest['valor_pedido'] = round(float(nova_qtd) * cue_item, 2)

                    # Marcar distribuicao do CD no item
                    if 'transferencias_cd' not in item_dest:
                        item_dest['transferencias_cd'] = []
                    item_dest['transferencias_cd'].append({
                        'cd_origem': cd_do_item,
                        'qtd': qtd_distribuir
                    })

                    if nova_qtd == 0:
                        item_dest['deve_pedir'] = False

                    estoque_restante_cd -= qtd_distribuir
                    total_distribuido_cd += qtd_distribuir
                    valor_economia_cd += qtd_distribuir * cue_item

                    transferencias_cd.append({
                        'cod_produto': codigo,
                        'descricao': item_dest.get('descricao', ''),
                        'curva_abc': item_dest.get('curva_abc', ''),
                        'cd_origem': cd_do_item,
                        'loja_destino': receptor['cod_loja'],
                        'nome_loja_destino': receptor['nome_loja'],
                        'estoque_destino': int(receptor.get('item_ref', {}).get('estoque_atual', 0) or 0),
                        'cobertura_destino_dias': round(receptor.get('cobertura_dias', 0), 1),
                        'qtd_distribuida': qtd_distribuir,
                        'urgencia': receptor['faixa'],
                        'cue': cue_item
                    })

                    if qtd_antes != nova_qtd:
                        total_itens_distribuidos += 1

            info_distribuicao_cd = {
                'total_distribuido': total_distribuido_cd,
                'total_transferencias': len(transferencias_cd),
                'total_itens_atendidos': total_itens_distribuidos,
                'valor_economia': round(valor_economia_cd, 2),
                'cds_utilizados': sorted(cds_detectados)
            }

            print(f"  [CD V29/V30] Distribuicao concluida: {total_distribuido_cd} un distribuidas em {len(transferencias_cd)} transferencias")
            print(f"  [CD V29/V30] Economia estimada: R$ {valor_economia_cd:,.2f}")

            # Recalcular itens_pedido apos distribuicao CD (alguns podem ter zerado)
            itens_pedido = [r for r in resultados if r.get('deve_pedir') and not r.get('bloqueado') and (r.get('quantidade_pedido', 0) or 0) > 0]
            itens_ok = [r for r in resultados if not r.get('deve_pedir') and not r.get('bloqueado')]

            # Recalcular agregacao apos distribuicao CD
            agregacao = agregar_por_fornecedor(resultados)

        # ==============================================================================
        # ETAPA 3: TRANSFERENCIAS ENTRE LOJAS (V25)
        # ==============================================================================
        # Apos distribuir estoque do CD, verificar se ha lojas com excesso
        # que podem transferir para lojas que AINDA tem falta
        transferencias_sugeridas = []
        valor_economia_transferencias = 0

        print(f"\n  [DEBUG] is_pedido_multiloja={is_pedido_multiloja}, len(lojas_demanda)={len(lojas_demanda)}")
        print(f"  [DEBUG] lojas_demanda={lojas_demanda}")
        print(f"  [DEBUG] Total resultados antes transferencias: {len(resultados)}")

        if is_pedido_multiloja and len(lojas_demanda) > 1:
            print(f"\n  [TRANSFERENCIAS] Analisando oportunidades de transferencia entre {len(lojas_demanda)} lojas...")
            print(f"  [TRANSFERENCIAS] Total de resultados a analisar: {len(resultados)}")

            # Reconectar ao banco para calcular transferencias
            conn_transf = get_db_connection()
            try:
                calc_transf = TransferenciaRegional(conn_transf)

                # ==============================================================================
                # CARREGAR GRUPOS REGIONAIS DE TRANSFERENCIA
                # ==============================================================================
                # Transferencias so podem ocorrer entre lojas do MESMO grupo regional
                cursor_grupos = conn_transf.cursor()
                cursor_grupos.execute("""
                    SELECT lg.cod_empresa, g.id as grupo_id, g.nome as grupo_nome
                    FROM lojas_grupo_transferencia lg
                    JOIN grupos_transferencia g ON lg.grupo_id = g.id
                    WHERE lg.ativo = TRUE AND g.ativo = TRUE
                """)
                grupos_por_loja = {}
                for row in cursor_grupos.fetchall():
                    cod_empresa, grupo_id, grupo_nome = row
                    grupos_por_loja[cod_empresa] = {'grupo_id': grupo_id, 'grupo_nome': grupo_nome}
                cursor_grupos.close()

                print(f"  [TRANSFERENCIAS] Grupos regionais carregados: {len(set(g['grupo_id'] for g in grupos_por_loja.values()))} grupos, {len(grupos_por_loja)} lojas mapeadas")

                # Agrupar resultados por codigo de produto
                itens_por_codigo = {}
                for item in resultados:
                    codigo = item.get('codigo')
                    if codigo not in itens_por_codigo:
                        itens_por_codigo[codigo] = []
                    itens_por_codigo[codigo].append(item)

                print(f"  [TRANSFERENCIAS] Produtos unicos encontrados: {len(itens_por_codigo)}")

                # DEBUG: Mostrar primeiro produto com multiplas lojas
                for codigo, itens in list(itens_por_codigo.items())[:3]:
                    if len(itens) >= 2:
                        print(f"  [TRANSFERENCIAS] DEBUG produto {codigo}: {len(itens)} lojas")
                        for it in itens[:3]:
                            print(f"    Loja {it.get('cod_loja')}: cob={it.get('cobertura_atual_dias')}, dem={it.get('demanda_prevista_diaria')}, qtd_ped={it.get('quantidade_pedido')}")

                # ==============================================================================
                # CONFIGURACAO DE TRANSFERENCIAS v6.11
                # ==============================================================================
                # Cobertura minima: doador so pode doar se tiver > 90 dias de cobertura
                COBERTURA_MINIMA_DOADOR = 90  # dias

                # Faixas de prioridade para receptores (ordenadas por urgencia)
                # Lojas com cobertura > 90 dias NAO sao receptoras
                FAIXAS_PRIORIDADE = [
                    ('RUPTURA', 0, 0),      # estoque = 0, prioridade 0 (maxima)
                    ('CRITICA', 0, 30),     # 0-30 dias, prioridade 1
                    ('ALTA', 31, 60),       # 31-60 dias, prioridade 2
                    ('MEDIA', 61, 90),      # 61-90 dias, prioridade 3
                ]
                # > 90 dias = NAO RECEBE (pode ser doador)

                # Carregar multiplos de embalagem para arredondamento
                cursor_emb = conn_transf.cursor()
                cursor_emb.execute("""
                    SELECT codigo, qtd_embalagem
                    FROM embalagem_arredondamento
                    WHERE qtd_embalagem > 1
                """)
                embalagens_multiplo = {row[0]: row[1] for row in cursor_emb.fetchall()}
                cursor_emb.close()
                print(f"  [TRANSFERENCIAS] Embalagens carregadas: {len(embalagens_multiplo)} produtos com multiplo > 1")

                produtos_analisados = 0
                for codigo, itens_loja in itens_por_codigo.items():
                    if len(itens_loja) < 2:
                        continue  # Precisa de pelo menos 2 lojas

                    produtos_analisados += 1

                    # Identificar lojas com EXCESSO (pode doar) e FALTA (precisa receber)
                    lojas_com_excesso = []
                    lojas_com_falta = []

                    for item in itens_loja:
                        cobertura_atual = item.get('cobertura_atual_dias', 0) or 0
                        demanda_diaria = item.get('demanda_prevista_diaria', 0) or 0
                        estoque_atual = item.get('estoque_atual', 0) or 0
                        cod_loja = item.get('cod_loja', 0)
                        nome_loja = item.get('nome_loja', f'Loja {cod_loja}')
                        qtd_pedido = item.get('quantidade_pedido', 0) or 0
                        cue = item.get('cue', 0) or item.get('preco_custo', 0) or 0

                        # ==============================================================
                        # v6.11: NOVA LOGICA DE IDENTIFICACAO DE DOADORES E RECEPTORES
                        # ==============================================================
                        # DOADOR: Cobertura > 90 dias (so transfere o excesso acima de 90 dias)
                        # RECEPTOR: Cobertura <= 90 dias E tem pedido pendente
                        # Prioridade: RUPTURA > CRITICA(0-30d) > ALTA(31-60d) > MEDIA(61-90d)

                        # Buscar grupo regional da loja (usado em ambos os casos)
                        grupo_info = grupos_por_loja.get(cod_loja, {})

                        # DOADOR: So pode doar se cobertura > 90 dias
                        if cobertura_atual > COBERTURA_MINIMA_DOADOR and demanda_diaria > 0:
                            # Excesso = apenas o que esta ACIMA de 90 dias
                            excesso_dias = cobertura_atual - COBERTURA_MINIMA_DOADOR
                            excesso_unidades = int(excesso_dias * demanda_diaria)

                            # Estoque minimo que deve manter = 90 dias de cobertura
                            estoque_minimo = COBERTURA_MINIMA_DOADOR * demanda_diaria
                            disponivel_doar = max(0, int(estoque_atual - estoque_minimo))

                            if disponivel_doar > 0:
                                lojas_com_excesso.append({
                                    'cod_loja': cod_loja,
                                    'nome_loja': nome_loja,
                                    'estoque': estoque_atual,
                                    'cobertura_dias': cobertura_atual,
                                    'cobertura_minima': COBERTURA_MINIMA_DOADOR,
                                    'demanda_diaria': demanda_diaria,
                                    'excesso_unidades': min(excesso_unidades, disponivel_doar),
                                    'disponivel_doar': disponivel_doar,
                                    'cue': cue,
                                    'grupo_id': grupo_info.get('grupo_id'),
                                    'grupo_nome': grupo_info.get('grupo_nome'),
                                    'usado': False  # Flag para algoritmo de matching
                                })

                        # RECEPTOR: Cobertura <= 90 dias E tem pedido pendente
                        # (lojas com > 90 dias nao precisam receber, podem doar)
                        if cobertura_atual <= COBERTURA_MINIMA_DOADOR and qtd_pedido > 0:
                            # Calcular faixa de prioridade
                            prioridade = 99  # Default (nao deve receber)
                            faixa_nome = 'SEM_PRIORIDADE'

                            # Verificar se esta em RUPTURA (estoque = 0)
                            if estoque_atual == 0:
                                prioridade = 0
                                faixa_nome = 'RUPTURA'
                            else:
                                # Verificar faixas de cobertura
                                for faixa, min_dias, max_dias in FAIXAS_PRIORIDADE:
                                    if faixa == 'RUPTURA':
                                        continue  # Ja tratado acima
                                    if min_dias <= cobertura_atual <= max_dias:
                                        prioridade = FAIXAS_PRIORIDADE.index((faixa, min_dias, max_dias))
                                        faixa_nome = faixa
                                        break

                            # Apenas adiciona se tem prioridade valida (cobertura <= 90 dias)
                            if prioridade < 99:
                                lojas_com_falta.append({
                                    'cod_loja': cod_loja,
                                    'nome_loja': nome_loja,
                                    'estoque': estoque_atual,
                                    'cobertura_dias': cobertura_atual,
                                    'demanda_diaria': demanda_diaria,
                                    'necessidade': qtd_pedido,
                                    'necessidade_restante': qtd_pedido,  # Para controle do algoritmo
                                    'item_ref': item,
                                    'cue': cue,
                                    'grupo_id': grupo_info.get('grupo_id'),
                                    'grupo_nome': grupo_info.get('grupo_nome'),
                                    'prioridade': prioridade,
                                    'faixa': faixa_nome
                                })

                    # Debug: mostrar produtos com potencial de transferencia (v6.11)
                    if lojas_com_falta and not lojas_com_excesso:
                        # Mostrar por que nao ha doadoras
                        coberturas_lojas = [(item.get('cod_loja'), item.get('cobertura_atual_dias', 0)) for item in itens_loja]
                        max_cobertura = max([c[1] for c in coberturas_lojas]) if coberturas_lojas else 0
                        if max_cobertura > 0 and max_cobertura <= COBERTURA_MINIMA_DOADOR:
                            print(f"    Produto {codigo}: SEM DOADORAS - cobertura max={max_cobertura:.1f}d (< {COBERTURA_MINIMA_DOADOR}d minimo), {len(lojas_com_falta)} lojas precisam")
                    elif lojas_com_excesso or lojas_com_falta:
                        # Mostrar grupos e faixas para debug
                        grupos_exc = set(l.get('grupo_nome', 'SEM_GRUPO') for l in lojas_com_excesso)
                        faixas_falta = [f"{l['faixa']}({l['cod_loja']})" for l in lojas_com_falta]
                        print(f"    Produto {codigo}: {len(lojas_com_excesso)} doadoras(>{COBERTURA_MINIMA_DOADOR}d) {grupos_exc}, receptoras: {faixas_falta[:5]}")

                    # ==============================================================
                    # ALGORITMO DE MATCHING OTIMIZADO v6.11
                    # ==============================================================
                    # Objetivo: Minimizar fragmentacao de transferencias (otimizar frete)
                    # Estrategia: Preferir 1 doador → 1 receptor (match completo)
                    # Ordenacao: Receptores por PRIORIDADE (ruptura primeiro), doadores por EXCESSO

                    if lojas_com_excesso and lojas_com_falta:
                        # Ordenar receptores por prioridade (menor = mais urgente)
                        # Empate: ordenar por menor cobertura
                        lojas_com_falta.sort(key=lambda x: (x['prioridade'], x['cobertura_dias']))

                        # Ordenar doadores por excesso disponivel (maior primeiro)
                        lojas_com_excesso.sort(key=lambda x: -x['disponivel_doar'])

                        descricao_prod = itens_loja[0].get('descricao', '')

                        # Buscar multiplo de embalagem do produto
                        multiplo_emb = embalagens_multiplo.get(codigo, 1)

                        for destino in lojas_com_falta:
                            if destino['necessidade_restante'] <= 0:
                                continue

                            grupo_destino = destino.get('grupo_id')
                            if grupo_destino is None:
                                continue  # Loja nao esta em nenhum grupo

                            # ==============================================================
                            # BUSCAR MELHOR DOADOR (OTIMIZACAO DE FRETE)
                            # ==============================================================
                            # Preferir doador que consegue cobrir 100% da necessidade
                            # para evitar fragmentacao (1 doador para varios receptores)

                            melhor_doador = None
                            for origem in lojas_com_excesso:
                                if origem['usado'] or origem['disponivel_doar'] <= 0:
                                    continue
                                if origem['cod_loja'] == destino['cod_loja']:
                                    continue  # Mesma loja

                                # Validar grupo regional (mesmo grupo apenas)
                                grupo_origem = origem.get('grupo_id')
                                if grupo_origem is None or grupo_origem != grupo_destino:
                                    continue

                                # Preferir doador que cobre 100% da necessidade
                                if origem['disponivel_doar'] >= destino['necessidade_restante']:
                                    melhor_doador = origem
                                    break  # Match perfeito encontrado

                                # Senao, guardar maior disponivel do mesmo grupo
                                if melhor_doador is None:
                                    melhor_doador = origem
                                elif origem['disponivel_doar'] > melhor_doador['disponivel_doar']:
                                    melhor_doador = origem

                            if melhor_doador is None:
                                continue  # Nenhum doador disponivel no mesmo grupo

                            # ==============================================================
                            # CALCULAR QUANTIDADE COM MULTIPLO DE EMBALAGEM
                            # ==============================================================
                            qtd_bruta = min(destino['necessidade_restante'], melhor_doador['disponivel_doar'])

                            # Arredondar para BAIXO (caixas fechadas apenas)
                            qtd_transferir = (qtd_bruta // multiplo_emb) * multiplo_emb

                            # Se arredondou para 0, nao transfere
                            if qtd_transferir <= 0:
                                continue

                            # Calcular valor
                            cue_usar = melhor_doador['cue'] if melhor_doador['cue'] > 0 else destino['cue']
                            valor_transf = round(qtd_transferir * cue_usar, 2)

                            # Registrar transferencia
                            transferencias_sugeridas.append({
                                'cod_produto': codigo,
                                'descricao': descricao_prod,
                                'loja_origem': melhor_doador['cod_loja'],
                                'nome_loja_origem': melhor_doador['nome_loja'],
                                'estoque_origem': melhor_doador['estoque'],
                                'cobertura_origem_dias': round(melhor_doador['cobertura_dias'], 1),
                                'loja_destino': destino['cod_loja'],
                                'nome_loja_destino': destino['nome_loja'],
                                'estoque_destino': destino['estoque'],
                                'cobertura_destino_dias': round(destino['cobertura_dias'], 1),
                                'qtd_sugerida': qtd_transferir,
                                'valor_estimado': valor_transf,
                                'cue': cue_usar,
                                'urgencia': destino['faixa'],  # v6.11: usar faixa de prioridade
                                'multiplo_embalagem': multiplo_emb
                            })

                            # Atualizar controles
                            valor_economia_transferencias += valor_transf
                            melhor_doador['disponivel_doar'] -= qtd_transferir
                            destino['necessidade_restante'] -= qtd_transferir

                            # Marcar doador como usado se esgotou capacidade
                            if melhor_doador['disponivel_doar'] <= 0:
                                melhor_doador['usado'] = True

                            # IMPORTANTE: Reduzir quantidade do pedido do item destino
                            item_destino = destino['item_ref']
                            qtd_pedido_atual = item_destino.get('quantidade_pedido', 0) or 0
                            nova_qtd = max(0, qtd_pedido_atual - qtd_transferir)

                            # Arredondar para multiplo de caixa apos transferencia
                            multiplo_caixa = item_destino.get('multiplo_caixa', 1) or 1
                            nova_qtd = arredondar_para_multiplo(nova_qtd, multiplo_caixa, 'cima')

                            item_destino['quantidade_pedido'] = nova_qtd
                            item_destino['valor_pedido'] = round(nova_qtd * cue_usar, 2)

                            # Marcar que tem transferencia
                            if 'transferencias_receber' not in item_destino:
                                item_destino['transferencias_receber'] = []
                            item_destino['transferencias_receber'].append({
                                'loja_origem': melhor_doador['cod_loja'],
                                'nome_loja_origem': melhor_doador['nome_loja'],
                                'qtd': qtd_transferir
                            })

                            # Se pedido zerou, nao precisa mais pedir
                            if nova_qtd == 0:
                                item_destino['deve_pedir'] = False

                print(f"  [TRANSFERENCIAS] Produtos analisados: {produtos_analisados}")
                print(f"  [TRANSFERENCIAS] Encontradas {len(transferencias_sugeridas)} oportunidades")
                print(f"  [TRANSFERENCIAS] Economia estimada: R$ {valor_economia_transferencias:,.2f}")

            except Exception as e:
                print(f"  [TRANSFERENCIAS] Erro ao calcular: {e}")
            finally:
                conn_transf.close()

            # Recalcular itens_pedido apos transferencias (alguns podem ter zerado)
            itens_pedido = [r for r in resultados if r.get('deve_pedir') and not r.get('bloqueado') and (r.get('quantidade_pedido', 0) or 0) > 0]
            itens_ok = [r for r in resultados if not r.get('deve_pedir') and not r.get('bloqueado')]

            # ==============================================================================
            # SALVAR TRANSFERENCIAS NO BANCO DE DADOS
            # ==============================================================================
            if transferencias_sugeridas:
                try:
                    conn_save = get_db_connection()
                    cursor_save = conn_save.cursor()

                    # Limpar transferencias antigas (mais de 24h)
                    cursor_save.execute("""
                        DELETE FROM oportunidades_transferencia
                        WHERE data_calculo < NOW() - INTERVAL '24 hours'
                    """)

                    # Inserir novas transferencias (usando colunas da tabela real)
                    for transf in transferencias_sugeridas:
                        cursor_save.execute("""
                            INSERT INTO oportunidades_transferencia (
                                cod_produto, descricao_produto, curva_abc,
                                loja_origem, nome_loja_origem, estoque_origem, cobertura_origem_dias,
                                loja_destino, nome_loja_destino, estoque_destino, cobertura_destino_dias,
                                qtd_sugerida, valor_estimado, cue, urgencia, data_calculo, status
                            ) VALUES (
                                %s, %s, %s,
                                %s, %s, %s, %s,
                                %s, %s, %s, %s,
                                %s, %s, %s, %s, NOW(), 'pendente'
                            )
                        """, (
                            transf.get('cod_produto', 0),
                            transf.get('descricao', '')[:300] if transf.get('descricao') else '',
                            transf.get('curva_abc', 'B'),
                            transf.get('loja_origem', 0),
                            transf.get('nome_loja_origem', '')[:100] if transf.get('nome_loja_origem') else '',
                            transf.get('estoque_origem', 0),
                            transf.get('cobertura_origem_dias', 0),
                            transf.get('loja_destino', 0),
                            transf.get('nome_loja_destino', '')[:100] if transf.get('nome_loja_destino') else '',
                            transf.get('estoque_destino', 0),
                            transf.get('cobertura_destino_dias', 0),
                            transf.get('qtd_sugerida', 0),
                            transf.get('valor_estimado', 0),
                            transf.get('cue', 0),
                            transf.get('urgencia', 'MEDIA')
                        ))

                    conn_save.commit()
                    cursor_save.close()
                    conn_save.close()
                    print(f"  [TRANSFERENCIAS] {len(transferencias_sugeridas)} oportunidades salvas no banco")
                except Exception as e_save:
                    print(f"  [TRANSFERENCIAS] Erro ao salvar no banco: {e_save}")

        # ==============================================================================
        # SALVAR TRANSFERENCIAS CD -> LOJAS NO BANCO DE DADOS (V30)
        # ==============================================================================
        if transferencias_cd:
            try:
                conn_save_cd = get_db_connection()
                cursor_save_cd = conn_save_cd.cursor()

                # Nomes dos CDs para exibicao
                NOMES_CD = {80: 'CD MDC', 81: 'CD OBC', 82: 'CD OSAL', 83: 'CD 83',
                            92: 'CD CAB', 93: 'CD ALH', 94: 'CD LAU'}

                for transf in transferencias_cd:
                    cd_orig = transf.get('cd_origem', 0)
                    cod_prod = transf.get('cod_produto', 0)
                    est_cd_info = estoque_cd.get(cod_prod, {})

                    cursor_save_cd.execute("""
                        INSERT INTO oportunidades_transferencia (
                            cod_produto, descricao_produto, curva_abc,
                            loja_origem, nome_loja_origem, estoque_origem, cobertura_origem_dias,
                            loja_destino, nome_loja_destino, estoque_destino, cobertura_destino_dias,
                            qtd_sugerida, valor_estimado, cue, urgencia, data_calculo, status
                        ) VALUES (
                            %s, %s, %s,
                            %s, %s, %s, %s,
                            %s, %s, %s, %s,
                            %s, %s, %s, %s, NOW(), 'pendente'
                        )
                    """, (
                        cod_prod,
                        transf.get('descricao', '')[:300],
                        transf.get('curva_abc', ''),
                        cd_orig,
                        NOMES_CD.get(cd_orig, f'CD {cd_orig}'),
                        int(est_cd_info.get('estoque', 0)),
                        0,  # cobertura_origem - CD nao tem demanda propria
                        transf.get('loja_destino', 0),
                        transf.get('nome_loja_destino', '')[:100],
                        transf.get('estoque_destino', 0),
                        transf.get('cobertura_destino_dias', 0),
                        transf.get('qtd_distribuida', 0),
                        round(transf.get('qtd_distribuida', 0) * transf.get('cue', 0), 2),
                        transf.get('cue', 0),
                        transf.get('urgencia', 'MEDIA')
                    ))

                conn_save_cd.commit()
                cursor_save_cd.close()
                conn_save_cd.close()
                print(f"  [CD V30] {len(transferencias_cd)} transferencias CD->lojas salvas no banco")
            except Exception as e_save_cd:
                print(f"  [CD V30] Erro ao salvar transferencias CD no banco: {e_save_cd}")

        # Validacao de pedido minimo
        validacao_pedido_minimo = {}
        alertas_pedido_minimo = []

        for item in itens_pedido:
            cod_fornecedor = item.get('codigo_fornecedor', '')
            cod_loja = item.get('cod_loja', 0)
            nome_fornecedor = item.get('nome_fornecedor', '')
            nome_loja = item.get('nome_loja', f'Loja {cod_loja}')
            valor_pedido = item.get('valor_pedido', 0) or 0
            pedido_minimo = item.get('pedido_minimo_fornecedor', 0) or 0

            chave = (cod_fornecedor, cod_loja)

            if chave not in validacao_pedido_minimo:
                validacao_pedido_minimo[chave] = {
                    'cod_fornecedor': cod_fornecedor,
                    'nome_fornecedor': nome_fornecedor,
                    'cod_loja': cod_loja,
                    'nome_loja': nome_loja,
                    'valor_total': 0,
                    'pedido_minimo': pedido_minimo,
                    'abaixo_minimo': False,
                    'diferenca': 0
                }

            validacao_pedido_minimo[chave]['valor_total'] += valor_pedido

        for chave, info in validacao_pedido_minimo.items():
            pedido_min = info['pedido_minimo']
            valor_total = info['valor_total']
            if pedido_min > 0 and valor_total < pedido_min:
                info['abaixo_minimo'] = True
                info['diferenca'] = pedido_min - valor_total
                alertas_pedido_minimo.append(info)

        for item in itens_pedido:
            cod_fornecedor = item.get('codigo_fornecedor', '')
            cod_loja = item.get('cod_loja', 0)
            chave = (cod_fornecedor, cod_loja)
            if chave in validacao_pedido_minimo:
                info = validacao_pedido_minimo[chave]
                item['pedido_abaixo_minimo'] = info['abaixo_minimo']
                if info['abaixo_minimo']:
                    item['critica_pedido_minimo'] = f"Abaixo do pedido minimo"

        # Estatisticas
        itens_nao_bloqueados = [r for r in resultados if not r.get('bloqueado')]
        coberturas_atuais = [r.get('cobertura_atual_dias', 0) for r in itens_nao_bloqueados if r.get('cobertura_atual_dias', 0) < 900]
        coberturas_pos = [r.get('cobertura_pos_pedido_dias', 0) for r in itens_pedido if r.get('cobertura_pos_pedido_dias', 0) < 900]

        cobertura_media_atual = round(np.mean(coberturas_atuais), 1) if coberturas_atuais else 0
        cobertura_media_pos = round(np.mean(coberturas_pos), 1) if coberturas_pos else 0

        if math.isnan(cobertura_media_atual) or math.isinf(cobertura_media_atual):
            cobertura_media_atual = 0
        if math.isnan(cobertura_media_pos) or math.isinf(cobertura_media_pos):
            cobertura_media_pos = 0

        estatisticas = {
            'total_itens_analisados': len(resultados),
            'total_itens_pedido': len(itens_pedido),
            'total_itens_bloqueados': len(itens_bloqueados),
            'total_itens_ok': len(itens_ok),
            'valor_total_pedido': round(sum(r.get('valor_pedido', 0) for r in itens_pedido), 2),
            'itens_ruptura_iminente': len([r for r in itens_nao_bloqueados if r.get('ruptura_iminente')]),
            'cobertura_media_atual': cobertura_media_atual,
            'cobertura_media_pos_pedido': cobertura_media_pos,
            # Estatisticas de transferencias
            'total_transferencias_sugeridas': len(transferencias_sugeridas),
            'valor_economia_transferencias': round(valor_economia_transferencias, 2),
            'produtos_com_transferencia': len(set(t['cod_produto'] for t in transferencias_sugeridas)) if transferencias_sugeridas else 0
        }

        return jsonify(converter_tipos_json({
            'success': True,
            'estatisticas': estatisticas,
            'agregacao_fornecedor': agregacao,
            'itens_pedido': itens_pedido,
            'itens_bloqueados': itens_bloqueados,
            'itens_ok': itens_ok,
            'is_pedido_multiloja': is_pedido_multiloja,
            'lojas_info': [{'cod_loja': cod, 'nome_loja': nome} for cod, nome in mapa_nomes_lojas.items()] if is_pedido_multiloja else [],
            'alertas_pedido_minimo': [
                {
                    'nome_fornecedor': a['nome_fornecedor'],
                    'nome_loja': a['nome_loja'],
                    'valor_total': a['valor_total'],
                    'pedido_minimo': a['pedido_minimo'],
                    'diferenca': a['diferenca']
                }
                for a in alertas_pedido_minimo
            ],
            'tem_alertas_pedido_minimo': len(alertas_pedido_minimo) > 0,
            # Transferencias sugeridas entre lojas
            'transferencias_sugeridas': transferencias_sugeridas,
            'tem_transferencias': len(transferencias_sugeridas) > 0,
            'resumo_transferencias': {
                'total': len(transferencias_sugeridas),
                'valor_total': round(valor_economia_transferencias, 2),
                'produtos_unicos': len(set(t['cod_produto'] for t in transferencias_sugeridas)) if transferencias_sugeridas else 0,
                'urgentes': len([t for t in transferencias_sugeridas if t.get('urgencia') == 'ALTA'])
            } if transferencias_sugeridas else None,
            # Distribuicao do CD para lojas (V29)
            'distribuicao_cd': transferencias_cd,
            'tem_distribuicao_cd': len(transferencias_cd) > 0,
            'resumo_distribuicao_cd': info_distribuicao_cd if info_distribuicao_cd else None,
            'is_destino_cd': is_destino_cd,
            'cds_detectados_v29': sorted(cds_detectados) if cds_detectados else [],
            'cd_por_item': {str(k): v for k, v in cd_por_item.items()} if cd_por_item else None,
            'estoque_cd_info': {
                str(cod): {
                    'cd': info.get('cod_cd', 0),
                    'estoque': info.get('estoque', 0),
                    'qtd_pendente': info.get('qtd_pendente', 0),
                    'qtd_pend_transf': info.get('qtd_pend_transf', 0)
                } for cod, info in estoque_cd.items()
            } if estoque_cd else None
        }))

    except Exception as e:
        import traceback
        print(f"[ERRO] {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'erro': str(e)
        }), 500


@pedido_fornecedor_bp.route('/api/pedido_fornecedor_integrado/exportar', methods=['POST'])
def api_pedido_fornecedor_integrado_exportar():
    """Exporta o pedido ao fornecedor para Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        dados = request.get_json()

        if not dados:
            return jsonify({'success': False, 'erro': 'Dados nao fornecidos'}), 400

        itens_pedido = dados.get('itens_pedido', [])
        itens_bloqueados = dados.get('itens_bloqueados', [])
        estatisticas = dados.get('estatisticas', {})
        agregacao = dados.get('agregacao_fornecedor', {})

        # Debug: mostrar quantos itens chegaram
        print(f"[EXPORT] Recebido: {len(itens_pedido)} itens_pedido, {len(itens_bloqueados)} itens_bloqueados")

        # Permitir exportar mesmo sem itens de pedido, se houver bloqueados
        if not itens_pedido and not itens_bloqueados:
            return jsonify({'success': False, 'erro': 'Nenhum item para exportar'}), 400

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        critica_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        critica_font = Font(color='B45309')

        # Aba Resumo
        ws_resumo = wb.active
        ws_resumo.title = 'Resumo'

        ws_resumo['A1'] = 'PEDIDO AO FORNECEDOR - RESUMO'
        ws_resumo['A1'].font = Font(bold=True, size=14)
        ws_resumo.merge_cells('A1:D1')

        ws_resumo['A2'] = f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}'

        ws_resumo['A4'] = 'Estatisticas'
        ws_resumo['A4'].font = Font(bold=True)

        stats_data = [
            ('Total de Itens a Pedir', estatisticas.get('total_itens_pedido', 0)),
            ('Valor Total do Pedido', f"R$ {estatisticas.get('valor_total_pedido', 0):,.2f}"),
            ('Itens Analisados', estatisticas.get('total_itens_analisados', 0)),
            ('Total de Fornecedores', agregacao.get('total_fornecedores', 0)),
        ]

        for i, (label, valor) in enumerate(stats_data, start=5):
            ws_resumo[f'A{i}'] = label
            ws_resumo[f'B{i}'] = valor

        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20

        # Aba Itens
        ws_itens = wb.create_sheet('Itens Pedido')

        # Mapeamento de codigo de filial para nome abreviado
        NOMES_FILIAIS = {
            1: 'GUS', 2: 'IMB', 3: 'PAL', 4: 'TAM', 5: 'AJU',
            6: 'JPA', 7: 'PNG', 8: 'CAU', 9: 'BAR', 11: 'FRT',
            80: 'MDC', 81: 'OBC', 82: 'OSAL', 91: 'CA2',
            92: 'CAB', 93: 'ALH', 94: 'LAU'
        }

        # Carregar mapeamento de codigos DIG
        codigos_dig = {}
        try:
            conn_dig = get_db_connection()
            cursor_dig = conn_dig.cursor()
            cursor_dig.execute("SELECT codigo, codigo_dig FROM codigo_dig")
            codigos_dig = {str(row[0]): str(row[1]) for row in cursor_dig.fetchall()}
            cursor_dig.close()
            conn_dig.close()
            print(f"[EXPORT] Codigos DIG carregados: {len(codigos_dig)} registros")
        except Exception as e:
            print(f"[EXPORT] Erro ao carregar codigos DIG: {e}")

        headers = [
            'Cod Filial', 'Filial', 'Cod Item', 'Cod DIG', 'Descricao Item',
            'CNPJ Fornecedor', 'Nome Fornecedor', 'Qtd Pedido', 'CUE', 'Critica'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_itens.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Filtrar apenas itens com pedido > 0 e ordenar
        itens_com_pedido = [item for item in itens_pedido if (item.get('quantidade_pedido', 0) or 0) > 0]
        itens_ordenados = sorted(itens_com_pedido, key=lambda x: (
            x.get('cod_loja', 0) or 0,
            x.get('cnpj_fornecedor', '') or ''
        ))

        # DEBUG: mostrar primeiro item para verificar estrutura
        if itens_ordenados:
            primeiro = itens_ordenados[0]
            print(f"[EXPORT] Primeiro item - cod_loja: {primeiro.get('cod_loja')} (tipo: {type(primeiro.get('cod_loja')).__name__})")
            print(f"[EXPORT] Primeiro item - codigo: {primeiro.get('codigo')} (tipo: {type(primeiro.get('codigo')).__name__})")

        for row_idx, item in enumerate(itens_ordenados, start=2):
            cod_loja = item.get('cod_loja', 0)
            # Garantir que cod_loja e inteiro (pode vir como numpy.int64 ou string)
            try:
                cod_loja_int = int(cod_loja) if cod_loja else 0
            except (ValueError, TypeError):
                cod_loja_int = 0
            ws_itens.cell(row=row_idx, column=1, value=cod_loja_int).border = border
            # Nome abreviado da filial
            nome_abrev = NOMES_FILIAIS.get(cod_loja_int, '')
            ws_itens.cell(row=row_idx, column=2, value=nome_abrev).border = border
            cod_item = str(item.get('codigo', ''))
            ws_itens.cell(row=row_idx, column=3, value=cod_item).border = border
            # Codigo DIG
            cod_dig = codigos_dig.get(cod_item, 'N/D')
            ws_itens.cell(row=row_idx, column=4, value=cod_dig).border = border

            # DEBUG: mostrar primeiros 3 itens
            if row_idx <= 4:
                print(f"[EXPORT] Item {row_idx-1}: cod_loja={cod_loja_int}, filial={nome_abrev}, codigo={cod_item}, dig={cod_dig}")
            ws_itens.cell(row=row_idx, column=5, value=item.get('descricao', '')[:60]).border = border
            ws_itens.cell(row=row_idx, column=6, value=item.get('cnpj_fornecedor', item.get('codigo_fornecedor', ''))).border = border
            ws_itens.cell(row=row_idx, column=7, value=item.get('nome_fornecedor', '')).border = border
            ws_itens.cell(row=row_idx, column=8, value=item.get('quantidade_pedido', 0)).border = border
            cue = item.get('cue', item.get('preco_custo', 0))
            ws_itens.cell(row=row_idx, column=9, value=cue).border = border
            ws_itens.cell(row=row_idx, column=9).number_format = '#,##0.00'
            critica = item.get('critica_pedido_minimo', '')
            cell_critica = ws_itens.cell(row=row_idx, column=10, value=critica)
            cell_critica.border = border
            if critica:
                cell_critica.fill = critica_fill
                cell_critica.font = critica_font

        col_widths = [10, 8, 12, 12, 50, 18, 30, 12, 12, 45]
        for i, width in enumerate(col_widths, start=1):
            ws_itens.column_dimensions[chr(64 + i)].width = width

        ws_itens.freeze_panes = 'A2'

        # Aba Itens Bloqueados (Situacao FL, NC, EN, etc.)
        if itens_bloqueados:
            ws_bloq = wb.create_sheet('Itens Bloqueados')

            # Estilo para itens bloqueados (vermelho claro)
            bloq_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            bloq_header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

            headers_bloq = [
                'Cod Filial', 'Nome Filial', 'Cod Item', 'Descricao Item',
                'Nome Fornecedor', 'Situacao', 'Motivo Bloqueio', 'Estoque Atual'
            ]

            for col, header in enumerate(headers_bloq, start=1):
                cell = ws_bloq.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = bloq_header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Ordenar por loja e fornecedor
            itens_bloq_ordenados = sorted(itens_bloqueados, key=lambda x: (
                x.get('cod_loja', 0) or 0,
                x.get('nome_fornecedor', '') or ''
            ))

            for row_idx, item in enumerate(itens_bloq_ordenados, start=2):
                ws_bloq.cell(row=row_idx, column=1, value=item.get('cod_loja', '')).border = border
                ws_bloq.cell(row=row_idx, column=2, value=item.get('nome_loja', '')).border = border
                ws_bloq.cell(row=row_idx, column=3, value=item.get('codigo', '')).border = border
                ws_bloq.cell(row=row_idx, column=4, value=item.get('descricao', '')[:60]).border = border
                ws_bloq.cell(row=row_idx, column=5, value=item.get('nome_fornecedor', '')).border = border

                # Situacao de compra (FL, NC, EN, etc.)
                situacao = item.get('sit_compra', item.get('situacao_compra', ''))
                ws_bloq.cell(row=row_idx, column=6, value=situacao).border = border

                # Motivo do bloqueio
                motivo = item.get('motivo_bloqueio', '')
                if not motivo:
                    # Gerar motivo baseado na situacao se nao vier preenchido
                    motivos_por_situacao = {
                        'FL': 'Fora de Linha - Item descontinuado',
                        'NC': 'Nao Comprar - Bloqueado para compras',
                        'EN': 'Encomenda - Apenas sob demanda',
                        'FF': 'Falta no Fornecedor',
                        'CO': 'Compra Oportunidade'
                    }
                    motivo = motivos_por_situacao.get(situacao, f'Bloqueado por situacao {situacao}')

                cell_motivo = ws_bloq.cell(row=row_idx, column=7, value=motivo)
                cell_motivo.border = border
                cell_motivo.fill = bloq_fill

                # Estoque atual
                estoque = item.get('estoque_atual', 0)
                ws_bloq.cell(row=row_idx, column=8, value=estoque).border = border

            # Largura das colunas
            col_widths_bloq = [12, 25, 12, 50, 30, 12, 45, 15]
            for i, width in enumerate(col_widths_bloq, start=1):
                ws_bloq.column_dimensions[chr(64 + i)].width = width

            ws_bloq.freeze_panes = 'A2'

            # Atualizar resumo com info de bloqueados
            ws_resumo['A10'] = 'Itens Bloqueados'
            ws_resumo['A10'].font = Font(bold=True)
            ws_resumo['A11'] = 'Total Bloqueados'
            ws_resumo['B11'] = len(itens_bloqueados)

            # Contar por situacao
            situacoes_count = {}
            for item in itens_bloqueados:
                sit = item.get('sit_compra', item.get('situacao_compra', 'N/A'))
                situacoes_count[sit] = situacoes_count.get(sit, 0) + 1

            row_sit = 12
            for sit, count in sorted(situacoes_count.items()):
                ws_resumo[f'A{row_sit}'] = f'  - {sit}'
                ws_resumo[f'B{row_sit}'] = count
                row_sit += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pedido_fornecedor_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500
