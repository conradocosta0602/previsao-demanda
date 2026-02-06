"""
Blueprint: Pedido Planejado
Rotas: /api/pedido_planejado, /api/pedido_planejado/exportar
"""

import io
import math
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from flask import Blueprint, request, jsonify, send_file
from psycopg2.extras import RealDictCursor

from app.utils.db_connection import get_db_connection
from app.utils.demanda_pre_calculada import calcular_proporcoes_vendas_por_loja
from decimal import Decimal

pedido_planejado_bp = Blueprint('pedido_planejado', __name__)


def to_float(value, default=0.0):
    """Converte qualquer valor numerico (incluindo Decimal) para float."""
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def to_int(value, default=0):
    """Converte qualquer valor numerico (incluindo Decimal) para int."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


@pedido_planejado_bp.route('/api/pedido_planejado', methods=['POST'])
def api_pedido_planejado():
    """
    Gera pedido planejado para um periodo futuro.

    Diferenca do pedido normal (reabastecimento):
    - Usa demanda validada para o periodo especificado
    - Calcula data de emissao com offset de lead time
    - Nao considera estoque atual (assume que sera consumido ate la)
    """
    try:
        from core.pedido_fornecedor_integrado import (
            PedidoFornecedorIntegrado,
            agregar_por_fornecedor
        )
        from datetime import datetime as dt, timedelta

        dados = request.get_json()

        # Parametros de filtro
        fornecedor_filtro = dados.get('fornecedor', 'TODOS')
        linha1_filtro = dados.get('linha1', 'TODAS')
        linha3_filtro = dados.get('linha3', 'TODAS')
        destino_tipo = dados.get('destino_tipo', 'LOJA')
        cod_empresa = dados.get('cod_empresa', 'TODAS')

        # Normalizar filtros - converter arrays de 1 elemento para string
        def normalizar_filtro(filtro, valor_padrao='TODAS'):
            if filtro is None:
                return valor_padrao
            if isinstance(filtro, str):
                return filtro
            try:
                filtro_lista = list(filtro)
                if len(filtro_lista) == 0:
                    return valor_padrao
                elif len(filtro_lista) == 1:
                    valor = filtro_lista[0]
                    return str(valor) if valor is not None else valor_padrao
                else:
                    return [str(v) for v in filtro_lista]
            except (TypeError, ValueError):
                pass
            return filtro

        linha1_filtro = normalizar_filtro(linha1_filtro, 'TODAS')
        linha3_filtro = normalizar_filtro(linha3_filtro, 'TODAS')

        # Parametros especificos do pedido planejado
        periodo_inicio = dados.get('periodo_inicio')
        periodo_fim = dados.get('periodo_fim')
        usar_demanda_validada = dados.get('usar_demanda_validada', True)

        if not periodo_inicio or not periodo_fim:
            return jsonify({'erro': 'Parametros obrigatorios: periodo_inicio, periodo_fim'}), 400

        print(f"\n{'='*60}")
        print("PEDIDO PLANEJADO")
        print(f"{'='*60}")
        print(f"  Periodo: {periodo_inicio} ate {periodo_fim}")
        print(f"  Fornecedor: {fornecedor_filtro}")
        print(f"  Destino: {destino_tipo}")
        print(f"  Empresa: {cod_empresa}")
        print(f"  Usar demanda validada: {usar_demanda_validada}")

        conn = get_db_connection()

        # Normalizar cod_empresa
        lojas_selecionadas = []
        if isinstance(cod_empresa, list):
            lojas_selecionadas = [int(x) for x in cod_empresa]
        elif cod_empresa != 'TODAS':
            lojas_selecionadas = [int(cod_empresa)]

        # Buscar lojas se nao especificadas
        if not lojas_selecionadas:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            tipo_filtro = 'Loja' if destino_tipo.upper() == 'LOJA' else 'CD'
            cursor.execute("""
                SELECT cod_empresa FROM cadastro_lojas
                WHERE tipo = %s AND ativo = TRUE
                ORDER BY cod_empresa
            """, [tipo_filtro])
            lojas_selecionadas = [r['cod_empresa'] for r in cursor.fetchall()]
            cursor.close()

        if not lojas_selecionadas:
            conn.close()
            return jsonify({'success': False, 'erro': 'Nenhuma loja encontrada'}), 400

        # Calcular dias do periodo
        data_inicio = dt.strptime(periodo_inicio, '%Y-%m-%d')
        data_fim = dt.strptime(periodo_fim, '%Y-%m-%d')
        dias_periodo = (data_fim - data_inicio).days + 1

        print(f"  Lojas: {lojas_selecionadas}")
        print(f"  Dias do periodo: {dias_periodo}")

        # =================================================================
        # PRE-CALCULAR PROPORCOES DE VENDAS POR LOJA (RATEIO PROPORCIONAL)
        # =================================================================
        # Buscar todos os produtos para calcular proporcoes
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        query_todos_produtos = """
            SELECT DISTINCT p.cod_produto as codigo
            FROM cadastro_produtos_completo p
            WHERE p.ativo = TRUE
        """
        params_produtos = []

        if fornecedor_filtro != 'TODOS':
            if isinstance(fornecedor_filtro, list):
                placeholders = ','.join(['%s'] * len(fornecedor_filtro))
                query_todos_produtos += f" AND p.nome_fornecedor IN ({placeholders})"
                params_produtos.extend(fornecedor_filtro)
            else:
                query_todos_produtos += " AND p.nome_fornecedor = %s"
                params_produtos.append(fornecedor_filtro)

        cursor.execute(query_todos_produtos, params_produtos if params_produtos else None)
        todos_produtos = [str(r['codigo']) for r in cursor.fetchall()]
        cursor.close()

        # Calcular proporcoes de vendas para cada produto/loja
        proporcoes_vendas = {}
        if len(lojas_selecionadas) > 1 and todos_produtos:
            print(f"  Calculando proporcoes de vendas para {len(todos_produtos)} produtos...")
            proporcoes_vendas = calcular_proporcoes_vendas_por_loja(
                conn,
                todos_produtos,
                lojas_selecionadas,
                dias_historico=365
            )
            print(f"  Proporcoes calculadas: {len(proporcoes_vendas)} combinacoes produto/loja")

        num_lojas = len(lojas_selecionadas)

        # Processar por loja
        processador = PedidoFornecedorIntegrado(conn)
        resultados = []
        itens_com_demanda_validada = 0

        for cod_loja in lojas_selecionadas:
            # Buscar demanda validada se habilitado
            demanda_validada = {}
            if usar_demanda_validada:
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables
                        WHERE table_name = 'demanda_validada'
                    )
                """)
                if cursor.fetchone()['exists']:
                    cursor.execute("""
                        SELECT cod_produto, demanda_diaria, demanda_total_periodo
                        FROM demanda_validada
                        WHERE cod_loja = %s
                          AND ativo = TRUE
                          AND status = 'validado'
                          AND data_inicio <= %s
                          AND data_fim >= %s
                    """, [cod_loja, periodo_inicio, periodo_fim])

                    for r in cursor.fetchall():
                        demanda_validada[str(r['cod_produto'])] = {
                            'demanda_diaria': to_float(r['demanda_diaria']),
                            'demanda_total_periodo': to_float(r['demanda_total_periodo'])
                        }
                cursor.close()

            # Buscar produtos do fornecedor
            query_produtos = """
                SELECT DISTINCT
                    p.cod_produto as codigo,
                    p.descricao,
                    p.categoria as linha1,
                    p.codigo_linha as linha3,
                    p.cnpj_fornecedor,
                    p.nome_fornecedor,
                    COALESCE(pf.lead_time_dias, 15) as lead_time_dias,
                    COALESCE(pf.ciclo_pedido_dias, 7) as ciclo_pedido_dias,
                    COALESCE(pf.pedido_minimo_valor, 0) as pedido_minimo_valor
                FROM cadastro_produtos_completo p
                LEFT JOIN parametros_fornecedor pf ON p.cnpj_fornecedor = pf.cnpj_fornecedor
                    AND pf.cod_empresa = %s AND pf.ativo = TRUE
                WHERE p.ativo = TRUE
            """
            params = [cod_loja]

            if fornecedor_filtro != 'TODOS':
                if isinstance(fornecedor_filtro, list):
                    placeholders = ','.join(['%s'] * len(fornecedor_filtro))
                    query_produtos += f" AND p.nome_fornecedor IN ({placeholders})"
                    params.extend(fornecedor_filtro)
                else:
                    query_produtos += " AND p.nome_fornecedor = %s"
                    params.append(fornecedor_filtro)

            if linha1_filtro != 'TODAS':
                is_lista = isinstance(linha1_filtro, (list, tuple))
                if is_lista and len(linha1_filtro) > 1:
                    placeholders = ','.join(['%s'] * len(linha1_filtro))
                    query_produtos += f" AND p.categoria IN ({placeholders})"
                    params.extend([str(v) for v in linha1_filtro])
                else:
                    query_produtos += " AND p.categoria = %s"
                    valor = linha1_filtro[0] if is_lista else linha1_filtro
                    params.append(str(valor) if not isinstance(valor, str) else valor)

            if linha3_filtro != 'TODAS':
                is_lista = isinstance(linha3_filtro, (list, tuple))
                if is_lista and len(linha3_filtro) > 1:
                    placeholders = ','.join(['%s'] * len(linha3_filtro))
                    query_produtos += f" AND p.codigo_linha IN ({placeholders})"
                    params.extend([str(v) for v in linha3_filtro])
                else:
                    query_produtos += " AND p.codigo_linha = %s"
                    valor = linha3_filtro[0] if is_lista else linha3_filtro
                    params.append(str(valor) if not isinstance(valor, str) else valor)

            df_produtos = pd.read_sql(query_produtos, conn, params=params)

            # Buscar nome da loja
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("SELECT nome_loja FROM cadastro_lojas WHERE cod_empresa = %s", [cod_loja])
            loja_info = cursor.fetchone()
            nome_loja = loja_info['nome_loja'] if loja_info else f'Loja {cod_loja}'
            cursor.close()

            for _, row in df_produtos.iterrows():
                codigo = str(row['codigo'])
                lead_time = to_int(row.get('lead_time_dias', 15), default=15)

                # Verificar se tem demanda validada
                origem_demanda = 'previsao'
                demanda_diaria = 0
                demanda_periodo = 0
                proporcao_loja = 1.0  # Default: 100% se loja unica

                if codigo in demanda_validada:
                    demanda_diaria = to_float(demanda_validada[codigo]['demanda_diaria'])
                    demanda_periodo = to_float(demanda_validada[codigo]['demanda_total_periodo'])
                    origem_demanda = 'validada'
                    itens_com_demanda_validada += 1

                # FALLBACK: Usar demanda pre-calculada se nao houver validada
                if demanda_diaria <= 0:
                    cursor = conn.cursor(cursor_factory=RealDictCursor)
                    cursor.execute("""
                        SELECT demanda_diaria_base
                        FROM demanda_pre_calculada
                        WHERE cod_produto = %s
                          AND cnpj_fornecedor = %s
                        ORDER BY data_calculo DESC
                        LIMIT 1
                    """, [codigo, row.get('cnpj_fornecedor', '')])
                    demanda_pre = cursor.fetchone()
                    cursor.close()

                    if demanda_pre:
                        demanda_diaria_consolidada = to_float(demanda_pre['demanda_diaria_base'])

                        # =================================================================
                        # APLICAR RATEIO PROPORCIONAL: Distribuir demanda por loja
                        # =================================================================
                        # Se ha multiplas lojas, aplicar proporcao baseada no historico de vendas
                        if num_lojas > 1:
                            # Obter proporcao da loja (fallback: divisao uniforme)
                            proporcao_loja = proporcoes_vendas.get((codigo, cod_loja), 1.0 / num_lojas)
                            demanda_diaria = demanda_diaria_consolidada * proporcao_loja
                            origem_demanda = f'pre_calculada_rateio_{proporcao_loja:.2%}'
                        else:
                            # Loja unica: usa demanda completa
                            demanda_diaria = demanda_diaria_consolidada
                            origem_demanda = 'pre_calculada'

                        demanda_periodo = demanda_diaria * dias_periodo

                # FALLBACK 2: Usar historico de vendas como ultimo recurso
                if demanda_diaria <= 0:
                    cursor = conn.cursor(cursor_factory=RealDictCursor)
                    cursor.execute("""
                        SELECT COALESCE(SUM(qtd_venda), 0) / NULLIF(COUNT(DISTINCT data), 0) as media_diaria
                        FROM historico_vendas_diario
                        WHERE codigo = %s
                          AND cod_empresa = %s
                          AND data >= CURRENT_DATE - INTERVAL '90 days'
                    """, [int(codigo), cod_loja])
                    hist = cursor.fetchone()
                    cursor.close()

                    if hist and hist['media_diaria']:
                        demanda_diaria = to_float(hist['media_diaria'])
                        demanda_periodo = demanda_diaria * dias_periodo
                        origem_demanda = 'historico'

                if demanda_diaria <= 0 and demanda_periodo <= 0:
                    continue

                # Recalcular demanda_periodo se necessario
                if demanda_periodo <= 0:
                    demanda_periodo = demanda_diaria * dias_periodo

                # Calcular data de emissao sugerida (lead_time dias antes do inicio)
                data_emissao = data_inicio - timedelta(days=lead_time)

                # Buscar estoque atual e CUE
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                cursor.execute("""
                    SELECT
                        COALESCE(estoque, 0) as estoque_disponivel,
                        COALESCE(qtd_pendente, 0) + COALESCE(qtd_pend_transf, 0) as estoque_transito,
                        COALESCE(cue, 0) as cue
                    FROM estoque_posicao_atual
                    WHERE codigo = %s AND cod_empresa = %s
                """, [int(codigo), cod_loja])
                estoque_info = cursor.fetchone()
                cursor.close()

                # Converter todos os valores para float/int para evitar erros com Decimal
                estoque_atual = to_float(estoque_info['estoque_disponivel']) if estoque_info else 0.0
                estoque_transito = to_float(estoque_info['estoque_transito']) if estoque_info else 0.0
                cue = to_float(estoque_info['cue']) if estoque_info else 0.0

                # Calcular quantidade do pedido
                quantidade_pedido = to_int(demanda_periodo)

                # Calcular valor
                valor_pedido = quantidade_pedido * cue

                # Calcular cobertura (demanda_diaria ja e float apos to_float)
                cobertura_atual = (estoque_atual / demanda_diaria) if demanda_diaria > 0 else 999.0
                cobertura_com_transito = ((estoque_atual + estoque_transito) / demanda_diaria) if demanda_diaria > 0 else 999.0

                resultados.append({
                    'codigo': codigo,
                    'descricao': row['descricao'],
                    'cod_loja': cod_loja,
                    'nome_loja': nome_loja,
                    'cnpj_fornecedor': row['cnpj_fornecedor'],
                    'codigo_fornecedor': row.get('cnpj_fornecedor', ''),
                    'nome_fornecedor': row['nome_fornecedor'],
                    'demanda_diaria': round(demanda_diaria, 2),
                    'demanda_periodo': round(demanda_periodo, 0),
                    'quantidade_pedido': quantidade_pedido,
                    'numero_caixas': quantidade_pedido,  # Alias para agregar_por_fornecedor
                    'cue': cue,
                    'valor_pedido': round(valor_pedido, 2),
                    'estoque_atual': to_int(estoque_atual),
                    'estoque_transito': to_int(estoque_transito),
                    'cobertura_atual_dias': round(cobertura_atual, 1),
                    'cobertura_com_transito_dias': round(cobertura_com_transito, 1),
                    'lead_time_dias': lead_time,
                    'data_emissao_sugerida': data_emissao.strftime('%Y-%m-%d'),
                    'periodo_inicio': periodo_inicio,
                    'periodo_fim': periodo_fim,
                    'origem_demanda': origem_demanda,
                    'proporcao_loja': round(proporcao_loja, 4),  # Proporcao usada no rateio
                    'deve_pedir': quantidade_pedido > 0,
                    'bloqueado': False
                })

        conn.close()

        if not resultados:
            return jsonify({
                'success': False,
                'erro': 'Nenhum item com demanda para o periodo'
            }), 400

        # Agregar por fornecedor
        agregacao = agregar_por_fornecedor(resultados)

        # Estatisticas
        estatisticas = {
            'total_itens': len(resultados),
            'valor_total': sum(r.get('valor_pedido', 0) for r in resultados),
            'lojas_processadas': len(lojas_selecionadas),
            'fornecedores_unicos': len(set(r.get('nome_fornecedor') for r in resultados)),
            'itens_com_demanda_validada': itens_com_demanda_validada
        }

        print(f"\n  Resultado: {estatisticas['total_itens']} itens, R$ {estatisticas['valor_total']:,.2f}")

        return jsonify({
            'success': True,
            'itens_pedido': resultados,
            'estatisticas': estatisticas,
            'agregado_fornecedor': agregacao
        })

    except Exception as e:
        import traceback
        print(f"[ERRO] api_pedido_planejado: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@pedido_planejado_bp.route('/api/pedido_planejado/exportar', methods=['POST'])
def api_pedido_planejado_exportar():
    """Exporta o pedido planejado para Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        dados = request.get_json()

        if not dados:
            return jsonify({'success': False, 'erro': 'Dados nao fornecidos'}), 400

        itens_pedido = dados.get('itens_pedido', [])
        estatisticas = dados.get('estatisticas', {})

        if not itens_pedido:
            return jsonify({'success': False, 'erro': 'Nenhum item para exportar'}), 400

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aba Resumo
        ws_resumo = wb.active
        ws_resumo.title = 'Resumo'

        ws_resumo['A1'] = 'PEDIDO PLANEJADO - RESUMO'
        ws_resumo['A1'].font = Font(bold=True, size=14)
        ws_resumo.merge_cells('A1:D1')

        ws_resumo['A2'] = f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}'

        if itens_pedido:
            periodo_inicio = itens_pedido[0].get('periodo_inicio', '')
            periodo_fim = itens_pedido[0].get('periodo_fim', '')
            ws_resumo['A3'] = f'Periodo: {periodo_inicio} ate {periodo_fim}'

        ws_resumo['A5'] = 'Estatisticas'
        ws_resumo['A5'].font = Font(bold=True)

        stats_data = [
            ('Total de Itens', estatisticas.get('total_itens', len(itens_pedido))),
            ('Valor Total do Pedido', f"R$ {estatisticas.get('valor_total', sum(i.get('valor_pedido', 0) for i in itens_pedido)):,.2f}"),
            ('Lojas Processadas', estatisticas.get('lojas_processadas', len(set(i.get('cod_loja') for i in itens_pedido)))),
            ('Fornecedores', estatisticas.get('fornecedores_unicos', len(set(i.get('nome_fornecedor') for i in itens_pedido)))),
            ('Itens com Demanda Validada', estatisticas.get('itens_com_demanda_validada', 0)),
        ]

        for i, (label, valor) in enumerate(stats_data, start=6):
            ws_resumo[f'A{i}'] = label
            ws_resumo[f'B{i}'] = valor

        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20

        # Aba Itens
        ws_itens = wb.create_sheet('Itens Pedido')

        headers = [
            'Cod Filial', 'Nome Filial', 'Cod Item', 'Descricao Item',
            'CNPJ Fornecedor', 'Nome Fornecedor', 'Demanda Diaria', 'Demanda Periodo',
            'Qtd Pedido', 'CUE', 'Valor Pedido', 'Estoque Atual', 'Estoque Transito',
            'Cobertura Atual', 'Data Emissao', 'Origem Demanda'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_itens.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        itens_ordenados = sorted(itens_pedido, key=lambda x: (
            x.get('cod_loja', 0) or 0,
            x.get('cnpj_fornecedor', '') or ''
        ))

        for row_idx, item in enumerate(itens_ordenados, start=2):
            ws_itens.cell(row=row_idx, column=1, value=item.get('cod_loja', '')).border = border
            ws_itens.cell(row=row_idx, column=2, value=item.get('nome_loja', '')).border = border
            ws_itens.cell(row=row_idx, column=3, value=item.get('codigo', '')).border = border
            ws_itens.cell(row=row_idx, column=4, value=item.get('descricao', '')[:60]).border = border
            ws_itens.cell(row=row_idx, column=5, value=item.get('cnpj_fornecedor', '')).border = border
            ws_itens.cell(row=row_idx, column=6, value=item.get('nome_fornecedor', '')).border = border
            ws_itens.cell(row=row_idx, column=7, value=item.get('demanda_diaria', 0)).border = border
            ws_itens.cell(row=row_idx, column=7).number_format = '#,##0.00'
            ws_itens.cell(row=row_idx, column=8, value=item.get('demanda_periodo', 0)).border = border
            ws_itens.cell(row=row_idx, column=8).number_format = '#,##0'
            ws_itens.cell(row=row_idx, column=9, value=item.get('quantidade_pedido', 0)).border = border
            ws_itens.cell(row=row_idx, column=9).number_format = '#,##0'
            ws_itens.cell(row=row_idx, column=10, value=item.get('cue', 0)).border = border
            ws_itens.cell(row=row_idx, column=10).number_format = '#,##0.00'
            ws_itens.cell(row=row_idx, column=11, value=item.get('valor_pedido', 0)).border = border
            ws_itens.cell(row=row_idx, column=11).number_format = '#,##0.00'
            ws_itens.cell(row=row_idx, column=12, value=item.get('estoque_atual', 0)).border = border
            ws_itens.cell(row=row_idx, column=13, value=item.get('estoque_transito', 0)).border = border
            ws_itens.cell(row=row_idx, column=14, value=item.get('cobertura_atual_dias', 0)).border = border
            ws_itens.cell(row=row_idx, column=14).number_format = '#,##0.0'
            ws_itens.cell(row=row_idx, column=15, value=item.get('data_emissao_sugerida', '')).border = border
            ws_itens.cell(row=row_idx, column=16, value=item.get('origem_demanda', '')).border = border

        col_widths = [10, 20, 10, 40, 18, 25, 12, 12, 10, 10, 12, 10, 10, 10, 12, 12]
        for i, width in enumerate(col_widths, start=1):
            ws_itens.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

        ws_itens.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pedido_planejado_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(f"[ERRO] api_pedido_planejado_exportar: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500
