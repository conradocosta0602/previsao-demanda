"""
Blueprint: Compra Planejada (Forward Buying)
Rotas: /compra_planejada, /api/compra_planejada/calcular, /api/compra_planejada/exportar

Permite calcular compras futuras para negociacao com fornecedores.
- 1a data de entrega: calculo completo (estoque, transferencias, ES)
- 2a data em diante: demanda bruta (sem estoque, pois e impreciso a meses)
"""

import io
import math
from datetime import datetime, date, timedelta
from calendar import monthrange
import pandas as pd
import numpy as np
from flask import Blueprint, render_template, request, jsonify, send_file

from app.utils.db_connection import get_db_connection
from app.utils.demanda_pre_calculada import (
    precarregar_demanda_em_lote,
    obter_demanda_do_cache,
    calcular_proporcoes_vendas_por_loja
)

compra_planejada_bp = Blueprint('compra_planejada', __name__)


# =====================================================================
# PAGINA
# =====================================================================

@compra_planejada_bp.route('/compra_planejada')
def compra_planejada_page():
    """Pagina de compra planejada (forward buying)."""
    return render_template('compra_planejada.html')


# =====================================================================
# API: LEAD TIME DO FORNECEDOR (para data padrao dinamica)
# =====================================================================

@compra_planejada_bp.route('/api/compra_planejada/lead_time', methods=['GET'])
def api_compra_planejada_lead_time():
    """
    Retorna o lead time maximo entre os fornecedores selecionados.
    Usado para calcular a data padrao de entrega na interface.

    Query params:
        fornecedores: nomes dos fornecedores separados por virgula
    """
    try:
        fornecedores_param = request.args.get('fornecedores', '')

        if not fornecedores_param:
            return jsonify({'lead_time_dias': 15})

        nomes = [f.strip() for f in fornecedores_param.split(',') if f.strip()]
        if not nomes:
            return jsonify({'lead_time_dias': 15})

        conn = get_db_connection()

        placeholders = ','.join(['%s'] * len(nomes))
        query = f"""
            SELECT DISTINCT p.cnpj_fornecedor, p.nome_fornecedor,
                   COALESCE(pf.lead_time_dias, 15) as lead_time_dias
            FROM cadastro_produtos_completo p
            LEFT JOIN parametros_fornecedor pf ON p.cnpj_fornecedor = pf.cnpj_fornecedor
                AND pf.ativo = TRUE
            WHERE p.nome_fornecedor IN ({placeholders})
              AND p.ativo = TRUE
            ORDER BY lead_time_dias DESC
        """
        df = pd.read_sql(query, conn, params=nomes)
        conn.close()

        lead_time_max = int(df['lead_time_dias'].max()) if not df.empty else 15

        return jsonify({
            'lead_time_dias': lead_time_max,
            'data_minima_entrega': (
                date.today() + timedelta(days=lead_time_max)
            ).isoformat()
        })

    except Exception as e:
        return jsonify({'lead_time_dias': 15, 'erro': str(e)})


# =====================================================================
# API: CALCULAR COMPRA PLANEJADA
# =====================================================================

@compra_planejada_bp.route('/api/compra_planejada/calcular', methods=['POST'])
def api_compra_planejada_calcular():
    """
    Calcula compra planejada para periodos futuros.

    Request JSON:
    {
        "fornecedor": ["NOME_FORNECEDOR"] ou "NOME",
        "cod_empresa": [1, 2, 3] ou "TODAS",
        "destino_tipo": "LOJA" ou "CD",
        "datas_entrega": ["2026-04-15", "2026-05-15", "2026-06-15"],
        "linha1": "TODAS" ou filtro,
        "linha3": "TODAS" ou filtro
    }

    Fase 1 (1a data): calculo completo via processar_item (estoque, ES, transferencias)
    Fase 2+ (demais): demanda bruta × dias do periodo, arredondada para multiplo de caixa
    """
    try:
        dados = request.get_json()

        # ---- Parametros ----
        fornecedor_filtro = dados.get('fornecedor', 'TODOS')
        linha1_filtro = dados.get('linha1', 'TODAS')
        linha3_filtro = dados.get('linha3', 'TODAS')
        destino_tipo = dados.get('destino_tipo', 'LOJA')
        cod_empresa = dados.get('cod_empresa', 'TODAS')
        datas_entrega_raw = dados.get('datas_entrega', [])
        cobertura_dias = dados.get('cobertura_dias', None)  # None = ABC automatica

        # Validar datas de entrega
        if not datas_entrega_raw or len(datas_entrega_raw) == 0:
            return jsonify({'success': False, 'erro': 'Informe pelo menos uma data de entrega.'}), 400

        # Parsear e ordenar datas
        datas_entrega = []
        for d in datas_entrega_raw:
            try:
                dt = datetime.strptime(d, '%Y-%m-%d').date()
                datas_entrega.append(dt)
            except ValueError:
                return jsonify({'success': False, 'erro': f'Data invalida: {d}. Use formato YYYY-MM-DD.'}), 400

        datas_entrega.sort()

        # Validar que datas sao futuras
        hoje = date.today()
        if datas_entrega[0] <= hoje:
            return jsonify({'success': False, 'erro': 'Todas as datas de entrega devem ser futuras.'}), 400

        # Normalizar filtros
        def normalizar_filtro(filtro, padrao='TODAS'):
            if filtro is None:
                return padrao
            if isinstance(filtro, list):
                if len(filtro) == 0:
                    return padrao
                if len(filtro) == 1:
                    return filtro[0]
                return filtro
            return filtro

        fornecedor_filtro = normalizar_filtro(fornecedor_filtro, 'TODOS')
        linha1_filtro = normalizar_filtro(linha1_filtro)
        linha3_filtro = normalizar_filtro(linha3_filtro)

        print(f"\n{'='*60}")
        print("COMPRA PLANEJADA (Forward Buying)")
        print(f"{'='*60}")
        print(f"  Fornecedor: {fornecedor_filtro}")
        print(f"  Destino: {destino_tipo}")
        print(f"  Datas de entrega: {[d.isoformat() for d in datas_entrega]}")
        print(f"  Cobertura: {cobertura_dias if cobertura_dias else 'ABC (automatica)'}")

        conn = get_db_connection()

        # ---- Normalizar lojas ----
        lojas_selecionadas = []
        if isinstance(cod_empresa, list):
            lojas_selecionadas = [int(x) for x in cod_empresa]
        elif cod_empresa != 'TODAS':
            lojas_selecionadas = [int(cod_empresa)]

        if lojas_selecionadas:
            is_destino_cd = destino_tipo == 'CD' or any(l >= 80 for l in lojas_selecionadas)
            cod_destino = lojas_selecionadas[0]
        else:
            is_destino_cd = destino_tipo == 'CD'
            cod_destino = 80

        # ---- Buscar fornecedores ----
        fornecedores_a_processar = []
        if isinstance(fornecedor_filtro, list):
            if 'TODOS' in fornecedor_filtro:
                fornecedor_filtro = 'TODOS'
            else:
                fornecedores_a_processar = fornecedor_filtro

        if not fornecedores_a_processar:
            if fornecedor_filtro == 'TODOS':
                query = """
                    SELECT DISTINCT nome_fornecedor
                    FROM cadastro_produtos_completo
                    WHERE ativo = TRUE AND nome_fornecedor IS NOT NULL AND TRIM(nome_fornecedor) != ''
                """
                params = []
                if linha1_filtro != 'TODAS':
                    if isinstance(linha1_filtro, list):
                        ph = ','.join(['%s'] * len(linha1_filtro))
                        query += f" AND categoria IN ({ph})"
                        params.extend(linha1_filtro)
                    else:
                        query += " AND categoria = %s"
                        params.append(linha1_filtro)
                if linha3_filtro != 'TODAS':
                    if isinstance(linha3_filtro, list):
                        ph = ','.join(['%s'] * len(linha3_filtro))
                        query += f" AND codigo_linha IN ({ph})"
                        params.extend(linha3_filtro)
                    else:
                        query += " AND codigo_linha = %s"
                        params.append(linha3_filtro)
                query += " ORDER BY nome_fornecedor"
                df_forn = pd.read_sql(query, conn, params=params if params else None)
                fornecedores_a_processar = df_forn['nome_fornecedor'].tolist()
            else:
                fornecedores_a_processar = [fornecedor_filtro]

        if not fornecedores_a_processar:
            conn.close()
            return jsonify({'success': False, 'erro': 'Nenhum fornecedor encontrado.'}), 400

        # ---- Buscar produtos ----
        df_produtos_total = pd.DataFrame()
        for nome_fornecedor in fornecedores_a_processar:
            query_p = """
                SELECT DISTINCT
                    p.cod_produto as codigo,
                    p.descricao,
                    p.categoria as linha1,
                    p.codigo_linha as linha3,
                    p.descricao_linha as linha3_descricao,
                    'B' as curva_abc,
                    p.cnpj_fornecedor as codigo_fornecedor,
                    p.nome_fornecedor,
                    COALESCE(pf.lead_time_dias, 15) as lead_time_dias,
                    COALESCE(pf.ciclo_pedido_dias, 7) as ciclo_pedido_dias,
                    COALESCE(pf.pedido_minimo_valor, 0) as pedido_minimo_valor,
                    CASE WHEN pf.id IS NOT NULL THEN TRUE ELSE FALSE END as fornecedor_cadastrado
                FROM cadastro_produtos_completo p
                LEFT JOIN parametros_fornecedor pf ON p.cnpj_fornecedor = pf.cnpj_fornecedor
                    AND pf.cod_empresa = %s AND pf.ativo = TRUE
                WHERE p.ativo = TRUE AND p.nome_fornecedor = %s
            """
            params_p = [cod_destino, nome_fornecedor]

            if linha1_filtro != 'TODAS':
                if isinstance(linha1_filtro, list):
                    ph = ','.join(['%s'] * len(linha1_filtro))
                    query_p += f" AND p.categoria IN ({ph})"
                    params_p.extend(linha1_filtro)
                else:
                    query_p += " AND p.categoria = %s"
                    params_p.append(linha1_filtro)

            if linha3_filtro != 'TODAS':
                if isinstance(linha3_filtro, list):
                    ph = ','.join(['%s'] * len(linha3_filtro))
                    query_p += f" AND p.codigo_linha IN ({ph})"
                    params_p.extend(linha3_filtro)
                else:
                    query_p += " AND p.codigo_linha = %s"
                    params_p.append(linha3_filtro)

            query_p += " ORDER BY p.cod_produto"
            df_f = pd.read_sql(query_p, conn, params=params_p)
            if not df_f.empty:
                df_produtos_total = pd.concat([df_produtos_total, df_f], ignore_index=True)

        df_produtos = df_produtos_total.drop_duplicates(subset=['codigo'], keep='first')

        if df_produtos.empty:
            conn.close()
            return jsonify({'success': False, 'erro': 'Nenhum item encontrado com os filtros.'}), 400

        # ---- Configurar lojas ----
        mapa_nomes_lojas = {}
        if is_destino_cd:
            df_lojas = pd.read_sql(
                "SELECT cod_empresa, nome_loja FROM cadastro_lojas WHERE ativo = TRUE AND cod_empresa < 80", conn
            )
            lojas_demanda = df_lojas['cod_empresa'].tolist()
            mapa_nomes_lojas = dict(zip(df_lojas['cod_empresa'], df_lojas['nome_loja']))
            is_pedido_multiloja = len(lojas_demanda) > 1
        else:
            if lojas_selecionadas:
                lojas_demanda = lojas_selecionadas
                is_pedido_multiloja = len(lojas_selecionadas) > 1
                if is_pedido_multiloja:
                    ph = ','.join(['%s'] * len(lojas_selecionadas))
                    df_nl = pd.read_sql(
                        f"SELECT cod_empresa, nome_loja FROM cadastro_lojas WHERE cod_empresa IN ({ph})",
                        conn, params=lojas_selecionadas
                    )
                    mapa_nomes_lojas = dict(zip(df_nl['cod_empresa'], df_nl['nome_loja']))
            else:
                df_tl = pd.read_sql(
                    "SELECT cod_empresa, nome_loja FROM cadastro_lojas WHERE ativo = TRUE AND cod_empresa < 80", conn
                )
                lojas_demanda = df_tl['cod_empresa'].tolist()
                mapa_nomes_lojas = dict(zip(df_tl['cod_empresa'], df_tl['nome_loja']))
                is_pedido_multiloja = len(lojas_demanda) > 1

        codigos_produtos = df_produtos['codigo'].astype(int).tolist()

        # ---- Calcular lead time e transit time para cobertura ----
        lead_times_list = df_produtos['lead_time_dias'].astype(int).tolist()
        lt_max_forn = max(lead_times_list) if lead_times_list else 15
        # Transit time CD (se destino = CD)
        dias_transferencia_cd_pre = 0
        if is_destino_cd:
            from core.padrao_compra import get_dias_transferencia
            dias_transferencia_cd_pre = get_dias_transferencia(conn)
        cobertura_ultima_fase = lt_max_forn + dias_transferencia_cd_pre + 5 + 7 + 4

        # ---- Calcular periodos entre datas de entrega ----
        # Periodos definem o intervalo entre entregas (para calculo de demanda sazonal)
        # A cobertura-alvo (OUL) e aplicada separadamente no calculo do pedido
        periodos = []
        for i, dt in enumerate(datas_entrega):
            if i < len(datas_entrega) - 1:
                dt_fim = datas_entrega[i + 1] - timedelta(days=1)
            else:
                # Ultima fase: usar 30 dias para calculo de demanda sazonal
                dt_fim = dt + timedelta(days=29)
            dias = (dt_fim - dt).days + 1
            periodos.append({
                'fase': i + 1,
                'data_entrega': dt,
                'data_inicio_cobertura': dt,
                'data_fim_cobertura': dt_fim,
                'dias': dias
            })

        print(f"  Periodos calculados:")
        for p in periodos:
            print(f"    Fase {p['fase']}: {p['data_entrega']} -> {p['data_fim_cobertura']} ({p['dias']} dias)")

        # ---- Verificar demanda pre-calculada ----
        # Buscar demanda para todos os meses cobertos
        meses_necessarios = set()
        for p in periodos:
            dt_atual = p['data_inicio_cobertura']
            while dt_atual <= p['data_fim_cobertura']:
                meses_necessarios.add((dt_atual.year, dt_atual.month))
                # Pular para proximo mes
                if dt_atual.month == 12:
                    dt_atual = date(dt_atual.year + 1, 1, 1)
                else:
                    dt_atual = date(dt_atual.year, dt_atual.month + 1, 1)

        print(f"  Meses necessarios: {sorted(meses_necessarios)}")

        # Pre-carregar demanda para cada mes necessario
        cache_demanda_por_mes = {}
        cnpjs_unicos = df_produtos['codigo_fornecedor'].unique()

        for ano_m, mes_m in sorted(meses_necessarios):
            cache_mes = {}
            for cnpj_forn in cnpjs_unicos:
                if cnpj_forn:
                    produtos_forn = df_produtos[df_produtos['codigo_fornecedor'] == cnpj_forn]['codigo'].tolist()
                    cache = precarregar_demanda_em_lote(
                        conn, produtos_forn, cnpj_forn,
                        cod_empresas=lojas_demanda if is_pedido_multiloja else None,
                        ano=ano_m, mes=mes_m
                    )
                    cache_mes.update(cache)
            cache_demanda_por_mes[(ano_m, mes_m)] = cache_mes

        # Verificar se ha demanda
        tem_demanda = any(len(c) > 0 for c in cache_demanda_por_mes.values())
        if not tem_demanda:
            conn.close()
            return jsonify({
                'success': False,
                'erro': 'Fornecedor sem demanda calculada. Execute o calculo de demanda na Tela de Demanda primeiro.'
            }), 400

        # Pre-calcular proporcoes
        cache_proporcoes = {}
        if is_pedido_multiloja:
            cache_proporcoes = calcular_proporcoes_vendas_por_loja(
                conn, codigos_produtos, lojas_demanda, dias_historico=365
            )

        # ================================================================
        # FASE 1: Chama internamente a API do Pedido ao Fornecedor
        # Garante resultado 100% identico (V25, V29, V30, V31, V32, V34)
        # ================================================================
        from flask import current_app

        print(f"  [FASE 1] Chamando API do pedido ao fornecedor (resultado identico)...")

        # Montar payload identico ao que a tela de pedido envia
        payload_pedido = {
            'fornecedor': fornecedor_filtro,
            'linha1': linha1_filtro,
            'linha3': linha3_filtro,
            'destino_tipo': destino_tipo,
            'cod_empresa': cod_empresa,
            'cobertura_dias': cobertura_dias,
        }

        # Chamar a API internamente via test_client
        with current_app.test_client() as client:
            resp = client.post(
                '/api/pedido_fornecedor_integrado',
                json=payload_pedido
            )

        resp_data = resp.get_json()

        if not resp_data or not resp_data.get('success'):
            conn.close()
            erro_msg = resp_data.get('erro', 'Erro ao calcular pedido base') if resp_data else 'Sem resposta da API'
            return jsonify({'success': False, 'erro': f'Fase 1: {erro_msg}'}), 400

        # Extrair resultados do pedido normal
        itens_pedido_normal = resp_data.get('itens_pedido', [])
        itens_bloqueados_normal = resp_data.get('itens_bloqueados', [])
        itens_ok_normal = resp_data.get('itens_ok', [])
        estatisticas_normal = resp_data.get('estatisticas', {})

        print(f"  [FASE 1] API retornou: {len(itens_pedido_normal)} itens pedido, {len(itens_bloqueados_normal)} bloqueados, {len(itens_ok_normal)} OK")

        # Consolidar itens por codigo (para multiloja, agrupar resultados de diferentes lojas)
        itens_fase1 = []
        itens_bloqueados = set()
        embalagens_cache = {}
        cue_cache = {}
        is_pedido_multiloja_resp = resp_data.get('is_pedido_multiloja', False)

        # Codigos de itens bloqueados
        codigos_bloqueados_set = set()
        for item in itens_bloqueados_normal:
            codigos_bloqueados_set.add(int(item.get('codigo', 0)))

        # Itens que nao tem pedido em nenhuma loja e estao bloqueados em todas
        # (apenas itens que so aparecem na lista de bloqueados)
        codigos_com_pedido = set(int(i.get('codigo', 0)) for i in itens_pedido_normal)
        codigos_ok = set(int(i.get('codigo', 0)) for i in itens_ok_normal)
        itens_bloqueados = codigos_bloqueados_set - codigos_com_pedido - codigos_ok

        if is_pedido_multiloja_resp:
            # Agrupar por codigo de produto
            resultados_por_item = {}
            for item in itens_pedido_normal:
                cod = int(item.get('codigo', 0))
                if cod not in resultados_por_item:
                    resultados_por_item[cod] = []
                resultados_por_item[cod].append(item)

            for codigo, resultados_lojas in resultados_por_item.items():
                if not resultados_lojas:
                    continue
                qtd_total = sum(float(r.get('quantidade_pedido', 0) or 0) for r in resultados_lojas)
                demanda_total = sum(float(r.get('demanda_prevista', 0) or 0) for r in resultados_lojas)
                estoque_total = sum(float(r.get('estoque_efetivo', r.get('estoque_atual', 0)) or 0) for r in resultados_lojas)
                es_total = sum(float(r.get('estoque_seguranca', 0) or 0) for r in resultados_lojas)
                demanda_diaria_total = sum(float(r.get('demanda_prevista_diaria', r.get('demanda_diaria', 0)) or 0) for r in resultados_lojas)
                r0 = resultados_lojas[0]

                cue_val = float(r0.get('cue', r0.get('preco_custo', 0)) or 0)
                multiplo = int(r0.get('multiplo_caixa', 1) or 1)
                embalagens_cache[codigo] = multiplo
                cue_cache[codigo] = cue_val

                itens_fase1.append({
                    'codigo': codigo,
                    'descricao': r0.get('descricao', ''),
                    'cnpj_fornecedor': r0.get('codigo_fornecedor', r0.get('cnpj_fornecedor', '')),
                    'nome_fornecedor': r0.get('nome_fornecedor', ''),
                    'curva_abc': r0.get('curva_abc', 'B'),
                    'metodo': r0.get('metodo_previsao', ''),
                    'demanda_diaria': round(demanda_diaria_total, 2),
                    'demanda_periodo': round(demanda_total, 0),
                    'estoque_atual': round(estoque_total, 0),
                    'estoque_seguranca': round(es_total, 0),
                    'quantidade_pedido': round(qtd_total, 0),
                    'multiplo_caixa': multiplo,
                    'numero_caixas': math.ceil(qtd_total / max(multiplo, 1)),
                    'cue': round(cue_val, 2),
                    'valor_total': round(qtd_total * cue_val, 2),
                    'fator_sazonal': r0.get('fator_sazonal', 1.0),
                    'tipo_calculo': 'completo',
                })
        else:
            for item in itens_pedido_normal:
                codigo = int(item.get('codigo', 0))
                qtd = float(item.get('quantidade_pedido', 0) or 0)
                cue_val = float(item.get('cue', item.get('preco_custo', 0)) or 0)
                multiplo = int(item.get('multiplo_caixa', 1) or 1)
                embalagens_cache[codigo] = multiplo
                cue_cache[codigo] = cue_val

                itens_fase1.append({
                    'codigo': codigo,
                    'descricao': item.get('descricao', ''),
                    'cnpj_fornecedor': item.get('codigo_fornecedor', item.get('cnpj_fornecedor', '')),
                    'nome_fornecedor': item.get('nome_fornecedor', ''),
                    'curva_abc': item.get('curva_abc', 'B'),
                    'metodo': item.get('metodo_previsao', ''),
                    'demanda_diaria': round(float(item.get('demanda_prevista_diaria', item.get('demanda_diaria', 0)) or 0), 2),
                    'demanda_periodo': round(float(item.get('demanda_prevista', 0) or 0), 0),
                    'estoque_atual': round(float(item.get('estoque_efetivo', item.get('estoque_atual', 0)) or 0), 0),
                    'estoque_seguranca': round(float(item.get('estoque_seguranca', 0) or 0), 0),
                    'quantidade_pedido': round(qtd, 0),
                    'multiplo_caixa': multiplo,
                    'numero_caixas': math.ceil(qtd / max(multiplo, 1)),
                    'cue': round(cue_val, 2),
                    'valor_total': round(qtd * cue_val, 2),
                    'fator_sazonal': item.get('fator_sazonal', 1.0),
                    'tipo_calculo': 'completo',
                })

        # Tambem incluir itens OK (sem pedido mas nao bloqueados) no cache
        for item in itens_ok_normal:
            codigo = int(item.get('codigo', 0))
            cue_val = float(item.get('cue', item.get('preco_custo', 0)) or 0)
            multiplo = int(item.get('multiplo_caixa', 1) or 1)
            if codigo not in embalagens_cache:
                embalagens_cache[codigo] = multiplo
            if codigo not in cue_cache:
                cue_cache[codigo] = cue_val

        print(f"  [FASE 1] {len(itens_fase1)} itens consolidados, {len(itens_bloqueados)} bloqueados")

        # Periodo da fase 1 = cobertura selecionada ou ABC
        if cobertura_dias:
            cobertura_display = int(cobertura_dias)
        else:
            cobertura_display = cobertura_ultima_fase

        # ================================================================
        # FASES 2+: Modelo Order-Up-To-Level (OUL)
        # ================================================================
        # Cada pedido repoe estoque para manter cobertura-alvo (do seletor da tela)
        # target_estoque = demanda_diaria × cobertura_alvo
        # necessidade = max(0, target_estoque - estoque_projetado)
        target_cobertura = cobertura_display  # Mesma cobertura do Pedido 1
        print(f"  [FASES 2+] Modelo OUL: cobertura-alvo = {target_cobertura} dias")

        # Montar cache de estoque atual consolidado por item (todas as lojas)
        # IMPORTANTE: Incluir itens_ok (estoque suficiente) e itens_pedido
        estoque_consolidado = {}
        pedidos_acumulados = {}
        demanda_diaria_cache = {}

        # Itens com pedido (fase 1)
        for item in itens_fase1:
            cod = item['codigo']
            estoque_consolidado[cod] = float(item.get('estoque_atual', 0) or 0)
            pedidos_acumulados[cod] = float(item.get('quantidade_pedido', 0) or 0)
            demanda_diaria_cache[cod] = float(item.get('demanda_diaria', 0) or 0)

        # Itens OK (sem pedido mas com estoque) - CRUCIAL para fases 2+
        # Sem isso, estoque desses itens = 0 nas fases futuras, inflando pedidos
        if is_pedido_multiloja_resp:
            # Agrupar itens_ok por codigo
            ok_por_codigo = {}
            for item in itens_ok_normal:
                cod = int(item.get('codigo', 0))
                if cod not in ok_por_codigo:
                    ok_por_codigo[cod] = []
                ok_por_codigo[cod].append(item)
            for cod, itens in ok_por_codigo.items():
                if cod not in estoque_consolidado:
                    estoque_consolidado[cod] = sum(float(i.get('estoque_efetivo', i.get('estoque_atual', 0)) or 0) for i in itens)
                    pedidos_acumulados[cod] = 0
                    demanda_diaria_cache[cod] = sum(float(i.get('demanda_prevista_diaria', i.get('demanda_diaria', 0)) or 0) for i in itens)
        else:
            for item in itens_ok_normal:
                cod = int(item.get('codigo', 0))
                if cod not in estoque_consolidado:
                    estoque_consolidado[cod] = float(item.get('estoque_efetivo', item.get('estoque_atual', 0)) or 0)
                    pedidos_acumulados[cod] = 0
                    demanda_diaria_cache[cod] = float(item.get('demanda_prevista_diaria', item.get('demanda_diaria', 0)) or 0)

        print(f"  [FASES 2+] Cache: {len(estoque_consolidado)} itens com estoque ({len(itens_fase1)} pedido + {len(estoque_consolidado) - len(itens_fase1)} OK)")

        # Buscar demanda do mes da 1a entrega (fallback para fases 2+)
        mes_fase1 = (datas_entrega[0].year, datas_entrega[0].month)
        cache_demanda_fase1 = cache_demanda_por_mes.get(mes_fase1, {})
        if not cache_demanda_fase1:
            for chave_mes in sorted(cache_demanda_por_mes.keys()):
                if cache_demanda_por_mes[chave_mes]:
                    cache_demanda_fase1 = cache_demanda_por_mes[chave_mes]
                    break

        # Carregar embalagens que podem faltar no cache (itens OK que nao tem pedido)
        if not embalagens_cache:
            try:
                cursor_emb = conn.cursor()
                cursor_emb.execute("""
                    SELECT codigo, qtd_embalagem FROM embalagem_arredondamento
                    WHERE codigo = ANY(%s) AND qtd_embalagem > 0
                """, (codigos_produtos,))
                for row_emb in cursor_emb.fetchall():
                    embalagens_cache[int(row_emb[0])] = int(row_emb[1])
                cursor_emb.close()
            except Exception:
                pass

        fases_resultado = [{
            'fase': 1,
            'data_entrega': periodos[0]['data_entrega'].isoformat(),
            'data_fim_cobertura': periodos[0]['data_fim_cobertura'].isoformat(),
            'periodo_dias': cobertura_display,
            'tipo_calculo': 'completo',
            'itens': itens_fase1,
            'total_valor': sum(i.get('valor_total', 0) for i in itens_fase1),
            'total_itens': len([i for i in itens_fase1 if i.get('quantidade_pedido', 0) > 0]),
            'total_quantidade': sum(i.get('quantidade_pedido', 0) for i in itens_fase1),
        }]

        for fase_idx in range(1, len(periodos)):
            periodo = periodos[fase_idx]
            itens_fase = []

            # Dias desde hoje ate a data de entrega desta fase
            dias_ate_entrega = (periodo['data_entrega'] - hoje).days

            for idx, row in df_produtos.iterrows():
                codigo = int(row['codigo'])
                cnpj_forn = row.get('codigo_fornecedor', '')

                # Pular itens bloqueados
                if codigo in itens_bloqueados:
                    continue

                # Demanda diaria para este periodo (ponderada por mes/sazonalidade)
                demanda_diaria_periodo = _calcular_demanda_diaria_periodo(
                    cache_demanda_por_mes, cache_proporcoes, cache_demanda_fase1,
                    str(codigo), periodo, lojas_demanda, is_pedido_multiloja
                )

                if demanda_diaria_periodo <= 0:
                    continue

                # Atualizar cache de demanda diaria (pode variar por sazonalidade)
                if codigo not in demanda_diaria_cache:
                    demanda_diaria_cache[codigo] = demanda_diaria_periodo

                multiplo = embalagens_cache.get(codigo, 1)
                if multiplo <= 0:
                    multiplo = 1

                # === Modelo Order-Up-To-Level (OUL) ===
                # Estoque projetado na data de entrega desta fase
                estoque_atual = estoque_consolidado.get(codigo, 0)
                pedidos_ant = pedidos_acumulados.get(codigo, 0)
                dem_diaria_media = demanda_diaria_cache.get(codigo, demanda_diaria_periodo)
                consumo_ate_entrega = dem_diaria_media * dias_ate_entrega
                estoque_projetado = estoque_atual + pedidos_ant - consumo_ate_entrega

                # Target = estoque necessario para manter cobertura-alvo
                target_estoque = demanda_diaria_periodo * target_cobertura

                # Necessidade = quanto falta para atingir o target
                necessidade = target_estoque - max(0, estoque_projetado)

                if necessidade <= 0:
                    qtd_pedido = 0
                else:
                    qtd_pedido = math.ceil(necessidade / multiplo) * multiplo

                # Debug para primeiros itens (apenas na fase 2 para diagnostico)
                if fase_idx == 1 and len(itens_fase) < 3 and qtd_pedido > 0:
                    print(f"    [OUL F{fase_idx+1}] cod={codigo}: est_hoje={estoque_atual:.0f} + ped_ant={pedidos_ant:.0f} - consumo({dem_diaria_media:.1f}x{dias_ate_entrega}d={consumo_ate_entrega:.0f}) = est_proj={estoque_projetado:.0f} | target={target_estoque:.0f} | nec={necessidade:.0f} -> pedido={qtd_pedido}")

                cue = cue_cache.get(codigo, 0)

                # Fator sazonal
                fator_saz = 1.0
                mes_principal = (periodo['data_entrega'].year, periodo['data_entrega'].month)
                cache_mes_p = cache_demanda_por_mes.get(mes_principal, {})
                if cache_mes_p:
                    key_saz = (str(codigo), None)
                    dados_saz = cache_mes_p.get(key_saz, {})
                    if dados_saz:
                        fator_saz = dados_saz.get('fator_sazonal', 1.0) or 1.0

                # Demanda bruta para exibicao (demanda do intervalo entre entregas)
                demanda_bruta_display = demanda_diaria_periodo * periodo['dias']

                itens_fase.append({
                    'codigo': codigo,
                    'descricao': row.get('descricao', ''),
                    'cnpj_fornecedor': cnpj_forn,
                    'nome_fornecedor': row.get('nome_fornecedor', ''),
                    'curva_abc': 'B',
                    'metodo': '',
                    'demanda_diaria': round(demanda_diaria_periodo, 2),
                    'demanda_periodo': round(demanda_bruta_display, 0),
                    'estoque_atual': round(max(0, estoque_projetado), 0),
                    'estoque_seguranca': 0,
                    'quantidade_pedido': qtd_pedido,
                    'multiplo_caixa': multiplo,
                    'numero_caixas': math.ceil(qtd_pedido / multiplo) if qtd_pedido > 0 else 0,
                    'cue': round(cue, 2),
                    'valor_total': round(qtd_pedido * cue, 2),
                    'fator_sazonal': fator_saz,
                    'tipo_calculo': 'projetado',
                })

                # Acumular pedido desta fase para as proximas
                pedidos_acumulados[codigo] = pedidos_ant + qtd_pedido

            print(f"  [FASE {fase_idx + 1}] {len(itens_fase)} itens calculados (OUL target={target_cobertura}d, intervalo={periodo['dias']}d)")

            fases_resultado.append({
                'fase': fase_idx + 1,
                'data_entrega': periodo['data_entrega'].isoformat(),
                'data_fim_cobertura': periodo['data_fim_cobertura'].isoformat(),
                'periodo_dias': target_cobertura,
                'tipo_calculo': 'projetado',
                'itens': itens_fase,
                'total_valor': sum(i.get('valor_total', 0) for i in itens_fase),
                'total_itens': len([i for i in itens_fase if i.get('quantidade_pedido', 0) > 0]),
                'total_quantidade': sum(i.get('quantidade_pedido', 0) for i in itens_fase),
            })

        conn.close()

        # Consolidado
        total_valor_geral = sum(f['total_valor'] for f in fases_resultado)
        total_qtd_geral = sum(f['total_quantidade'] for f in fases_resultado)

        # Consolidado por item (todas as fases)
        consolidado_itens = {}
        for fase in fases_resultado:
            for item in fase['itens']:
                cod = item['codigo']
                if cod not in consolidado_itens:
                    consolidado_itens[cod] = {
                        'codigo': cod,
                        'descricao': item['descricao'],
                        'cnpj_fornecedor': item['cnpj_fornecedor'],
                        'nome_fornecedor': item['nome_fornecedor'],
                        'cue': item['cue'],
                        'multiplo_caixa': item['multiplo_caixa'],
                        'fases': {}
                    }
                consolidado_itens[cod]['fases'][fase['fase']] = {
                    'quantidade': item['quantidade_pedido'],
                    'valor': item['valor_total'],
                    'demanda_diaria': item['demanda_diaria'],
                }

        # Calcular totais no consolidado
        for cod, info in consolidado_itens.items():
            info['total_quantidade'] = sum(f['quantidade'] for f in info['fases'].values())
            info['total_valor'] = sum(f['valor'] for f in info['fases'].values())
            info['total_caixas'] = math.ceil(
                info['total_quantidade'] / max(info['multiplo_caixa'], 1)
            )

        return jsonify({
            'success': True,
            'fases': fases_resultado,
            'consolidado': {
                'total_valor': round(total_valor_geral, 2),
                'total_quantidade': round(total_qtd_geral, 0),
                'total_itens_unicos': len(consolidado_itens),
                'num_fases': len(fases_resultado),
                'itens': list(consolidado_itens.values()),
            },
            'is_pedido_multiloja': is_pedido_multiloja,
            'is_destino_cd': is_destino_cd,
            'fornecedores': fornecedores_a_processar,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


def _calcular_demanda_diaria_periodo(
    cache_demanda_por_mes, cache_proporcoes, cache_fallback,
    cod_produto, periodo, lojas_demanda, is_multiloja
):
    """
    Calcula demanda diaria media ponderada para um periodo que pode cruzar meses.
    Pondera pela quantidade de dias de cada mes no periodo.
    """
    dt_inicio = periodo['data_inicio_cobertura']
    dt_fim = periodo['data_fim_cobertura']

    # Decompor o periodo em segmentos por mes
    segmentos = []
    dt_atual = dt_inicio
    while dt_atual <= dt_fim:
        ultimo_dia_mes = monthrange(dt_atual.year, dt_atual.month)[1]
        fim_segmento = min(dt_fim, date(dt_atual.year, dt_atual.month, ultimo_dia_mes))
        dias_segmento = (fim_segmento - dt_atual).days + 1
        segmentos.append({
            'ano': dt_atual.year,
            'mes': dt_atual.month,
            'dias': dias_segmento
        })
        # Proximo mes
        if dt_atual.month == 12:
            dt_atual = date(dt_atual.year + 1, 1, 1)
        else:
            dt_atual = date(dt_atual.year, dt_atual.month + 1, 1)

    # Calcular demanda ponderada
    demanda_total = 0
    dias_total = 0

    for seg in segmentos:
        cache_mes = cache_demanda_por_mes.get((seg['ano'], seg['mes']), {})
        if not cache_mes:
            cache_mes = cache_fallback  # Fallback para ultimo cache disponivel

        if is_multiloja:
            demanda_diaria_seg = 0
            num_lojas = len(lojas_demanda)
            for loja in lojas_demanda:
                prop = cache_proporcoes.get((cod_produto, loja))
                dd, _, meta = obter_demanda_do_cache(
                    cache_mes, cod_produto, cod_empresa=loja,
                    num_lojas=num_lojas, proporcao_loja=prop
                )
                demanda_diaria_seg += dd
        else:
            demanda_diaria_seg, _, _ = obter_demanda_do_cache(cache_mes, cod_produto)

        demanda_total += demanda_diaria_seg * seg['dias']
        dias_total += seg['dias']

    return demanda_total / dias_total if dias_total > 0 else 0


# =====================================================================
# API: EXPORTAR EXCEL
# =====================================================================

@compra_planejada_bp.route('/api/compra_planejada/exportar', methods=['POST'])
def api_compra_planejada_exportar():
    """Exporta compra planejada para Excel com abas por fase + consolidado."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        dados = request.get_json()
        if not dados:
            return jsonify({'success': False, 'erro': 'Dados nao fornecidos'}), 400

        fases = dados.get('fases', [])
        consolidado = dados.get('consolidado', {})
        fornecedores = dados.get('fornecedores', [])

        if not fases:
            return jsonify({'success': False, 'erro': 'Nenhuma fase para exportar'}), 400

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill_azul = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_fill_verde = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_fill_laranja = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
        titulo_font = Font(bold=True, size=14)
        subtitulo_font = Font(bold=True, size=11)
        moeda_format = '#,##0.00'
        inteiro_format = '#,##0'
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Mapeamento de filiais
        NOMES_FILIAIS = {
            1: 'GUS', 2: 'IMB', 3: 'PAL', 4: 'TAM', 5: 'AJU',
            6: 'JPA', 7: 'PNG', 8: 'CAU', 9: 'BAR', 11: 'FRT',
            80: 'MDC', 81: 'OBC', 82: 'OSAL'
        }

        # Carregar codigos DIG
        codigos_dig = {}
        try:
            conn_dig = get_db_connection()
            cursor_dig = conn_dig.cursor()
            cursor_dig.execute("SELECT codigo, codigo_dig FROM codigo_dig")
            codigos_dig = {str(row[0]): str(row[1]) for row in cursor_dig.fetchall()}
            cursor_dig.close()
            conn_dig.close()
        except Exception:
            pass

        # ============== ABA RESUMO ==============
        ws_resumo = wb.active
        ws_resumo.title = 'Resumo'

        ws_resumo['A1'] = 'COMPRA PLANEJADA - RESUMO'
        ws_resumo['A1'].font = titulo_font
        ws_resumo.merge_cells('A1:D1')

        ws_resumo['A2'] = f'Data de geracao: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws_resumo['A3'] = f'Fornecedor(es): {", ".join(fornecedores[:5])}{"..." if len(fornecedores) > 5 else ""}'

        ws_resumo['A5'] = 'RESUMO POR FASE'
        ws_resumo['A5'].font = subtitulo_font

        # Cabecalho da tabela de resumo
        resumo_headers = ['Fase', 'Data Entrega', 'Periodo (dias)', 'Tipo Calculo', 'Itens', 'Quantidade', 'Valor (R$)']
        for col, h in enumerate(resumo_headers, 1):
            cell = ws_resumo.cell(row=6, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill_azul
            cell.border = border

        for i, fase in enumerate(fases, 7):
            ws_resumo.cell(row=i, column=1, value=fase.get('fase', '')).border = border
            ws_resumo.cell(row=i, column=2, value=fase.get('data_entrega', '')).border = border
            ws_resumo.cell(row=i, column=3, value=fase.get('periodo_dias', 0)).border = border
            tipo = 'Completo (estoque+ES+transf)' if fase.get('tipo_calculo') == 'completo' else 'Projetado (estoque projetado)'
            ws_resumo.cell(row=i, column=4, value=tipo).border = border
            ws_resumo.cell(row=i, column=5, value=fase.get('total_itens', 0)).border = border
            c_qtd = ws_resumo.cell(row=i, column=6, value=fase.get('total_quantidade', 0))
            c_qtd.number_format = inteiro_format
            c_qtd.border = border
            c_val = ws_resumo.cell(row=i, column=7, value=fase.get('total_valor', 0))
            c_val.number_format = moeda_format
            c_val.border = border

        # Totais
        row_total = 7 + len(fases)
        ws_resumo.cell(row=row_total, column=1, value='TOTAL').font = Font(bold=True)
        c_tq = ws_resumo.cell(row=row_total, column=6, value=consolidado.get('total_quantidade', 0))
        c_tq.font = Font(bold=True)
        c_tq.number_format = inteiro_format
        c_tv = ws_resumo.cell(row=row_total, column=7, value=consolidado.get('total_valor', 0))
        c_tv.font = Font(bold=True)
        c_tv.number_format = moeda_format

        # Ajustar larguras
        for col_idx in range(1, 8):
            ws_resumo.column_dimensions[get_column_letter(col_idx)].width = 18

        # ============== ABAS POR FASE ==============
        for fase in fases:
            fase_num = fase.get('fase', 1)
            tipo = fase.get('tipo_calculo', '')
            nome_aba = f'Fase {fase_num}'
            if tipo == 'completo':
                nome_aba += ' (Completo)'

            ws = wb.create_sheet(nome_aba)

            dt_entrega = fase.get('data_entrega', '')
            ws['A1'] = f'FASE {fase_num} - Entrega: {dt_entrega}'
            ws['A1'].font = titulo_font
            ws.merge_cells('A1:F1')

            if tipo == 'completo':
                ws['A2'] = 'Calculo completo: considera estoque, estoque de seguranca e transferencias'
            else:
                ws['A2'] = 'Estoque projetado: demanda do periodo descontando estoque projetado e pedidos anteriores'

            ws['A2'].font = Font(italic=True, color='666666')

            # Headers
            if tipo == 'completo':
                headers = ['Cod Item', 'Cod DIG', 'Descricao', 'Fornecedor', 'ABC',
                           'Demanda/dia', 'Demanda Periodo', 'Estoque', 'ES',
                           'Qtd Pedido', 'Caixas', 'CUE (R$)', 'Valor Total (R$)']
                fill = header_fill_verde
            else:
                headers = ['Cod Item', 'Cod DIG', 'Descricao', 'Fornecedor',
                           'Demanda/dia', 'Dias', 'Demanda Periodo', 'Est. Projetado',
                           'Qtd Pedido', 'Caixas', 'CUE (R$)', 'Valor Total (R$)']
                fill = header_fill_laranja

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = header_font
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            itens = fase.get('itens', [])
            itens_ordenados = sorted(itens, key=lambda x: x.get('valor_total', 0), reverse=True)

            for row_idx, item in enumerate(itens_ordenados, 5):
                cod_str = str(item.get('codigo', ''))
                cod_dig = codigos_dig.get(cod_str, '')

                if tipo == 'completo':
                    valores = [
                        item.get('codigo', ''), cod_dig, item.get('descricao', ''),
                        item.get('nome_fornecedor', ''), item.get('curva_abc', ''),
                        item.get('demanda_diaria', 0), item.get('demanda_periodo', 0),
                        item.get('estoque_atual', 0), item.get('estoque_seguranca', 0),
                        item.get('quantidade_pedido', 0),
                        item.get('numero_caixas', 0), item.get('cue', 0),
                        item.get('valor_total', 0)
                    ]
                else:
                    valores = [
                        item.get('codigo', ''), cod_dig, item.get('descricao', ''),
                        item.get('nome_fornecedor', ''),
                        item.get('demanda_diaria', 0), fase.get('periodo_dias', 0),
                        item.get('demanda_periodo', 0), item.get('estoque_atual', 0),
                        item.get('quantidade_pedido', 0),
                        item.get('numero_caixas', 0), item.get('cue', 0),
                        item.get('valor_total', 0)
                    ]

                for col, val in enumerate(valores, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = border
                    # Formatar colunas numericas
                    if isinstance(val, float):
                        cell.number_format = moeda_format

            # Totais
            row_t = 5 + len(itens_ordenados)
            ws.cell(row=row_t, column=1, value='TOTAL').font = Font(bold=True)
            col_qtd = len(headers) - 3
            col_val = len(headers)
            c = ws.cell(row=row_t, column=col_qtd, value=fase.get('total_quantidade', 0))
            c.font = Font(bold=True)
            c.number_format = inteiro_format
            c = ws.cell(row=row_t, column=col_val, value=fase.get('total_valor', 0))
            c.font = Font(bold=True)
            c.number_format = moeda_format

            # Ajustar larguras
            widths = {'A': 10, 'B': 10, 'C': 35, 'D': 25}
            for letra, w in widths.items():
                ws.column_dimensions[letra].width = w
            for c_idx in range(5, len(headers) + 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = 14

        # ============== ABA CONSOLIDADO ==============
        ws_cons = wb.create_sheet('Consolidado')
        ws_cons['A1'] = 'CONSOLIDADO - TODAS AS FASES'
        ws_cons['A1'].font = titulo_font
        ws_cons.merge_cells('A1:F1')

        # Headers dinamicos baseados no numero de fases
        cons_headers = ['Cod Item', 'Cod DIG', 'Descricao', 'Fornecedor', 'CUE (R$)']
        for fase in fases:
            f_num = fase['fase']
            cons_headers.append(f'Fase {f_num} (un)')
            cons_headers.append(f'Fase {f_num} (R$)')
        cons_headers.extend(['Total (un)', 'Total Caixas', 'Total (R$)'])

        for col, h in enumerate(cons_headers, 1):
            cell = ws_cons.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill_azul
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        itens_cons = consolidado.get('itens', [])
        itens_cons_ord = sorted(itens_cons, key=lambda x: x.get('total_valor', 0), reverse=True)

        for row_idx, item in enumerate(itens_cons_ord, 4):
            cod_str = str(item.get('codigo', ''))
            cod_dig = codigos_dig.get(cod_str, '')

            col = 1
            ws_cons.cell(row=row_idx, column=col, value=item.get('codigo', '')).border = border
            col += 1
            ws_cons.cell(row=row_idx, column=col, value=cod_dig).border = border
            col += 1
            ws_cons.cell(row=row_idx, column=col, value=item.get('descricao', '')).border = border
            col += 1
            ws_cons.cell(row=row_idx, column=col, value=item.get('nome_fornecedor', '')).border = border
            col += 1
            c = ws_cons.cell(row=row_idx, column=col, value=item.get('cue', 0))
            c.number_format = moeda_format
            c.border = border
            col += 1

            fases_item = item.get('fases', {})
            for fase in fases:
                f_num = fase['fase']
                f_data = fases_item.get(f_num, fases_item.get(str(f_num), {}))
                c = ws_cons.cell(row=row_idx, column=col, value=f_data.get('quantidade', 0))
                c.number_format = inteiro_format
                c.border = border
                col += 1
                c = ws_cons.cell(row=row_idx, column=col, value=f_data.get('valor', 0))
                c.number_format = moeda_format
                c.border = border
                col += 1

            c = ws_cons.cell(row=row_idx, column=col, value=item.get('total_quantidade', 0))
            c.number_format = inteiro_format
            c.font = Font(bold=True)
            c.border = border
            col += 1
            c = ws_cons.cell(row=row_idx, column=col, value=item.get('total_caixas', 0))
            c.number_format = inteiro_format
            c.border = border
            col += 1
            c = ws_cons.cell(row=row_idx, column=col, value=item.get('total_valor', 0))
            c.number_format = moeda_format
            c.font = Font(bold=True)
            c.border = border

        # Ajustar larguras consolidado
        ws_cons.column_dimensions['A'].width = 10
        ws_cons.column_dimensions['B'].width = 10
        ws_cons.column_dimensions['C'].width = 35
        ws_cons.column_dimensions['D'].width = 25
        for c_idx in range(5, len(cons_headers) + 1):
            ws_cons.column_dimensions[get_column_letter(c_idx)].width = 14

        # Salvar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nome_forn = fornecedores[0] if fornecedores else 'geral'
        nome_forn_clean = nome_forn.replace(' ', '_').replace('/', '_')[:30]
        filename = f'compra_planejada_{nome_forn_clean}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500
