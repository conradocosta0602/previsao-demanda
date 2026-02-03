"""
Blueprint: Padrao de Compra
Rotas: /api/padrao-compra/*
"""

import io
from datetime import datetime
import pandas as pd
from flask import Blueprint, request, jsonify, send_file
from psycopg2.extras import RealDictCursor

from app.utils.db_connection import get_db_connection

padrao_compra_bp = Blueprint('padrao_compra', __name__)


@padrao_compra_bp.route('/api/padrao-compra/buscar', methods=['GET', 'POST'])
def api_padrao_compra_buscar():
    """
    Busca padroes de compra cadastrados.

    GET Query params:
        codigo: Codigo do produto (opcional)
        cod_empresa: Codigo da loja de venda (opcional)
        cod_empresa_destino: Codigo da loja destino (opcional)
        limit: Limite de resultados (padrao: 100)

    POST JSON body:
        codigos: Lista de codigos de produtos para buscar
    """
    try:
        conn = get_db_connection()

        # Verificar se e POST com lista de codigos
        if request.method == 'POST':
            dados = request.get_json() or {}
            codigos = dados.get('codigos', [])

            if not codigos:
                return jsonify({'success': True, 'total': 0, 'padroes': []})

            placeholders = ','.join(['%s'] * len(codigos))
            query = f"""
                SELECT
                    pci.codigo,
                    pci.cod_empresa_venda,
                    pci.cod_empresa_destino
                FROM padrao_compra_item pci
                WHERE pci.codigo IN ({placeholders})
                ORDER BY pci.codigo, pci.cod_empresa_venda
            """

            df = pd.read_sql(query, conn, params=codigos)
            conn.close()

            return jsonify({
                'success': True,
                'total': len(df),
                'padroes': df.to_dict('records')
            })

        # GET
        codigo = request.args.get('codigo', type=int)
        cod_empresa = request.args.get('cod_empresa', type=int)
        cod_empresa_destino = request.args.get('cod_empresa_destino', type=int)
        limit = request.args.get('limit', 100, type=int)

        query = """
            SELECT
                pci.codigo,
                cp.descricao as descricao_produto,
                pci.cod_empresa_venda,
                lv.nome_loja as loja_venda,
                pci.cod_empresa_destino,
                ld.nome_loja as loja_destino,
                CASE
                    WHEN pci.cod_empresa_venda = pci.cod_empresa_destino THEN 'Direto'
                    ELSE 'Centralizado'
                END as tipo_padrao,
                pci.data_referencia,
                pci.data_atualizacao
            FROM padrao_compra_item pci
            LEFT JOIN cadastro_produtos cp ON pci.codigo = cp.codigo
            LEFT JOIN cadastro_lojas lv ON pci.cod_empresa_venda = lv.cod_empresa
            LEFT JOIN cadastro_lojas ld ON pci.cod_empresa_destino = ld.cod_empresa
            WHERE 1=1
        """
        params = []

        if codigo is not None:
            query += " AND pci.codigo = %s"
            params.append(codigo)

        if cod_empresa is not None:
            query += " AND pci.cod_empresa_venda = %s"
            params.append(cod_empresa)

        if cod_empresa_destino is not None:
            query += " AND pci.cod_empresa_destino = %s"
            params.append(cod_empresa_destino)

        query += f" ORDER BY pci.codigo, pci.cod_empresa_venda LIMIT {limit}"

        df = pd.read_sql(query, conn, params=params if params else None)
        conn.close()

        return jsonify({
            'success': True,
            'total': len(df),
            'padroes': df.to_dict('records')
        })

    except Exception as e:
        import traceback
        print(f"[ERRO] api_padrao_compra_buscar: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@padrao_compra_bp.route('/api/padrao-compra/estatisticas', methods=['GET'])
def api_padrao_compra_estatisticas():
    """Retorna estatisticas dos padroes de compra cadastrados."""
    try:
        conn = get_db_connection()

        stats = {}

        # Total de registros
        df = pd.read_sql("SELECT COUNT(*) as total FROM padrao_compra_item", conn)
        stats['total_registros'] = int(df.iloc[0]['total'])

        # Produtos unicos
        df = pd.read_sql("SELECT COUNT(DISTINCT codigo) as total FROM padrao_compra_item", conn)
        stats['produtos_unicos'] = int(df.iloc[0]['total'])

        # Lojas de venda
        df = pd.read_sql("SELECT COUNT(DISTINCT cod_empresa_venda) as total FROM padrao_compra_item", conn)
        stats['lojas_venda'] = int(df.iloc[0]['total'])

        # Lojas destino
        df = pd.read_sql("SELECT COUNT(DISTINCT cod_empresa_destino) as total FROM padrao_compra_item", conn)
        stats['lojas_destino'] = int(df.iloc[0]['total'])

        # Compras diretas vs centralizadas
        df = pd.read_sql("""
            SELECT
                COUNT(*) FILTER (WHERE cod_empresa_venda = cod_empresa_destino) as diretas,
                COUNT(*) FILTER (WHERE cod_empresa_venda != cod_empresa_destino) as centralizadas
            FROM padrao_compra_item
        """, conn)
        stats['compras_diretas'] = int(df.iloc[0]['diretas'])
        stats['compras_centralizadas'] = int(df.iloc[0]['centralizadas'])

        # Top lojas centralizadoras
        df = pd.read_sql("""
            SELECT cod_empresa_destino, COUNT(*) as qtd
            FROM padrao_compra_item
            WHERE cod_empresa_venda != cod_empresa_destino
            GROUP BY cod_empresa_destino
            ORDER BY qtd DESC
            LIMIT 5
        """, conn)
        stats['top_centralizadoras'] = df.to_dict('records')

        # Parametro de dias de transferencia
        df = pd.read_sql("""
            SELECT valor FROM parametros_globais
            WHERE chave = 'dias_transferencia_padrao_compra' AND ativo = TRUE
        """, conn)
        stats['dias_transferencia'] = int(df.iloc[0]['valor']) if not df.empty else 10

        conn.close()

        return jsonify({
            'success': True,
            'estatisticas': stats
        })

    except Exception as e:
        import traceback
        print(f"[ERRO] api_padrao_compra_estatisticas: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@padrao_compra_bp.route('/api/padrao-compra/itens-sem-padrao', methods=['GET'])
def api_padrao_compra_itens_sem_padrao():
    """Busca itens que tem vendas mas nao tem padrao de compra definido."""
    try:
        fornecedor = request.args.get('fornecedor')
        cod_empresa = request.args.get('cod_empresa', type=int)
        limit = request.args.get('limit', 100, type=int)

        conn = get_db_connection()

        query = """
            SELECT DISTINCT
                hvd.codigo,
                cp.descricao as descricao_produto,
                hvd.cod_empresa,
                cl.nome_loja,
                cpc.nome_fornecedor,
                cpc.cnpj_fornecedor
            FROM historico_vendas_diario hvd
            LEFT JOIN cadastro_produtos cp ON hvd.codigo = cp.codigo
            LEFT JOIN cadastro_lojas cl ON hvd.cod_empresa = cl.cod_empresa
            LEFT JOIN cadastro_produtos_completo cpc ON hvd.codigo::text = cpc.cod_produto
            WHERE NOT EXISTS (
                SELECT 1 FROM padrao_compra_item pci
                WHERE pci.codigo = hvd.codigo
                  AND pci.cod_empresa_venda = hvd.cod_empresa
            )
            AND hvd.data >= CURRENT_DATE - INTERVAL '90 days'
        """
        params = []

        if fornecedor:
            query += " AND cpc.nome_fornecedor = %s"
            params.append(fornecedor)

        if cod_empresa is not None:
            query += " AND hvd.cod_empresa = %s"
            params.append(cod_empresa)

        query += f" ORDER BY cpc.nome_fornecedor, hvd.codigo LIMIT {limit}"

        df = pd.read_sql(query, conn, params=params if params else None)
        conn.close()

        return jsonify({
            'success': True,
            'total': len(df),
            'itens': df.to_dict('records'),
            'critica': 'sem_padrao_compra' if len(df) > 0 else None
        })

    except Exception as e:
        import traceback
        print(f"[ERRO] api_padrao_compra_itens_sem_padrao: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@padrao_compra_bp.route('/api/padrao-compra/processar-pedidos', methods=['POST'])
def api_padrao_compra_processar_pedidos():
    """
    Processa pedidos aplicando logica de padrao de compra.
    Recebe lista de itens e reagrupa por loja destino.
    """
    try:
        from core.padrao_compra import (
            PadraoCompraCalculator,
            agregar_pedidos_por_destino_fornecedor,
            formatar_pedido_para_excel
        )

        dados = request.get_json()
        itens = dados.get('itens', [])
        lead_times = dados.get('lead_times', {})
        agrupar_por_fornecedor = dados.get('agrupar_por_fornecedor', True)

        if not itens:
            return jsonify({
                'success': False,
                'erro': 'Nenhum item para processar'
            }), 400

        conn = get_db_connection()
        calculator = PadraoCompraCalculator(conn)

        # Processar itens
        resultados = calculator.processar_itens(itens, lead_times)

        # Agregar por destino/fornecedor
        agregado = agregar_pedidos_por_destino_fornecedor(
            resultados,
            agrupar_por_fornecedor=agrupar_por_fornecedor
        )

        conn.close()

        return jsonify({
            'success': True,
            'resultados': resultados,
            'agregado': agregado,
            'total_itens': len(resultados)
        })

    except Exception as e:
        import traceback
        print(f"[ERRO] api_padrao_compra_processar_pedidos: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@padrao_compra_bp.route('/api/padrao-compra/exportar', methods=['POST'])
def api_padrao_compra_exportar():
    """
    Exporta pedidos com padrao de compra para Excel.

    O formato do Excel e: uma linha por item x filial destino.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        dados = request.get_json()
        linhas = dados.get('linhas_excel', [])
        estatisticas = dados.get('estatisticas', {})
        criticas = dados.get('criticas', [])

        if not linhas:
            return jsonify({'success': False, 'erro': 'Nenhum dado para exportar'}), 400

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
        critica_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # =====================================================================
        # ABA 1: POR PADRAO DE COMPRA (agregado por filial destino)
        # =====================================================================
        ws = wb.active
        ws.title = 'Por Padrao de Compra'

        headers = [
            'Codigo', 'Descricao', 'Categoria', 'Curva ABC',
            'Fornecedor', 'CNPJ Fornecedor',
            'Filial Destino', 'Quantidade', 'CUE', 'Valor Pedido',
            'Lead Time Efetivo', 'Lojas Origem', 'Criticas'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, linha in enumerate(linhas, 2):
            ws.cell(row=row_idx, column=1, value=linha.get('codigo')).border = border
            ws.cell(row=row_idx, column=2, value=linha.get('descricao', '')).border = border
            ws.cell(row=row_idx, column=3, value=linha.get('categoria', '')).border = border
            ws.cell(row=row_idx, column=4, value=linha.get('curva_abc', 'B')).border = border
            ws.cell(row=row_idx, column=5, value=linha.get('nome_fornecedor', '')).border = border
            ws.cell(row=row_idx, column=6, value=linha.get('codigo_fornecedor', '')).border = border
            ws.cell(row=row_idx, column=7, value=linha.get('cod_empresa_destino')).border = border
            ws.cell(row=row_idx, column=8, value=linha.get('quantidade_pedido', 0)).border = border
            ws.cell(row=row_idx, column=9, value=linha.get('cue', 0)).border = border
            ws.cell(row=row_idx, column=10, value=linha.get('valor_pedido', 0)).border = border
            ws.cell(row=row_idx, column=11, value=linha.get('lead_time_efetivo', 0)).border = border
            ws.cell(row=row_idx, column=12, value=linha.get('lojas_origem', '')).border = border

            critica_texto = linha.get('criticas', '')
            cell_critica = ws.cell(row=row_idx, column=13, value=critica_texto)
            cell_critica.border = border
            if critica_texto:
                cell_critica.fill = critica_fill

        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 30
        ws.column_dimensions['M'].width = 40

        # =====================================================================
        # ABA 2: CRITICAS (se houver)
        # =====================================================================
        if criticas:
            ws_criticas = wb.create_sheet('Criticas')

            headers_criticas = ['Tipo', 'Codigo', 'Loja', 'Mensagem']
            for col, header in enumerate(headers_criticas, 1):
                cell = ws_criticas.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='E53E3E', end_color='E53E3E', fill_type='solid')
                cell.border = border

            for row_idx, critica in enumerate(criticas, 2):
                ws_criticas.cell(row=row_idx, column=1, value=critica.get('tipo', '')).border = border
                ws_criticas.cell(row=row_idx, column=2, value=critica.get('codigo', '')).border = border
                ws_criticas.cell(row=row_idx, column=3, value=critica.get('cod_empresa', '')).border = border
                ws_criticas.cell(row=row_idx, column=4, value=critica.get('mensagem', '')).border = border

        # Salvar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pedido_padrao_compra_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(f"[ERRO] api_padrao_compra_exportar: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500
