"""
Blueprint: Transferencias entre Lojas
Rotas: /transferencias, /api/transferencias/*
"""

import os
import io
from datetime import datetime
import pandas as pd
from flask import Blueprint, render_template, request, jsonify, send_file

from app.utils.db_connection import get_db_connection

transferencias_bp = Blueprint('transferencias', __name__)


@transferencias_bp.route('/transferencias')
def transferencias():
    """Pagina de transferencias entre lojas - Oportunidades calculadas automaticamente"""
    return render_template('transferencias.html')


@transferencias_bp.route('/api/transferencias/oportunidades', methods=['GET'])
def api_transferencias_oportunidades():
    """
    Retorna oportunidades de transferencia calculadas.
    Query params:
        - grupo_id: Filtrar por grupo (opcional)
        - horas: Limitar a oportunidades das ultimas X horas (default: 24)
    """
    try:
        from core.transferencia_regional import TransferenciaRegional
        from psycopg2.extras import RealDictCursor

        grupo_id = request.args.get('grupo_id', type=int)
        horas = request.args.get('horas', default=24, type=int)

        conn = get_db_connection()

        # Verificar se tabelas de transferencia existem
        cursor_check = conn.cursor()
        cursor_check.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'oportunidades_transferencia'
            )
        """)
        tabela_existe = cursor_check.fetchone()[0]
        cursor_check.close()

        if not tabela_existe:
            conn.close()
            return jsonify({
                'success': True,
                'oportunidades': [],
                'resumo': {'total': 0, 'criticas': 0, 'altas': 0, 'valor_total': 0, 'produtos_unicos': 0},
                'aviso': 'Tabela de transferencias ainda nao foi criada.'
            })

        cursor = conn.cursor(cursor_factory=RealDictCursor)

        query = """
            SELECT *
            FROM oportunidades_transferencia
            WHERE data_calculo >= NOW() - INTERVAL '%s hours'
        """
        params = [horas]

        if grupo_id:
            query += " AND grupo_id = %s"
            params.append(grupo_id)

        query += " ORDER BY urgencia DESC, data_calculo DESC"

        cursor.execute(query, params)
        oportunidades = cursor.fetchall()

        # Calcular resumo
        oportunidades_list = [dict(o) for o in oportunidades]
        for o in oportunidades_list:
            if o.get('data_calculo'):
                o['data_calculo'] = str(o['data_calculo'])

        resumo = {
            'total': len(oportunidades_list),
            'criticas': len([o for o in oportunidades_list if o.get('urgencia') == 'CRITICA']),
            'altas': len([o for o in oportunidades_list if o.get('urgencia') == 'ALTA']),
            'valor_total': sum(float(o.get('valor_estimado', 0)) for o in oportunidades_list),
            'produtos_unicos': len(set(o.get('cod_produto') for o in oportunidades_list))
        }

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'oportunidades': oportunidades_list,
            'resumo': resumo
        })

    except Exception as e:
        print(f"Erro ao buscar oportunidades: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@transferencias_bp.route('/api/transferencias/grupos', methods=['GET'])
def api_transferencias_grupos():
    """Retorna grupos regionais configurados."""
    try:
        from psycopg2.extras import RealDictCursor

        conn = get_db_connection()

        cursor_check = conn.cursor()
        cursor_check.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'grupos_transferencia'
            )
        """)
        tabela_existe = cursor_check.fetchone()[0]
        cursor_check.close()

        if not tabela_existe:
            conn.close()
            return jsonify({
                'success': True,
                'grupos': [],
                'aviso': 'Tabela de grupos ainda nao foi criada.'
            })

        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT id, nome, cd_principal, lojas, ativo, created_at
            FROM grupos_transferencia
            WHERE ativo = TRUE
            ORDER BY nome
        """)
        grupos = cursor.fetchall()

        grupos_list = []
        for g in grupos:
            grupo_dict = dict(g)
            if grupo_dict.get('created_at'):
                grupo_dict['created_at'] = str(grupo_dict['created_at'])
            grupos_list.append(grupo_dict)

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'grupos': grupos_list
        })

    except Exception as e:
        print(f"Erro ao buscar grupos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@transferencias_bp.route('/api/transferencias/exportar', methods=['GET'])
def api_transferencias_exportar():
    """Exporta oportunidades de transferencia para Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from psycopg2.extras import RealDictCursor

        grupo_id = request.args.get('grupo_id', type=int)
        horas = request.args.get('horas', default=24, type=int)

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        query = """
            SELECT *
            FROM oportunidades_transferencia
            WHERE data_calculo >= NOW() - INTERVAL '%s hours'
        """
        params = [horas]

        if grupo_id:
            query += " AND grupo_id = %s"
            params.append(grupo_id)

        query += " ORDER BY urgencia DESC, data_calculo DESC"

        cursor.execute(query, params)
        oportunidades = cursor.fetchall()

        cursor.close()
        conn.close()

        if not oportunidades:
            return jsonify({'success': False, 'erro': 'Nenhuma oportunidade para exportar'}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = 'Transferencias'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Mapeamento de codigo de filial para nome abreviado
        NOMES_FILIAIS = {
            1: 'GUS', 2: 'IMB', 3: 'PAL', 4: 'TAM', 5: 'AJU',
            6: 'JPA', 7: 'PNG', 8: 'CAU', 9: 'BAR', 11: 'FRT',
            80: 'MDC', 81: 'OBC', 82: 'OSAL', 91: 'CA2',
            92: 'CAB', 93: 'ALH', 94: 'LAU'
        }

        # Carregar mapeamento de codigos DIG
        codigos_dig = {}
        try:
            conn_dig = get_db_connection()
            cursor_dig = conn_dig.cursor()
            cursor_dig.execute("SELECT codigo, codigo_dig FROM codigo_dig")
            codigos_dig = {str(row[0]): str(row[1]) for row in cursor_dig.fetchall()}
            cursor_dig.close()
            conn_dig.close()
        except Exception:
            pass  # Tabela pode nao existir ainda

        headers = [
            'Cod Destino', 'Filial Dest', 'Codigo', 'Cod DIG', 'Descricao', 'Curva ABC',
            'Cod Origem', 'Filial Orig', 'Estoque Origem', 'Cobertura Origem',
            'Estoque Destino', 'Cobertura Destino',
            'Qtd Sugerida', 'Valor Estimado', 'Urgencia', 'Data Calculo'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_idx, o in enumerate(oportunidades, start=2):
            # Codigo e nome abreviado da loja destino (primeiras colunas)
            # NOTA: A tabela usa 'loja_destino', nao 'cod_loja_destino'
            cod_destino = o.get('loja_destino', 0)
            if cod_destino is None:
                cod_destino = 0
            ws.cell(row=row_idx, column=1, value=cod_destino).border = border
            nome_destino_abrev = NOMES_FILIAIS.get(int(cod_destino), '') if cod_destino else ''
            ws.cell(row=row_idx, column=2, value=nome_destino_abrev).border = border
            cod_produto = str(o.get('cod_produto', ''))
            ws.cell(row=row_idx, column=3, value=cod_produto).border = border
            # Codigo DIG
            cod_dig = codigos_dig.get(cod_produto, 'N/D')
            ws.cell(row=row_idx, column=4, value=cod_dig).border = border
            ws.cell(row=row_idx, column=5, value=o.get('descricao_produto', '')).border = border
            ws.cell(row=row_idx, column=6, value=o.get('curva_abc', '')).border = border
            # Codigo e nome abreviado da loja origem
            # NOTA: A tabela usa 'loja_origem', nao 'cod_loja_origem'
            cod_origem = o.get('loja_origem', 0)
            if cod_origem is None:
                cod_origem = 0
            ws.cell(row=row_idx, column=7, value=cod_origem).border = border
            nome_origem_abrev = NOMES_FILIAIS.get(int(cod_origem), '') if cod_origem else ''
            ws.cell(row=row_idx, column=8, value=nome_origem_abrev).border = border
            ws.cell(row=row_idx, column=9, value=o.get('estoque_origem', 0)).border = border
            ws.cell(row=row_idx, column=10, value=round(o.get('cobertura_origem_dias', 0), 1)).border = border
            ws.cell(row=row_idx, column=11, value=o.get('estoque_destino', 0)).border = border
            ws.cell(row=row_idx, column=12, value=round(o.get('cobertura_destino_dias', 0), 1)).border = border
            ws.cell(row=row_idx, column=13, value=o.get('qtd_sugerida', 0)).border = border
            ws.cell(row=row_idx, column=14, value=o.get('valor_estimado', 0)).border = border
            ws.cell(row=row_idx, column=14).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=15, value=o.get('urgencia', '')).border = border
            ws.cell(row=row_idx, column=16, value=str(o.get('data_calculo', ''))).border = border

        col_widths = [10, 10, 12, 12, 40, 8, 10, 10, 12, 12, 12, 12, 12, 14, 10, 20]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'transferencias_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar transferencias: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@transferencias_bp.route('/api/transferencias/limpar', methods=['DELETE'])
def api_transferencias_limpar():
    """Limpa historico de transferencias."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM oportunidades_transferencia")
        total_antes = cursor.fetchone()[0]

        cursor.execute("DELETE FROM oportunidades_transferencia")
        conn.commit()

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'registros_removidos': total_antes,
            'mensagem': f'{total_antes} registros removidos com sucesso'
        })

    except Exception as e:
        print(f"Erro ao limpar historico: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500
