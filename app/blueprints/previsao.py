"""
Blueprint: Previsao de Demanda (APIs)
Rotas: /api/gerar_previsao_banco*, /api/exportar_tabela_comparativa
"""

import os
import io
import math
from datetime import datetime
import pandas as pd
import numpy as np
from flask import Blueprint, request, jsonify, send_file, current_app
from werkzeug.utils import secure_filename

from app.utils.previsao_helper import processar_previsao, set_ultima_previsao_data

previsao_bp = Blueprint('previsao', __name__)


@previsao_bp.route('/api/processar', methods=['POST'])
def processar():
    """
    Processa dados e gera previsoes.
    Aceita arquivo Excel via upload.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'erro': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'success': False, 'erro': 'Nenhum arquivo selecionado'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'erro': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400

        # Obter parametros
        meses_previsao = request.form.get('meses_previsao', 6, type=int)
        granularidade = request.form.get('granularidade', 'mensal')

        # Filtros opcionais
        filiais_filtro = request.form.getlist('filiais')
        produtos_filtro = request.form.getlist('produtos')

        # Salvar arquivo temporariamente
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Processar previsao
        resultado = processar_previsao(
            arquivo_excel=filepath,
            meses_previsao=meses_previsao,
            granularidade=granularidade,
            filiais_filtro=filiais_filtro if filiais_filtro else None,
            produtos_filtro=produtos_filtro if produtos_filtro else None,
            output_folder=current_app.config['OUTPUT_FOLDER']
        )

        # Armazenar dados da ultima previsao para o simulador
        if resultado.get('success') and 'grafico_data' in resultado:
            set_ultima_previsao_data(resultado['grafico_data'])

        # Limpar arquivo temporario
        try:
            os.remove(filepath)
        except:
            pass

        return jsonify(resultado)

    except ValueError as e:
        return jsonify({'success': False, 'erro': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': f'Erro: {str(e)}'}), 500


@previsao_bp.route('/api/gerar_previsao_banco', methods=['POST'])
def api_gerar_previsao_banco():
    """
    Gera previsao de demanda a partir de dados do banco de dados.
    Redireciona para a versao Bottom-Up que retorna WMAPE, BIAS e grafico de linha.
    """
    # Redirecionar para a API V2 interna que tem a estrutura correta
    return _api_gerar_previsao_banco_v2_interno()


def _api_gerar_previsao_banco_v2_interno():
    """
    BOTTOM-UP FORECASTING V2 - Conforme app_original.py

    Usa DemandCalculator.calcular_demanda_diaria_unificada() que implementa
    os 6 metodos estatisticos de forma otimizada para dados diarios:
    - media_simples: Media Movel Simples
    - media_ponderada: Media Movel Ponderada (WMA)
    - media_movel_exp: Suavizacao Exponencial (EMA)
    - tendencia: Regressao com Tendencia
    - sazonal: Decomposicao Sazonal
    - tsb: TSB para demanda intermitente

    Calcula previsao para cada item individualmente, depois agrega para o total.
    Retorna estrutura com WMAPE, BIAS e dados para grafico de linha.
    """
    import psycopg2
    from psycopg2.extras import RealDictCursor
    from datetime import datetime as dt, timedelta
    from calendar import monthrange
    from app.utils.db_connection import get_db_connection, DB_CONFIG

    try:
        dados = request.get_json()

        # Extrair parametros
        loja = dados.get('loja', 'TODAS')
        categoria = dados.get('categoria', 'TODAS')
        produto = dados.get('produto', 'TODOS')
        granularidade = dados.get('granularidade', 'mensal')

        # Parametros adicionais de filtro
        fornecedor = dados.get('fornecedor', 'TODOS')
        linha = dados.get('linha', 'TODAS')
        sublinha = dados.get('sublinha', 'TODAS')

        # Parametros de periodo
        mes_inicio = dados.get('mes_inicio')
        ano_inicio = dados.get('ano_inicio')
        mes_fim = dados.get('mes_fim')
        ano_fim = dados.get('ano_fim')
        semana_inicio = dados.get('semana_inicio')
        semana_fim = dados.get('semana_fim')
        ano_semana_inicio = dados.get('ano_semana_inicio')
        ano_semana_fim = dados.get('ano_semana_fim')
        data_inicio = dados.get('data_inicio')
        data_fim = dados.get('data_fim')

        # Calcular numero de periodos de previsao
        meses_previsao = 6
        dias_periodo_total = 181

        if granularidade == 'mensal' and mes_inicio and ano_inicio and mes_fim and ano_fim:
            meses_previsao = (int(ano_fim) - int(ano_inicio)) * 12 + (int(mes_fim) - int(mes_inicio)) + 1
            dias_periodo_total = 0
            for i in range(meses_previsao):
                mes_atual = int(mes_inicio) + i
                ano_atual = int(ano_inicio)
                while mes_atual > 12:
                    mes_atual -= 12
                    ano_atual += 1
                dias_periodo_total += monthrange(ano_atual, mes_atual)[1]

        elif granularidade == 'semanal' and semana_inicio and ano_semana_inicio and semana_fim and ano_semana_fim:
            semanas = (int(ano_semana_fim) - int(ano_semana_inicio)) * 52 + (int(semana_fim) - int(semana_inicio)) + 1
            meses_previsao = semanas
            dias_periodo_total = semanas * 7

        elif granularidade == 'diario' and data_inicio and data_fim:
            d1 = dt.strptime(data_inicio, '%Y-%m-%d')
            d2 = dt.strptime(data_fim, '%Y-%m-%d')
            dias = (d2 - d1).days + 1
            meses_previsao = dias
            dias_periodo_total = dias

        periodos_previsao = meses_previsao

        # Conectar ao banco
        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Construir filtros SQL
        where_conditions = []
        params = []

        def add_filter(field, value, exclude_value, convert_int=False):
            if value and value != exclude_value:
                if isinstance(value, list):
                    # Filtrar valores de exclusao da lista
                    filtered = [v for v in value if v != exclude_value]
                    if not filtered:
                        return  # Todos os valores eram de exclusao
                    if len(filtered) == 1:
                        where_conditions.append(f"{field} = %s")
                        params.append(int(filtered[0]) if convert_int else filtered[0])
                    elif len(filtered) > 1:
                        placeholders = ', '.join(['%s'] * len(filtered))
                        where_conditions.append(f"{field} IN ({placeholders})")
                        params.extend([int(v) if convert_int else v for v in filtered])
                else:
                    where_conditions.append(f"{field} = %s")
                    params.append(int(value) if convert_int else value)

        add_filter("h.cod_empresa", loja, 'TODAS', convert_int=True)
        add_filter("p.nome_fornecedor", fornecedor, 'TODOS')
        add_filter("p.categoria", linha, 'TODAS')
        add_filter("p.codigo_linha", sublinha, 'TODAS')
        add_filter("h.codigo", produto, 'TODOS', convert_int=True)

        where_sql = ""
        if where_conditions:
            where_sql = "AND " + " AND ".join(where_conditions)

        # =====================================================
        # CORTE DE DATA: Exclui dados do periodo de previsao
        # Se usuario quer prever Jan/2026, dados de Jan/2026
        # NAO devem ser usados no calculo do historico
        # =====================================================
        corte_sql = ""
        params_com_corte = list(params)

        if granularidade == 'mensal' and mes_inicio and ano_inicio:
            data_corte = f"{int(ano_inicio):04d}-{int(mes_inicio):02d}-01"
            corte_sql = "AND h.data < %s"
            params_com_corte.append(data_corte)
        elif granularidade == 'semanal' and semana_inicio and ano_semana_inicio:
            try:
                from datetime import datetime as dt_calc
                data_base = dt_calc.strptime(f'{int(ano_semana_inicio)}-W{int(semana_inicio):02d}-1', '%G-W%V-%u')
                corte_sql = "AND h.data < %s"
                params_com_corte.append(data_base.strftime('%Y-%m-%d'))
            except:
                pass
        elif granularidade == 'diario' and data_inicio:
            corte_sql = "AND h.data < %s"
            params_com_corte.append(data_inicio)

        # Buscar lista de itens (incluindo situacao de compra para filtro EN/FL v6.2)
        # A situacao e agregada por item - se tiver EN ou FL em alguma loja, mostra
        query_itens = f"""
            SELECT DISTINCT
                h.codigo as cod_produto,
                p.descricao,
                p.nome_fornecedor,
                p.cnpj_fornecedor,
                p.categoria,
                p.descricao_linha,
                COALESCE(
                    (SELECT string_agg(DISTINCT s.sit_compra, ',')
                     FROM situacao_compra_itens s
                     WHERE s.codigo = h.codigo
                       AND s.sit_compra IN ('EN', 'FL')),
                    'ATIVO'
                ) as situacao_compra
            FROM historico_vendas_diario h
            JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
            WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
            {where_sql}
            ORDER BY p.nome_fornecedor, p.descricao
        """
        cursor.execute(query_itens, params)
        lista_itens = cursor.fetchall()

        if not lista_itens:
            cursor.close()
            conn.close()
            return jsonify({'erro': 'Nenhum item encontrado com os filtros selecionados'}), 400

        # Buscar vendas historicas (COM corte de data para excluir periodo de previsao)
        query_vendas = f"""
            SELECT
                h.codigo as cod_produto,
                h.data as periodo,
                SUM(h.qtd_venda) as qtd_venda
            FROM historico_vendas_diario h
            JOIN cadastro_produtos_completo p ON h.codigo::text = p.cod_produto
            WHERE h.data >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '2 years')
            {where_sql}
            {corte_sql}
            GROUP BY h.codigo, h.data
            ORDER BY h.codigo, h.data
        """
        cursor.execute(query_vendas, params_com_corte)
        vendas_por_item_raw = cursor.fetchall()

        # =====================================================
        # BUSCAR CUE (Custo Unitario de Estoque) POR ITEM
        # Media ponderada pelo estoque de cada loja
        # =====================================================
        query_cue = """
            SELECT codigo as cod_produto,
                   CASE WHEN SUM(estoque) > 0
                        THEN SUM(cue * estoque) / SUM(estoque)
                        ELSE MAX(cue)
                   END as cue_medio
            FROM estoque_posicao_atual
            WHERE cue > 0
            GROUP BY codigo
        """
        cursor.execute(query_cue)
        cue_rows = cursor.fetchall()
        cue_por_item = {}
        for row in cue_rows:
            cue_por_item[row['cod_produto']] = float(row['cue_medio'] or 0)

        # Organizar vendas por item
        vendas_item_dict = {}
        todas_datas = set()
        for venda in vendas_por_item_raw:
            cod = venda['cod_produto']
            if cod not in vendas_item_dict:
                vendas_item_dict[cod] = []
            data_str = venda['periodo'].strftime('%Y-%m-%d')
            todas_datas.add(data_str)
            vendas_item_dict[cod].append({
                'periodo': data_str,
                'qtd_venda': float(venda['qtd_venda'] or 0)
            })

        datas_ordenadas = sorted(list(todas_datas))

        # Agregar vendas para serie historica
        vendas_agregadas = {}
        for cod, vendas in vendas_item_dict.items():
            for v in vendas:
                periodo = v['periodo']
                if periodo not in vendas_agregadas:
                    vendas_agregadas[periodo] = 0
                vendas_agregadas[periodo] += v['qtd_venda']

        # Agregar para granularidade de exibicao
        vendas_agregadas_por_periodo = {}
        for data_str in datas_ordenadas:
            valor = vendas_agregadas.get(data_str, 0)
            try:
                data_obj = dt.strptime(data_str, '%Y-%m-%d')
                if granularidade == 'mensal':
                    chave_periodo = f"{data_obj.year:04d}-{data_obj.month:02d}-01"
                elif granularidade == 'semanal':
                    ano_iso, semana_iso, _ = data_obj.isocalendar()
                    chave_periodo = f"{ano_iso}-S{semana_iso:02d}"
                else:
                    chave_periodo = data_str

                if chave_periodo not in vendas_agregadas_por_periodo:
                    vendas_agregadas_por_periodo[chave_periodo] = 0
                vendas_agregadas_por_periodo[chave_periodo] += valor
            except:
                pass

        datas_historicas = sorted(vendas_agregadas_por_periodo.keys())
        serie_historica_agregada = [vendas_agregadas_por_periodo[d] for d in datas_historicas]

        # Dividir serie em BASE (75%) e TESTE (25%)
        split_index = int(len(datas_historicas) * 0.75)
        datas_base = datas_historicas[:split_index]
        valores_base = serie_historica_agregada[:split_index]
        datas_teste = datas_historicas[split_index:]
        valores_teste = serie_historica_agregada[split_index:]

        # Gerar datas de previsao
        datas_previsao = []
        if granularidade == 'mensal' and mes_inicio and ano_inicio:
            for i in range(periodos_previsao):
                mes_atual = int(mes_inicio) + i
                ano_atual = int(ano_inicio)
                while mes_atual > 12:
                    mes_atual -= 12
                    ano_atual += 1
                datas_previsao.append(f"{ano_atual:04d}-{mes_atual:02d}-01")
        elif granularidade == 'semanal' and semana_inicio and ano_semana_inicio:
            try:
                data_base = dt.strptime(f'{int(ano_semana_inicio)}-W{int(semana_inicio):02d}-1', '%G-W%V-%u')
            except ValueError:
                data_base = dt(int(ano_semana_inicio), 1, 1) + timedelta(weeks=int(semana_inicio) - 1)
            for i in range(periodos_previsao):
                data_prev = data_base + timedelta(weeks=i)
                semana_atual = data_prev.isocalendar()[1]
                ano_atual = data_prev.isocalendar()[0]
                datas_previsao.append(f"{ano_atual}-S{semana_atual:02d}")
        elif granularidade == 'diario' and data_inicio:
            data_base = dt.strptime(data_inicio, '%Y-%m-%d')
            for i in range(periodos_previsao):
                data_prev = data_base + timedelta(days=i)
                datas_previsao.append(data_prev.strftime('%Y-%m-%d'))
        else:
            from dateutil.relativedelta import relativedelta
            ultima_data = dt.strptime(datas_historicas[-1], '%Y-%m-%d') if datas_historicas else dt.now()
            for i in range(periodos_previsao):
                if granularidade == 'mensal':
                    data_prev = ultima_data + relativedelta(months=i+1)
                    datas_previsao.append(data_prev.strftime('%Y-%m-01'))
                else:
                    data_prev = ultima_data + timedelta(days=(i+1)*7)
                    datas_previsao.append(data_prev.strftime('%Y-%m-%d'))

        # =====================================================
        # BUSCAR DEMANDA PRE-CALCULADA (V54)
        # Fonte primaria de previsao - ja inclui saneamento V53 por loja,
        # fatores sazonais corrigidos e bypass do limitador V11
        # =====================================================
        relatorio_itens = []
        previsoes_agregadas_por_periodo = {d: 0 for d in datas_previsao}
        contagem_modelos = {}

        demanda_pre_calc = {}  # {(cod_produto_str, periodo_str): {campos...}}

        if granularidade in ('mensal', 'semanal') and datas_previsao:
            try:
                if granularidade == 'mensal':
                    periodos_query = []
                    for dp in datas_previsao:
                        periodos_query.append((int(dp[:4]), int(dp[5:7])))

                    if periodos_query:
                        periodo_conditions = ' OR '.join(
                            [f'(d.ano = %s AND d.mes = %s)' for _ in periodos_query]
                        )
                        params_periodos = []
                        for a, m in periodos_query:
                            params_periodos.extend([a, m])

                        cursor.execute(f"""
                            SELECT d.cod_produto, d.ano, d.mes,
                                   COALESCE(d.ajuste_manual, d.demanda_prevista) as demanda_efetiva,
                                   d.demanda_prevista,
                                   d.demanda_diaria_base,
                                   d.fator_sazonal,
                                   d.fator_tendencia_yoy,
                                   d.valor_ano_anterior,
                                   d.metodo_usado,
                                   d.classificacao_tendencia,
                                   d.desvio_padrao,
                                   d.taxa_disponibilidade,
                                   d.demanda_censurada_corrigida,
                                   d.ajuste_manual,
                                   d.limitador_aplicado
                            FROM demanda_pre_calculada d
                            WHERE d.cod_empresa IS NULL
                              AND d.tipo_granularidade = 'mensal'
                              AND ({periodo_conditions})
                        """, params_periodos)

                        for row in cursor.fetchall():
                            periodo_str = f"{row['ano']:04d}-{int(row['mes']):02d}-01"
                            demanda_pre_calc[(str(row['cod_produto']), periodo_str)] = dict(row)

                elif granularidade == 'semanal':
                    periodos_query = []
                    for dp in datas_previsao:
                        if '-S' in dp:
                            partes = dp.split('-S')
                            periodos_query.append((int(partes[0]), int(partes[1])))

                    if periodos_query:
                        periodo_conditions = ' OR '.join(
                            [f'(d.ano = %s AND d.semana = %s)' for _ in periodos_query]
                        )
                        params_periodos = []
                        for a, s in periodos_query:
                            params_periodos.extend([a, s])

                        cursor.execute(f"""
                            SELECT d.cod_produto, d.ano, d.semana,
                                   COALESCE(d.ajuste_manual, d.demanda_prevista) as demanda_efetiva,
                                   d.demanda_prevista,
                                   d.demanda_diaria_base,
                                   d.fator_sazonal,
                                   d.fator_tendencia_yoy,
                                   d.valor_ano_anterior,
                                   d.metodo_usado,
                                   d.classificacao_tendencia,
                                   d.desvio_padrao,
                                   d.taxa_disponibilidade,
                                   d.demanda_censurada_corrigida,
                                   d.ajuste_manual,
                                   d.limitador_aplicado
                            FROM demanda_pre_calculada d
                            WHERE d.cod_empresa IS NULL
                              AND d.tipo_granularidade = 'semanal'
                              AND ({periodo_conditions})
                        """, params_periodos)

                        for row in cursor.fetchall():
                            periodo_str = f"{row['ano']}-S{int(row['semana']):02d}"
                            demanda_pre_calc[(str(row['cod_produto']), periodo_str)] = dict(row)

                print(f"  [V54] Carregados {len(demanda_pre_calc)} registros de demanda_pre_calculada", flush=True)
            except Exception as e_pc:
                print(f"  [V54] Erro ao carregar demanda_pre_calculada: {e_pc}", flush=True)

        # Mapeamento de metodos para nomes simplificados
        mapeamento_metodos = {
            'media_simples': 'SMA',
            'media_ponderada': 'WMA',
            'media_movel_exp': 'EMA',
            'tendencia': 'Regressao com Tendencia',
            'sazonal': 'Decomposicao Sazonal',
            'tsb': 'TSB',
            'tsb_simplificado': 'TSB',
            'wma': 'WMA',
            'ema': 'EMA',
            'wma_adaptativo_estavel': 'WMA',
            'wma_adaptativo_crescente': 'Regressao com Tendencia',
            'wma_adaptativo_decrescente': 'Regressao com Tendencia',
            'tsb_intermitente_extremo': 'TSB',
            'sma': 'SMA',
            'tendencia_linear': 'Tendencia',
            'sazonal_decomposicao': 'Sazonal'
        }

        for item in lista_itens:
            cod_produto = item['cod_produto']
            situacao_compra = item.get('situacao_compra', 'ATIVO')

            # =====================================================
            # FILTRO EN/FL v6.2: Itens com situacao EN ou FL
            # Nao calcular demanda, mas incluir na tabela com demanda zerada
            # =====================================================
            if situacao_compra and situacao_compra != 'ATIVO' and ('EN' in situacao_compra or 'FL' in situacao_compra):
                sit_principal = 'FL' if 'FL' in situacao_compra else 'EN'
                sit_descricao = 'Fora de Linha' if sit_principal == 'FL' else 'Em Negociacao'

                relatorio_itens.append({
                    'cod_produto': cod_produto,
                    'descricao': item['descricao'],
                    'nome_fornecedor': item['nome_fornecedor'] or 'SEM FORNECEDOR',
                    'categoria': item['categoria'],
                    'situacao_compra': situacao_compra,
                    'sit_compra': sit_principal,
                    'sit_compra_descricao': sit_descricao,
                    'demanda_prevista_total': 0,
                    'demanda_ano_anterior': 0,
                    'variacao_percentual': 0,
                    'sinal_emoji': '⚫',
                    'metodo_estatistico': 'BLOQUEADO',
                    'fator_tendencia_yoy': 0,
                    'classificacao_tendencia': 'bloqueado',
                    'previsao_por_periodo': [],
                    'cue': round(cue_por_item.get(cod_produto, 0), 4)
                })
                continue

            # =====================================================
            # V54: Usar demanda_pre_calculada como fonte primaria
            # =====================================================
            previsao_item_periodos = []
            total_previsao_item = 0
            total_ano_anterior_item = 0
            metodo_usado = None
            fator_tendencia_yoy = 1.0
            classificacao_tendencia = 'nao_calculado'
            desvio_diario_base = 0
            tem_dados_pre_calc = False

            for data_prev in datas_previsao:
                chave_pc = (str(cod_produto), data_prev)
                dados_pc = demanda_pre_calc.get(chave_pc)

                if dados_pc:
                    tem_dados_pre_calc = True
                    previsao_periodo_int = round(float(dados_pc['demanda_efetiva'] or 0))
                    valor_aa_periodo = round(float(dados_pc['valor_ano_anterior'] or 0))
                    tem_ajuste = dados_pc['ajuste_manual'] is not None

                    # Capturar metadados do primeiro periodo encontrado
                    if metodo_usado is None:
                        metodo_usado = dados_pc.get('metodo_usado', 'sma')
                        fator_tendencia_yoy = float(dados_pc.get('fator_tendencia_yoy') or 1.0)
                        classificacao_tendencia = dados_pc.get('classificacao_tendencia') or 'nao_calculado'
                        desvio_diario_base = float(dados_pc.get('desvio_padrao') or 0)
                else:
                    previsao_periodo_int = 0
                    valor_aa_periodo = 0
                    tem_ajuste = False

                previsao_item_periodos.append({
                    'periodo': data_prev,
                    'previsao': previsao_periodo_int,
                    'ano_anterior': valor_aa_periodo,
                    'ajuste_manual': tem_ajuste
                })
                previsoes_agregadas_por_periodo[data_prev] += previsao_periodo_int
                total_previsao_item += previsao_periodo_int
                total_ano_anterior_item += valor_aa_periodo

            # Se nao tem dados pre-calculados E nao tem vendas historicas, pular item
            if not tem_dados_pre_calc:
                vendas_hist = vendas_item_dict.get(cod_produto, [])
                if not vendas_hist:
                    continue

            if metodo_usado is None:
                metodo_usado = 'sem_dados'

            # =====================================================
            # APLICAR EVENTOS (promocoes, sazonais, etc.)
            # =====================================================
            try:
                from core.event_manager_v2 import EventManagerV2
                _evt_mgr = EventManagerV2()
                cnpj_forn = item.get('cnpj_fornecedor')
                cat_item = item.get('categoria')
                linha3_item = item.get('descricao_linha')

                total_previsao_item = 0
                for pp in previsao_item_periodos:
                    try:
                        data_prev_str = pp['periodo']
                        if granularidade == 'mensal':
                            data_inicio_evt = dt.strptime(data_prev_str, '%Y-%m-%d').date()
                            dias_mes = monthrange(data_inicio_evt.year, data_inicio_evt.month)[1]
                            data_fim_evt = data_inicio_evt.replace(day=dias_mes)
                        else:
                            data_inicio_evt = dt.strptime(data_prev_str[:10], '%Y-%m-%d').date()
                            data_fim_evt = data_inicio_evt + timedelta(days=6)

                        resultado_evt = _evt_mgr.calcular_fator_eventos(
                            codigo=int(cod_produto),
                            cod_empresa=0,
                            cod_fornecedor=cnpj_forn,
                            linha1=cat_item,
                            linha3=linha3_item,
                            data_inicio=data_inicio_evt,
                            data_fim=data_fim_evt
                        )
                        fator_evt = resultado_evt.get('fator_total', 1.0)
                        if fator_evt != 1.0:
                            pp['previsao_original'] = pp['previsao']
                            pp['previsao'] = round(pp['previsao'] * fator_evt)
                            pp['evento_fator'] = fator_evt
                            pp['evento_nomes'] = ', '.join(
                                e.get('nome', '') for e in resultado_evt.get('eventos_aplicados', [])
                            )
                    except Exception:
                        pass
                    total_previsao_item += pp['previsao']

                for pp in previsao_item_periodos:
                    if pp.get('evento_fator') and pp.get('previsao_original') is not None:
                        diff = pp['previsao'] - pp['previsao_original']
                        previsoes_agregadas_por_periodo[pp['periodo']] += diff
            except Exception as e_evt:
                print(f"[EVENTOS] Erro ao aplicar eventos para item {cod_produto}: {e_evt}")

            # Metodo de exibicao
            chave_metodo = metodo_usado.split('+')[-1] if '+' in metodo_usado else metodo_usado
            metodo_exibicao = mapeamento_metodos.get(chave_metodo, mapeamento_metodos.get(metodo_usado, chave_metodo.upper()))

            tem_algum_ajuste = any(p.get('ajuste_manual') for p in previsao_item_periodos)
            if tem_algum_ajuste:
                metodo_exibicao = 'Ajuste Manual'

            # Variacao percentual
            variacao_pct = 0
            if total_ano_anterior_item > 0:
                variacao_pct = ((total_previsao_item - total_ano_anterior_item) / total_ano_anterior_item) * 100

            # CV e sinal de alerta
            num_periodos_prev = len(previsao_item_periodos) if previsao_item_periodos else 1
            media_previsao_mensal = total_previsao_item / num_periodos_prev if num_periodos_prev > 0 else 0
            desvio_mensal = desvio_diario_base * math.sqrt(30)
            cv = desvio_mensal / media_previsao_mensal if media_previsao_mensal > 0 else 0

            if media_previsao_mensal == 0:
                sinal_emoji = '⚪'
            elif cv > 0.80:
                sinal_emoji = '🔴'
            elif cv > 0.50:
                sinal_emoji = '🟡'
            elif cv > 0.30:
                sinal_emoji = '🔵'
            else:
                sinal_emoji = '🟢'

            sinal_alerta_map = {
                '⚪': 'cinza', '🔴': 'vermelho', '🟡': 'amarelo',
                '🔵': 'azul', '🟢': 'verde'
            }
            sinal_alerta = sinal_alerta_map.get(sinal_emoji, 'cinza')

            relatorio_itens.append({
                'cod_produto': cod_produto,
                'cnpj_fornecedor': item.get('cnpj_fornecedor', ''),
                'descricao': item['descricao'],
                'nome_fornecedor': item['nome_fornecedor'] or 'SEM FORNECEDOR',
                'categoria': item['categoria'],
                'situacao_compra': item.get('situacao_compra', 'ATIVO'),
                'demanda_prevista_total': total_previsao_item,
                'demanda_ano_anterior': round(total_ano_anterior_item, 2),
                'variacao_percentual': round(variacao_pct, 1),
                'sinal_emoji': sinal_emoji,
                'sinal_alerta': sinal_alerta,
                'cv': round(cv, 4),
                'desvio_padrao': round(desvio_mensal, 2),
                'metodo_estatistico': metodo_exibicao,
                'fator_tendencia_yoy': round(fator_tendencia_yoy, 4),
                'classificacao_tendencia': classificacao_tendencia,
                'previsao_por_periodo': previsao_item_periodos,
                'cue': round(cue_por_item.get(cod_produto, 0), 4)
            })

            if metodo_exibicao not in contagem_modelos:
                contagem_modelos[metodo_exibicao] = 0
            contagem_modelos[metodo_exibicao] += 1

        # =====================================================
        # CALCULAR BACKTEST - USAR ANO ANTERIOR + TENDENCIA
        # Conforme app_original.py - abordagem mais realista
        # =====================================================
        from dateutil.relativedelta import relativedelta

        # Calcular fatores sazonais para o grafico
        fatores_sazonais = {}
        if len(serie_historica_agregada) >= 12:
            vendas_por_chave_saz = {}
            for i, data_str in enumerate(datas_historicas):
                try:
                    if '-S' in data_str:
                        chave_saz = int(data_str.split('-S')[1])
                    else:
                        data_obj = dt.strptime(data_str, '%Y-%m-%d')
                        if granularidade == 'mensal':
                            chave_saz = data_obj.month
                        else:
                            chave_saz = data_obj.isocalendar()[1]
                    if chave_saz not in vendas_por_chave_saz:
                        vendas_por_chave_saz[chave_saz] = []
                    vendas_por_chave_saz[chave_saz].append(serie_historica_agregada[i])
                except:
                    pass
            media_geral_saz = sum(serie_historica_agregada) / len(serie_historica_agregada) if serie_historica_agregada else 1
            for chave, valores_saz in vendas_por_chave_saz.items():
                media_chave = sum(valores_saz) / len(valores_saz)
                fatores_sazonais[chave] = media_chave / media_geral_saz if media_geral_saz > 0 else 1.0

        # =====================================================
        # BACKTEST - Mesma metodologia do app_original.py
        # Para cada periodo de teste: buscar mesmo periodo do ano anterior
        # e aplicar fator de tendencia observado na base
        # =====================================================
        previsao_teste = []
        wmape_backtest = 0
        bias_backtest = 0

        if len(datas_teste) > 0 and len(valores_base) >= 3:
            # Criar indice de vendas agregadas por periodo para lookup rapido
            vendas_por_data = {d: v for d, v in zip(datas_historicas, serie_historica_agregada)}

            # =====================================================
            # CALCULAR FATOR DE TENDENCIA REAL (ano atual vs ano anterior)
            # Comparar periodos equivalentes dentro da base para detectar
            # crescimento ou queda real entre os anos
            # =====================================================
            soma_valores_ano_recente = 0
            soma_valores_ano_anterior = 0
            pares_encontrados = 0

            for i, data_str in enumerate(datas_base):
                try:
                    if '-S' in data_str:
                        ano_str, sem_str = data_str.split('-S')
                        ano = int(ano_str)
                        chave = int(sem_str)
                        data_aa_str = f"{ano - 1}-S{chave:02d}"
                    else:
                        data_obj = dt.strptime(data_str, '%Y-%m-%d')
                        if granularidade == 'mensal':
                            data_aa_str = f"{data_obj.year - 1:04d}-{data_obj.month:02d}-01"
                        else:
                            data_aa_str = f"{data_obj.year - 1}-{data_obj.month:02d}-{data_obj.day:02d}"

                    # Se encontrar o periodo correspondente do ano anterior na base
                    valor_atual = valores_base[i]
                    valor_aa = vendas_por_data.get(data_aa_str, 0)

                    if valor_aa > 0 and valor_atual > 0:
                        soma_valores_ano_recente += valor_atual
                        soma_valores_ano_anterior += valor_aa
                        pares_encontrados += 1
                except:
                    pass

            # Calcular fator de tendencia baseado em periodos equivalentes
            if pares_encontrados >= 3 and soma_valores_ano_anterior > 0:
                fator_tendencia_geral = soma_valores_ano_recente / soma_valores_ano_anterior
                # Limitar entre 0.6 (-40%) e 1.5 (+50%)
                fator_tendencia_geral = max(0.6, min(1.5, fator_tendencia_geral))
            else:
                # Fallback: comparar metades da base
                meio_base = len(valores_base) // 2
                if meio_base > 0:
                    media_primeira_metade = sum(valores_base[:meio_base]) / meio_base
                    media_segunda_metade = sum(valores_base[meio_base:]) / (len(valores_base) - meio_base)
                    if media_primeira_metade > 0:
                        fator_tendencia_geral = media_segunda_metade / media_primeira_metade
                        fator_tendencia_geral = max(0.6, min(1.5, fator_tendencia_geral))
                    else:
                        fator_tendencia_geral = 1.0
                else:
                    fator_tendencia_geral = 1.0

            for data_teste_str in datas_teste:
                previsao_periodo = 0

                try:
                    # Determinar chave sazonal e buscar periodo correspondente do ano anterior
                    if '-S' in data_teste_str:
                        # Formato semanal: YYYY-SWW
                        ano_str, sem_str = data_teste_str.split('-S')
                        chave_teste = int(sem_str)
                        ano_teste = int(ano_str)
                        # Buscar mesma semana do ano anterior
                        ano_anterior = ano_teste - 1
                        data_aa_str = f"{ano_anterior}-S{chave_teste:02d}"
                    else:
                        data_obj = dt.strptime(data_teste_str, '%Y-%m-%d')
                        if granularidade == 'mensal':
                            chave_teste = data_obj.month
                            # Buscar mesmo mes do ano anterior (formato: YYYY-MM-01)
                            data_aa_str = f"{data_obj.year - 1:04d}-{data_obj.month:02d}-01"
                        elif granularidade == 'semanal':
                            chave_teste = data_obj.isocalendar()[1]
                            # Buscar mesma semana do ano anterior
                            ano_aa = data_obj.year - 1
                            data_aa_str = f"{ano_aa}-S{chave_teste:02d}"
                        else:
                            chave_teste = data_obj.weekday()
                            data_aa_str = f"{data_obj.year - 1}-{data_obj.month:02d}-{data_obj.day:02d}"

                    # Tentar encontrar valor do ano anterior
                    valor_aa = vendas_por_data.get(data_aa_str, 0)

                    if valor_aa > 0:
                        # Usar valor do ano anterior com ajuste de tendencia
                        previsao_periodo = valor_aa * fator_tendencia_geral
                    else:
                        # Fallback: usar fator sazonal agregado
                        media_base = sum(valores_base) / len(valores_base)
                        fator = fatores_sazonais.get(chave_teste, 1.0)
                        previsao_periodo = media_base * fator

                except Exception as e:
                    # Fallback simples
                    media_base = sum(valores_base) / len(valores_base)
                    previsao_periodo = media_base

                previsao_teste.append(round(previsao_periodo, 2))

        # =====================================================
        # CALCULAR WMAPE e BIAS
        # WMAPE (Weighted MAPE) = soma(|erro|) / soma(|real|) * 100
        # BIAS = soma(erro com sinal) / soma(real) * 100
        # =====================================================
        if previsao_teste and valores_teste:
            soma_erros_abs = 0
            soma_erros_signed = 0
            soma_reais = 0

            for i in range(min(len(previsao_teste), len(valores_teste))):
                real = valores_teste[i]
                previsto = previsao_teste[i]
                erro = previsto - real
                if real > 0:
                    soma_erros_abs += abs(erro)
                    soma_erros_signed += erro
                    soma_reais += real

            if soma_reais > 0:
                wmape_backtest = (soma_erros_abs / soma_reais) * 100
                bias_backtest = (soma_erros_signed / soma_reais) * 100

        # Fallback se nao calculou backtest
        if not previsao_teste and len(valores_base) >= 3 and len(datas_teste) > 0:
            media_base_fb = sum(valores_base) / len(valores_base)
            previsao_teste = [round(media_base_fb, 2)] * len(datas_teste)

        # =====================================================
        # PREVISAO FUTURA - SOMA DAS PREVISOES POR ITEM (Bottom-Up)
        # A previsao ja foi calculada por item usando DemandCalculator
        # Aqui apenas agregamos os valores por periodo
        # =====================================================
        previsao_agregada = []

        # A previsao ja foi calculada na secao anterior (por item)
        # Apenas arredondar os valores agregados
        for data_prev in datas_previsao:
            previsao_base = previsoes_agregadas_por_periodo.get(data_prev, 0)
            previsao_agregada.append(round(previsao_base, 2))

        # Buscar ano anterior
        ano_anterior_valores = []
        ano_anterior_datas = []
        ano_anterior_total = 0

        for data_prev in datas_previsao:
            try:
                if '-S' in data_prev:
                    ano_str, sem_str = data_prev.split('-S')
                    ano_anterior_str = f"{int(ano_str)-1}-S{sem_str}"
                else:
                    data_obj = dt.strptime(data_prev, '%Y-%m-%d')
                    data_aa = data_obj.replace(year=data_obj.year - 1)
                    if granularidade == 'mensal':
                        ano_anterior_str = f"{data_aa.year:04d}-{data_aa.month:02d}-01"
                    else:
                        ano_anterior_str = data_aa.strftime('%Y-%m-%d')

                valor_aa = vendas_agregadas_por_periodo.get(ano_anterior_str, 0)
                ano_anterior_valores.append(valor_aa)
                ano_anterior_datas.append(ano_anterior_str)
                ano_anterior_total += valor_aa
            except:
                ano_anterior_valores.append(0)
                ano_anterior_datas.append('')

        # Formatar periodos
        periodos_formatados = []
        for data_str in datas_previsao:
            try:
                if '-S' in data_str:
                    ano_str, sem_str = data_str.split('-S')
                    periodos_formatados.append({'semana': int(sem_str), 'ano': int(ano_str)})
                else:
                    data_obj = dt.strptime(data_str, '%Y-%m-%d')
                    if granularidade == 'mensal':
                        periodos_formatados.append({'mes': data_obj.month, 'ano': data_obj.year})
                    else:
                        periodos_formatados.append(data_str)
            except:
                periodos_formatados.append(data_str)

        # =====================================================
        # AGREGAR POR FORNECEDOR - Comparacao YoY por Fornecedor
        # =====================================================
        # Funcao auxiliar para formatar periodo igual ao frontend
        meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

        def formatar_periodo_display(periodo_str, gran):
            """Formata periodo para exibicao (igual ao frontend)"""
            try:
                if gran == 'semanal':
                    if '-S' in periodo_str:
                        ano, sem = periodo_str.split('-S')
                        return f"S{int(sem)}/{ano}"
                    else:
                        data_obj = dt.strptime(periodo_str, '%Y-%m-%d')
                        semana = data_obj.isocalendar()[1]
                        return f"S{semana}/{data_obj.year}"
                elif gran == 'diario' or gran == 'diaria':
                    data_obj = dt.strptime(periodo_str, '%Y-%m-%d')
                    return f"{data_obj.day:02d}/{data_obj.month:02d}/{data_obj.year}"
                else:  # mensal
                    data_obj = dt.strptime(periodo_str, '%Y-%m-%d')
                    return f"{meses_nomes[data_obj.month - 1]}/{data_obj.year}"
            except:
                return periodo_str

        comparacao_por_fornecedor = {}
        for item in relatorio_itens:
            nome_forn = item.get('nome_fornecedor') or 'SEM FORNECEDOR'

            if nome_forn not in comparacao_por_fornecedor:
                comparacao_por_fornecedor[nome_forn] = {
                    'nome_fornecedor': nome_forn,
                    'previsao_total': 0,
                    'ano_anterior_total': 0,
                    'previsao_por_periodo': {},
                    'ano_anterior_por_periodo': {}
                }

            comparacao_por_fornecedor[nome_forn]['previsao_total'] += item.get('demanda_prevista_total', 0)
            comparacao_por_fornecedor[nome_forn]['ano_anterior_total'] += item.get('demanda_ano_anterior', 0)

            # Agregar por periodo (usando formato de exibicao para matching com frontend)
            if item.get('previsao_por_periodo'):
                for p in item['previsao_por_periodo']:
                    periodo_raw = p.get('periodo')
                    if periodo_raw:
                        # Formatar periodo igual ao frontend para matching correto
                        periodo = formatar_periodo_display(periodo_raw, granularidade)
                        if periodo not in comparacao_por_fornecedor[nome_forn]['previsao_por_periodo']:
                            comparacao_por_fornecedor[nome_forn]['previsao_por_periodo'][periodo] = 0
                            comparacao_por_fornecedor[nome_forn]['ano_anterior_por_periodo'][periodo] = 0
                        comparacao_por_fornecedor[nome_forn]['previsao_por_periodo'][periodo] += p.get('previsao', 0)
                        comparacao_por_fornecedor[nome_forn]['ano_anterior_por_periodo'][periodo] += p.get('ano_anterior', 0)

        # Calcular variacao percentual e ordenar
        comparacao_yoy_por_fornecedor = []
        for nome_forn, dados in comparacao_por_fornecedor.items():
            prev = dados['previsao_total']
            ant = dados['ano_anterior_total']
            variacao = ((prev - ant) / ant * 100) if ant > 0 else 0

            comparacao_yoy_por_fornecedor.append({
                'nome_fornecedor': nome_forn,
                'previsao_total': round(prev, 2),
                'ano_anterior_total': round(ant, 2),
                'variacao_percentual': round(variacao, 1),
                'previsao_por_periodo': dados['previsao_por_periodo'],
                'ano_anterior_por_periodo': dados['ano_anterior_por_periodo']
            })

        # Ordenar por previsao total (descendente)
        comparacao_yoy_por_fornecedor = sorted(
            comparacao_yoy_por_fornecedor,
            key=lambda x: x['previsao_total'],
            reverse=True
        )

        print(f"  Comparacao YoY montada para {len(comparacao_yoy_por_fornecedor)} fornecedores", flush=True)

        # Montar resposta
        resposta = {
            'sucesso': True,
            'versao': 'v2_bottom_up',
            'granularidade': granularidade,
            'filtros': {
                'loja': loja,
                'fornecedor': fornecedor,
                'linha': linha,
                'sublinha': sublinha,
                'produto': produto
            },
            'serie_temporal': {
                'datas': datas_historicas,
                'valores': serie_historica_agregada
            },
            'historico_base': {
                'datas': datas_base,
                'valores': valores_base
            },
            'historico_teste': {
                'datas': datas_teste,
                'valores': valores_teste
            },
            'melhor_modelo': 'Bottom-Up (Individual por Item)',
            'modelos': {
                'Bottom-Up (Individual por Item)': {
                    'metricas': {
                        'wmape': round(wmape_backtest, 2),
                        'bias': round(bias_backtest, 2)
                    },
                    'teste': {
                        'datas': datas_teste,
                        'valores': previsao_teste
                    },
                    'futuro': {
                        'datas': datas_previsao,
                        'valores': [round(v, 2) for v in previsao_agregada]
                    }
                }
            },
            'ano_anterior': {
                'datas': ano_anterior_datas,
                'valores': ano_anterior_valores,
                'total': ano_anterior_total
            },
            'relatorio_detalhado': {
                'itens': relatorio_itens,  # Todos os itens para consistencia com totais
                'periodos_previsao': periodos_formatados,
                'total_itens': len(relatorio_itens),
                'granularidade': granularidade
            },
            'estatisticas_modelos': contagem_modelos,
            'periodos_previsao_formatados': datas_previsao,
            'periodos_previsao_display': [formatar_periodo_display(d, granularidade) for d in datas_previsao],
            'comparacao_yoy_por_fornecedor': comparacao_yoy_por_fornecedor
        }

        cursor.close()
        conn.close()

        return jsonify(resposta)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@previsao_bp.route('/api/demanda/salvar_tela', methods=['POST'])
def api_salvar_demanda_tela():
    """
    Salva os valores exibidos na Tela de Demanda diretamente em demanda_pre_calculada.

    Recebe os dados ja calculados (relatorio_detalhado.itens) do frontend e grava
    cada item/periodo como ajuste_manual, substituindo qualquer valor anterior
    (inclusive o calculado pelo cronjob). O cronjob preserva registros com
    ajuste_manual IS NOT NULL.

    Request JSON:
    {
        "itens": [
            {
                "cod_produto": "123456",
                "cnpj_fornecedor": "61413282000143",
                "previsao_por_periodo": [
                    {"periodo": "2026-07-01", "previsao": 350},
                    ...
                ]
            },
            ...
        ],
        "usuario": "valter.lino"
    }
    """
    try:
        from app.utils.db_connection import get_db_connection

        dados = request.get_json()
        itens = dados.get('itens', [])
        usuario = dados.get('usuario', 'tela_demanda')

        if not itens:
            return jsonify({'success': False, 'erro': 'Nenhum item recebido'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        total_registros = 0
        total_itens = 0

        for item in itens:
            cod_produto = str(item.get('cod_produto', ''))
            cnpj_fornecedor = str(item.get('cnpj_fornecedor', ''))
            periodos = item.get('previsao_por_periodo', [])

            if not cod_produto or not cnpj_fornecedor or not periodos:
                continue

            for pp in periodos:
                periodo_str = pp.get('periodo', '')  # '2026-07-01' (mensal) ou '2026-S11' (semanal)
                previsao_val = pp.get('previsao', 0)

                if not periodo_str:
                    continue

                # V50: Detectar formato semanal vs mensal
                is_semanal = '-S' in periodo_str

                if is_semanal:
                    # Formato semanal: '2026-S11' -> ano=2026, semana=11
                    try:
                        partes = periodo_str.split('-S')
                        ano = int(partes[0])
                        semana = int(partes[1])
                    except (ValueError, IndexError):
                        continue

                    # Calcular segunda-feira da semana ISO
                    from datetime import datetime as dt_aux
                    try:
                        data_inicio_semana = dt_aux.strptime(f'{ano}-W{semana:02d}-1', '%G-W%V-%u').date()
                    except ValueError:
                        continue

                    divisor = 7.0

                    # V56: Verificar se valor eh diferente do atual
                    cursor.execute("""
                        SELECT demanda_prevista FROM demanda_pre_calculada
                        WHERE cod_produto = %s AND cnpj_fornecedor = %s
                          AND cod_empresa IS NULL AND ano = %s AND semana = %s
                          AND tipo_granularidade = 'semanal'
                    """, (cod_produto, cnpj_fornecedor, ano, semana))
                    row_sem = cursor.fetchone()
                    valor_atual_sem = float(row_sem[0]) if row_sem and row_sem[0] is not None else None

                    if row_sem:
                        if valor_atual_sem is not None and round(float(previsao_val)) == round(valor_atual_sem):
                            continue  # Valor nao mudou
                        else:
                            cursor.execute("""
                                UPDATE demanda_pre_calculada
                                SET demanda_prevista      = %s,
                                    demanda_diaria_base   = %s,
                                    ajuste_manual         = %s,
                                    ajuste_manual_data    = NOW(),
                                    ajuste_manual_usuario = %s,
                                    ajuste_manual_motivo  = 'salvo_tela_demanda',
                                    data_calculo          = NOW()
                                WHERE cod_produto = %s
                                  AND cnpj_fornecedor = %s
                                  AND cod_empresa IS NULL
                                  AND ano = %s
                                  AND semana = %s
                                  AND tipo_granularidade = 'semanal'
                            """, (
                                float(previsao_val), float(previsao_val) / divisor,
                                float(previsao_val), usuario,
                                cod_produto, cnpj_fornecedor, ano, semana
                            ))
                    else:
                        cursor.execute("""
                            INSERT INTO demanda_pre_calculada (
                                cod_produto, cnpj_fornecedor, cod_empresa, ano, mes,
                                semana, tipo_granularidade, data_inicio_semana,
                                demanda_prevista, demanda_diaria_base,
                                ajuste_manual, ajuste_manual_data,
                                ajuste_manual_usuario, ajuste_manual_motivo,
                                data_calculo
                            ) VALUES (
                                %s, %s, NULL, %s, NULL,
                                %s, 'semanal', %s,
                                %s, %s,
                                %s, NOW(), %s, 'salvo_tela_demanda',
                                NOW()
                            )
                        """, (
                            cod_produto, cnpj_fornecedor, ano,
                            semana, data_inicio_semana,
                            float(previsao_val), float(previsao_val) / divisor,
                            float(previsao_val), usuario
                        ))
                else:
                    # Formato mensal: '2026-07-01' ou '2026-07'
                    try:
                        ano = int(periodo_str[:4])
                        mes = int(periodo_str[5:7])
                    except (ValueError, IndexError):
                        continue

                    # V56: Verificar se valor eh diferente do demanda_prevista atual
                    # So marcar ajuste_manual se usuario realmente alterou o valor
                    cursor.execute("""
                        SELECT demanda_prevista FROM demanda_pre_calculada
                        WHERE cod_produto = %s AND cnpj_fornecedor = %s
                          AND cod_empresa IS NULL AND ano = %s AND mes = %s
                          AND tipo_granularidade = 'mensal'
                    """, (cod_produto, cnpj_fornecedor, ano, mes))
                    row_existente = cursor.fetchone()
                    valor_atual = float(row_existente[0]) if row_existente and row_existente[0] is not None else None

                    if row_existente:
                        if valor_atual is not None and round(float(previsao_val)) == round(valor_atual):
                            # Valor nao mudou - nao marcar como ajuste manual
                            # Se ja tinha ajuste_manual igual ao previsto, limpar
                            continue
                        else:
                            # Valor mudou - marcar como ajuste manual
                            cursor.execute("""
                                UPDATE demanda_pre_calculada
                                SET demanda_prevista      = %s,
                                    demanda_diaria_base   = %s,
                                    ajuste_manual         = %s,
                                    ajuste_manual_data    = NOW(),
                                    ajuste_manual_usuario = %s,
                                    ajuste_manual_motivo  = 'salvo_tela_demanda',
                                    data_calculo          = NOW()
                                WHERE cod_produto = %s
                                  AND cnpj_fornecedor = %s
                                  AND cod_empresa IS NULL
                                  AND ano = %s
                                  AND mes = %s
                                  AND tipo_granularidade = 'mensal'
                            """, (
                                float(previsao_val), float(previsao_val) / 30.0,
                                float(previsao_val), usuario,
                                cod_produto, cnpj_fornecedor, ano, mes
                            ))
                    else:
                        # Registro nao existe - inserir com ajuste
                        cursor.execute("""
                            INSERT INTO demanda_pre_calculada (
                                cod_produto, cnpj_fornecedor, cod_empresa, ano, mes,
                                tipo_granularidade,
                                demanda_prevista, demanda_diaria_base,
                                ajuste_manual, ajuste_manual_data,
                                ajuste_manual_usuario, ajuste_manual_motivo,
                                data_calculo
                            ) VALUES (
                                %s, %s, NULL, %s, %s,
                                'mensal',
                                %s, %s,
                                %s, NOW(), %s, 'salvo_tela_demanda',
                                NOW()
                            )
                        """, (
                            cod_produto, cnpj_fornecedor, ano, mes,
                            float(previsao_val), float(previsao_val) / 30.0,
                            float(previsao_val), usuario
                        ))

                total_registros += 1

            total_itens += 1

        conn.commit()
        conn.close()

        print(f"[SalvarTela] {total_itens} itens, {total_registros} registros salvos por {usuario}")

        return jsonify({
            'success': True,
            'total_itens': total_itens,
            'total_registros': total_registros,
            'mensagem': f'{total_itens} itens salvos ({total_registros} registros)'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@previsao_bp.route('/api/gerar_previsao_banco_v2', methods=['POST'])
def api_gerar_previsao_banco_v2():
    """
    Versao 2 da API de previsao: Bottom-Up (Loja -> Produto).
    Gera previsao por loja/produto e agrega para fornecedor.

    Parametros JSON:
        - loja / cod_empresa: Codigo da loja ou 'TODAS'
        - fornecedor: Nome do fornecedor ou 'TODOS'
        - linha / linha1: Categoria (Linha 1)
        - sublinha / linha3: Sublinha (Linha 3)
        - granularidade: 'mensal', 'semanal' ou 'diario'
        - periodos: Numero de periodos de previsao
        - mes_inicio, ano_inicio, mes_fim, ano_fim: Para granularidade mensal
        - semana_inicio, semana_fim, ano_semana_inicio, ano_semana_fim: Para semanal
        - data_inicio, data_fim: Para granularidade diaria
    """
    try:
        from core.previsao_v2 import PrevisaoV2Engine

        dados = request.get_json()

        # Parametros de filtro (aceita ambos os nomes para compatibilidade)
        cod_empresa = dados.get('cod_empresa') or dados.get('loja', 'TODAS')
        fornecedor = dados.get('fornecedor', 'TODOS')
        linha1 = dados.get('linha1') or dados.get('linha', 'TODAS')
        linha3 = dados.get('linha3') or dados.get('sublinha', 'TODAS')

        # Parametros de previsao
        granularidade = dados.get('granularidade', 'mensal')
        periodos = dados.get('periodos', 6)

        # Parametros de periodo - mensal
        mes_inicio = dados.get('mes_inicio')
        ano_inicio = dados.get('ano_inicio')
        mes_fim = dados.get('mes_fim')
        ano_fim = dados.get('ano_fim')

        # Parametros de periodo - semanal
        semana_inicio = dados.get('semana_inicio')
        semana_fim = dados.get('semana_fim')
        ano_semana_inicio = dados.get('ano_semana_inicio')
        ano_semana_fim = dados.get('ano_semana_fim')

        # Parametros de periodo - diario
        data_inicio = dados.get('data_inicio')
        data_fim = dados.get('data_fim')

        # Criar engine e gerar previsao
        engine = PrevisaoV2Engine()

        resultado = engine.gerar_previsao_bottom_up(
            fornecedor=fornecedor,
            cod_empresa=cod_empresa,
            linha1=linha1,
            linha3=linha3,
            granularidade=granularidade,
            periodos=periodos,
            mes_inicio=mes_inicio,
            ano_inicio=ano_inicio,
            mes_fim=mes_fim,
            ano_fim=ano_fim,
            semana_inicio=semana_inicio,
            semana_fim=semana_fim,
            ano_semana_inicio=ano_semana_inicio,
            ano_semana_fim=ano_semana_fim,
            data_inicio=data_inicio,
            data_fim=data_fim
        )

        # Verificar se houve erro
        if 'erro' in resultado:
            return jsonify({
                'success': False,
                'erro': resultado['erro']
            }), 400

        return jsonify(resultado)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@previsao_bp.route('/api/exportar_tabela_comparativa', methods=['POST'])
def api_exportar_tabela_comparativa():
    """Exporta tabela comparativa YoY para Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        dados = request.get_json()

        if not dados:
            return jsonify({'success': False, 'erro': 'Dados nao fornecidos'}), 400

        # Suportar dois formatos: novo (periodos, previsao, real) e antigo (dados)
        if 'periodos' in dados:
            # Novo formato V2 do frontend
            periodos = dados.get('periodos', [])
            previsao = dados.get('previsao', [])
            real = dados.get('real', [])
            variacao = dados.get('variacao', [])
            granularidade = dados.get('granularidade', 'mensal')
            total_previsao = dados.get('total_previsao', 0)
            total_real = dados.get('total_real', 0)
            variacao_total = dados.get('variacao_total', 0)
            filtros = dados.get('filtros', {})
            fornecedores = dados.get('fornecedores', [])

            if not periodos or not previsao:
                return jsonify({'success': False, 'erro': 'Periodos ou previsao vazios'}), 400

            wb = Workbook()
            ws = wb.active
            ws.title = 'Comparativo YoY'

            # Estilos
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
            total_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            row_num = 1

            # Informacoes de filtros (se houver)
            if filtros:
                ws.cell(row=row_num, column=1, value='Filtros Aplicados:').font = Font(bold=True)
                row_num += 1
                for chave, valor in filtros.items():
                    ws.cell(row=row_num, column=1, value=f'  {chave.title()}: {valor}')
                    row_num += 1
                row_num += 1

            # ===== TABELA CONSOLIDADA =====
            ws.cell(row=row_num, column=1, value='TABELA COMPARATIVA - CONSOLIDADO').font = Font(bold=True, size=12)
            row_num += 1

            # Cabecalho: Tipo | Periodo1 | Periodo2 | ... | TOTAL | Var %
            header_row = row_num
            ws.cell(row=header_row, column=1, value='Tipo').font = header_font
            ws.cell(row=header_row, column=1).fill = header_fill
            ws.cell(row=header_row, column=1).border = border

            for col_idx, periodo in enumerate(periodos, start=2):
                cell = ws.cell(row=header_row, column=col_idx, value=periodo)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            col_total = len(periodos) + 2
            col_var = len(periodos) + 3

            ws.cell(row=header_row, column=col_total, value='TOTAL').font = header_font
            ws.cell(row=header_row, column=col_total).fill = header_fill
            ws.cell(row=header_row, column=col_total).border = border

            ws.cell(row=header_row, column=col_var, value='Var %').font = header_font
            ws.cell(row=header_row, column=col_var).fill = header_fill
            ws.cell(row=header_row, column=col_var).border = border

            # Linha de Previsao
            row_num += 1
            ws.cell(row=row_num, column=1, value='Previsao').border = border
            for col_idx, valor in enumerate(previsao, start=2):
                cell = ws.cell(row=row_num, column=col_idx, value=valor)
                cell.border = border
                cell.number_format = '#,##0'
            ws.cell(row=row_num, column=col_total, value=total_previsao).border = border
            ws.cell(row=row_num, column=col_total).number_format = '#,##0'

            # Linha de Ano Anterior
            row_num += 1
            ws.cell(row=row_num, column=1, value='Ano Ant.').border = border
            for col_idx, valor in enumerate(real, start=2):
                cell = ws.cell(row=row_num, column=col_idx, value=valor)
                cell.border = border
                cell.number_format = '#,##0'
            ws.cell(row=row_num, column=col_total, value=total_real).border = border
            ws.cell(row=row_num, column=col_total).number_format = '#,##0'

            # Linha de Variacao
            row_num += 1
            ws.cell(row=row_num, column=1, value='Var %').border = border
            for col_idx, valor in enumerate(variacao, start=2):
                cell = ws.cell(row=row_num, column=col_idx, value=f'{valor:.1f}%' if valor else '-')
                cell.border = border
            ws.cell(row=row_num, column=col_var, value=f'{variacao_total:.1f}%').border = border

            row_num += 2

            # ===== DETALHAMENTO POR FORNECEDOR (se houver) =====
            # Debug: log quantidade de fornecedores
            print(f"[Exportar] Fornecedores recebidos: {len(fornecedores) if fornecedores else 0}")
            if fornecedores:
                for f in fornecedores:
                    print(f"[Exportar]   - {f.get('nome_fornecedor', 'SEM NOME')}: {f.get('previsao_total', 0)}")

            if fornecedores and len(fornecedores) >= 1:
                ws.cell(row=row_num, column=1, value='DETALHAMENTO POR FORNECEDOR').font = Font(bold=True, size=12)
                row_num += 2

                for forn in fornecedores:
                    nome_forn = forn.get('nome_fornecedor', 'SEM NOME')
                    previsao_total_forn = forn.get('previsao_total', 0)
                    ano_ant_total_forn = forn.get('ano_anterior_total', 0)
                    variacao_forn = forn.get('variacao_percentual', 0)
                    previsao_por_periodo = forn.get('previsao_por_periodo', {})
                    ano_ant_por_periodo = forn.get('ano_anterior_por_periodo', {})

                    # Nome do fornecedor como titulo
                    ws.cell(row=row_num, column=1, value=nome_forn).font = Font(bold=True, size=11)
                    ws.cell(row=row_num, column=1).fill = total_fill
                    # Mesclar celula do nome ou deixar mais largo
                    row_num += 1

                    # Cabecalho do fornecedor
                    ws.cell(row=row_num, column=1, value='Tipo').font = header_font
                    ws.cell(row=row_num, column=1).fill = header_fill
                    ws.cell(row=row_num, column=1).border = border

                    for col_idx, periodo in enumerate(periodos, start=2):
                        cell = ws.cell(row=row_num, column=col_idx, value=periodo)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                    ws.cell(row=row_num, column=col_total, value='TOTAL').font = header_font
                    ws.cell(row=row_num, column=col_total).fill = header_fill
                    ws.cell(row=row_num, column=col_total).border = border

                    ws.cell(row=row_num, column=col_var, value='Var %').font = header_font
                    ws.cell(row=row_num, column=col_var).fill = header_fill
                    ws.cell(row=row_num, column=col_var).border = border

                    # Linha de Previsao do fornecedor
                    row_num += 1
                    ws.cell(row=row_num, column=1, value='Previsao').border = border
                    for col_idx, periodo in enumerate(periodos, start=2):
                        valor = previsao_por_periodo.get(periodo, 0)
                        cell = ws.cell(row=row_num, column=col_idx, value=round(valor) if valor else 0)
                        cell.border = border
                        cell.number_format = '#,##0'
                    ws.cell(row=row_num, column=col_total, value=round(previsao_total_forn)).border = border
                    ws.cell(row=row_num, column=col_total).number_format = '#,##0'

                    # Linha de Ano Anterior do fornecedor
                    row_num += 1
                    ws.cell(row=row_num, column=1, value='Ano Ant.').border = border
                    for col_idx, periodo in enumerate(periodos, start=2):
                        valor = ano_ant_por_periodo.get(periodo, 0)
                        cell = ws.cell(row=row_num, column=col_idx, value=round(valor) if valor else 0)
                        cell.border = border
                        cell.number_format = '#,##0'
                    ws.cell(row=row_num, column=col_total, value=round(ano_ant_total_forn)).border = border
                    ws.cell(row=row_num, column=col_total).number_format = '#,##0'

                    # Linha de Variacao do fornecedor
                    row_num += 1
                    ws.cell(row=row_num, column=1, value='Var %').border = border
                    for col_idx, periodo in enumerate(periodos, start=2):
                        prev = previsao_por_periodo.get(periodo, 0)
                        ant = ano_ant_por_periodo.get(periodo, 0)
                        var_periodo = ((prev - ant) / ant * 100) if ant > 0 else 0
                        cell = ws.cell(row=row_num, column=col_idx, value=f'{var_periodo:.1f}%' if ant > 0 else '-')
                        cell.border = border
                    ws.cell(row=row_num, column=col_var, value=f'{variacao_forn:.1f}%').border = border

                    row_num += 2

            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 15
            for col_idx in range(2, col_var + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 12

            ws.freeze_panes = 'B1'

        else:
            # Formato antigo (compatibilidade)
            if 'dados' not in dados:
                return jsonify({'success': False, 'erro': 'Dados nao fornecidos no formato esperado'}), 400

            dados_tabela = dados['dados']

            if not dados_tabela:
                return jsonify({'success': False, 'erro': 'Tabela vazia'}), 400

            wb = Workbook()
            ws = wb.active
            ws.title = 'Comparativo YoY'

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Cabecalho
            headers = ['Mes', 'Ano Anterior', 'Previsao Atual', 'Variacao (%)']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Dados
            for row_idx, item in enumerate(dados_tabela, start=2):
                ws.cell(row=row_idx, column=1, value=item.get('mes_nome', '')).border = border
                ws.cell(row=row_idx, column=2, value=item.get('demanda_ano_anterior', 0)).border = border
                ws.cell(row=row_idx, column=2).number_format = '#,##0.0'
                ws.cell(row=row_idx, column=3, value=item.get('previsao_atual', 0)).border = border
                ws.cell(row=row_idx, column=3).number_format = '#,##0.0'
                variacao = item.get('variacao_percentual')
                ws.cell(row=row_idx, column=4, value=variacao if variacao is not None else '-').border = border
                if variacao is not None:
                    ws.cell(row=row_idx, column=4).number_format = '+0.0%;-0.0%;0%'

            col_widths = [15, 18, 18, 15]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width

            ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'comparativo_yoy_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


# =====================================================
# API DE AJUSTES MANUAIS DE PREVISAO
# =====================================================

@previsao_bp.route('/api/ajuste_previsao/salvar', methods=['POST'])
def api_salvar_ajuste_previsao():
    """
    Salva ajustes manuais de previsao.
    Recebe uma lista de ajustes por item/periodo.
    """
    try:
        dados = request.get_json()

        if not dados or 'ajustes' not in dados:
            return jsonify({'success': False, 'erro': 'Dados de ajuste nao fornecidos'}), 400

        ajustes = dados.get('ajustes', [])
        motivo = dados.get('motivo', None)
        usuario = dados.get('usuario', 'sistema')

        if not ajustes:
            return jsonify({'success': False, 'erro': 'Nenhum ajuste fornecido'}), 400

        # Conectar ao banco
        from app.utils.db_connection import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()

        ajustes_salvos = 0

        for ajuste in ajustes:
            cod_produto = ajuste.get('cod_produto')
            periodo = ajuste.get('periodo')
            granularidade = ajuste.get('granularidade', 'mensal')
            valor_original = ajuste.get('valor_original', 0)
            valor_ajustado = ajuste.get('valor_ajustado', 0)
            cod_fornecedor = ajuste.get('cod_fornecedor')
            nome_fornecedor = ajuste.get('nome_fornecedor')
            metodo_estatistico = ajuste.get('metodo_estatistico')

            if not cod_produto or not periodo:
                continue

            # Calcular diferenca e percentual
            diferenca = valor_ajustado - valor_original
            percentual_ajuste = 0
            if valor_original > 0:
                percentual_ajuste = ((valor_ajustado - valor_original) / valor_original) * 100

            # Inserir ajuste (trigger desativa ajustes antigos automaticamente)
            cursor.execute("""
                INSERT INTO ajuste_previsao (
                    cod_produto, cod_fornecedor, nome_fornecedor,
                    periodo, granularidade,
                    valor_original, valor_ajustado, diferenca, percentual_ajuste,
                    motivo, usuario_ajuste, metodo_estatistico
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                cod_produto, cod_fornecedor, nome_fornecedor,
                periodo, granularidade,
                valor_original, valor_ajustado, diferenca, percentual_ajuste,
                motivo, usuario, metodo_estatistico
            ))

            ajustes_salvos += 1

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'mensagem': f'{ajustes_salvos} ajuste(s) salvo(s) com sucesso',
            'ajustes_salvos': ajustes_salvos
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@previsao_bp.route('/api/ajuste_previsao/carregar', methods=['GET'])
def api_carregar_ajustes_previsao():
    """
    Carrega ajustes ativos para um produto ou lista de produtos.
    Parametros: cod_produto, cod_fornecedor, granularidade
    """
    try:
        cod_produto = request.args.get('cod_produto')
        cod_fornecedor = request.args.get('cod_fornecedor')
        granularidade = request.args.get('granularidade')

        # Conectar ao banco
        from app.utils.db_connection import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()

        # Construir query dinamicamente
        query = """
            SELECT
                id, cod_produto, cod_fornecedor, nome_fornecedor,
                periodo, granularidade,
                valor_original, valor_ajustado, diferenca, percentual_ajuste,
                motivo, usuario_ajuste, data_ajuste, metodo_estatistico
            FROM ajuste_previsao
            WHERE ativo = TRUE
        """
        params = []

        if cod_produto:
            query += " AND cod_produto = %s"
            params.append(cod_produto)

        if cod_fornecedor:
            query += " AND cod_fornecedor = %s"
            params.append(cod_fornecedor)

        if granularidade:
            query += " AND granularidade = %s"
            params.append(granularidade)

        query += " ORDER BY cod_produto, periodo"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Formatar resultados
        ajustes = []
        for row in rows:
            ajustes.append({
                'id': row[0],
                'cod_produto': row[1],
                'cod_fornecedor': row[2],
                'nome_fornecedor': row[3],
                'periodo': row[4],
                'granularidade': row[5],
                'valor_original': float(row[6]) if row[6] else 0,
                'valor_ajustado': float(row[7]) if row[7] else 0,
                'diferenca': float(row[8]) if row[8] else 0,
                'percentual_ajuste': float(row[9]) if row[9] else 0,
                'motivo': row[10],
                'usuario_ajuste': row[11],
                'data_ajuste': row[12].isoformat() if row[12] else None,
                'metodo_estatistico': row[13]
            })

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'ajustes': ajustes,
            'total': len(ajustes)
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@previsao_bp.route('/api/ajuste_previsao/historico', methods=['GET'])
def api_historico_ajustes():
    """
    Retorna historico de ajustes para um produto.
    """
    try:
        cod_produto = request.args.get('cod_produto')

        if not cod_produto:
            return jsonify({'success': False, 'erro': 'cod_produto e obrigatorio'}), 400

        # Conectar ao banco
        from app.utils.db_connection import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                h.id, h.cod_produto, h.periodo, h.granularidade,
                h.valor_original, h.valor_ajustado, h.motivo,
                h.usuario_acao, h.tipo_acao, h.data_acao
            FROM ajuste_previsao_historico h
            WHERE h.cod_produto = %s
            ORDER BY h.data_acao DESC
            LIMIT 100
        """, (cod_produto,))

        rows = cursor.fetchall()

        historico = []
        for row in rows:
            historico.append({
                'id': row[0],
                'cod_produto': row[1],
                'periodo': row[2],
                'granularidade': row[3],
                'valor_original': float(row[4]) if row[4] else 0,
                'valor_ajustado': float(row[5]) if row[5] else 0,
                'motivo': row[6],
                'usuario_acao': row[7],
                'tipo_acao': row[8],
                'data_acao': row[9].isoformat() if row[9] else None
            })

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'historico': historico,
            'total': len(historico)
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@previsao_bp.route('/api/ajuste_previsao/excluir', methods=['POST'])
def api_excluir_ajuste():
    """
    Exclui (desativa) um ajuste especifico.
    """
    try:
        dados = request.get_json()
        ajuste_id = dados.get('id')
        cod_produto = dados.get('cod_produto')
        periodo = dados.get('periodo')

        if not ajuste_id and not (cod_produto and periodo):
            return jsonify({'success': False, 'erro': 'Informe id ou (cod_produto + periodo)'}), 400

        # Conectar ao banco
        from app.utils.db_connection import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()

        if ajuste_id:
            cursor.execute("""
                UPDATE ajuste_previsao
                SET ativo = FALSE, updated_at = NOW()
                WHERE id = %s
            """, (ajuste_id,))
        else:
            cursor.execute("""
                UPDATE ajuste_previsao
                SET ativo = FALSE, updated_at = NOW()
                WHERE cod_produto = %s AND periodo = %s AND ativo = TRUE
            """, (cod_produto, periodo))

        conn.commit()
        registros_afetados = cursor.rowcount

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'mensagem': f'{registros_afetados} ajuste(s) excluido(s)',
            'registros_afetados': registros_afetados
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


@previsao_bp.route('/api/historico_item', methods=['GET'])
def api_historico_item():
    """
    Retorna historico de vendas de um item especifico para exibicao no grafico drill-down.
    Retorna dados agregados por mes/semana/dia dos ultimos 2 anos.
    """
    try:
        cod_produto = request.args.get('cod_produto')
        cod_loja = request.args.get('cod_loja')
        granularidade = request.args.get('granularidade', 'mensal')

        if not cod_produto:
            return jsonify({'success': False, 'erro': 'cod_produto e obrigatorio'}), 400

        from app.utils.db_connection import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()

        # Buscar vendas dos ultimos 2 anos da tabela historico_vendas_diario
        # Nota: a tabela usa 'codigo' (INTEGER) e 'data' (DATE)
        query = """
            SELECT
                h.data as data_venda,
                SUM(h.qtd_venda) as qtd_venda
            FROM historico_vendas_diario h
            WHERE h.codigo::text = %s
              AND h.data >= CURRENT_DATE - INTERVAL '2 years'
        """
        params = [str(cod_produto)]

        if cod_loja:
            query += " AND h.cod_loja = %s"
            params.append(cod_loja)

        query += " GROUP BY h.data ORDER BY h.data"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        cursor.close()
        conn.close()

        # Agregar por granularidade
        dados_agregados = {}

        for row in rows:
            data_venda = row[0]
            qtd = float(row[1]) if row[1] else 0

            if granularidade == 'mensal':
                # Formato: YYYY-MM
                chave = data_venda.strftime('%Y-%m')
            elif granularidade == 'semanal':
                # Formato: YYYY-SWW
                ano_iso, semana_iso, _ = data_venda.isocalendar()
                chave = f"{ano_iso}-S{semana_iso:02d}"
            else:
                # Diario: YYYY-MM-DD
                chave = data_venda.strftime('%Y-%m-%d')

            if chave not in dados_agregados:
                dados_agregados[chave] = 0
            dados_agregados[chave] += qtd

        # Converter para lista ordenada
        historico = []
        for periodo in sorted(dados_agregados.keys()):
            historico.append({
                'periodo': periodo,
                'valor': round(dados_agregados[periodo], 2)
            })

        return jsonify({
            'success': True,
            'cod_produto': cod_produto,
            'granularidade': granularidade,
            'historico': historico,
            'total_periodos': len(historico)
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500


# =====================================================
# API DE EXPORTACAO DO RELATORIO DETALHADO DE ITENS
# =====================================================

@previsao_bp.route('/api/exportar_relatorio_itens', methods=['POST'])
def api_exportar_relatorio_itens():
    """
    Exporta o relatorio detalhado de itens para Excel.
    Recebe os dados do relatorio_detalhado e gera arquivo Excel.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        dados = request.get_json()

        if not dados:
            return jsonify({'success': False, 'erro': 'Dados nao fornecidos'}), 400

        itens = dados.get('itens', [])
        periodos = dados.get('periodos', [])
        granularidade = dados.get('granularidade', 'mensal')
        filtros = dados.get('filtros', {})

        if not itens:
            return jsonify({'success': False, 'erro': 'Nenhum item para exportar'}), 400

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Relatorio Detalhado'

        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='10B981', end_color='059669', fill_type='solid')
        fornecedor_fill = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')
        total_fill = PatternFill(start_color='047857', end_color='047857', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row_num = 1

        # Informacoes de filtros (se houver)
        if filtros:
            ws.cell(row=row_num, column=1, value='Filtros Aplicados:').font = Font(bold=True)
            row_num += 1
            for chave, valor in filtros.items():
                if valor:
                    ws.cell(row=row_num, column=1, value=f'  {chave.title()}: {valor}')
                    row_num += 1
            row_num += 1

        # Funcao auxiliar para formatar periodo
        meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

        def formatar_periodo_header(periodo):
            if not periodo:
                return '-'
            if granularidade == 'semanal':
                if isinstance(periodo, dict):
                    return f"S{periodo.get('semana', '?')}/{periodo.get('ano', '?')}"
                elif isinstance(periodo, str) and '-S' in periodo:
                    ano, sem = periodo.split('-S')
                    return f"S{int(sem)}/{ano}"
                return str(periodo)
            elif granularidade in ['diario', 'diaria']:
                if isinstance(periodo, str):
                    partes = periodo.split('-')
                    if len(partes) == 3:
                        return f"{partes[2]}/{partes[1]}"
                return str(periodo)
            else:
                # Mensal
                if isinstance(periodo, dict):
                    mes_idx = periodo.get('mes', 1) - 1
                    return f"{meses_nomes[mes_idx]}/{periodo.get('ano', '?')}"
                return str(periodo)

        # Cabecalho
        headers = ['Codigo', 'Descricao']
        for periodo in periodos:
            headers.append(formatar_periodo_header(periodo))
        headers.extend(['Total Prev.', 'Ano Ant.', 'Var. %', 'Alerta', 'Metodo', 'CUE'])

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row_num += 1

        # Agrupar itens por fornecedor
        itens_por_fornecedor = {}
        for item in itens:
            fornecedor = item.get('nome_fornecedor') or 'SEM FORNECEDOR'
            if fornecedor not in itens_por_fornecedor:
                itens_por_fornecedor[fornecedor] = []
            itens_por_fornecedor[fornecedor].append(item)

        # Ordenar fornecedores alfabeticamente
        fornecedores_ordenados = sorted(itens_por_fornecedor.keys())

        for fornecedor in fornecedores_ordenados:
            itens_fornecedor = itens_por_fornecedor[fornecedor]

            # Calcular totais do fornecedor
            total_prev_forn = 0
            total_aa_forn = 0
            totais_por_periodo = [0] * len(periodos)

            for item in itens_fornecedor:
                total_prev_forn += item.get('demanda_prevista_total', 0)
                total_aa_forn += item.get('demanda_ano_anterior', 0)

                if item.get('previsao_por_periodo'):
                    for idx, p in enumerate(item['previsao_por_periodo']):
                        if idx < len(periodos):
                            totais_por_periodo[idx] += p.get('previsao', 0)

            variacao_forn = ((total_prev_forn - total_aa_forn) / total_aa_forn * 100) if total_aa_forn > 0 else 0

            # Linha do fornecedor (agregada)
            ws.cell(row=row_num, column=1, value='').fill = fornecedor_fill
            ws.cell(row=row_num, column=1).border = border
            ws.cell(row=row_num, column=2, value=f'{fornecedor} ({len(itens_fornecedor)} itens)').font = Font(bold=True)
            ws.cell(row=row_num, column=2).fill = fornecedor_fill
            ws.cell(row=row_num, column=2).border = border

            for idx, total in enumerate(totais_por_periodo, start=3):
                cell = ws.cell(row=row_num, column=idx, value=round(total))
                cell.fill = fornecedor_fill
                cell.border = border
                cell.number_format = '#,##0'
                cell.font = Font(bold=True)

            col_offset = 3 + len(periodos)
            ws.cell(row=row_num, column=col_offset, value=round(total_prev_forn)).fill = fornecedor_fill
            ws.cell(row=row_num, column=col_offset).border = border
            ws.cell(row=row_num, column=col_offset).number_format = '#,##0'
            ws.cell(row=row_num, column=col_offset).font = Font(bold=True)

            ws.cell(row=row_num, column=col_offset + 1, value=round(total_aa_forn)).fill = fornecedor_fill
            ws.cell(row=row_num, column=col_offset + 1).border = border
            ws.cell(row=row_num, column=col_offset + 1).number_format = '#,##0'
            ws.cell(row=row_num, column=col_offset + 1).font = Font(bold=True)

            ws.cell(row=row_num, column=col_offset + 2, value=f'{variacao_forn:.1f}%').fill = fornecedor_fill
            ws.cell(row=row_num, column=col_offset + 2).border = border
            ws.cell(row=row_num, column=col_offset + 2).font = Font(bold=True)

            ws.cell(row=row_num, column=col_offset + 3, value='').fill = fornecedor_fill
            ws.cell(row=row_num, column=col_offset + 3).border = border

            ws.cell(row=row_num, column=col_offset + 4, value='').fill = fornecedor_fill
            ws.cell(row=row_num, column=col_offset + 4).border = border

            ws.cell(row=row_num, column=col_offset + 5, value='').fill = fornecedor_fill
            ws.cell(row=row_num, column=col_offset + 5).border = border

            row_num += 1

            # Linhas dos itens do fornecedor
            for item in itens_fornecedor:
                ws.cell(row=row_num, column=1, value=item.get('cod_produto', '')).border = border
                ws.cell(row=row_num, column=2, value=item.get('descricao', '')[:50]).border = border

                # Valores por periodo
                if item.get('previsao_por_periodo'):
                    for idx, p in enumerate(item['previsao_por_periodo'], start=3):
                        if idx - 3 < len(periodos):
                            cell = ws.cell(row=row_num, column=idx, value=round(p.get('previsao', 0)))
                            cell.border = border
                            cell.number_format = '#,##0'

                col_offset = 3 + len(periodos)
                ws.cell(row=row_num, column=col_offset, value=round(item.get('demanda_prevista_total', 0))).border = border
                ws.cell(row=row_num, column=col_offset).number_format = '#,##0'

                ws.cell(row=row_num, column=col_offset + 1, value=round(item.get('demanda_ano_anterior', 0))).border = border
                ws.cell(row=row_num, column=col_offset + 1).number_format = '#,##0'

                variacao_item = item.get('variacao_percentual', 0)
                ws.cell(row=row_num, column=col_offset + 2, value=f'{variacao_item:.1f}%').border = border

                # Emoji de alerta (converter para texto) - V50: baseado em CV
                sinal = item.get('sinal_emoji', '')
                alerta_texto = {
                    '🔴': 'RISCO CRITICO',
                    '🟡': 'ATENCAO',
                    '🔵': 'BAIXO RISCO',
                    '🟢': 'CONFIAVEL',
                    '⚪': 'SEM DADOS'
                }.get(sinal, '-')
                ws.cell(row=row_num, column=col_offset + 3, value=alerta_texto).border = border

                ws.cell(row=row_num, column=col_offset + 4, value=item.get('metodo_estatistico', '')).border = border

                cue_valor = item.get('cue', 0)
                ws.cell(row=row_num, column=col_offset + 5, value=cue_valor).border = border
                ws.cell(row=row_num, column=col_offset + 5).number_format = '#,##0.0000'

                row_num += 1

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        for col_idx in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        ws.freeze_panes = 'C2'

        # =====================================================
        # ABA 2: RELATORIO EM VALOR DE CUSTO (CUE)
        # =====================================================
        ws_custo = wb.create_sheet(title='Custo (CUE)')

        custo_header_fill = PatternFill(start_color='3B82F6', end_color='1D4ED8', fill_type='solid')
        custo_fornecedor_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        custo_total_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')

        row_custo = 1

        # Filtros (se houver)
        if filtros:
            ws_custo.cell(row=row_custo, column=1, value='Filtros Aplicados:').font = Font(bold=True)
            row_custo += 1
            for chave, valor in filtros.items():
                if valor:
                    ws_custo.cell(row=row_custo, column=1, value=f'  {chave.title()}: {valor}')
                    row_custo += 1
            row_custo += 1

        # Cabecalho
        headers_custo = ['Codigo', 'Descricao']
        for periodo in periodos:
            headers_custo.append(formatar_periodo_header(periodo))
        headers_custo.extend(['Total Prev. R$', 'CUE Unit.'])

        for col, header in enumerate(headers_custo, start=1):
            cell = ws_custo.cell(row=row_custo, column=col, value=header)
            cell.font = header_font
            cell.fill = custo_header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row_custo += 1

        # Agrupar e ordenar por fornecedor (mesmo da aba 1)
        for fornecedor in fornecedores_ordenados:
            itens_fornecedor = itens_por_fornecedor[fornecedor]

            # Totais de custo do fornecedor
            total_custo_forn = 0
            totais_custo_periodo = [0] * len(periodos)

            for item in itens_fornecedor:
                cue = item.get('cue', 0)
                total_custo_forn += item.get('demanda_prevista_total', 0) * cue

                if item.get('previsao_por_periodo'):
                    for idx, p in enumerate(item['previsao_por_periodo']):
                        if idx < len(periodos):
                            totais_custo_periodo[idx] += p.get('previsao', 0) * cue

            # Linha do fornecedor
            ws_custo.cell(row=row_custo, column=1, value='').fill = custo_fornecedor_fill
            ws_custo.cell(row=row_custo, column=1).border = border
            ws_custo.cell(row=row_custo, column=2, value=f'{fornecedor} ({len(itens_fornecedor)} itens)').font = Font(bold=True)
            ws_custo.cell(row=row_custo, column=2).fill = custo_fornecedor_fill
            ws_custo.cell(row=row_custo, column=2).border = border

            for idx, total in enumerate(totais_custo_periodo, start=3):
                cell = ws_custo.cell(row=row_custo, column=idx, value=round(total, 2))
                cell.fill = custo_fornecedor_fill
                cell.border = border
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True)

            col_custo_offset = 3 + len(periodos)
            ws_custo.cell(row=row_custo, column=col_custo_offset, value=round(total_custo_forn, 2)).fill = custo_fornecedor_fill
            ws_custo.cell(row=row_custo, column=col_custo_offset).border = border
            ws_custo.cell(row=row_custo, column=col_custo_offset).number_format = '#,##0.00'
            ws_custo.cell(row=row_custo, column=col_custo_offset).font = Font(bold=True)

            ws_custo.cell(row=row_custo, column=col_custo_offset + 1, value='').fill = custo_fornecedor_fill
            ws_custo.cell(row=row_custo, column=col_custo_offset + 1).border = border

            row_custo += 1

            # Linhas dos itens
            for item in itens_fornecedor:
                cue = item.get('cue', 0)

                ws_custo.cell(row=row_custo, column=1, value=item.get('cod_produto', '')).border = border
                ws_custo.cell(row=row_custo, column=2, value=item.get('descricao', '')[:50]).border = border

                if item.get('previsao_por_periodo'):
                    for idx, p in enumerate(item['previsao_por_periodo'], start=3):
                        if idx - 3 < len(periodos):
                            valor_custo = round(p.get('previsao', 0) * cue, 2)
                            cell = ws_custo.cell(row=row_custo, column=idx, value=valor_custo)
                            cell.border = border
                            cell.number_format = '#,##0.00'

                col_custo_offset = 3 + len(periodos)
                total_item_custo = round(item.get('demanda_prevista_total', 0) * cue, 2)
                ws_custo.cell(row=row_custo, column=col_custo_offset, value=total_item_custo).border = border
                ws_custo.cell(row=row_custo, column=col_custo_offset).number_format = '#,##0.00'

                ws_custo.cell(row=row_custo, column=col_custo_offset + 1, value=round(cue, 4)).border = border
                ws_custo.cell(row=row_custo, column=col_custo_offset + 1).number_format = '#,##0.0000'

                row_custo += 1

        # Ajustar largura
        ws_custo.column_dimensions['A'].width = 12
        ws_custo.column_dimensions['B'].width = 35
        for col_idx in range(3, len(headers_custo) + 1):
            ws_custo.column_dimensions[get_column_letter(col_idx)].width = 15

        ws_custo.freeze_panes = 'C2'

        # Salvar em memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'relatorio_detalhado_itens_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'erro': str(e)}), 500
