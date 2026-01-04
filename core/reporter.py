"""
Módulo de Geração de Relatórios
Gera arquivos Excel com múltiplas abas contendo todas as informações de previsão
"""

import pandas as pd
import numpy as np
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os


class ExcelReporter:
    """
    Gera relatórios Excel formatados com múltiplas abas
    """

    # Estilos
    HEADER_FILL = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALTERNATE_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self):
        self.workbook = None
        self.resultados = {}

    def gerar_relatorio(self, resultados: Dict, caminho_saida: str) -> str:
        """
        Gera relatório Excel completo

        Args:
            resultados: Dicionário com todos os resultados do processamento
            caminho_saida: Caminho para salvar o arquivo

        Returns:
            Caminho do arquivo gerado
        """
        self.resultados = resultados
        self.workbook = Workbook()

        # Remover planilha padrão
        self.workbook.remove(self.workbook.active)

        # Criar abas
        self._criar_aba_previsoes_unificada()
        self._criar_aba_metodos()
        self._criar_aba_rupturas()
        self._criar_aba_caracteristicas()
        self._criar_aba_resumo()

        # Salvar arquivo
        self.workbook.save(caminho_saida)

        return caminho_saida

    def _criar_aba_previsoes_unificada(self):
        """
        Cria aba PREVISOES com previsões de lojas e CD unificadas
        """
        dfs = []

        # Adicionar previsões das lojas
        if 'previsoes_lojas' in self.resultados and not self.resultados['previsoes_lojas'].empty:
            df_lojas = self.resultados['previsoes_lojas'].copy()
            # Padronizar colunas
            df_lojas = df_lojas.rename(columns={'Mes_Previsao': 'Periodo', 'Previsao': 'Previsao'})
            if 'Loja' not in df_lojas.columns:
                df_lojas['Loja'] = ''
            dfs.append(df_lojas)

        # Adicionar previsões do CD
        if 'previsoes_cd' in self.resultados and not self.resultados['previsoes_cd'].empty:
            df_cd = self.resultados['previsoes_cd'].copy()
            # Padronizar colunas para combinar com lojas
            df_cd['Loja'] = 'CD'
            df_cd = df_cd.rename(columns={'Mes_Previsao': 'Periodo', 'Previsao_CD': 'Previsao'})
            # Selecionar colunas compatíveis
            colunas = ['Loja', 'SKU', 'Periodo', 'Previsao', 'Metodo', 'Confianca']
            df_cd = df_cd[[c for c in colunas if c in df_cd.columns]]
            dfs.append(df_cd)

        if not dfs:
            return

        # Unir DataFrames
        df_unificado = pd.concat(dfs, ignore_index=True)

        # Ordenar: primeiro por SKU, depois Loja (CD por último), depois Periodo
        df_unificado['Ordem_Loja'] = df_unificado['Loja'].apply(lambda x: 'ZZZ' if x == 'CD' else x)
        df_unificado = df_unificado.sort_values(['SKU', 'Ordem_Loja', 'Periodo'])
        df_unificado = df_unificado.drop(columns=['Ordem_Loja'])

        ws = self.workbook.create_sheet("PREVISOES")

        # Escrever dados
        self._escrever_dataframe(ws, df_unificado)

    def _criar_aba_metodos(self):
        """
        Cria aba METODOS_UTILIZADOS com detalhes dos métodos selecionados
        """
        if 'metodos_utilizados' not in self.resultados:
            return

        df = self.resultados['metodos_utilizados']

        if df.empty:
            return

        ws = self.workbook.create_sheet("METODOS_UTILIZADOS")

        # Ordenar dados
        df = df.sort_values(['Local', 'SKU'])

        # Escrever dados
        self._escrever_dataframe(ws, df)

    def _criar_aba_rupturas(self):
        """
        Cria aba ANALISE_RUPTURAS com métricas de stockout
        """
        if 'analise_rupturas' not in self.resultados:
            return

        df = self.resultados['analise_rupturas']

        if df.empty:
            return

        ws = self.workbook.create_sheet("ANALISE_RUPTURAS")

        # Ordenar por taxa de ruptura (maior primeiro)
        if 'taxa_ruptura' in df.columns:
            df = df.sort_values('taxa_ruptura', ascending=False)

        # Escrever dados
        self._escrever_dataframe(ws, df)

    def _criar_aba_caracteristicas(self):
        """
        Cria aba CARACTERISTICAS_DEMANDA com análise das séries
        """
        if 'caracteristicas' not in self.resultados:
            return

        df = self.resultados['caracteristicas']

        if df.empty:
            return

        ws = self.workbook.create_sheet("CARACTERISTICAS_DEMANDA")

        # Escrever dados
        self._escrever_dataframe(ws, df)

    def _criar_aba_resumo(self):
        """
        Cria aba RESUMO_EXECUTIVO com estatísticas gerais
        """
        ws = self.workbook.create_sheet("RESUMO_EXECUTIVO", 0)  # Primeira aba

        resumo = self.resultados.get('resumo', {})

        # Título
        ws['A1'] = "RESUMO EXECUTIVO - PREVISÃO DE DEMANDA"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # Data de geração
        ws['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(italic=True)

        # Métricas
        row = 5
        metricas = [
            ("Total de SKUs processados:", resumo.get('total_skus', 0)),
            ("Total de lojas:", resumo.get('total_lojas', 9)),
            ("Combinações Loja+SKU:", resumo.get('total_combinacoes', 0)),
            ("Período histórico:", resumo.get('periodo_historico', 'N/A')),
            ("Meses de previsão:", resumo.get('meses_previsao', 0)),
            ("", ""),
            ("ANÁLISE DE RUPTURAS", ""),
            ("Meses com ruptura (total):", resumo.get('meses_com_ruptura', 0)),
            ("Taxa média de ruptura:", f"{resumo.get('taxa_ruptura_media', 0)*100:.1f}%"),
            ("Vendas perdidas estimadas:", f"{resumo.get('vendas_perdidas', 0):,.0f}"),
            ("", ""),
            ("MÉTODOS UTILIZADOS", ""),
        ]

        for label, valor in metricas:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if label and not label.endswith(':') else Font()
            ws[f'B{row}'] = valor
            row += 1

        # Contagem de métodos
        if 'contagem_metodos' in resumo:
            for metodo, count in resumo['contagem_metodos'].items():
                ws[f'A{row}'] = f"  {metodo}:"
                ws[f'B{row}'] = count
                row += 1

        # Alertas
        row += 1
        ws[f'A{row}'] = "ALERTAS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        alertas = self.resultados.get('alertas', [])
        if alertas:
            for alerta in alertas:
                ws[f'A{row}'] = f"• {alerta.get('mensagem', alerta)}"
                row += 1
        else:
            ws[f'A{row}'] = "Nenhum alerta gerado."

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30

    def _escrever_dataframe(self, ws, df: pd.DataFrame):
        """
        Escreve um DataFrame na planilha com formatação

        Args:
            ws: Worksheet
            df: DataFrame a ser escrito
        """
        # Converter datetime para string para evitar problemas
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m')

        # Escrever dados
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Formatação do cabeçalho
                if r_idx == 1:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Linhas alternadas
                    if r_idx % 2 == 0:
                        cell.fill = self.ALTERNATE_FILL

                    # Alinhar números à direita
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')

                cell.border = self.BORDER

        # Auto-ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def gerar_nome_arquivo():
    """
    Gera nome único para o arquivo de saída

    Returns:
        Nome do arquivo com timestamp
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"previsao_demanda_{timestamp}.xlsx"


def preparar_resultados(
    df_previsoes_lojas: pd.DataFrame,
    df_previsoes_cd: pd.DataFrame,
    df_metodos: pd.DataFrame,
    df_rupturas: pd.DataFrame,
    df_caracteristicas: pd.DataFrame,
    resumo: Dict,
    alertas: List
) -> Dict:
    """
    Prepara dicionário de resultados para o reporter

    Args:
        df_previsoes_lojas: Previsões por loja
        df_previsoes_cd: Previsões do CD
        df_metodos: Métodos utilizados
        df_rupturas: Análise de rupturas
        df_caracteristicas: Características das séries
        resumo: Dicionário com resumo
        alertas: Lista de alertas

    Returns:
        Dicionário formatado para o reporter
    """
    return {
        'previsoes_lojas': df_previsoes_lojas,
        'previsoes_cd': df_previsoes_cd,
        'metodos_utilizados': df_metodos,
        'analise_rupturas': df_rupturas,
        'caracteristicas': df_caracteristicas,
        'resumo': resumo,
        'alertas': alertas
    }
