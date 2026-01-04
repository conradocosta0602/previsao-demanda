"""
Módulo de Simulação de Cenários

Permite ajustar previsões para simular diferentes cenários de negócio:
- Cenários globais (aumento/redução geral)
- Cenários por segmento (loja, SKU, fornecedor)
- Cenários pré-definidos (otimista, pessimista, sazonal)
- Comparação com baseline
"""

from typing import List, Dict, Optional
import pandas as pd
import numpy as np
from datetime import datetime
import copy


class ScenarioSimulator:
    """
    Simulador de cenários de demanda

    Permite ajustar previsões base para simular diferentes cenários
    e calcular o impacto nas métricas de negócio.
    """

    def __init__(self, previsoes_base: pd.DataFrame):
        """
        Inicializa o simulador com as previsões base

        Args:
            previsoes_base: DataFrame com colunas [Loja, SKU, Mes_Previsao, Previsao, ...]
        """
        self.previsoes_base = previsoes_base.copy()
        self.previsoes_simuladas = previsoes_base.copy()
        self.ajustes_aplicados = []
        self.cenario_nome = "Baseline (Original)"

    def aplicar_ajuste_global(self, percentual: float, descricao: str = None):
        """
        Aplica ajuste percentual em todas as previsões

        Args:
            percentual: Percentual de ajuste (ex: 20 para +20%, -10 para -10%)
            descricao: Descrição do ajuste
        """
        fator = 1 + (percentual / 100)

        self.previsoes_simuladas['Previsao'] = self.previsoes_base['Previsao'] * fator

        ajuste = {
            'tipo': 'global',
            'percentual': percentual,
            'fator': fator,
            'descricao': descricao or f"Ajuste global de {percentual:+.1f}%",
            'timestamp': datetime.now().isoformat()
        }

        self.ajustes_aplicados.append(ajuste)
        self._atualizar_nome_cenario()

    def aplicar_ajuste_por_loja(self, loja: str, percentual: float, descricao: str = None):
        """
        Aplica ajuste percentual em uma loja específica

        Args:
            loja: Código da loja
            percentual: Percentual de ajuste
            descricao: Descrição do ajuste
        """
        fator = 1 + (percentual / 100)

        # Filtrar por loja e aplicar ajuste
        mask = self.previsoes_simuladas['Loja'] == loja
        self.previsoes_simuladas.loc[mask, 'Previsao'] = (
            self.previsoes_base.loc[mask, 'Previsao'] * fator
        )

        ajuste = {
            'tipo': 'por_loja',
            'loja': loja,
            'percentual': percentual,
            'fator': fator,
            'descricao': descricao or f"Loja {loja}: {percentual:+.1f}%",
            'timestamp': datetime.now().isoformat()
        }

        self.ajustes_aplicados.append(ajuste)
        self._atualizar_nome_cenario()

    def aplicar_ajuste_por_sku(self, sku: str, percentual: float, descricao: str = None):
        """
        Aplica ajuste percentual em um SKU específico

        Args:
            sku: Código do SKU
            percentual: Percentual de ajuste
            descricao: Descrição do ajuste
        """
        fator = 1 + (percentual / 100)

        # Filtrar por SKU e aplicar ajuste
        mask = self.previsoes_simuladas['SKU'] == sku
        self.previsoes_simuladas.loc[mask, 'Previsao'] = (
            self.previsoes_base.loc[mask, 'Previsao'] * fator
        )

        ajuste = {
            'tipo': 'por_sku',
            'sku': sku,
            'percentual': percentual,
            'fator': fator,
            'descricao': descricao or f"SKU {sku}: {percentual:+.1f}%",
            'timestamp': datetime.now().isoformat()
        }

        self.ajustes_aplicados.append(ajuste)
        self._atualizar_nome_cenario()

    def aplicar_ajuste_por_mes(self, mes_numero: int, percentual: float, descricao: str = None):
        """
        Aplica ajuste percentual em um mês específico (sazonalidade)

        Args:
            mes_numero: Número do mês (1-12)
            percentual: Percentual de ajuste
            descricao: Descrição do ajuste
        """
        fator = 1 + (percentual / 100)

        # Filtrar por mês e aplicar ajuste
        self.previsoes_simuladas['Mes_Num'] = pd.to_datetime(
            self.previsoes_simuladas['Mes_Previsao']
        ).dt.month

        mask = self.previsoes_simuladas['Mes_Num'] == mes_numero
        self.previsoes_simuladas.loc[mask, 'Previsao'] = (
            self.previsoes_base.loc[mask, 'Previsao'] * fator
        )

        # Remover coluna auxiliar
        self.previsoes_simuladas.drop('Mes_Num', axis=1, inplace=True)

        meses = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

        ajuste = {
            'tipo': 'por_mes',
            'mes': mes_numero,
            'mes_nome': meses[mes_numero],
            'percentual': percentual,
            'fator': fator,
            'descricao': descricao or f"{meses[mes_numero]}: {percentual:+.1f}%",
            'timestamp': datetime.now().isoformat()
        }

        self.ajustes_aplicados.append(ajuste)
        self._atualizar_nome_cenario()

    def aplicar_cenario_predefinido(self, tipo: str):
        """
        Aplica um cenário pré-definido

        Args:
            tipo: Tipo de cenário ('otimista', 'pessimista', 'conservador', 'black_friday')
        """
        self.resetar()

        if tipo == 'otimista':
            self.aplicar_ajuste_global(20, "Cenário Otimista (+20%)")
            self.cenario_nome = "📈 Otimista (+20%)"

        elif tipo == 'pessimista':
            self.aplicar_ajuste_global(-20, "Cenário Pessimista (-20%)")
            self.cenario_nome = "📉 Pessimista (-20%)"

        elif tipo == 'conservador':
            self.aplicar_ajuste_global(-10, "Cenário Conservador (-10%)")
            self.cenario_nome = "📊 Conservador (-10%)"

        elif tipo == 'black_friday':
            # Novembro: +50%, Dezembro: +30%
            self.aplicar_ajuste_por_mes(11, 50, "Black Friday (Nov +50%)")
            self.aplicar_ajuste_por_mes(12, 30, "Natal (Dez +30%)")
            self.cenario_nome = "🎄 Black Friday + Natal"

        elif tipo == 'verao':
            # Jan/Fev: +25%
            self.aplicar_ajuste_por_mes(1, 25, "Verão (Jan +25%)")
            self.aplicar_ajuste_por_mes(2, 25, "Verão (Fev +25%)")
            self.cenario_nome = "☀️ Pico de Verão"

        else:
            raise ValueError(f"Cenário '{tipo}' não reconhecido")

    def resetar(self):
        """Reseta para as previsões base (remove todos os ajustes)"""
        self.previsoes_simuladas = self.previsoes_base.copy()
        self.ajustes_aplicados = []
        self.cenario_nome = "Baseline (Original)"

    def calcular_impacto(self) -> Dict:
        """
        Calcula o impacto dos ajustes comparando com baseline

        Returns:
            Dicionário com métricas de impacto
        """
        # Total base vs simulado
        total_base = self.previsoes_base['Previsao'].sum()
        total_simulado = self.previsoes_simuladas['Previsao'].sum()

        diferenca_absoluta = total_simulado - total_base
        diferenca_percentual = ((total_simulado / total_base) - 1) * 100 if total_base > 0 else 0

        # Por loja
        impacto_lojas = []
        for loja in self.previsoes_base['Loja'].unique():
            base_loja = self.previsoes_base[self.previsoes_base['Loja'] == loja]['Previsao'].sum()
            sim_loja = self.previsoes_simuladas[self.previsoes_simuladas['Loja'] == loja]['Previsao'].sum()

            impacto_lojas.append({
                'loja': loja,
                'base': float(base_loja),
                'simulado': float(sim_loja),
                'diferenca': float(sim_loja - base_loja),
                'percentual': float(((sim_loja / base_loja) - 1) * 100) if base_loja > 0 else 0
            })

        # Por SKU
        impacto_skus = []
        for sku in self.previsoes_base['SKU'].unique():
            base_sku = self.previsoes_base[self.previsoes_base['SKU'] == sku]['Previsao'].sum()
            sim_sku = self.previsoes_simuladas[self.previsoes_simuladas['SKU'] == sku]['Previsao'].sum()

            impacto_skus.append({
                'sku': sku,
                'base': float(base_sku),
                'simulado': float(sim_sku),
                'diferenca': float(sim_sku - base_sku),
                'percentual': float(((sim_sku / base_sku) - 1) * 100) if base_sku > 0 else 0
            })

        # Por mês
        impacto_meses = []
        for mes in self.previsoes_base['Mes_Previsao'].unique():
            base_mes = self.previsoes_base[self.previsoes_base['Mes_Previsao'] == mes]['Previsao'].sum()
            sim_mes = self.previsoes_simuladas[self.previsoes_simuladas['Mes_Previsao'] == mes]['Previsao'].sum()

            impacto_meses.append({
                'mes': mes.strftime('%Y-%m') if hasattr(mes, 'strftime') else str(mes),
                'base': float(base_mes),
                'simulado': float(sim_mes),
                'diferenca': float(sim_mes - base_mes),
                'percentual': float(((sim_mes / base_mes) - 1) * 100) if base_mes > 0 else 0
            })

        return {
            'total': {
                'base': float(total_base),
                'simulado': float(total_simulado),
                'diferenca_absoluta': float(diferenca_absoluta),
                'diferenca_percentual': float(diferenca_percentual)
            },
            'por_loja': impacto_lojas,
            'por_sku': impacto_skus,
            'por_mes': sorted(impacto_meses, key=lambda x: x['mes'])
        }

    def get_previsoes_simuladas(self) -> pd.DataFrame:
        """Retorna DataFrame com previsões simuladas"""
        df = self.previsoes_simuladas.copy()
        df['Cenario'] = self.cenario_nome
        return df

    def get_comparacao(self) -> pd.DataFrame:
        """
        Retorna DataFrame comparativo (base vs simulado)

        Returns:
            DataFrame com colunas adicionais de comparação
        """
        df_comp = self.previsoes_base.copy()
        df_comp = df_comp.rename(columns={'Previsao': 'Previsao_Base'})

        df_comp['Previsao_Simulada'] = self.previsoes_simuladas['Previsao'].values
        df_comp['Diferenca_Absoluta'] = df_comp['Previsao_Simulada'] - df_comp['Previsao_Base']
        df_comp['Diferenca_Percentual'] = (
            ((df_comp['Previsao_Simulada'] / df_comp['Previsao_Base']) - 1) * 100
        ).fillna(0)
        df_comp['Cenario'] = self.cenario_nome

        return df_comp

    def _atualizar_nome_cenario(self):
        """Atualiza nome do cenário baseado nos ajustes"""
        if len(self.ajustes_aplicados) == 0:
            self.cenario_nome = "Baseline (Original)"
        elif len(self.ajustes_aplicados) == 1:
            self.cenario_nome = f"🎯 {self.ajustes_aplicados[0]['descricao']}"
        else:
            self.cenario_nome = f"🎯 Cenário Personalizado ({len(self.ajustes_aplicados)} ajustes)"

    def exportar_para_excel(self, filepath: str):
        """
        Exporta comparação para Excel

        Args:
            filepath: Caminho do arquivo Excel
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Aba 1: Comparação
        ws_comp = wb.active
        ws_comp.title = "Comparação"

        df_comp = self.get_comparacao()

        # Cabeçalhos
        headers = list(df_comp.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws_comp.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Dados
        for row_num, row_data in enumerate(df_comp.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws_comp.cell(row=row_num, column=col_num, value=value)

        # Aba 2: Resumo de Impacto
        ws_impacto = wb.create_sheet("Impacto")
        impacto = self.calcular_impacto()

        ws_impacto.cell(1, 1, "RESUMO DE IMPACTO").font = Font(size=14, bold=True)
        ws_impacto.cell(3, 1, "Métrica")
        ws_impacto.cell(3, 2, "Base")
        ws_impacto.cell(3, 3, "Simulado")
        ws_impacto.cell(3, 4, "Diferença")
        ws_impacto.cell(3, 5, "% Variação")

        ws_impacto.cell(4, 1, "TOTAL GERAL")
        ws_impacto.cell(4, 2, impacto['total']['base'])
        ws_impacto.cell(4, 3, impacto['total']['simulado'])
        ws_impacto.cell(4, 4, impacto['total']['diferenca_absoluta'])
        ws_impacto.cell(4, 5, f"{impacto['total']['diferenca_percentual']:.2f}%")

        wb.save(filepath)

        return filepath


def criar_simulador_de_dados(previsoes_lojas: List[Dict]) -> ScenarioSimulator:
    """
    Cria um simulador a partir de dados de previsão

    Args:
        previsoes_lojas: Lista de dicionários com previsões

    Returns:
        ScenarioSimulator inicializado
    """
    df = pd.DataFrame(previsoes_lojas)
    return ScenarioSimulator(df)
